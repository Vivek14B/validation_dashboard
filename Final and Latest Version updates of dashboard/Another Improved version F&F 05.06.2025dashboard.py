import streamlit as st
import pandas as pd
import sqlite3
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np 
import logging
import os
import json 

# Helper function to serialize objects not recognized by default json.dumps
def json_serializer_default(obj):
    if isinstance(obj, datetime): 
        if hasattr(obj, 'tzinfo') and obj.tzinfo is not None and obj.tzinfo.utcoffset(obj) is not None:
            if isinstance(obj, pd.Timestamp) and pd.isna(obj):
                 return None
            return obj.astimezone(pd.timezone('UTC')).isoformat()
        else:
            if isinstance(obj, pd.Timestamp) and pd.isna(obj):
                 return None
            return obj.isoformat()
    elif isinstance(obj, np.datetime64):
        if np.isnat(obj):
            return None
        return pd.Timestamp(obj).isoformat()
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        if np.isnan(obj) or np.isinf(obj):
            return str(obj) 
        return float(obj)
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, pd.Series):
        return obj.to_dict() 
    elif isinstance(obj, (pd.Interval, pd.Period, pd.Categorical)):
        return str(obj)
    if hasattr(obj, 'isoformat') and callable(getattr(obj, 'isoformat')):
        try:
            return obj.isoformat()
        except Exception:
            pass 
    try:
        return str(obj)
    except Exception:
        return f"Unserializable object: {type(obj).__name__}"

# Page configuration
st.set_page_config(
    page_title="Data Validation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    body { /* Apply Inter to the whole body for wider reach */
        font-family: 'Inter', sans-serif;
    }
    .main {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    .main > div { /* This targets the main content blocks */
        background: white;
        border-radius: 20px;
        margin: 1rem;
        padding: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    h1 {
        color: #2d3748;
        font-weight: 700;
        text-align: center;
        margin-bottom: 1rem; /* Adjusted for underline */
        font-size: 2.5rem;
        position: relative; 
        padding-bottom: 0.5rem; 
    }
    h1::after { /* Graphical underline for main title */
        content: '';
        display: block;
        width: 100px; 
        height: 4px; 
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        margin: 0.5rem auto 1rem auto; 
        border-radius: 2px;
    }
    h2, h3 {
        color: #4a5568;
        font-weight: 600;
    }
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin: 0.5rem;
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.3);
        transition: transform 0.3s ease;
    }
    .metric-container:hover {
        transform: translateY(-5px);
    }
    .metric-title {
        font-size: 0.9rem;
        font-weight: 500;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    .upload-section {
        border: 3px dashed #667eea;
        border-radius: 20px;
        padding: 3rem;
        text-align: center;
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        margin: 2rem 0;
        transition: all 0.3s ease;
    }
    .upload-section:hover {
        border-color: #764ba2;
        background: linear-gradient(135deg, #edf2f7 0%, #e2e8f0 100%);
    }
    .upload-title {
        color: #2d3748;
        font-size: 1.5rem;
        font-weight: 600;
        margin-bottom: 1rem;
    }
    .upload-subtitle {
        color: #718096;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }
    .success-box {
        background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(72, 187, 120, 0.3);
    }
    .error-box {
        background: linear-gradient(135deg, #f56565 0%, #e53e3e 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(245, 101, 101, 0.3);
    }
    .warning-box {
        background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(237, 137, 54, 0.3);
    }
    .info-box {
        background: linear-gradient(135deg, #4299e1 0%, #3182ce 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(66, 153, 225, 0.3);
    }
    .css-1d391kg { 
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    .css-1d391kg .css-1v0mbm { 
        color: white;
    }
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 3px 10px rgba(102, 126, 234, 0.3);
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
    }
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        border-radius: 10px;
        font-weight: 600;
    }
    .stProgress .st-bo { 
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

DB_TIMEOUT = 30

# Font dictionary for Plotly charts
PLOTLY_FONT = dict(family="Inter, sans-serif", size=12, color="#2d3748")
PLOTLY_TITLE_FONT = dict(family="Inter, sans-serif", size=16, color="#2d3748")


class DatabaseManager:
    def __init__(self, db_path="validation_dashboard.db"):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS validation_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT NOT NULL,
                    upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP, total_records INTEGER,
                    total_exceptions INTEGER, status TEXT DEFAULT 'completed', file_size INTEGER)''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS exceptions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, department TEXT,
                    sub_department TEXT, created_user TEXT, modified_user TEXT,
                    exception_reason TEXT, severity INTEGER, net_amount REAL, location TEXT,
                    crop TEXT, activity TEXT, function_name TEXT, vertical_name TEXT,
                    region_name TEXT, zone_name TEXT, business_unit TEXT,
                    original_row_data TEXT, 
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            
            cursor.execute("PRAGMA table_info(exceptions)")
            columns = [info[1] for info in cursor.fetchall()]
            if 'function_name' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN function_name TEXT')
                logging.info("Added function_name column to exceptions table.")
            if 'original_row_data' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN original_row_data TEXT')
                logging.info("Added original_row_data column to exceptions table.")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS department_summary (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, department TEXT,
                    total_records INTEGER, exception_records INTEGER, exception_rate REAL,
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS user_performance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, user TEXT,
                    total_records INTEGER, exception_records INTEGER, exception_rate REAL,
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Database initialization error: {e}", exc_info=True)
            if conn: conn.rollback()
        finally:
            if conn: conn.close()

    def save_validation_run(self, filename, total_records, total_exceptions, file_size):
        conn = None; run_id = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO validation_runs (filename, total_records, total_exceptions, file_size)
                              VALUES (?, ?, ?, ?)''', (filename, total_records, total_exceptions, file_size))
            run_id = cursor.lastrowid
            conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Error in save_validation_run: {e}", exc_info=True)
            if conn: conn.rollback()
            raise
        finally:
            if conn: conn.close()
        return run_id

    def save_exceptions(self, run_id, exceptions_df):
        conn = None
        if exceptions_df.empty: return
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            
            data_to_insert = []
            for _, row_series in exceptions_df.iterrows():
                original_row_data_dict = row_series.to_dict()
                try:
                    serialized_row_data = json.dumps(original_row_data_dict, default=json_serializer_default)
                except Exception as json_e: 
                    logging.error(f"Could not serialize row for run_id {run_id}. Row data: {original_row_data_dict}. Error: {json_e}", exc_info=True)
                    try:
                        serialized_row_data = json.dumps({"error": "Failed to serialize original row", "details": str(original_row_data_dict)})
                    except: 
                        serialized_row_data = '{"error": "Critical serialization failure for original row"}'

                data_to_insert.append(
                    (run_id,
                     row_series.get('Department.Name', ''), 
                     row_series.get('Sub Department.Name', ''),
                     row_series.get('Created user', ''), 
                     row_series.get('Modified user', ''),
                     str(row_series.get('Exception Reasons', '')),
                     row_series.get('Severity', 0),
                     row_series.get('Net amount', 0.0), 
                     row_series.get('Location.Name', ''),
                     row_series.get('Crop.Name', ''), 
                     row_series.get('Activity.Name', ''),
                     row_series.get('Function.Name', ''), 
                     row_series.get('FC-Vertical.Name', ''),
                     row_series.get('Region.Name', ''), 
                     row_series.get('Zone.Name', ''),
                     row_series.get('Business Unit.Name', ''),
                     serialized_row_data 
                    )
                )
            
            if data_to_insert:
                cursor.executemany('''INSERT INTO exceptions (
                    run_id, department, sub_department, created_user, modified_user, exception_reason,
                    severity, net_amount, location, crop, activity, function_name, vertical_name,
                    region_name, zone_name, business_unit, original_row_data) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', data_to_insert)
                conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Error in save_exceptions (SQLite): {e}", exc_info=True)
            if conn: conn.rollback()
            st.error(f"Database error while saving exceptions: {e}")
            raise 
        except Exception as ex: 
            logging.error(f"Unexpected error in save_exceptions: {ex}", exc_info=True)
            st.error(f"An unexpected error occurred while saving exceptions: {ex}")
            if conn and getattr(conn, 'in_transaction', False): conn.rollback() 
        finally:
            if conn: conn.close()

    def save_user_performance(self, run_id, df, exceptions_df):
        conn = None
        try:
            if 'Created user' not in df.columns:
                st.markdown('<div class="error-box"><strong>❌ Error:</strong> Created user column missing in input for user performance.</div>', unsafe_allow_html=True)
                logging.error("Created user column missing in input df for save_user_performance.")
                return
            
            df['Created user'] = df['Created user'].dropna().astype(str).str.strip()
            valid_users = df['Created user'][df['Created user'] != ''].unique()

            if len(valid_users) == 0:
                st.markdown('<div class="warning-box"><strong>⚠ Warning:</strong> No valid Created user values for user performance.</div>', unsafe_allow_html=True)
                logging.warning("No valid Created user values in input df for save_user_performance.")
                return
            
            if exceptions_df.empty or 'Created user' not in exceptions_df.columns:
                user_stats = pd.DataFrame({'Created user': valid_users, 'exception_records': 0})
            else:
                exceptions_df_copy = exceptions_df.copy()
                exceptions_df_copy['Created user'] = exceptions_df_copy['Created user'].astype(str).str.strip()
                user_stats = exceptions_df_copy.groupby('Created user').size().reset_index(name='exception_records')
                user_stats = user_stats[user_stats['Created user'].isin(valid_users)]
                users_with_no_exceptions_list = [u for u in valid_users if u not in user_stats['Created user'].tolist()]
                if users_with_no_exceptions_list:
                    users_with_no_exceptions_df = pd.DataFrame({'Created user': users_with_no_exceptions_list, 'exception_records': 0})
                    user_stats = pd.concat([user_stats, users_with_no_exceptions_df], ignore_index=True)
            
            total_records_user_df = df[df['Created user'].isin(valid_users)].groupby('Created user').size().reset_index(name='total_records')
            
            user_stats['Created user'] = user_stats['Created user'].astype(str)
            total_records_user_df['Created user'] = total_records_user_df['Created user'].astype(str)
            user_stats = pd.merge(user_stats, total_records_user_df, on='Created user', how='left').fillna({'total_records': 0, 'exception_records': 0})
            
            user_stats['exception_rate'] = 0.0
            mask = user_stats['total_records'] > 0
            user_stats.loc[mask, 'exception_rate'] = (user_stats.loc[mask, 'exception_records'] / user_stats.loc[mask, 'total_records'] * 100).round(2)
            user_stats['exception_rate'] = user_stats['exception_rate'].fillna(0)

            if user_stats.empty: return

            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            data_to_insert_perf = [
                (run_id, row['Created user'], int(row['total_records']), int(row['exception_records']), row['exception_rate'])
                for _, row in user_stats.iterrows()
            ]
            if data_to_insert_perf:
                cursor.executemany('''INSERT INTO user_performance (run_id, user, total_records, exception_records, exception_rate)
                                      VALUES (?,?,?,?,?)''', data_to_insert_perf)
                conn.commit()
        except sqlite3.Error as e:
            logging.error(f"SQLite error in save_user_performance: {e}", exc_info=True)
            if conn: conn.rollback()
            raise
        except Exception as ex_gen: 
            logging.error(f"General error in save_user_performance: {ex_gen}", exc_info=True)
            st.error(f"An unexpected error occurred while saving user performance: {ex_gen}")
        finally:
            if conn: conn.close()

    def get_validation_history(self):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            return pd.read_sql_query('SELECT * FROM validation_runs ORDER BY upload_time DESC', conn)
        except sqlite3.Error as e:
            logging.error(f"Error in get_validation_history: {e}", exc_info=True); return pd.DataFrame()
        finally:
            if conn: conn.close()

    def get_exceptions_by_run(self, run_id):
        conn = None
        all_records = []
        all_json_keys_for_run = set()
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            raw_exceptions_df = pd.read_sql_query('SELECT * FROM exceptions WHERE run_id = ?', conn, params=(run_id,))

            if not raw_exceptions_df.empty:
                for _, db_row in raw_exceptions_df.iterrows():
                    current_row_json_data = {}
                    if 'original_row_data' in db_row.index and pd.notna(db_row['original_row_data']):
                        try:
                            current_row_json_data = json.loads(db_row['original_row_data'])
                            all_json_keys_for_run.update(current_row_json_data.keys())
                        except json.JSONDecodeError:
                            logging.warning(f"Failed to parse original_row_data for exception id {db_row.get('id')}")
                    
                    record_data = current_row_json_data.copy()
                    record_data['id'] = db_row.get('id')
                    record_data['run_id'] = db_row.get('run_id')
                    record_data['Exception Reasons'] = db_row.get('exception_reason')
                    record_data['Severity'] = db_row.get('severity')

                    mapped_db_fields_to_display_names = {
                        'department': 'Department.Name', 'sub_department': 'Sub Department.Name',
                        'created_user': 'Created user', 'modified_user': 'Modified user',
                        'net_amount': 'Net amount', 'location': 'Location.Name',
                        'crop': 'Crop.Name', 'activity': 'Activity.Name',
                        'function_name': 'Function.Name', 'vertical_name': 'FC-Vertical.Name',
                        'region_name': 'Region.Name', 'zone_name': 'Zone.Name',
                        'business_unit': 'Business Unit.Name'
                    }
                    for db_key, display_key in mapped_db_fields_to_display_names.items():
                        if display_key not in record_data: 
                            record_data[display_key] = db_row.get(db_key)
                    all_records.append(record_data)
            final_df = pd.DataFrame(all_records)
            if final_df.empty:
                ordered_empty_cols = ['id', 'run_id']
                other_json_keys = sorted(list(all_json_keys_for_run - {'id', 'run_id', 'Exception Reasons', 'Severity'}))
                ordered_empty_cols.extend(other_json_keys)
                ordered_empty_cols.extend(['Exception Reasons', 'Severity'])
                final_ordered_empty_cols = []
                for item in ordered_empty_cols:
                    if item not in final_ordered_empty_cols:
                        final_ordered_empty_cols.append(item)
                return pd.DataFrame(columns=final_ordered_empty_cols)

            all_columns_in_final_df = list(final_df.columns)
            first_cols = ['id', 'run_id']
            last_cols = ['Exception Reasons', 'Severity']
            final_ordered_columns = []
            for col in first_cols:
                if col in all_columns_in_final_df:
                    final_ordered_columns.append(col)
            middle_cols = []
            for col_name in sorted(all_columns_in_final_df):
                if col_name not in first_cols and \
                   col_name not in last_cols and \
                   col_name != 'original_row_data' and \
                   col_name not in final_ordered_columns: 
                    middle_cols.append(col_name)
            final_ordered_columns.extend(middle_cols)
            for col in last_cols:
                if col in all_columns_in_final_df and col not in final_ordered_columns:
                    final_ordered_columns.append(col)
            for col_from_df in all_columns_in_final_df:
                if col_from_df not in final_ordered_columns and col_from_df != 'original_row_data':
                    final_ordered_columns.append(col_from_df)
            final_df = final_df.reindex(columns=final_ordered_columns, fill_value=pd.NA)
            return final_df
        except sqlite3.Error as e:
            st.error(f"Database error retrieving exceptions for run {run_id}: {e}")
            logging.error(f"get_exceptions_by_run SQLite error for run_id {run_id}: {e}", exc_info=True)
            return pd.DataFrame(columns=['id', 'run_id', 'Exception Reasons', 'Severity']) 
        except Exception as ex:
            st.error(f"Unexpected error retrieving exceptions for run {run_id}: {ex}")
            logging.error(f"get_exceptions_by_run general error for run_id {run_id}: {ex}", exc_info=True)
            return pd.DataFrame(columns=['id', 'run_id', 'Exception Reasons', 'Severity'])
        finally:
            if conn: conn.close()


@st.cache_resource
def get_database_manager():
    return DatabaseManager()

db_manager = get_database_manager()

class DataValidator:
    def __init__(self, base_ref_path="reference_data"):
        self.base_ref_path = base_ref_path
        self.no_crop_check = {
            "Finance & Account", "Human Resource", "Administration",
            "Information Technology", "Legal", "Accounts Receivable & MIS"
        }
        self.no_activity_check = self.no_crop_check.copy()
        self.no_activity_check.update({"Production", "Processing", "Parent Seed"})
        self.ref_files = self._load_reference_data()
        self.training_map = {
            'Incorrect Location Name': 'Review Location Name Guidelines: Ensure locations are valid (e.g., Bandamailaram).',
            'Incorrect Activity Name': 'Complete Activity Name Training: Use only approved activities for your department.',
            'Incorrect Crop Name': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'FC-Vertical Name cannot be blank': 'Ensure Valid FC-Vertical Name.',
            'Crop Name cannot be blank': 'Ensure Crop Name is provided when required.',
            'Incorrect Crop Name starting with ZZ': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'Incorrect Crop Name for FC-field crop Vertical': 'Ensure Valid Crop Name for FC-field crop.',
            'Incorrect Crop Name for VC-Veg Crop Vertical': 'Ensure Valid Crop Name for VC-Veg Crop.',
            'Incorrect Crop Name for Fruit Crop Vertical': 'Ensure Valid Crop Name for Fruit Crop.',
            'Incorrect Crop Name for Common vertical': 'Ensure Valid Crop Name for Common vertical.',
            'Incorrect Sub Department Name': 'Verify Sub-Department Standards.',
            'Incorrect Function Name': 'Check Function Name Guidelines.',
            'Incorrect FC-Vertical Name': 'Ensure Valid FC-Vertical Name.', 
            'Need to Update Processing Location': 'Use Approved Processing Locations.',
            'Incorrect Activity Name for Lab QC': 'Use Lab QA Approved Activities.',
            'Incorrect Activity Name for Field QA': 'Use Field QA Approved Activities.',
            'Incorrect Activity Name for Bio Tech Services': 'Use Bio Tech Approved Activities.',
            'Sub Department should be blank': 'Ensure Sub-Department is blank for this department.',
            'Activity Name cannot be blank or start with ZZ': 'Complete Activity Name Training: Use only approved activities.',
            'Incorrect Activity Name for Biotech - Markers': 'Use Approved Activities for Biotech - Markers.',
            'Incorrect Activity Name for Biotech - Tissue Culture': 'Use Approved Activities for Biotech - Tissue Culture.',
            'Incorrect Activity Name for Biotech - Mutation': 'Use Approved Activities for Biotech - Mutation.',
            'Incorrect Activity Name for Entomology': 'Use Approved Activities for Entomology.',
            'Incorrect Activity Name for Pathology': 'Use Approved Activities for Pathology.', 
            'Incorrect Activity Name for Bioinformatics': 'Use Approved Activities for Bioinformatics.',
            'Incorrect Activity Name for Biochemistry': 'Use Approved Activities for Biochemistry.',
            'Incorrect Activity Name for Common': 'Use Approved Activities for Common sub-department in Breeding Support.',
            'Need to update Zone can not left Blank': 'Ensure Zone is specified.',
            'Incorrect Zone Name for FC-field crop Vertical': 'Use Approved Zone Names for FC-field crop.',
            'Incorrect Zone Name for VC-Veg Crop Vertical': 'Use Approved Zone Names for VC-Veg Crop.',
            'Need to update Zone Name can not left Blank': 'Ensure Zone is specified for Common vertical in Production.',
            'Need to update Business Unit can not left Blank': 'Ensure Business Unit is specified.',
            'Incorrect Business Unit Name for FC-field crop Vertical': 'Use Approved Business Unit Names for FC-field crop.',
            'Incorrect Business Unit Name for VC-Veg Crop Vertical': 'Use Approved Business Unit Names for VC-Veg Crop.',
            'Need to update Region Name can not left Blank': 'Ensure Region is specified.',
            'Incorrect Region Name for FC-field crop Vertical': 'Use Approved Region Names for FC-field crop.',
            'Incorrect Region Name for VC-Veg Crop Vertical': 'Use Approved Region Names for VC-Veg Crop.',
            'Region, Zone, BU need to check for Root Stock': 'Ensure Region, Zone, and BU are blank for Root Stock vertical in Sales/Marketing.',
            'Incorrect Activity Name for Sales': 'Use Approved Activities for Sales.',
            'Incorrect Activity Name for Marketing': 'Use Approved Activities for Marketing.'
        }

    def _load_reference_data(self):
        loaded_ref_files = {}
        ref_file_mappings = {
            "FC_Crop": ("FC-field crop.xlsx", "Crop.Name"),
            "VC_Crop": ("VC-Veg Crop.xlsx", "Crop.Name"),
            "SBFC_Region": ("SBFC-Region.xlsx", "Region.Name"),
            "SBVC_Region": ("SBVC-Region.xlsx", "Region.Name"),
            "SaleFC_Zone": ("SaleFC-Zone.xlsx", "Zone.Name"),
            "SaleVC_Zone": ("SaleVC-Zone.xlsx", "Zone.Name"),
            "FC_BU": ("FC-BU.xlsx", "Business Unit.Name"),
            "VC_BU": ("VC-BU.xlsx", "Business Unit.Name"),
            "Fruit_Crop": ("Fruit Crop.xlsx", "Crop.Name"),
            "Common_Crop": ("Common crop.xlsx", "Crop.Name"),
            "ProductionFC_Zone": ("ProductionFC-Zone.xlsx", "Zone.Name"),
            "ProductionVC_Zone": ("ProductionVC-Zone.xlsx", "Zone.Name"),
            "SalesActivity": ("SalesActivity.xlsx", "Activity.Name"),
            "MarketingActivity": ("MarketingActivity.xlsx", "Activity.Name"),
            "RS_BU": ("RS-BU.xlsx", "Business Unit.Name"),
            "SaleRS_Zone": ("SaleRS-Zone.xlsx", "Zone.Name"),
            "SBRS_Region": ("SBRS-Region.xlsx", "Region.Name"),
            "Root Stock_Crop": ("Root Stock Crop.xlsx", "Crop.Name"),
            "Region_Excluded_Accounts": ("Region.Name excluded.xlsx", "Account.Code"),
            "Zone_Excluded_Accounts": ("Zone.Name excluded.xlsx", "Account.Code"),
        }

        if not os.path.isdir(self.base_ref_path):
            st.error(f"Reference data directory not found: '{self.base_ref_path}'. Please create it and add reference Excel files. Validations will be highly inaccurate.")
            logging.critical(f"Reference data directory not found: '{self.base_ref_path}'. Cannot load reference data.")
            return {key: [] for key in ref_file_mappings.keys()}

        all_files_loaded_successfully = True
        for key, (filename, col_name) in ref_file_mappings.items():
            try:
                file_path = os.path.join(self.base_ref_path, filename)
                if os.path.exists(file_path):
                    df_ref = pd.read_excel(file_path, engine='openpyxl')
                    if col_name in df_ref.columns:
                        if key in ["Region_Excluded_Accounts", "Zone_Excluded_Accounts"]:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        else:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        logging.info(f"Loaded {len(loaded_ref_files[key])} items for '{key}' from '{filename}'.")
                    else:
                        st.warning(f"Column '{col_name}' not found in reference file '{filename}'. '{key}' will be empty.")
                        logging.warning(f"Column '{col_name}' not found in '{file_path}' for key '{key}'.")
                        loaded_ref_files[key] = []
                        all_files_loaded_successfully = False
                else:
                    st.warning(f"Reference file '{filename}' not found in '{self.base_ref_path}'. Validations for '{key}' may be inaccurate.")
                    logging.warning(f"Reference file '{filename}' not found for key '{key}' at path '{file_path}'.")
                    loaded_ref_files[key] = []
                    all_files_loaded_successfully = False
            except Exception as e:
                st.error(f"Error loading reference file '{filename}' for '{key}': {e}")
                logging.error(f"Error loading reference file '{filename}' for key '{key}': {e}", exc_info=True)
                loaded_ref_files[key] = []
                all_files_loaded_successfully = False

        if not all_files_loaded_successfully:
            st.error("One or more reference data files could not be loaded correctly. Please check logs and file paths. Validation accuracy will be affected.")
        elif not any(loaded_ref_files.values()): 
            st.error("All reference data lists are empty after attempting to load files. This indicates a problem with file contents or loading logic. Validations will be highly inaccurate.")
            logging.critical("All loaded reference file lists are empty. Check file contents and parsing logic.")
        else:
            logging.info("Reference data loading process completed.")
        return loaded_ref_files

    def is_not_blank(self, value):
        if pd.isna(value) or value is None: return False
        val = str(value).strip().replace("\u00A0", "").replace("\u200B", "")
        return val != "" and val.upper() not in ["N/A", "NULL", "NONE", "NA", "0", "-"]

    def is_blank(self, value):
        return not self.is_not_blank(value)

    def validate_row(self, dept, row):
        reasons = []
        sub_dept = str(row.get("Sub Department.Name", "") or "").strip().replace("\u00A0", "").replace("\u200B", "")
        func = str(row.get("Function.Name", "") or "").strip()
        vertical = str(row.get("FC-Vertical.Name", "") or "").strip()
        loc = str(row.get("Location.Name", "") or "").strip()
        crop = str(row.get("Crop.Name", "") or "").strip()
        act = str(row.get("Activity.Name", "") or "").strip()
        
        region_val = row.get("Region.Name", "") 
        zone_val = row.get("Zone.Name", "")     
        bu_val = row.get("Business Unit.Name", "") 
        
        region = str(region_val or "").strip()
        zone = str(zone_val or "").strip()
        bu = str(bu_val or "").strip()
        account_code = str(row.get("Account.Code", "") or "").strip()

        if self.is_blank(loc) or loc.startswith("ZZ"):
            reasons.append("Incorrect Location Name")
        
        if dept not in self.no_activity_check and \
           dept not in ["Breeding", "Trialing & PD", "Sales", "Marketing", "Breeding Support"]:
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Incorrect Activity Name") 

        if dept not in self.no_crop_check:
            if self.is_blank(vertical):
                reasons.append("FC-Vertical Name cannot be blank")
            
            if self.is_blank(crop):
                reasons.append("Crop Name cannot be blank")
            elif crop.startswith("ZZ"):
                reasons.append("Incorrect Crop Name starting with ZZ")
            elif not self.is_blank(vertical): 
                fc_crop_list = self.ref_files.get("FC_Crop", [])
                vc_crop_list = self.ref_files.get("VC_Crop", [])
                fruit_crop_list = self.ref_files.get("Fruit_Crop", [])
                common_crop_list = self.ref_files.get("Common_Crop", [])
                root_stock_crop_list = self.ref_files.get("Root Stock_Crop", [])

                if vertical == "FC-field crop" and (not fc_crop_list or crop not in fc_crop_list):
                    reasons.append("Incorrect Crop Name for FC-field crop Vertical")
                elif vertical == "VC-Veg Crop" and (not vc_crop_list or crop not in vc_crop_list):
                    reasons.append("Incorrect Crop Name for VC-Veg Crop Vertical")
                elif vertical == "Fruit Crop" and (not fruit_crop_list or crop not in fruit_crop_list):
                    reasons.append("Incorrect Crop Name for Fruit Crop Vertical")
                elif vertical == "Common" and (not common_crop_list or crop not in common_crop_list):
                    reasons.append("Incorrect Crop Name for Common vertical")
                elif vertical == "Root Stock" and (not root_stock_crop_list or crop not in root_stock_crop_list):
                    reasons.append("Incorrect Crop Name for Root Stock Crop Vertical")

        region_excluded_accounts_list = self.ref_files.get("Region_Excluded_Accounts", [])
        zone_excluded_accounts_list = self.ref_files.get("Zone_Excluded_Accounts", [])
        if account_code in region_excluded_accounts_list and self.is_not_blank(region_val): 
            reasons.append("Region Name should be blank for this Account Code")
        if account_code in zone_excluded_accounts_list and self.is_not_blank(zone_val): 
            reasons.append("Zone Name should be blank for this Account Code")
        
        if dept == "Parent Seed":
            if sub_dept not in ["Breeder Seed Production", "Foundation Seed Production", "Processing FS"]: reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Production":
            if sub_dept not in ["Commercial Seed Production", "Seed Production Research"]: reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            
            prod_fc_zone_list = self.ref_files.get("ProductionFC_Zone", [])
            prod_vc_zone_list = self.ref_files.get("ProductionVC_Zone", [])
            if sub_dept == "Commercial Seed Production":
                if vertical == "FC-field crop":
                    if self.is_blank(zone_val): reasons.append("Need to update Zone can not left Blank") 
                    elif not prod_fc_zone_list or zone not in prod_fc_zone_list: reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(zone_val): reasons.append("Need to update Zone can not left Blank")
                    elif not prod_vc_zone_list or zone not in prod_vc_zone_list: reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                elif vertical == "Common" and self.is_blank(zone_val): reasons.append("Need to update Zone Name can not left Blank")

        elif dept == "Processing":
            if sub_dept not in ["Processing", "Warehousing", "Project & Maintenance"]: reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            if loc not in ["Bandamailaram", "Deorjhal", "Boriya"]: reasons.append("Need to Update Processing Location")

        elif dept == "Quality Assurance":
            if sub_dept not in ["Field QA", "Lab QC", "Bio Tech Services"]: reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            if sub_dept == "Lab QC" and act not in ["Lab Operations QA", "All Activity"]: reasons.append("Incorrect Activity Name for Lab QC")
            if sub_dept == "Field QA" and act not in ["Field Operations QA", "All Activity"]: reasons.append("Incorrect Activity Name for Field QA")
            if sub_dept == "Bio Tech Services" and act not in ["Molecular", "All Activity"]: reasons.append("Incorrect Activity Name for Bio Tech Services")
        
        elif dept == "Seed Tech":
            if sub_dept not in ["Aging Test", "Pelleting", "Priming", "Common"]: reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "In Licensing & Procurement":
            if self.is_not_blank(sub_dept): reasons.append("Sub Department should be blank")
            if func != "Supply Chain": reasons.append("Incorrect Function Name")
            if vertical == "" or vertical == "N/A" or vertical == "Common":
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Breeding":
            if self.is_not_blank(sub_dept): reasons.append("Sub Department should be blank")
            if func != "Research and Development": reasons.append("Incorrect Function Name")
            if vertical == "" or vertical == "N/A":
                reasons.append("Incorrect FC-Vertical Name")
            allowed_activities_breeding = ["Breeding", "All Activity", "Trialing", "Pre Breeding", "Germplasm Maintainance", "Experimental Seed Production"]
            if act not in allowed_activities_breeding:
                 reasons.append("Incorrect Activity Name")

        elif dept == "Breeding Support":
            if sub_dept not in ["Pathology", "Biotech - Tissue Culture", "Biotech - Mutation", "Biotech - Markers", "Bioinformatics", "Biochemistry", "Entomology", "Common"]: reasons.append("Incorrect Sub Department Name")
            if func != "Research and Development": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Activity Name cannot be blank or start with ZZ")
            else: 
                if sub_dept == "Biotech - Markers" and act not in ["Molecular", "Grain Quality", "Seed Treatment", "All Activity"]: reasons.append("Incorrect Activity Name for Biotech - Markers")
                elif sub_dept == "Biotech - Tissue Culture" and act not in ["Tissue Culture", "All Activity"]: reasons.append("Incorrect Activity Name for Biotech - Tissue Culture")
                elif sub_dept == "Biotech - Mutation" and act not in ["Mutation", "All Activity"]: reasons.append("Incorrect Activity Name for Biotech - Mutation")
                elif sub_dept == "Entomology" and act not in ["Entomology", "All Activity"]: reasons.append("Incorrect Activity Name for Entomology")
                elif sub_dept == "Pathology" and act not in ["Pathalogy", "All Activity"]: reasons.append("Incorrect Activity Name for Pathology") 
                elif sub_dept == "Bioinformatics" and act not in ["Bioinformatics", "All Activity"]: reasons.append("Incorrect Activity Name for Bioinformatics")
                elif sub_dept == "Biochemistry" and act not in ["Biochemistry", "All Activity"]: reasons.append("Incorrect Activity Name for Biochemistry")
                elif sub_dept == "Common" and act not in ["All Activity"]: reasons.append("Incorrect Activity Name for Common")

        elif dept == "Trialing & PD":
            if self.is_not_blank(sub_dept): reasons.append("Sub Department should be blank")
            if func != "Research and Development": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            allowed_activities_trialing = ["CT", "All Activity", "Trialing", "RST"]
            if act not in allowed_activities_trialing:
                 reasons.append("Incorrect Activity Name")

        elif dept == "Sales":
            if sub_dept not in ["Sales Brand", "Sales Export", "Sales Institutional & Govt"]: reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            
            sales_activity_list = self.ref_files.get("SalesActivity", [])
            if self.is_blank(act) or act.startswith("ZZ") or (not sales_activity_list or act not in sales_activity_list):
                reasons.append("Incorrect Activity Name for Sales")

            fc_bu_list = self.ref_files.get("FC_BU", [])
            sale_fc_zone_list = self.ref_files.get("SaleFC_Zone", [])
            sbfc_region_list = self.ref_files.get("SBFC_Region", [])
            vc_bu_list = self.ref_files.get("VC_BU", [])
            sale_vc_zone_list = self.ref_files.get("SaleVC_Zone", [])
            sbvc_region_list = self.ref_files.get("SBVC_Region", [])
            rs_bu_list = self.ref_files.get("RS_BU", [])
            sale_rs_zone_list = self.ref_files.get("SaleRS_Zone", [])
            sbrs_region_list = self.ref_files.get("SBRS_Region", [])

            if sub_dept == "Sales Brand":
                if vertical == "FC-field crop":
                    if self.is_blank(bu_val): reasons.append("Need to update Business Unit can not left Blank")
                    elif not fc_bu_list or bu not in fc_bu_list: reasons.append("Incorrect Business Unit Name for FC-field crop Vertical")
                    if self.is_blank(zone_val): reasons.append("Need to update Zone can not left Blank")
                    elif not sale_fc_zone_list or zone not in sale_fc_zone_list: reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                    if self.is_blank(region_val): reasons.append("Need to update Region Name can not left Blank")
                    elif not sbfc_region_list or region not in sbfc_region_list: reasons.append("Incorrect Region Name for FC-field crop Vertical")
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(bu_val): reasons.append("Need to update Business Unit can not left Blank")
                    elif not vc_bu_list or bu not in vc_bu_list: reasons.append("Incorrect Business Unit Name for VC-Veg Crop Vertical")
                    if self.is_blank(zone_val): reasons.append("Need to update Zone can not left Blank")
                    elif not sale_vc_zone_list or zone not in sale_vc_zone_list: reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                    if self.is_blank(region_val): reasons.append("Need to update Region Name can not left Blank")
                    elif not sbvc_region_list or region not in sbvc_region_list: reasons.append("Incorrect Region Name for VC-Veg Crop Vertical")
                elif vertical == "Root Stock": 
                    if self.is_blank(bu_val): reasons.append("Need to update Business Unit can not left Blank")
                    elif not rs_bu_list or bu not in rs_bu_list: reasons.append("Incorrect Business Unit Name for Root Stock Crop Vertical")
                    if self.is_blank(zone_val): reasons.append("Need to update Zone can not left Blank")
                    elif not sale_rs_zone_list or zone not in sale_rs_zone_list: reasons.append("Incorrect Zone Name for Root Stock Crop Vertical")
                    if self.is_blank(region_val): reasons.append("Need to update Region Name can not left Blank")
                    elif not sbrs_region_list or region not in sbrs_region_list: reasons.append("Incorrect Region Name for Root Stock Crop Vertical")

        elif dept == "Marketing":
            if sub_dept not in ["Business Development", "Digital Marketing", "Product Management"]: reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
            elif vertical == "Root Stock" and any(self.is_not_blank(x) for x in [region_val, zone_val, bu_val]):
                reasons.append("Region, Zone, BU need to check for Root Stock")
            
            marketing_activity_list = self.ref_files.get("MarketingActivity", [])
            if self.is_blank(act) or act.startswith("ZZ") or \
               (not marketing_activity_list or act not in marketing_activity_list):
                reasons.append("Incorrect Activity Name for Marketing")
        
        if dept == "Finance & Account":
            if sub_dept not in ["Accounts", "Finance", "Analytics, Internal Control & Budget", "Purchase ops", "Secretarial", "Document Management System", "Automation", "Group Company"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Human Resource":
            if sub_dept not in ["Compliances", "HR Ops", "Recruitment", "Team Welfare", "Training", "Common"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Administration":
            if sub_dept not in ["Events", "Maintenance", "Travel Desk", "Common"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Information Technology":
            if sub_dept not in ["ERP Support", "Infra & Hardware", "Application Development"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Legal":
            if sub_dept not in ["Compliances", "Litigation", "Common"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Accounts Receivable & MIS":
            if sub_dept not in ["Branch and C&F Ops", "Commercial & AR Management", "Common", "Order Processing", "Transport & Logistic"]: reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Management":
            if self.is_not_blank(sub_dept): reasons.append("Sub Department should be blank")
            if func != "Management": reasons.append("Incorrect Function Name")
            if self.is_blank(vertical): reasons.append("Incorrect FC-Vertical Name")
        
        unique_reasons = sorted(list(set(reasons))) 
        severity = len(unique_reasons) * 2 
        return unique_reasons, severity

    def validate_dataframe(self, df):
        exceptions = [] 
        department_stats = {}
        
        null_like_values = [pd.NA, "N/A", "NaN", "null", "NONE", "", " ", "-", "\u00A0", None, 0, "0"]
        if 'Sub Department.Name' in df.columns:
            df['Sub Department.Name'] = df['Sub Department.Name'].replace(null_like_values, "").astype(str).str.strip()
        else:
            df['Sub Department.Name'] = "" 

        input_columns = df.columns.tolist() 
        
        if 'Department.Name' not in df.columns:
            st.error("Critical Error: 'Department.Name' column is missing from the input file.")
            logging.critical("Critical Error: 'Department.Name' column is missing in validate_dataframe.")
            return pd.DataFrame(columns=input_columns + ['Exception Reasons', 'Severity']), {}

        for dept_name_val in df['Department.Name'].dropna().unique():
            dept = str(dept_name_val) 
            dept_df = df[df['Department.Name'] == dept].copy() 
            dept_exception_rows_count = 0 
            
            for index, row in dept_df.iterrows():
                reasons, severity = self.validate_row(dept, row) 
                if reasons: 
                    record = row.to_dict() 
                    record['Exception Reasons'] = "; ".join(reasons) 
                    record['Severity'] = severity 
                    exceptions.append(record) 
                    dept_exception_rows_count += 1
            
            department_stats[dept] = {
                'total_records': len(dept_df),
                'exception_records': dept_exception_rows_count, 
                'exception_rate': (dept_exception_rows_count / len(dept_df) * 100) if len(dept_df) > 0 else 0
            }
        
        exceptions_df_output = pd.DataFrame()
        output_columns_with_exceptions = input_columns + ['Exception Reasons', 'Severity']

        if exceptions: 
            exceptions_df_output = pd.DataFrame(exceptions)
            for col in output_columns_with_exceptions:
                if col not in exceptions_df_output.columns:
                    exceptions_df_output[col] = pd.NA 
            exceptions_df_output = exceptions_df_output.reindex(columns=output_columns_with_exceptions, fill_value=pd.NA)
        else: 
            exceptions_df_output = pd.DataFrame(columns=output_columns_with_exceptions)
        
        return exceptions_df_output, department_stats


def display_metric(title, value, delta=None, container=None):
    target_container = container if container else st
    delta_html = f'<div class="metric-delta" style="font-size:0.8rem; color:#e2e8f0;">{delta}</div>' if delta else "" 
    target_container.markdown(f"""
    <div class="metric-container">
        <div class="metric-title">{title}</div>
        <div class="metric-value">{value}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

def create_excel_report(exceptions_df, dept_stats, filename_for_logging="ExcelReport"):
    output = io.BytesIO()
    try:
        if exceptions_df is None or exceptions_df.empty:
            logging.info(f"create_excel_report: exceptions_df is empty for {filename_for_logging}. Creating empty Exceptions sheet.")
            exceptions_df_prepared = exceptions_df.copy() if exceptions_df is not None else pd.DataFrame()
        else:
            exceptions_df_prepared = exceptions_df.copy()

        if 'Exception Reasons' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Exception Reasons'] = exceptions_df_prepared['Exception Reasons'].astype(str).replace('<NA>', '').replace('nan', '').replace('None','')
        if 'Severity' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Severity'] = pd.to_numeric(exceptions_df_prepared['Severity'], errors='coerce').fillna(0)
            if exceptions_df_prepared['Severity'].notna().all(): 
                 try: 
                    exceptions_df_prepared['Severity'] = exceptions_df_prepared['Severity'].astype(int)
                 except ValueError: 
                    pass 

        numeric_cols_to_preserve = ['Net amount', 'Severity', 'id', 'run_id']
        if dept_stats: 
             numeric_cols_to_preserve.extend(['Total Records', 'Exception Records', 'Exception Rate (%)'])

        for col in exceptions_df_prepared.columns:
            if col in numeric_cols_to_preserve:
                if col == 'Net amount':
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce').fillna(0.0)
                elif col in ['id', 'run_id']: 
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce')
            else: 
                if not pd.api.types.is_numeric_dtype(exceptions_df_prepared[col]) and \
                   not pd.api.types.is_string_dtype(exceptions_df_prepared[col]): 
                    exceptions_df_prepared[col] = exceptions_df_prepared[col].astype(str).replace('<NA>', '').replace('nan', '').replace('None','').replace('NaT','')
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            exceptions_df_prepared.to_excel(writer, sheet_name='Exceptions', index=False)
            if not dept_stats:
                 dept_summary_df = pd.DataFrame(columns=['Department', 'Total Records', 'Exception Records', 'Exception Rate (%)'])
            else:
                dept_summary_df = pd.DataFrame([{'Department': dept, 'Total Records': stats.get('total_records', 0), 'Exception Records': stats.get('exception_records', 0), 'Exception Rate (%)': round(stats.get('exception_rate', 0), 2)} for dept, stats in dept_stats.items()])
            dept_summary_df.to_excel(writer, sheet_name='Department Summary', index=False)
            workbook = writer.book
            header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            thin_border_side = Side(style='thin'); cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.max_row == 0 : continue
                df_ref_for_headers = exceptions_df_prepared if sheet_name == 'Exceptions' else dept_summary_df
                if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                    if not df_ref_for_headers.columns.empty:
                        for c_idx, h_val in enumerate(df_ref_for_headers.columns, 1):
                            cell = ws.cell(row=1, column=c_idx, value=str(h_val)); cell.font=header_font; cell.fill=header_fill; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=cell_border
                    else: continue
                for cell in ws[1]:
                    if cell.value is not None: cell.value = str(cell.value); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = cell_border
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx); cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cell.border = cell_border
                        header_cell_value = str(ws.cell(row=1, column=col_idx).value or "")
                        if header_cell_value == "Net amount": cell.number_format = '#,##0.00'
                        elif header_cell_value == "Exception Rate (%)": cell.number_format = '0.00"%"'
                        elif header_cell_value == "Severity" or "Records" in header_cell_value or "ID" in header_cell_value or "id" in header_cell_value : cell.number_format = '0'
                for col_idx_letter_enum, column_cells_obj in enumerate(ws.columns, 1):
                    current_col_letter = get_column_letter(col_idx_letter_enum); header_val_str = str(ws.cell(row=1, column=col_idx_letter_enum).value or ""); max_length = len(header_val_str)
                    for cell_in_col_obj in column_cells_obj:
                        if cell_in_col_obj.row == 1: continue
                        try:
                            if cell_in_col_obj.value is not None: cell_str_val = str(cell_in_col_obj.value); cell_len = max(len(s) for s in cell_str_val.split('\n')) if '\n' in cell_str_val else len(cell_str_val); max_length = max(max_length, cell_len)
                        except: pass
                    adjusted_width = min(max_length + 5, 60); ws.column_dimensions[current_col_letter].width = adjusted_width
        output.seek(0)
        if output.getbuffer().nbytes == 0:
            logging.error(f"create_excel_report: Excel output buffer is empty for {filename_for_logging}")
            st.markdown('<div class="error-box"><strong>❌ Error:</strong> Failed to generate Excel report (output empty).</div>', unsafe_allow_html=True); return None
        return output
    except Exception as e:
        logging.exception(f"create_excel_report: Error generating Excel report for {filename_for_logging}: {e}")
        st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error generating Excel report: {e}</div>', unsafe_allow_html=True); return None


def process_uploaded_file(uploaded_file):
    try:
        df = pd.DataFrame() 
        with st.spinner(f"📖 Reading file: {uploaded_file.name}..."):
            try:
                df = pd.read_excel(uploaded_file, engine='openpyxl') 
                df.columns = df.columns.str.strip() 
            except Exception as e:
                st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Could not read Excel file "{uploaded_file.name}". Details: {str(e)}</div>', unsafe_allow_html=True)
                logging.error(f"Error reading Excel file {uploaded_file.name}: {e}", exc_info=True)
                return

        if df.empty:
            st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> The uploaded file "{uploaded_file.name}" is empty or could not be parsed.</div>', unsafe_allow_html=True)
            return

        required_columns_for_processing = ['Department.Name', 'Created user']
        validator_expected_columns = [
            'Department.Name', 'Sub Department.Name', 'Created user', 'Modified user',
            'Net amount', 'Location.Name', 'Crop.Name', 'Activity.Name', 'Function.Name',
            'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name'
        ]

        missing_core_cols = [col for col in required_columns_for_processing if col not in df.columns]
        if missing_core_cols:
            st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Missing essential columns for processing: {", ".join(missing_core_cols)}. Cannot proceed with file "{uploaded_file.name}".</div>', unsafe_allow_html=True)
            return
        
        missing_validator_cols_list = [col for col in validator_expected_columns if col not in df.columns]
        if missing_validator_cols_list:
             st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> For file "{uploaded_file.name}", the following columns are missing and might affect validation accuracy: {", ".join(missing_validator_cols_list)}. Missing columns will be treated as blank during validation.</div>', unsafe_allow_html=True)
             for col in missing_validator_cols_list:
                 df[col] = "" 

        st.markdown("#### 📊 File Information")
        col_info1, col_info2, col_info3, col_info4 = st.columns(4)
        display_metric("Total Records", f"{len(df):,}", container=col_info1)
        display_metric("Total Columns", len(df.columns), container=col_info2)
        display_metric("File Size", f"{uploaded_file.size / 1024:.1f} KB", container=col_info3)
        display_metric("Departments", df['Department.Name'].nunique() if 'Department.Name' in df.columns else "N/A", container=col_info4)
        
        exceptions_df_from_validation = pd.DataFrame() 
        department_statistics = {} 

        with st.spinner(f"🔍 Validating data for {uploaded_file.name}..."):
            validator_instance = DataValidator(base_ref_path="reference_data") 
            ref_data_loaded_check = any(isinstance(lst, list) and len(lst) > 0 for lst in validator_instance.ref_files.values())

            if not ref_data_loaded_check:
                st.error(f"CRITICAL ERROR for file '{uploaded_file.name}': Reference data files could not be loaded by the validator or are all empty. Validation results will be highly inaccurate. Please check the 'reference_data' directory and ensure Excel files are present, correctly named, and contain data. See logs for details.")
                logging.critical(f"Reference data files appear to be uninitialized or empty in DataValidator for file '{uploaded_file.name}'. Aborting full validation display logic.")
                run_id_ref_error = db_manager.save_validation_run(
                    filename=uploaded_file.name, total_records=len(df),
                    total_exceptions=0, 
                    file_size=uploaded_file.size
                )
                st.markdown(f"A validation run (ID: {run_id_ref_error}) was logged for '{uploaded_file.name}', but no exceptions were processed due to missing/empty reference data.")
                return 

            exceptions_df_from_validation, department_statistics = validator_instance.validate_dataframe(df.copy())

        current_run_id = db_manager.save_validation_run(
            filename=uploaded_file.name,
            total_records=len(df),
            total_exceptions=len(exceptions_df_from_validation), 
            file_size=uploaded_file.size
        )

        if not exceptions_df_from_validation.empty:
            db_manager.save_exceptions(current_run_id, exceptions_df_from_validation)
        
        exceptions_for_perf_calc = exceptions_df_from_validation
        if not exceptions_df_from_validation.empty and 'Created user' not in exceptions_df_from_validation.columns:
            logging.warning("'Created user' column missing in exceptions_df for user performance calculation. Adding it as blank.")
            exceptions_for_perf_calc = exceptions_df_from_validation.copy()
            exceptions_for_perf_calc['Created user'] = "" 
        db_manager.save_user_performance(current_run_id, df, exceptions_for_perf_calc)
        
        if department_statistics:
            conn_dept_save = None
            try:
                conn_dept_save = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                cursor_dept_save = conn_dept_save.cursor()
                for dept_key_save, stats_val_save in department_statistics.items():
                    cursor_dept_save.execute('''
                        INSERT INTO department_summary (run_id, department, total_records, exception_records, exception_rate)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (current_run_id, dept_key_save, stats_val_save['total_records'], stats_val_save['exception_records'], stats_val_save['exception_rate']))
                conn_dept_save.commit()
            except sqlite3.Error as e_dept_save:
                logging.error(f"Error saving department_summary for run {current_run_id}: {e_dept_save}", exc_info=True)
                if conn_dept_save: conn_dept_save.rollback()
                st.error(f"Failed to save department summary for '{uploaded_file.name}': {e_dept_save}")
            finally:
                if conn_dept_save: conn_dept_save.close()

        st.markdown("#### 🛠 Validation Results")
        if exceptions_df_from_validation.empty:
            st.markdown(f'<div class="success-box"><strong>✅ Perfect!</strong> No validation issues found in "{uploaded_file.name}"!</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> Found {len(exceptions_df_from_validation)} records with validation issues in "{uploaded_file.name}".</div>', unsafe_allow_html=True)
            
            col_res1, col_res2, col_res3 = st.columns(3)
            display_metric("Total Exceptions", f"{len(exceptions_df_from_validation):,}", container=col_res1)
            current_exception_rate = (len(exceptions_df_from_validation)/len(df)*100) if len(df) > 0 else 0
            display_metric("Exception Rate", f"{current_exception_rate:.2f}%", container=col_res2)
            
            avg_sev = 0.0
            if not exceptions_df_from_validation.empty and 'Severity' in exceptions_df_from_validation.columns and not exceptions_df_from_validation['Severity'].dropna().empty:
                avg_sev = exceptions_df_from_validation['Severity'].mean()
            display_metric("Average Severity", f"{avg_sev:.2f}", container=col_res3)

            st.markdown("##### 📋 Exception Records")
            display_cols_for_exceptions_ui = [col for col in validator_expected_columns if col in exceptions_df_from_validation.columns] + ['Exception Reasons', 'Severity']
            display_cols_for_exceptions_ui = [col for col in display_cols_for_exceptions_ui if col in exceptions_df_from_validation.columns]
            
            if display_cols_for_exceptions_ui: 
                st.dataframe(exceptions_df_from_validation[display_cols_for_exceptions_ui], use_container_width=True)
            else: 
                st.dataframe(exceptions_df_from_validation, use_container_width=True)
            
            excel_report_data = create_excel_report(exceptions_df_from_validation, department_statistics, uploaded_file.name)
            if excel_report_data:
                st.download_button(
                    label=f"📥 Download Validation Report for {uploaded_file.name}",
                    data=excel_report_data,
                    file_name=f"Validation_Report_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error(f"Could not generate the Excel report for download for '{uploaded_file.name}'.")
    except Exception as e_process: 
        st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Failed to process file "{uploaded_file.name if uploaded_file else "N/A"}": {str(e_process)}. Check logs for more details.</div>', unsafe_allow_html=True)
        logging.exception(f"Unhandled error processing uploaded file {uploaded_file.name if uploaded_file else 'N/A'}: {e_process}")


def show_upload_page():
    st.markdown("### 📁 File Upload & Validation")
    st.markdown("""<div class="upload-section"><div class="upload-title">📤 Drag and Drop Your Excel Files</div><div class="upload-subtitle">Supported formats: .xlsx, .xls | Max file size: 200MB per file</div></div>""", unsafe_allow_html=True)
    uploaded_files_list = st.file_uploader("Choose Excel files", type=['xlsx','xls'], accept_multiple_files=True, help="Upload your VNR Seeds expense report files for validation", label_visibility="collapsed")
    if uploaded_files_list:
        st.success(f"{len(uploaded_files_list)} file(s) selected. Processing...")
        for individual_uploaded_file in uploaded_files_list:
            with st.expander(f"📄 Process: {individual_uploaded_file.name}", expanded=True):
                process_uploaded_file(individual_uploaded_file)
    st.markdown("### 📝 Manual Data Entry & Validation")
    with st.form("manual_data_entry_form"):
        st.markdown("##### Enter Record Details (fields with * are mandatory):")
        cols_m1_row1, cols_m1_row2, cols_m1_row3 = st.columns(3)
        department_manual_input = cols_m1_row1.selectbox("Department*", [""] + sorted(["Parent Seed", "Production", "Processing", "Quality Assurance", "Finance & Account", "Human Resource", "Administration", "Information Technology", "Legal", "Accounts Receivable & MIS", "Seed Tech", "In Licensing & Procurement", "Breeding", "Breeding Support", "Trialing & PD", "Sales", "Marketing", "Management"]), key="manual_input_department", help="Select the department.")
        location_manual_input = cols_m1_row2.text_input("Location Name*", key="manual_input_location", help="Enter location name (e.g., Bandamailaram).")
        created_user_manual_input = cols_m1_row3.text_input("Created User*", key="manual_input_created_user", help="Enter the user ID of the creator.")
        cols_m2_row1, cols_m2_row2, cols_m2_row3 = st.columns(3)
        activity_manual_input = cols_m2_row1.text_input("Activity Name", key="manual_input_activity", help="Enter activity name if applicable.")
        net_amount_manual_input = cols_m2_row2.number_input("Net Amount", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="manual_input_net_amount", help="Enter the net amount.")
        sub_dept_manual_input = cols_m2_row3.text_input("Sub Department Name", value="", key="manual_input_sub_department", help="Enter sub-department if applicable.")
        with st.expander("Optional Fields for Manual Entry"):
            cols_m3_row1, cols_m3_row2, cols_m3_row3 = st.columns(3)
            modified_user_manual_input = cols_m3_row1.text_input("Modified User", value="", key="manual_input_modified_user")
            crop_manual_input = cols_m3_row2.text_input("Crop Name", value="", key="manual_input_crop") 
            function_manual_input = cols_m3_row3.text_input("Function Name", value="", key="manual_input_function")
            cols_m4_row1, cols_m4_row2, cols_m4_row3, cols_m4_row4 = st.columns(4)
            vertical_manual_input = cols_m4_row1.text_input("FC-Vertical Name", value="", key="manual_input_vertical")
            region_manual_input = cols_m4_row2.text_input("Region Name", value="", key="manual_input_region")
            zone_manual_input = cols_m4_row3.text_input("Zone Name", value="", key="manual_input_zone")
            business_unit_manual_input = cols_m4_row4.text_input("Business Unit Name", value="", key="manual_input_business_unit")
        submitted_manual_record = st.form_submit_button("Validate & Submit Manual Record")
        if submitted_manual_record:
            if not department_manual_input or not location_manual_input or not created_user_manual_input:
                st.error("Department, Location Name, and Created User are mandatory fields for manual entry.")
            else:
                manual_validator = DataValidator(base_ref_path="reference_data")
                manual_ref_data_loaded = any(isinstance(lst, list) and len(lst) > 0 for lst in manual_validator.ref_files.values())
                if not manual_ref_data_loaded:
                    st.error("CRITICAL: Reference data not loaded. Manual validation cannot be performed accurately.")
                    logging.critical("Reference data not loaded for manual validation.")
                else:
                    manual_row_data = pd.Series({'Department.Name': department_manual_input, 'Location.Name': location_manual_input, 'Activity.Name': activity_manual_input, 'Created user': created_user_manual_input, 'Net amount': net_amount_manual_input, 'Sub Department.Name': sub_dept_manual_input, 'Modified user': modified_user_manual_input, 'Crop.Name': crop_manual_input, 'Function.Name': function_manual_input, 'FC-Vertical.Name': vertical_manual_input, 'Region.Name': region_manual_input, 'Zone.Name': zone_manual_input, 'Business Unit.Name': business_unit_manual_input})
                    manual_reasons, manual_severity = manual_validator.validate_row(department_manual_input, manual_row_data)
                    manual_df_for_db = pd.DataFrame([manual_row_data])
                    if manual_reasons:
                        st.warning(f"Validation Issues for manual entry: {'; '.join(manual_reasons)} (Severity: {manual_severity})")
                        manual_df_for_db['Exception Reasons'] = "; ".join(manual_reasons)
                        manual_df_for_db['Severity'] = manual_severity
                        manual_entry_run_id = db_manager.save_validation_run(filename=f"Manual_Entry_Error_{datetime.now().strftime('%Y%m%d_%H%M%S')}", total_records=1, total_exceptions=1, file_size=0 )
                        db_manager.save_exceptions(manual_entry_run_id, manual_df_for_db)
                        db_manager.save_user_performance(manual_entry_run_id, pd.DataFrame([manual_row_data]), manual_df_for_db) 
                        st.info(f"Manual record submitted with noted validation issues (Run ID: {manual_entry_run_id}).")
                    else:
                        valid_manual_entry_run_id = db_manager.save_validation_run(filename=f"Manual_Entry_OK_{datetime.now().strftime('%Y%m%d_%H%M%S')}", total_records=1, total_exceptions=0, file_size=0)
                        empty_exceptions_for_valid_manual = pd.DataFrame(columns=['Created user', 'Exception Reasons', 'Severity']) 
                        db_manager.save_user_performance(valid_manual_entry_run_id, pd.DataFrame([manual_row_data]), empty_exceptions_for_valid_manual)
                        st.success(f"Manual record validated successfully and run logged (Run ID: {valid_manual_entry_run_id}).")

def show_analytics_page():
    st.markdown("### 📊 Dashboard Analytics")
    validation_history = db_manager.get_validation_history()
    if validation_history.empty:
        st.info("No validation runs found. Upload a file or enter data manually to see analytics.")
        return
    st.markdown("#### 📈 Overall Statistics")
    stat_col1, stat_col2, stat_col3 = st.columns(3)
    total_runs = len(validation_history)
    display_metric("Total Validation Runs", f"{total_runs:,}", container=stat_col1)
    total_recs_processed = validation_history['total_records'].sum() if 'total_records' in validation_history.columns else 0
    display_metric("Total Records Processed", f"{total_recs_processed:,}", container=stat_col2)
    total_excs_found = validation_history['total_exceptions'].sum() if 'total_exceptions' in validation_history.columns else 0
    display_metric("Total Exceptions Found", f"{total_excs_found:,}", container=stat_col3)

    if total_recs_processed > 0:
        st.markdown("#### 🔍 Overall Data Quality Snapshot")
        labels = ['Records with Exceptions', 'Records without Exceptions']
        values = [total_excs_found, total_recs_processed - total_excs_found]
        
        # Define a more attractive color scheme
        colors = ['#FF6B6B', '#6BCB77'] # A red for exceptions, a green for no exceptions

        fig_overall_quality = go.Figure(data=[go.Pie(labels=labels, 
                                                     values=values, 
                                                     hole=.4, 
                                                     marker_colors=colors,
                                                     hoverinfo='label+percent+value', 
                                                     textinfo='value+label',
                                                     insidetextorientation='radial',
                                                     pull=[0.05, 0] # Slightly pull out the exceptions slice
                                                     )])
        fig_overall_quality.update_layout(
            # title_text='Overall Record Quality', # Already have a markdown title
            annotations=[dict(text='Quality', x=0.5, y=0.5, font_size=20, showarrow=False, font=PLOTLY_FONT)],
            legend_title_text='Record Status',
            margin=dict(t=30, b=30, l=10, r=10),
            font=PLOTLY_FONT,
            paper_bgcolor='rgba(0,0,0,0)', 
            plot_bgcolor='rgba(0,0,0,0)',
            height=400
        )
        st.plotly_chart(fig_overall_quality, use_container_width=True)


    if not validation_history.empty and 'id' in validation_history.columns:
        latest_run_id_analytics = validation_history['id'].max()
        conn_analytics = None; dept_summary_analytics_df = pd.DataFrame()
        try:
            conn_analytics = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
            dept_summary_analytics_df = pd.read_sql_query("SELECT department, total_records, exception_records, exception_rate FROM department_summary WHERE run_id = ?", conn_analytics, params=(int(latest_run_id_analytics),))
        except sqlite3.Error as e_analytics_dept:
            logging.error(f"Error fetching department summary for analytics (run {latest_run_id_analytics}): {e_analytics_dept}", exc_info=True)
        finally:
            if conn_analytics: conn_analytics.close()

        if not dept_summary_analytics_df.empty:
            st.markdown(f"#### 🏭 Department Analysis (Latest Run: ID {latest_run_id_analytics})")
            
            # Prepare data for hover
            dept_summary_analytics_df_sorted = dept_summary_analytics_df.sort_values(by='exception_rate', ascending=False)

            fig_dept_analysis = px.bar(dept_summary_analytics_df_sorted,
                                       x='department', y='exception_rate',
                                       labels={'exception_rate': 'Exception Rate (%)', 'department': 'Department'},
                                       color='exception_rate',
                                       color_continuous_scale='Sunsetdark', # Changed color scale
                                       text_auto='.2f',
                                       hover_name='department',
                                       custom_data=['total_records', 'exception_records']
                                       )
            fig_dept_analysis.update_traces(hovertemplate="<b>%{hovertext}</b><br><br>" +
                                                        "Exception Rate: %{y:.2f}%<br>" +
                                                        "Total Records: %{customdata[0]:,}<br>" +
                                                        "Exception Records: %{customdata[1]:,}<extra></extra>")
            fig_dept_analysis.update_layout(
                title_text="Exception Rate by Department",
                title_x=0.5,
                title_font=PLOTLY_TITLE_FONT,
                xaxis_title="Department",
                yaxis_title="Exception Rate (%)",
                margin=dict(l=40, r=20, t=60, b=150), # Increased bottom margin
                xaxis_tickangle=-45,
                font=PLOTLY_FONT,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                yaxis=dict(gridcolor='#e9ecef'), 
                xaxis=dict(showgrid=False) 
            )
            st.plotly_chart(fig_dept_analysis, use_container_width=True)
            st.markdown("##### 📋 Department Summary Table (Latest Run)")
            st.dataframe(dept_summary_analytics_df.style.format({"exception_rate": "{:.2f}%", "total_records":"{:,}","exception_records":"{:,}"}), use_container_width=True, hide_index=True)
        else:
            st.info(f"No department summary found for the latest run (ID {latest_run_id_analytics}). This might occur if the run had no data or if summarization failed.")
    else:
        st.info("No run history available to determine the latest run for department analysis.")


def show_trends_page():
    st.markdown("### 📈 Trends & History")
    trends_history_df = db_manager.get_validation_history()
    if trends_history_df.empty:
        st.info("No historical data available. Upload files to see trends.")
        return
    trends_history_df['upload_time'] = pd.to_datetime(trends_history_df['upload_time'])
    trends_history_df = trends_history_df.sort_values(by='upload_time', ascending=True)
    
    fig_trends = go.Figure()
    theme_colors = {'exceptions': '#FF6B6B', 'records': '#6A89CC'}

    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'], 
                                    y=trends_history_df['total_exceptions'], 
                                    mode='lines+markers', name='Total Exceptions', 
                                    line=dict(color=theme_colors['exceptions'], width=2.5, shape='spline'), 
                                    marker=dict(symbol="circle", size=8, line=dict(width=1,color=theme_colors['exceptions'])),
                                    hovertemplate="<b>Total Exceptions</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>"))
    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'], 
                                    y=trends_history_df['total_records'], 
                                    mode='lines+markers', name='Total Records Processed', 
                                    line=dict(color=theme_colors['records'], width=2.5, shape='spline'), 
                                    marker=dict(symbol="square", size=8, line=dict(width=1,color=theme_colors['records'])),
                                    hovertemplate="<b>Total Records</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>",
                                    fill='tozeroy', 
                                    fillcolor='rgba(106, 137, 204, 0.2)'
                                    ))
    fig_trends.update_layout(
        title_text="Validation Trends Over Time",
        title_x=0.5,
        title_font=PLOTLY_TITLE_FONT,
        xaxis_title="Upload Date", 
        yaxis_title="Count", 
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, 
                    bgcolor='rgba(255,255,255,0.7)', bordercolor='rgba(0,0,0,0.05)', borderwidth=1),
        margin=dict(l=50, r=20, t=70, b=40), 
        hovermode="x unified",
        font=PLOTLY_FONT,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da'),
        xaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da', rangeslider=dict(visible=True, thickness=0.05), type="date")
    )
    st.plotly_chart(fig_trends, use_container_width=True)
    st.markdown("##### 📜 Validation History Log")
    display_history_log_df = trends_history_df[['id', 'filename', 'upload_time', 'total_records', 'total_exceptions', 'file_size']].copy()
    display_history_log_df.columns = ['Run ID', 'Filename', 'Upload Time', 'Total Records', 'Total Exceptions', 'File Size (Bytes)']
    display_history_log_df['Upload Time'] = display_history_log_df['Upload Time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    st.dataframe(display_history_log_df.sort_values(by='Upload Time', ascending=False), use_container_width=True, hide_index=True)


def show_exception_details_page():
    st.markdown("### 🔍 Exception Details Viewer")
    try:
        validation_runs_list = db_manager.get_validation_history()
        if validation_runs_list.empty:
            st.info("No validation runs found in the database to display details for.")
            return
        run_options_for_details = {f"Run {run_row['id']}: {run_row['filename']} ({pd.to_datetime(run_row['upload_time']).strftime('%Y-%m-%d %H:%M')})": run_row['id'] for _, run_row in validation_runs_list.iterrows()}
        selected_run_display_text = st.selectbox("Select Validation Run to View Details", options=list(run_options_for_details.keys()), index=0, key="exception_details_run_select")
        if selected_run_display_text:
            selected_run_id = run_options_for_details[selected_run_display_text]
            exceptions_for_run_df = db_manager.get_exceptions_by_run(selected_run_id) 
            if exceptions_for_run_df.empty:
                st.success(f"No exceptions found for the selected Run ID {selected_run_id}.")
                return
            st.markdown(f"#### Exceptions Found for Run ID: {selected_run_id}")
            st.dataframe(exceptions_for_run_df, use_container_width=True, height=400) 
            if 'Severity' in exceptions_for_run_df.columns and not exceptions_for_run_df['Severity'].dropna().empty:
                st.markdown("#### 📊 Severity Distribution")
                severity_value_counts = exceptions_for_run_df['Severity'].value_counts().reset_index()
                severity_value_counts.columns = ['Severity', 'Count'] 
                fig_severity_dist = px.bar(severity_value_counts.sort_values(by='Severity'), 
                                           x='Severity', y='Count', 
                                           title=f'Exception Severity Distribution (Run ID {selected_run_id})', 
                                           color='Severity', 
                                           color_continuous_scale=px.colors.sequential.OrRd, 
                                           text_auto=True,
                                           hover_name='Severity',
                                           labels={'Count': 'Number of Exceptions'})
                fig_severity_dist.update_traces(hovertemplate="Severity: %{x}<br>Count: %{y}<extra></extra>")
                fig_severity_dist.update_layout(
                    title_x=0.5, title_font=PLOTLY_TITLE_FONT,
                    xaxis_title="Severity Score", yaxis_title="Number of Exceptions",
                    font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False),
                    margin=dict(l=40, r=20, t=60, b=40)
                )
                st.plotly_chart(fig_severity_dist, use_container_width=True)
            else:
                st.info("Severity data not available or no exceptions with severity for this run.")
            conn_details_dept = None; dept_stats_for_excel_df = pd.DataFrame()
            try:
                conn_details_dept = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                dept_stats_for_excel_df = pd.read_sql_query("SELECT department, total_records, exception_records, exception_rate FROM department_summary WHERE run_id = ?", conn_details_dept, params=(selected_run_id,))
            except sqlite3.Error as e_details_dept_excel:
                logging.error(f"Error fetching department stats for Excel report (run {selected_run_id}): {e_details_dept_excel}", exc_info=True)
            finally:
                if conn_details_dept: conn_details_dept.close()
            dept_stats_dict_for_excel = {row_stat['department']: {'total_records': row_stat['total_records'], 'exception_records': row_stat['exception_records'], 'exception_rate': row_stat['exception_rate']} for _, row_stat in dept_stats_for_excel_df.iterrows()}
            try:
                original_filename_for_report = validation_runs_list[validation_runs_list['id'] == selected_run_id]['filename'].iloc[0]
                excel_report_name = f"Exceptions_Report_Run_{selected_run_id}_{original_filename_for_report}.xlsx"
                excel_binary_data = create_excel_report(exceptions_for_run_df, dept_stats_dict_for_excel, excel_report_name) 
                if excel_binary_data is None:
                    st.error("Failed to generate the Exceptions Report. Please check the application logs for more details.")
                    logging.error(f"Excel report generation returned None for run_id {selected_run_id}")
                else:
                    st.download_button(label="📥 Download Exceptions Report (Excel)", data=excel_binary_data, file_name=excel_report_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e_generate_report:
                st.error(f"Error generating or providing download for Excel report: {e_generate_report}")
                logging.exception(f"Error in Excel report generation/download setup for run_id {selected_run_id}: {e_generate_report}")
    except Exception as e_page_load:
        st.error(f"An error occurred while loading the exception details page: {e_page_load}")
        logging.exception(f"General error on exception details page: {e_page_load}")


def show_user_location_page():
    st.markdown("### 👤📍 User & Location Analysis")
    ul_history_df = db_manager.get_validation_history()
    if ul_history_df.empty: st.info("No validation runs found. Upload data to enable User & Location analysis."); return
    ul_run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in ul_history_df.iterrows()}
    ul_selected_run_display = st.selectbox("Select Validation Run for User/Location Analysis", options=list(ul_run_options.keys()), key="user_location_run_selector")
    ul_selected_run_id = ul_run_options[ul_selected_run_display] if ul_selected_run_display else None
    if not ul_selected_run_id: st.info("Please select a validation run to view analysis."); return
    ul_exceptions_df = db_manager.get_exceptions_by_run(ul_selected_run_id)
    if ul_exceptions_df.empty: st.success(f"No exceptions found for Run ID {ul_selected_run_id}. Nothing to analyze for user/location."); return
    
    st.markdown(f"#### 📊 Exceptions by User and Location (Run ID: {ul_selected_run_id})")
    ul_required_cols = ['Created user', 'Location.Name', 'Severity', 'Exception Reasons']
    for col_ul in ul_required_cols:
        if col_ul not in ul_exceptions_df.columns: 
            st.warning(f"Missing required column '{col_ul}' in exceptions data for this run. Analysis might be incomplete or inaccurate.")
            ul_exceptions_df[col_ul] = pd.NA 
    
    ul_summary_df = ul_exceptions_df.groupby(['Created user', 'Location.Name'], dropna=False).agg(
        Total_Severity_Score=('Severity', 'sum'), 
        Number_of_Exceptions=('Exception Reasons', lambda x: x.notna().sum())
    ).reset_index().rename(columns={'Created user': 'User', 'Location.Name': 'Location'})
    
    ul_metric_col1, ul_metric_col2, ul_metric_col3 = st.columns(3)
    display_metric("Total Exceptions", f"{ul_exceptions_df['Exception Reasons'].notna().sum():,}", container=ul_metric_col1)
    display_metric("Unique Users with Exceptions", f"{ul_exceptions_df['Created user'].nunique()}", container=ul_metric_col2)
    ul_avg_severity = ul_exceptions_df['Severity'].mean() if not ul_exceptions_df['Severity'].dropna().empty else 0.0
    display_metric("Average Severity", f"{ul_avg_severity:.2f}", container=ul_metric_col3)
    
    st.markdown("##### 📋 User-Location Exception Summary Table")
    st.dataframe(ul_summary_df.sort_values(by="Number_of_Exceptions", ascending=False), use_container_width=True, hide_index=True)
    
    if not ul_summary_df.empty:
        fig_ul_exc_count_chart = px.bar(ul_summary_df, x='User', y='Number_of_Exceptions', color='Location', 
                                        barmode='group', title="Exceptions by User and Location", 
                                        labels={'Number_of_Exceptions': 'Number of Exceptions'}, 
                                        color_discrete_sequence=px.colors.qualitative.Pastel) # Changed sequence
        fig_ul_exc_count_chart.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=120)
        )
        st.plotly_chart(fig_ul_exc_count_chart, use_container_width=True)
        
        fig_ul_severity_chart = px.bar(ul_summary_df, x='User', y='Total_Severity_Score', color='Location', 
                                       barmode='stack', title="Total Severity by User and Location (Stacked)", 
                                       labels={'Total_Severity_Score': 'Total Severity Score'}, 
                                       color_discrete_sequence=px.colors.qualitative.Antique) # Changed sequence
        fig_ul_severity_chart.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=120)
        )
        st.plotly_chart(fig_ul_severity_chart, use_container_width=True)
        
    ul_error_types_df = pd.DataFrame(columns=['Error Type', 'Count'])
    if 'Exception Reasons' in ul_exceptions_df.columns and ul_exceptions_df['Exception Reasons'].notna().any():
        st.markdown("##### 📊 Top 10 Common Error Types for this Run")
        ul_error_types_df = ul_exceptions_df['Exception Reasons'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        ul_error_types_df.columns = ['Error Type', 'Count']
        st.dataframe(ul_error_types_df.head(10), use_container_width=True, hide_index=True)
        
        fig_ul_top_errors = px.bar(ul_error_types_df.head(10), x='Error Type', y='Count', 
                                   title="Top 10 Error Types by Occurrence", 
                                   color='Count', color_continuous_scale=px.colors.sequential.Tealgrn, text_auto=True) # Changed scale
        fig_ul_top_errors.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=150)
        )
        st.plotly_chart(fig_ul_top_errors, use_container_width=True)
        
    ul_user_risk_df = pd.DataFrame(columns=['User', 'Total Exceptions by User', 'Average Severity Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)'])
    if 'Created user' in ul_exceptions_df.columns and ul_exceptions_df['Created user'].notna().any():
        st.markdown("##### 🎯 User Risk Analysis for this Run")
        ul_user_risk_df_agg = ul_exceptions_df.groupby('Created user', dropna=False).agg(Total_Exceptions_by_User=('Exception Reasons', lambda x: x.notna().sum()), Average_Severity_Score=('Severity', 'mean')).reset_index().rename(columns={'Created user': 'User'})
        ul_total_exceptions_in_run = ul_user_risk_df_agg['Total_Exceptions_by_User'].sum()
        ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] = (ul_user_risk_df_agg['Total_Exceptions_by_User'] / ul_total_exceptions_in_run * 100) if ul_total_exceptions_in_run > 0 else 0.0
        ul_user_risk_df_agg['Average_Severity_Score'] = ul_user_risk_df_agg['Average_Severity_Score'].fillna(0)
        ul_max_possible_severity = 10 
        ul_user_risk_df_agg['Normalized_Severity'] = (ul_user_risk_df_agg['Average_Severity_Score'] / ul_max_possible_severity).clip(0,1) * 100
        ul_user_risk_df_agg['Risk_Score (0-100)'] = (ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] * 0.6 + ul_user_risk_df_agg['Normalized_Severity'] * 0.4).round(2)
        ul_user_risk_df = ul_user_risk_df_agg[['User', 'Total_Exceptions_by_User', 'Average_Severity_Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)']]
        st.dataframe(ul_user_risk_df.sort_values(by="Risk_Score (0-100)", ascending=False), use_container_width=True, hide_index=True)
        
    ul_excel_output = io.BytesIO()
    with pd.ExcelWriter(ul_excel_output, engine='openpyxl') as ul_writer:
        ul_summary_df.to_excel(ul_writer, sheet_name='User_Location_Summary', index=False)
        ul_error_types_df.to_excel(ul_writer, sheet_name='Common_Error_Types', index=False)
        ul_user_risk_df.to_excel(ul_writer, sheet_name='User_Risk_Analysis', index=False)
    ul_excel_output.seek(0)
    if ul_excel_output.getbuffer().nbytes > 0:
        ul_report_filename_base = ul_history_df[ul_history_df['id'] == ul_selected_run_id]['filename'].iloc[0]
        st.download_button(label="📥 Download User-Location Analysis Report", data=ul_excel_output, file_name=f"User_Location_Report_Run_{ul_selected_run_id}_{ul_report_filename_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.warning("Could not generate the User-Location Analysis report for download (no data to write).")


def show_user_performance_page():
    st.markdown("### 👤📊 User Performance Dashboard")
    conn_fetch_users = None; up_users_from_exceptions = pd.DataFrame(); up_users_from_perf = pd.DataFrame()
    try:
        conn_fetch_users = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
        up_users_from_exceptions = pd.read_sql_query("SELECT DISTINCT created_user AS user FROM exceptions WHERE created_user IS NOT NULL AND created_user != ''", conn_fetch_users)
        up_users_from_perf = pd.read_sql_query("SELECT DISTINCT user FROM user_performance WHERE user IS NOT NULL AND user != ''", conn_fetch_users)
    except sqlite3.Error as e_fetch_users: logging.error(f"Error fetching user lists for performance page: {e_fetch_users}", exc_info=True)
    finally:
        if conn_fetch_users: conn_fetch_users.close()
    
    combined_users = pd.concat([up_users_from_exceptions['user'], up_users_from_perf['user']]).dropna().astype(str).str.strip()
    unique_sorted_users = sorted([user for user in combined_users.unique() if user])
    if not unique_sorted_users: st.info("No user performance data available. Upload files with 'Created user' information or ensure users have exceptions/performance records."); return
    
    selected_user_for_perf = st.selectbox("Select User to View Performance", options=[""] + unique_sorted_users, key="user_performance_selector", index=0)
    if not selected_user_for_perf: st.info("Please select a user to view their performance details."); return
    
    st.markdown(f"#### Performance Details for User: **{selected_user_for_perf}**")
    conn_user_perf_details = None; user_perf_summary_for_display_df = pd.DataFrame(); user_exceptions_history_df = pd.DataFrame()
    try:
        conn_user_perf_details = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
        user_perf_summary_for_display_df = pd.read_sql_query('''SELECT up.run_id, vr.upload_time, vr.filename, up.user, up.total_records, up.exception_records, up.exception_rate FROM user_performance up JOIN validation_runs vr ON up.run_id = vr.id WHERE up.user = ? ORDER BY vr.upload_time ASC''', conn_user_perf_details, params=(selected_user_for_perf,))
        user_exceptions_history_df = pd.read_sql_query('''SELECT e.run_id, vr.upload_time, vr.filename, e.exception_reason, e.severity FROM exceptions e JOIN validation_runs vr ON e.run_id = vr.id WHERE e.created_user = ? ORDER BY vr.upload_time DESC''', conn_user_perf_details, params=(selected_user_for_perf,))
    except sqlite3.Error as e_user_perf_fetch: logging.error(f"Error fetching performance details for user '{selected_user_for_perf}': {e_user_perf_fetch}", exc_info=True)
    finally:
        if conn_user_perf_details: conn_user_perf_details.close()
        
    if user_perf_summary_for_display_df.empty and user_exceptions_history_df.empty: st.info(f"No performance data or exception history found for user '{selected_user_for_perf}'."); return
    
    st.markdown("##### 📈 Key Performance Indicators")
    kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
    total_exceptions_for_selected_user = user_perf_summary_for_display_df['exception_records'].sum() if not user_perf_summary_for_display_df.empty else 0
    display_metric("Total Exceptions Logged", f"{total_exceptions_for_selected_user:,}", container=kpi_col1)
    avg_exc_rate_for_user = 0.0
    if not user_perf_summary_for_display_df.empty and 'exception_rate' in user_perf_summary_for_display_df.columns and user_perf_summary_for_display_df['exception_rate'].notna().any(): avg_exc_rate_for_user = user_perf_summary_for_display_df['exception_rate'].mean()
    display_metric("Avg. Exception Rate", f"{avg_exc_rate_for_user:.2f}%", container=kpi_col2)
    num_runs_involved_by_user = user_perf_summary_for_display_df['run_id'].nunique() if not user_perf_summary_for_display_df.empty else 0
    display_metric("Validation Runs Involved", f"{num_runs_involved_by_user}", container=kpi_col3)
    
    if not user_exceptions_history_df.empty and 'exception_reason' in user_exceptions_history_df.columns and user_exceptions_history_df['exception_reason'].notna().any():
        st.markdown("##### 🛠️ Common Mistake Analysis for this User")
        user_mistake_counts_df = user_exceptions_history_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        user_mistake_counts_df.columns = ['Mistake Type', 'Count']
        mistake_chart_cols = st.columns(2)
        with mistake_chart_cols[0]:
            fig_user_pie_mistakes = px.pie(user_mistake_counts_df.head(7), names='Mistake Type', values='Count', 
                                           title="Top Mistake Types Distribution", 
                                           color_discrete_sequence=px.colors.qualitative.Plotly) # Changed sequence
            fig_user_pie_mistakes.update_traces(textposition='inside', textinfo='percent+label', pull=[0.05 if i==0 else 0 for i in range(len(user_mistake_counts_df.head(7)))]) # Pull first slice
            fig_user_pie_mistakes.update_layout(
                title_x=0.5, title_font=PLOTLY_TITLE_FONT,
                font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                showlegend=True, legend_title_text='Mistake Types', height=450,
                margin=dict(t=70, b=20, l=20, r=20)
            )
            st.plotly_chart(fig_user_pie_mistakes, use_container_width=True)
        with mistake_chart_cols[1]:
            fig_user_bar_mistakes = px.bar(user_mistake_counts_df.head(10), x='Mistake Type', y='Count', 
                                           title="Top 10 Mistake Counts", 
                                           color='Count', color_continuous_scale=px.colors.sequential.Cividis_r, text_auto=True) # Changed scale
            fig_user_bar_mistakes.update_layout(
                title_x=0.5, title_font=PLOTLY_TITLE_FONT,
                font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
                height=450, margin=dict(t=70, b=170, l=40, r=20) # Adjusted margins
            )
            st.plotly_chart(fig_user_bar_mistakes, use_container_width=True)
            
        with st.expander("📋 View Detailed Mistake History for this User"):
            user_mistake_history_display_df = user_exceptions_history_df[['run_id', 'filename', 'upload_time', 'exception_reason', 'severity']].copy()
            user_mistake_history_display_df.columns = ['Run ID', 'Filename', 'Timestamp', 'Exception Details', 'Severity Score']
            user_mistake_history_display_df['Timestamp'] = pd.to_datetime(user_mistake_history_display_df['Timestamp']).dt.strftime('%Y-%m-%d %H:%M')
            st.dataframe(user_mistake_history_display_df.sort_values(by="Timestamp", ascending=False), use_container_width=True, hide_index=True)
    else: st.info("No specific mistake details (exception reasons) found for this user in their exception history.")
    
    if not user_perf_summary_for_display_df.empty and 'upload_time' in user_perf_summary_for_display_df.columns and 'exception_rate' in user_perf_summary_for_display_df.columns:
        st.markdown("##### 📉 User Exception Rate Trend Over Time")
        user_perf_summary_for_display_df['upload_time'] = pd.to_datetime(user_perf_summary_for_display_df['upload_time'])
        fig_user_exc_rate_trend = go.Figure()
        fig_user_exc_rate_trend.add_trace(go.Scatter(x=user_perf_summary_for_display_df['upload_time'], 
                                                     y=user_perf_summary_for_display_df['exception_rate'], 
                                                     mode='lines+markers', name='Exception Rate (%)', 
                                                     line=dict(color='#EF553B', width=2.5, shape='spline'), # Adjusted color
                                                     marker=dict(size=8, symbol="diamond-open"),
                                                     hovertemplate="Date: %{x|%Y-%m-%d}<br>Exception Rate: %{y:.2f}%<extra></extra>"
                                                     ))
        fig_user_exc_rate_trend.update_layout(
            title_text="User Exception Rate Trend", 
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            xaxis_title="Run Date", yaxis_title="Exception Rate (%)", yaxis_ticksuffix="%", 
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, type="date"),
            margin=dict(l=50, r=20, t=70, b=40), hovermode="x unified"
        )
        st.plotly_chart(fig_user_exc_rate_trend, use_container_width=True)
        
        with st.expander("📋 View User Performance Log (All Runs for this User)"):
            user_perf_log_display_df = user_perf_summary_for_display_df[['run_id', 'filename', 'upload_time', 'total_records', 'exception_records', 'exception_rate']].copy()
            user_perf_log_display_df.columns = ['Run ID', 'Filename', 'Timestamp', 'Total Records by User in Run', 'User Exceptions in Run', 'User Exception Rate (%)']
            user_perf_log_display_df['Timestamp'] = user_perf_log_display_df['Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
            st.dataframe(user_perf_log_display_df.sort_values(by='Timestamp', ascending=False), use_container_width=True, hide_index=True)
            
    if not user_exceptions_history_df.empty and 'exception_reason' in user_exceptions_history_df.columns and user_exceptions_history_df['exception_reason'].notna().any():
        st.markdown("##### 📚 Personalized Training Recommendations")
        user_error_series_for_training = user_exceptions_history_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts()
        training_validator_instance = DataValidator(base_ref_path="reference_data") 
        recommendations_provided = False
        for error_type_training, count_training in user_error_series_for_training.head(3).items():
            if error_type_training in training_validator_instance.training_map:
                st.markdown(f"- **Common Error ({count_training} times):** *{error_type_training}*")
                st.markdown(f"  - **Suggestion:** {training_validator_instance.training_map[error_type_training]}"); recommendations_provided = True
        if not recommendations_provided: st.markdown("No specific training recommendations available based on the top errors for this user, or errors are not mapped to training.")


def show_settings_page():
    st.markdown("### ⚙️ Application Settings")
    st.markdown("#### 🛠 Database Management")
    st.warning("🚨 **Caution:** Clearing the database is irreversible and will delete all stored validation history, exceptions, and performance data. This action cannot be undone.")
    if 'confirm_db_clear_pressed_once' not in st.session_state: st.session_state.confirm_db_clear_pressed_once = False
    if st.button("🗑️ Clear Entire Database", type="primary", key="clear_db_initial_button"): st.session_state.confirm_db_clear_pressed_once = True
    if st.session_state.confirm_db_clear_pressed_once:
        st.error("ARE YOU ABSOLUTELY SURE? This will delete all data.")
        if st.button("✅ Yes, I am sure. Delete all data.", key="clear_db_confirm_button"):
            conn_settings_clear = None
            try:
                conn_settings_clear = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                cursor_settings_clear = conn_settings_clear.cursor()
                tables_to_be_cleared = ["validation_runs", "exceptions", "department_summary", "user_performance"]
                for table_name in tables_to_be_cleared: cursor_settings_clear.execute(f"DELETE FROM {table_name}"); logging.info(f"Cleared table: {table_name}")
                conn_settings_clear.commit(); st.success("Database cleared successfully. Please refresh the application to see the changes across all pages."); logging.info("Database cleared successfully by user from settings page.")
                st.session_state.confirm_db_clear_pressed_once = False 
            except sqlite3.Error as e_settings_clear: st.error(f"Failed to clear database: {str(e_settings_clear)}"); logging.exception("Failed to clear database from settings page.");
            finally:
                if conn_settings_clear: conn_settings_clear.close()
        elif st.button("❌ No, cancel.", key="clear_db_cancel_button"): st.session_state.confirm_db_clear_pressed_once = False; st.info("Database clearing cancelled.")
    st.markdown("---")
    st.markdown("#### ℹ️ About This Dashboard")
    dashboard_version = "2.4.3" 
    st.markdown(f"""**Data Validation Dashboard - Version {dashboard_version}**\n\nThis application is designed to help users validate data from Excel files against a predefined set of business rules. It provides insights into data quality, tracks validation history, and analyzes user performance.\n\n- **Upload & Validate:** Upload your Excel files or manually enter data for instant validation.\n- **Dashboard Analytics:** Get an overview of validation statistics and department-level performance.\n- **Trends & History:** Track validation metrics over time.\n- **Exception Details:** Dive deep into specific validation runs and their exceptions.\n- **User & Location Analysis:** Analyze exceptions based on users and locations.\n- **User Performance:** Monitor individual user performance and get training recommendations.\n\nBuilt with Streamlit, Pandas, Plotly, and SQLite.""")
    st.markdown(f"**SQLite Database Path:** `{os.path.abspath(db_manager.db_path)}`")


def main():
    st.markdown("<h1>🎯 Data Validation Dashboard</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align:center;color:#718096;font-size:1.2rem;margin-bottom:2rem;'>Upload expense reports or enter data manually for validation and insights.</p>", unsafe_allow_html=True)
    
    st.sidebar.image("https://vnrseeds.com/wp-content/uploads/2022/11/logo.png", width=150) 
    st.sidebar.markdown(f"**Version**: 2.4.3") 
    st.sidebar.markdown("---")
    
    with st.sidebar:
        page_navigation_options = ["🏠 Upload & Validate", "📊 Dashboard Analytics", "📈 Trends & History", "📋 Exception Details", "👤📍 User & Location Analysis", "👤📊 User Performance", "⚙️ Settings"]
        selected_page = st.radio("Main Navigation:", page_navigation_options, label_visibility="collapsed")
    
    if selected_page == "🏠 Upload & Validate": show_upload_page()
    elif selected_page == "📊 Dashboard Analytics": show_analytics_page()
    elif selected_page == "📈 Trends & History": show_trends_page()
    elif selected_page == "📋 Exception Details": show_exception_details_page()
    elif selected_page == "👤📍 User & Location Analysis": show_user_location_page()
    elif selected_page == "👤📊 User Performance": show_user_performance_page()
    elif selected_page == "⚙️ Settings": show_settings_page()

if __name__ == "__main__":
    log_file_path = "dashboard.log"
    try:
        if os.path.exists(log_file_path) and not os.access(log_file_path, os.W_OK):
            raise IOError(f"Log file '{log_file_path}' exists but is not writable.")
        if not os.path.exists(log_file_path) and not os.access(os.path.dirname(os.path.abspath(log_file_path)) or '.', os.W_OK):
             raise IOError(f"Log file directory for '{log_file_path}' is not writable.")

        with open(log_file_path, "a") as f: 
            f.write(f"--- Log session started at {datetime.now()} ---\n") 
    except IOError as e:
        print(f"Warning: Log file '{log_file_path}' is not writable or cannot be created. Logging to console only. Error: {e}")
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[logging.StreamHandler()]
        )
    else:
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[
                logging.FileHandler(log_file_path, mode='a'), 
                logging.StreamHandler()
            ]
        )
    
    logging.info("Dashboard application started.")
    main()