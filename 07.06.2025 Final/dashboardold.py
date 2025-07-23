import streamlit as st
import pandas as pd
import sqlite3
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging

# Page configuration
st.set_page_config(
    page_title="Data Validation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    .main {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    .main > div {
        background: white;
        border-radius: 20px;
        margin: 1rem;
        padding: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    h1 {
        color: #2d3748;
        font-weight: 700;
        text-align: center;
        margin-bottom: 2rem;
        font-size: 2.5rem;
    }
    h2, h3 {
        color: #4a5568;
        font-weight: 600;
    }
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin: 0.5rem;
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.3);
        transition: transform 0.3s ease;
    }
    .metric-container:hover {
        transform: translateY(-5px);
    }
    .metric-title {
        font-size: 0.9rem;
        font-weight: 500;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    .upload-section {
        border: 3px dashed #667eea;
        border-radius: 20px;
        padding: 3rem;
        text-align: center;
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        margin: 2rem 0;
        transition: all 0.3s ease;
    }
    .upload-section:hover {
        border-color: #764ba2;
        background: linear-gradient(135deg, #edf2f7 0%, #e2e8f0 100%);
    }
    .upload-title {
        color: #2d3748;
        font-size: 1.5rem;
        font-weight: 600;
        margin-bottom: 1rem;
    }
    .upload-subtitle {
        color: #718096;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }
    .success-box {
        background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(72, 187, 120, 0.3);
    }
    .error-box {
        background: linear-gradient(135deg, #f56565 0%, #e53e3e 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(245, 101, 101, 0.3);
    }
    .warning-box {
        background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(237, 137, 54, 0.3);
    }
    .info-box {
        background: linear-gradient(135deg, #4299e1 0%, #3182ce 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(66, 153, 225, 0.3);
    }
    .css-1d391kg {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    .css-1d391kg .css-1v0mbm {
        color: white;
    }
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 3px 10px rgba(102, 126, 234, 0.3);
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
    }
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        border-radius: 10px;
        font-weight: 600;
    }
    .stProgress .st-bo {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

class DatabaseManager:
    def __init__(self, db_path="validation_dashboard.db"):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        """Initialize database tables and ensure schema is up-to-date"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
    
    # Create tables if they don't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS validation_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                total_records INTEGER,
                total_exceptions INTEGER,
                status TEXT DEFAULT 'completed',
                file_size INTEGER
        )
    ''')
    
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS exceptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER,
                department TEXT,
                sub_department TEXT,
                created_user TEXT,
                modified_user TEXT,
                exception_reason TEXT,
                severity INTEGER,
                net_amount REAL,
                location TEXT,
                crop TEXT,
                activity TEXT,
                function_name TEXT,
                vertical_name TEXT,
                region_name TEXT,
                zone_name TEXT,
                business_unit TEXT,
                FOREIGN KEY (run_id) REFERENCES validation_runs (id)
            )
        ''')
    
    # Check if function_name column exists in exceptions table
        cursor.execute("PRAGMA table_info(exceptions)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'function_name' not in columns:
            cursor.execute('''
                ALTER TABLE exceptions ADD COLUMN function_name TEXT
            ''')
            st.markdown('<div class="success-box"><strong>✅ Success!</strong> Added function_name column to exceptions table.</div>', unsafe_allow_html=True)
    
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS department_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER,
                department TEXT,
                total_records INTEGER,
                exception_records INTEGER,
                exception_rate REAL,
                FOREIGN KEY (run_id) REFERENCES validation_runs (id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_performance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER,
                user TEXT,
                total_records INTEGER,
                exception_records INTEGER,
                exception_rate REAL,
                FOREIGN KEY (run_id) REFERENCES validation_runs (id)
            )
        ''')
        
        conn.commit()
        conn.close()

    def save_validation_run(self, filename, total_records, total_exceptions, file_size):
        """Save validation run metadata"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO validation_runs (filename, total_records, total_exceptions, file_size)
            VALUES (?, ?, ?, ?)
        ''', (filename, total_records, total_exceptions, file_size))
        
        run_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return run_id

    def save_exceptions(self, run_id, exceptions_df):
        """Save exceptions data"""
        conn = sqlite3.connect(self.db_path)
        
        for _, row in exceptions_df.iterrows():
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO exceptions (
                    run_id, department, sub_department, created_user, 
                    modified_user, exception_reason, severity, net_amount, location, 
                    crop, activity, function_name, vertical_name, region_name, 
                    zone_name, business_unit
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                run_id,
                row.get('Department.Name', ''),
                row.get('Sub Department.Name', ''),
                row.get('Created user', ''),
                row.get('Modified user', ''),
                row.get('Exception Reasons', ''),
                row.get('Severity', 0),
                row.get('Net amount', 0.0),
                row.get('Location.Name', ''),
                row.get('Crop.Name', ''),
                row.get('Activity.Name', ''),
                row.get('Function.Name', ''),
                row.get('FC-Vertical.Name', ''),
                row.get('Region.Name', ''),
                row.get('Zone.Name', ''),
                row.get('Business Unit.Name', '')
            ))
        
        conn.commit()
        conn.close()

    def save_user_performance(self, run_id, df, exceptions_df):
        """Save user performance data"""
        conn = sqlite3.connect(self.db_path)
        
        if 'Created user' not in df.columns:
            st.markdown('<div class="error-box"><strong>❌ Error:</strong> Created user column missing in input file.</div>', unsafe_allow_html=True)
            conn.close()
            return
        
        valid_users = df['Created user'].dropna().str.strip()
        valid_users = valid_users[valid_users != ''].unique()
        
        invalid_users = df['Created user'].isna().sum() + (df['Created user'] == '').sum()
        st.write(f"Debug: Found {invalid_users} invalid Created user values (NaN or empty)")
        
        if len(valid_users) == 0:
            st.markdown('<div class="warning-box"><strong>⚠ Warning:</strong> No valid Created user values in input file.</div>', unsafe_allow_html=True)
            conn.close()
            return
        
        if exceptions_df.empty:
            user_stats = pd.DataFrame({'Created user': valid_users, 'exception_records': 0})
        else:
            user_stats = exceptions_df.groupby('Created user').size().reset_index(name='exception_records')
            user_stats = user_stats[user_stats['Created user'].isin(valid_users)]
        
        total_records = df[df['Created user'].isin(valid_users)].groupby('Created user').size().reset_index(name='total_records')
        user_stats = user_stats.merge(total_records, on='Created user', how='left').fillna({'total_records': 0, 'exception_records': 0})
        user_stats['exception_rate'] = (user_stats['exception_records'] / user_stats['total_records'] * 100).round(2).fillna(0)
        
        st.write(f"Debug: Saving {len(user_stats)} user performance records for run_id {run_id}")
        
        for _, row in user_stats.iterrows():
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO user_performance (run_id, user, total_records, exception_records, exception_rate)
                VALUES (?, ?, ?, ?, ?)
            ''', (run_id, row['Created user'], int(row['total_records']), int(row['exception_records']), row['exception_rate']))
        
        conn.commit()
        conn.close()

    def get_validation_history(self):
        """Get validation run history"""
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql_query('''
            SELECT * FROM validation_runs ORDER BY upload_time DESC
        ''', conn)
        conn.close()
        return df

    def get_exceptions_by_run(self, run_id):
        """Retrieve exceptions for a specific run"""
        try:
            conn = sqlite3.connect(self.db_path)
            query = "SELECT * FROM exceptions WHERE run_id = ?"
            df = pd.read_sql_query(query, conn, params=(run_id,))
            conn.close()
            
            column_mapping = {
                'id': 'id',
                'run_id': 'run_id',
                'department': 'Department.Name',
                'sub_department': 'Sub Department.Name',
                'created_user': 'Created user',
                'modified_user': 'Modified user',
                'exception_reason': 'Exception Reasons',
                'severity': 'Severity',
                'net_amount': 'Net amount',
                'location': 'Location.Name',
                'crop': 'Crop.Name',
                'activity': 'Activity.Name',
                'function_name': 'Function.Name',
                'vertical_name': 'FC-Vertical.Name',
                'region_name': 'Region.Name',
                'zone_name': 'Zone.Name',
                'business_unit': 'Business Unit.Name'
            }
            df = df.rename(columns=column_mapping)
            
            expected_columns = [
                'id', 'run_id', 'Department.Name', 'Sub Department.Name', 'Created user',
                'Modified user', 'Exception Reasons', 'Severity', 'Net amount', 'Location.Name',
                'Crop.Name', 'Activity.Name', 'Function.Name',
                'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name'
            ]
            
            if df.empty:
                logging.warning(f"get_exceptions_by_run: No exceptions found for run_id {run_id}")
                return pd.DataFrame(columns=expected_columns)
            
            for col in expected_columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df.reindex(columns=expected_columns)
            
            return df
        
        except Exception as e:
            st.markdown(
                f'<div class="error-box"><strong>❌ Error:</strong> Error retrieving exceptions for run {run_id}: {e}</div>',
                unsafe_allow_html=True
            )
            logging.error(f"get_exceptions_by_run: Error for run_id {run_id}: {e}")
            return pd.DataFrame(columns=[
                'id', 'run_id', 'Department.Name', 'Sub Department.Name', 'Created user',
                'Modified user', 'Exception Reasons', 'Severity', 'Net amount', 'Location.Name',
                'Crop.Name', 'Activity.Name', 'Function.Name',
                'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name'
            ])

    def get_user_performance(self):
        """Get user performance history"""
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql_query('''
            SELECT up.run_id, vr.upload_time, up.user, up.total_records, up.exception_records, up.exception_rate
            FROM user_performance up
            JOIN validation_runs vr ON up.run_id = vr.id
            ORDER BY vr.upload_time DESC
        ''', conn)
        conn.close()
        return df

@st.cache_resource
def get_database_manager():
    return DatabaseManager()

db_manager = get_database_manager()

class DataValidator:
    def __init__(self):
        self.no_crop_check = {
            "Finance & Account", "Human Resource", "Administration",
            "Information Technology", "Legal", "Accounts Receivable & MIS"
        }
        self.no_activity_check = self.no_crop_check.copy()
        self.no_activity_check.update({"Production", "Processing", "Parent Seed"})
        
        self.ref_files = {
            "FC_Crop": [], "VC_Crop": [], "SBFC_Region": [], "SBVC_Region": [],
            "SaleFC_Zone": [], "SaleVC_Zone": [], "FC_BU": [], "VC_BU": [],
            "Fruit_Crop": [], "Common_Crop": [], "ProductionFC_Zone": [],
            "ProductionVC_Zone": [], "SalesActivity": [], "MarketingActivity": []
        }
        
        self.training_map = {
            'Incorrect Location Name': 'Review Location Name Guidelines: Ensure locations are valid (e.g., Bandamailaram).',
            'Incorrect Activity Name': 'Complete Activity Name Training: Use only approved activities for your department.',
            'Incorrect Crop Name': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'FC-Vertical Name cannot be blank': 'Ensure Valid FC-Vertical Name.',
            'Incorrect Crop Name starting with ZZ': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'Incorrect Crop Name for FC-field crop Vertical': 'Ensure Valid Crop Name for FC-field crop.',
            'Incorrect Crop Name for VC-Veg Crop Vertical': 'Ensure Valid Crop Name for VC-Veg Crop.',
            'Incorrect Crop Name for Fruit Crop Vertical': 'Ensure Valid Crop Name for Fruit Crop.',
            'Incorrect Crop Name for Common vertical': 'Ensure Valid Crop Name for Common vertical.',
            'Incorrect Sub Department Name': 'Verify Sub-Department Standards.',
            'Incorrect Function Name': 'Check Function Name Guidelines.',
            'Need to Update Processing Location': 'Use Approved Processing Locations.',
            'Incorrect Activity Name for Lab QC': 'Use Lab QA Approved Activities.',
            'Incorrect Activity Name for Field QA': 'Use Field QA Approved Activities.',
            'Incorrect Activity Name for Bio Tech Services': 'Use Bio Tech Approved Activities.',
            'Sub Department should be blank': 'Ensure Sub-Department is blank for this department.',
            'Activity Name cannot be blank or start with ZZ': 'Complete Activity Name Training: Use only approved activities.',
            'Incorrect Activity Name for Biotech - Markers': 'Use Approved Activities for Biotech - Markers.',
            'Incorrect Activity Name for Biotech - Tissue Culture': 'Use Approved Activities for Biotech - Tissue Culture.',
            'Incorrect Activity Name for Biotech - Mutation': 'Use Approved Activities for Biotech - Mutation.',
            'Incorrect Activity Name for Entomology': 'Use Approved Activities for Entomology.',
            'Incorrect Activity Name for Pathology': 'Use Approved Activities for Pathology.',
            'Incorrect Activity Name for Bioinformatics': 'Use Approved Activities for Bioinformatics.',
            'Incorrect Activity Name for Biochemistry': 'Use Approved Activities for Biochemistry.',
            'Incorrect Activity Name for Common': 'Use Approved Activities for Common.',
            'Need to update Zone can not left Blank': 'Ensure Zone is specified.',
            'Incorrect Zone Name for FC-field crop Vertical': 'Use Approved Zone Names for FC-field crop.',
            'Incorrect Zone Name for VC-Veg Crop Vertical': 'Use Approved Zone Names for VC-Veg Crop.',
            'Need to update Zone Name can not left Blank': 'Ensure Zone is specified for Common vertical.',
            'Need to update Business Unit can not left Blank': 'Ensure Business Unit is specified.',
            'Incorrect Business Unit Name for FC-field crop Vertical': 'Use Approved Business Unit Names for FC-field crop.',
            'Incorrect Business Unit Name for VC-Veg Crop Vertical': 'Use Approved Business Unit Names for VC-Veg Crop.',
            'Need to update Region Name can not left Blank': 'Ensure Region is specified.',
            'Incorrect Region Name for FC-field crop Vertical': 'Use Approved Region Names for FC-field crop.',
            'Incorrect Region Name for VC-Veg Crop Vertical': 'Use Approved Region Names for VC-Veg Crop.',
            'Region, Zone, BU need to check for Root Stock': 'Ensure Region, Zone, and BU are blank for Root Stock.',
            'Incorrect Activity Name for Sales': 'Use Approved Activities for Sales.',
            'Incorrect Activity Name for Marketing': 'Use Approved Activities for Marketing.'
        }

    def is_not_blank(self, value):
        if pd.isna(value) or value is None:
            return False
        val = str(value).strip().replace("\u00A0", "").replace("\u200B", "")
        return val != "" and val.upper() not in ["N/A", "NULL", "NONE", "NA", "0", "-"]

    def is_blank(self, value):
        return not self.is_not_blank(value)

    def validate_row(self, dept, row):
        """Validate row with validation.py logic and dashboard.py severity scoring"""
        reasons = []
        severity = 0

        sub_dept = str(row.get("Sub Department.Name", "") or "").strip().replace("\u00A0", "").replace("\u200B", "")
        func = str(row.get("Function.Name", "") or "").strip()
        vertical = str(row.get("FC-Vertical.Name", "") or "").strip()
        loc = str(row.get("Location.Name", "") or "").strip()
        crop = str(row.get("Crop.Name", "") or "").strip()
        act = str(row.get("Activity.Name", "") or "").strip()
        region = str(row.get("Region.Name", "") or "").strip()
        zone = str(row.get("Zone.Name", "") or "").strip()
        bu = str(row.get("Business Unit.Name", "") or "").strip()

        # Generic checks
        if self.is_blank(loc) or loc.startswith("ZZ"):
            reasons.append("Incorrect Location Name")
            severity += 3
        if dept not in self.no_activity_check and dept not in ["Breeding", "Trialing & PD", "Sales", "Marketing", "Breeding Support"]:
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Incorrect Activity Name")
                severity += 3
        if dept not in self.no_crop_check:
            if self.is_blank(vertical):
                reasons.append("FC-Vertical Name cannot be blank")
                severity += 4
            if self.is_blank(crop):
                reasons.append("Incorrect Crop Name")
                severity += 4
            elif crop.startswith("ZZ"):
                reasons.append("Incorrect Crop Name starting with ZZ")
                severity += 4
            elif vertical == "FC-field crop" and crop not in self.ref_files["FC_Crop"]:
                reasons.append("Incorrect Crop Name for FC-field crop Vertical")
                severity += 4
            elif vertical == "VC-Veg Crop" and crop not in self.ref_files["VC_Crop"]:
                reasons.append("Incorrect Crop Name for VC-Veg Crop Vertical")
                severity += 4
            elif vertical == "Fruit Crop" and crop not in self.ref_files["Fruit_Crop"]:
                reasons.append("Incorrect Crop Name for Fruit Crop Vertical")
                severity += 4
            elif vertical == "Common" and crop not in self.ref_files["Common_Crop"]:
                reasons.append("Incorrect Crop Name for Common vertical")
                severity += 4

        # Department-specific checks
        if dept == "Parent Seed":
            if sub_dept not in ["Breeder Seed Production", "Foundation Seed Production", "Processing FS"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Production":
            if sub_dept not in ["Commercial Seed Production", "Seed Production Research"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if sub_dept == "Commercial Seed Production":
                if vertical == "FC-field crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                        severity += 3
                    elif zone not in self.ref_files["ProductionFC_Zone"]:
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                        severity += 3
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                        severity += 3
                    elif zone not in self.ref_files["ProductionVC_Zone"]:
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                        severity += 3
                elif vertical == "Common" and self.is_blank(zone):
                    reasons.append("Need to update Zone Name can not left Blank")
                    severity += 3

        elif dept == "Processing":
            if sub_dept not in ["Processing", "Warehousing", "Project & Maintenance"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if loc not in ["Bandamailaram", "Deorjhal", "Boriya"]:
                reasons.append("Need to Update Processing Location")
                severity += 3

        elif dept == "Quality Assurance":
            if sub_dept not in ["Field QA", "Lab QC", "Bio Tech Services"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if sub_dept == "Lab QC" and act not in ["Lab Operations QA", "All Activity"]:
                reasons.append("Incorrect Activity Name for Lab QC")
                severity += 3
            if sub_dept == "Field QA" and act not in ["Field Operations QA", "All Activity"]:
                reasons.append("Incorrect Activity Name for Field QA")
                severity += 3
            if sub_dept == "Bio Tech Services" and act not in ["Molecular", "All Activity"]:
                reasons.append("Incorrect Activity Name for Bio Tech Services")
                severity += 3

        elif dept == "Seed Tech":
            if sub_dept not in ["Aging Test", "Pelleting", "Priming", "Common"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "In Licensing & Procurement":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
                severity += 2
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
                severity += 2
            if vertical in ["", "N/A", "Common"]:
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Breeding":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
                severity += 2
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
                severity += 2
            if vertical in ["", "N/A"]:
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if dept not in self.no_activity_check and act not in ["Breeding", "All Activity", "Trialing", "Pre Breeding", "Germplasm Maintainance", "Experimental Seed Production"]:
                reasons.append("Incorrect Activity Name")
                severity += 3

        elif dept == "Breeding Support":
            if sub_dept not in ["Pathology", "Biotech - Tissue Culture", "Biotech - Mutation", "Biotech - Markers", "Bioinformatics", "Biochemistry", "Entomology", "Common"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Activity Name cannot be blank or start with ZZ")
                severity += 3
            else:
                if sub_dept == "Biotech - Markers" and act not in ["Molecular", "Grain Quality", "Seed Treatment", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Markers")
                    severity += 3
                elif sub_dept == "Biotech - Tissue Culture" and act not in ["Tissue Culture", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Tissue Culture")
                    severity += 3
                elif sub_dept == "Biotech - Mutation" and act not in ["Mutation", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Mutation")
                    severity += 3
                elif sub_dept == "Entomology" and act not in ["Entomology", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Entomology")
                    severity += 3
                elif sub_dept == "Pathology" and act not in ["Pathalogy", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Pathology")
                    severity += 3
                elif sub_dept == "Bioinformatics" and act not in ["Bioinformatics", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Bioinformatics")
                    severity += 3
                elif sub_dept == "Biochemistry" and act not in ["Biochemistry", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biochemistry")
                    severity += 3
                elif sub_dept == "Common" and act not in ["All Activity"]:
                    reasons.append("Incorrect Activity Name for Common")
                    severity += 3

        elif dept == "Trialing & PD":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
                severity += 2
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            if dept not in self.no_activity_check and act not in ["CT", "All Activity", "Trialing", "RST"]:
                reasons.append("Incorrect Activity Name")
                severity += 3

        elif dept == "Sales":
            valid_subs = ["Sales Brand", "Sales Export", "Sales Institutional & Govt"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            elif vertical == "Root Stock" and any(self.is_not_blank(x) for x in [region, zone, bu]):
                reasons.append("Region, Zone, BU need to check for Root Stock")
                severity += 3
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files["SalesActivity"]:
                reasons.append("Incorrect Activity Name for Sales")
                severity += 3
            if sub_dept == "Sales Brand":
                if vertical == "FC-field crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                        severity += 3
                    elif bu not in self.ref_files["FC_BU"]:
                        reasons.append("Incorrect Business Unit Name for FC-field crop Vertical")
                        severity += 3
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                        severity += 3
                    elif zone not in self.ref_files["SaleFC_Zone"]:
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                        severity += 3
                    if self.is_blank(region):
                        reasons.append("Need to update Region Name can not left Blank")
                        severity += 3
                    elif region not in self.ref_files["SBFC_Region"]:
                        reasons.append("Incorrect Region Name for FC-field crop Vertical")
                        severity += 3
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                        severity += 3
                    elif bu not in self.ref_files["VC_BU"]:
                        reasons.append("Incorrect Business Unit Name for VC-Veg Crop Vertical")
                        severity += 3
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                        severity += 3
                    elif zone not in self.ref_files["SaleVC_Zone"]:
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                        severity += 3
                    if self.is_blank(region):
                        reasons.append("Need to update Region Name can not left Blank")
                        severity += 3
                    elif region not in self.ref_files["SBVC_Region"]:
                        reasons.append("Incorrect Region Name for VC-Veg Crop Vertical")
                        severity += 3

        elif dept == "Marketing":
            valid_subs = ["Business Development", "Digital Marketing", "Product Management"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4
            elif vertical == "Root Stock" and any(self.is_not_blank(x) for x in [region, zone, bu]):
                reasons.append("Region, Zone, BU need to check for Root Stock")
                severity += 3
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files["MarketingActivity"]:
                reasons.append("Incorrect Activity Name for Marketing")
                severity += 3

        elif dept == "Finance & Account":
            if sub_dept not in ["Accounts", "Finance", "Analytics, Internal Control & Budget", "Purchase ops", "Secretarial", "Document Management System", "Automation", "Group Company"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Human Resource":
            if sub_dept not in ["Compliances", "HR Ops", "Recruitment", "Team Welfare", "Training"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Administration":
            if sub_dept not in ["Events", "Maintenance", "Travel Desk"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Information Technology":
            if sub_dept not in ["ERP Support", "Infra & Hardware", "Application Development"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Legal":
            if sub_dept not in ["Compliances", "Litigation"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Accounts Receivable & MIS":
            if sub_dept not in ["Branch and C&F Ops", "Commercial & AR Management", "Common", "Order Processing", "Transport & Logistic"]:
                reasons.append("Incorrect Sub Department Name")
                severity += 2
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        elif dept == "Management":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
                severity += 2
            if func != "Management":
                reasons.append("Incorrect Function Name")
                severity += 2
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
                severity += 4

        return reasons, severity

    def validate_dataframe(self, df):
        """Validate entire dataframe"""
        exceptions = []
        department_stats = {}
        
        null_like_values = [pd.NA, "N/A", "NaN", "null", "NONE", "", " ", "-", "\u00A0"]
        df['Sub Department.Name'] = df['Sub Department.Name'].replace(null_like_values, "").str.strip()
        
        input_columns = df.columns.tolist()
        
        for dept in df['Department.Name'].dropna().unique():
            dept_df = df[df['Department.Name'] == dept].copy()
            dept_exceptions = 0
            
            for _, row in dept_df.iterrows():
                reasons, severity = self.validate_row(dept, row)
                if reasons:
                    record = row.to_dict()
                    record['Exception Reasons'] = "; ".join(reasons)
                    record['Severity'] = severity
                    exceptions.append(record)
                    dept_exceptions += 1
            
            department_stats[dept] = {
                'total_records': len(dept_df),
                'exception_records': dept_exceptions,
                'exception_rate': (dept_exceptions / len(dept_df) * 100) if len(dept_df) > 0 else 0
            }
        
        if exceptions:
            exceptions_df = pd.DataFrame(exceptions)
            output_columns = input_columns + ['Exception Reasons', 'Severity']
            exceptions_df = exceptions_df.reindex(columns=output_columns, fill_value='')
        else:
            exceptions_df = pd.DataFrame(columns=input_columns + ['Exception Reasons', 'Severity'])
        
        return exceptions_df, department_stats

def display_metric(title, value, delta=None, container=None):
    """Display a custom styled metric"""
    if container:
        with container:
            if delta:
                st.markdown(f"""
                <div class="metric-container">
                    <div class="metric-title">{title}</div>
                    <div class="metric-value">{value}</div>
                    <div class="metric-delta">{delta}</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="metric-container">
                    <div class="metric-title">{title}</div>
                    <div class="metric-value">{value}</div>
                </div>
                """, unsafe_allow_html=True)

def create_excel_report(exceptions_df, dept_stats, filename):
    """Create a formatted Excel report"""
    output = io.BytesIO()
    
    try:
        expected_columns = [
            'id', 'run_id', 'Department.Name', 'Sub Department.Name', 'Created user',
            'Modified user', 'Exception Reasons', 'Severity', 'Net amount', 'Location.Name',
            'Crop.Name', 'Activity.Name', 'Function.Name',
            'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name'
        ]
        
        if exceptions_df is None or exceptions_df.empty:
            logging.warning(f"create_excel_report: exceptions_df is {exceptions_df} for filename {filename}")
            exceptions_df = pd.DataFrame(columns=expected_columns)
        else:
            for col in expected_columns:
                if col not in exceptions_df.columns:
                    exceptions_df[col] = ''
            exceptions_df = exceptions_df.reindex(columns=expected_columns)
            
            exceptions_df = exceptions_df.astype({
                col: str for col in exceptions_df.columns if col not in ['Net amount', 'Severity']
            })
            exceptions_df['Net amount'] = pd.to_numeric(exceptions_df['Net amount'], errors='coerce')
            exceptions_df['Severity'] = pd.to_numeric(exceptions_df['Severity'], errors='coerce')
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            exceptions_df.to_excel(writer, sheet_name='Exceptions', index=False)
            
            dept_summary_df = pd.DataFrame([
                {
                    'Department': dept,
                    'Total Records': stats.get('total_records', 0),
                    'Exception Records': stats.get('exception_records', 0),
                    'Exception Rate (%)': round(stats.get('exception_rate', 0), 2)
                }
                for dept, stats in dept_stats.items()
            ])
            if dept_summary_df.empty:
                logging.warning(f"create_excel_report: dept_stats is empty for filename {filename}")
                dept_summary_df = pd.DataFrame(columns=['Department', 'Total Records', 'Exception Records', 'Exception Rate (%)'])
            dept_summary_df.to_excel(writer, sheet_name='Department Summary', index=False)
            
            workbook = writer.book
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                if worksheet.max_row <= 1:
                    logging.info(f"Sheet '{sheet_name}' has no data rows; applying headers only")
                
                headers = worksheet[1]
                headers_list = [cell.value for cell in headers if cell.value is not None]
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                
                for col_idx, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                for row_idx in range(2, worksheet.max_row + 1):
                    for col_idx, col_name in enumerate(headers_list, start=1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = border
                        if col_name == "Net amount":
                            cell.number_format = '#,##0.00'
                        elif col_name == "Exception Rate (%)":
                            cell.number_format = '0.00'
                
                for col_idx, column in enumerate(worksheet.columns, start=1):
                    max_length = len(str(headers_list[col_idx - 1] or ""))
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        try:
                            if cell.value is not None:
                                cell_str = str(cell.value)
                                max_length = max(max_length, len(cell_str))
                        except (TypeError, UnicodeEncodeError) as e:
                            logging.warning(f"Error processing cell value in sheet '{sheet_name}', column {column_letter}: {e}")
                            continue
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        if output.getvalue() == b"":
            logging.error(f"create_excel_report: Output is empty for filename {filename}")
            st.markdown('<div class="error-box"><strong>❌ Error:</strong> Failed to generate Excel report; output is empty.</div>', unsafe_allow_html=True)
            return None
        
        return output
    
    except Exception as e:
        logging.error(f"create_excel_report: Error generating Excel report for filename {filename}: {e}")
        st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error generating Excel report: {e}</div>', unsafe_allow_html=True)
        return None

def process_uploaded_file(uploaded_file):
    """Process uploaded file and run validation"""
    try:
        with st.spinner("📖 Reading file..."):
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.strip()
        
        required_columns = [
            'Department.Name', 'Sub Department.Name', 'Created user', 'Modified user',
            'Net amount', 'Location.Name', 'Crop.Name', 'Activity.Name', 'Function.Name',
            'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.markdown(
                f'<div class="error-box"><strong>❌ Error!</strong> Missing required columns: {", ".join(missing_columns)}</div>',
                unsafe_allow_html=True
            )
            return

        st.markdown("#### 📊 File Information")
        col1, col2, col3, col4 = st.columns(4)
        display_metric("Total Records", f"{len(df):,}", container=col1)
        display_metric("Total Columns", len(df.columns), container=col2)
        display_metric("File Size", f"{uploaded_file.size / 1024:.1f} KB", container=col3)
        display_metric("Departments", df['Department.Name'].nunique() if 'Department.Name' in df.columns else "N/A", container=col4)
        
        with st.spinner("🔍 Validating data..."):
            validator = DataValidator()
            exceptions_df, dept_stats = validator.validate_dataframe(df)

        run_id = db_manager.save_validation_run(
            filename=uploaded_file.name,
            total_records=len(df),
            total_exceptions=len(exceptions_df),
            file_size=uploaded_file.size
        )

        if not exceptions_df.empty:
            db_manager.save_exceptions(run_id, exceptions_df)
        
        db_manager.save_user_performance(run_id, df, exceptions_df if not exceptions_df.empty else pd.DataFrame(columns=['Created user']))
        
        conn = sqlite3.connect(db_manager.db_path)
        for dept, stats in dept_stats.items():
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO department_summary (run_id, department, total_records, exception_records, exception_rate)
                VALUES (?, ?, ?, ?, ?)
            ''', (run_id, dept, stats['total_records'], stats['exception_records'], stats['exception_rate']))
        conn.commit()
        conn.close()

        st.markdown("#### 🛠 Validation Results")
        if exceptions_df.empty:
            st.markdown(
                '<div class="success-box"><strong>✅ Perfect!</strong> No validation issues found!</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                f'<div class="warning-box"><strong>⚠ Warning!</strong> Found {len(exceptions_df)} records with validation issues</div>',
                unsafe_allow_html=True
            )
            
            col1, col2, col3 = st.columns(3)
            display_metric("Total Exceptions", f"{len(exceptions_df):,}", container=col1)
            display_metric("Exception Rate", f"{(len(exceptions_df)/len(df)*100):.2f}%", container=col2)
            display_metric("Average Severity", f"{exceptions_df['Severity'].mean():.2f}", container=col3)
            
            st.markdown("##### 📋 Exception Records")
            st.dataframe(exceptions_df, use_container_width=True)
            
            excel_data = create_excel_report(exceptions_df, dept_stats, uploaded_file.name)
            st.download_button(
                label="📥 Download Validation Report",
                data=excel_data,
                file_name=f"Validation_Report_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.markdown(
            f'<div class="error-box"><strong>❌ Error!</strong> Failed to process file: {str(e)}</div>',
            unsafe_allow_html=True
        )

def show_upload_page():
    st.markdown("### 📁 File Upload & Validation")
    
    st.markdown("""
    <div class="upload-section">
        <div class="upload-title">📤 Drag and Drop Your Excel Files</div>
        <div class="upload-subtitle">Supported formats: .xlsx, .xls | Maximum file size: 200MB</div>
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader(
        "Choose Excel files",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="Upload your VNR Seeds expense report files for validation",
        label_visibility="collapsed"
    )
    
    if uploaded_files:
        st.markdown(f'<div class="success-box"><strong>✅ Success!</strong> {len(uploaded_files)} file(s) uploaded successfully!</div>', unsafe_allow_html=True)
        for uploaded_file in uploaded_files:
            with st.expander(f"📄 Process: {uploaded_file.name}", expanded=True):
                process_uploaded_file(uploaded_file)
    
    st.markdown("### 📝 Manual Data Entry")
    with st.form("data_entry_form"):
        department = st.selectbox("Department", options=[
            "Parent Seed", "Production", "Processing", "Quality Assurance",
            "Finance & Account", "Human Resource", "Administration", "Information Technology",
            "Legal", "Accounts Receivable & MIS", "Seed Tech", "In Licensing & Procurement",
            "Breeding", "Breeding Support", "Trialing & PD", "Sales", "Marketing", "Management"
        ])
        location = st.text_input("Location Name")
        activity = st.text_input("Activity Name")
        created_user = st.text_input("Created User")
        net_amount = st.number_input("Net Amount", min_value=0.0)
        sub_dept = st.text_input("Sub Department Name", value="")
        modified_user = st.text_input("Modified User", value="")
        crop = st.text_input("Crop Name", value="")
        function = st.text_input("Function Name", value="")
        vertical = st.text_input("FC-Vertical Name", value="")
        region = st.text_input("Region Name", value="")
        zone = st.text_input("Zone Name", value="")
        business_unit = st.text_input("Business Unit Name", value="")
        
        submit = st.form_submit_button("Submit Record")
        
        if submit:
            validator = DataValidator()
            row = {
                'Department.Name': department,
                'Location.Name': location,
                'Activity.Name': activity,
                'Created user': created_user,
                'Net amount': net_amount,
                'Sub Department.Name': sub_dept,
                'Modified user': modified_user,
                'Crop.Name': crop,
                'Function.Name': function,
                'FC-Vertical.Name': vertical,
                'Region.Name': region,
                'Zone.Name': zone,
                'Business Unit.Name': business_unit
            }
            reasons, severity = validator.validate_row(department, row)
            if reasons:
                st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Invalid data: {"; ".join(reasons)} (Severity: {severity})</div>', unsafe_allow_html=True)
            else:
                run_id = db_manager.save_validation_run(
                    filename=f"Manual_Entry_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                    total_records=1,
                    total_exceptions=0,
                    file_size=0
                )
                st.markdown('<div class="success-box"><strong>✅ Success!</strong> Record added successfully.</div>', unsafe_allow_html=True)

def show_analytics_page():
    """Display analytics dashboard with visualizations"""
    st.markdown("### 📊 Dashboard Analytics")
    
    history_df = db_manager.get_validation_history()
    
    if history_df.empty:
        st.markdown('<div class="info-box"><strong>ℹ Info:</strong> No validation runs found.</div>', unsafe_allow_html=True)
        return
    
    st.markdown("#### 📈 Overall Statistics")
    col1, col2, col3 = st.columns(3)
    display_metric("Total Validation Runs", len(history_df), container=col1)
    display_metric("Total Records Processed", f"{history_df['total_records'].sum():,}", container=col2)
    display_metric("Total Exceptions", f"{history_df['total_exceptions'].sum():,}", container=col3)
    
    latest_run_id = history_df['id'].max()
    conn = sqlite3.connect(db_manager.db_path)
    dept_summary_df = pd.read_sql_query('''
        SELECT department, total_records, exception_records, exception_rate
        FROM department_summary
        WHERE run_id = ?
    ''', conn, params=(latest_run_id,))
    conn.close()
    
    if not dept_summary_df.empty:
        st.markdown("#### 🏭 Department Analysis")
        fig = px.bar(
            dept_summary_df,
            x='department',
            y='exception_rate',
            title="Exception Rate by Department",
            labels={'exception_rate': 'Exception Rate (%)', 'department': 'Department'},
            color='exception_rate',
            color_continuous_scale='Reds'
        )
        fig.update_layout(
            xaxis_title="Department",
            yaxis_title="Exception Rate (%)",
            showlegend=False,
            margin=dict(l=20, r=20, t=50, b=20)
        )
        st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("##### 📋 Department Summary")
        st.dataframe(dept_summary_df, use_container_width=True)

def show_trends_page():
    """Display trends and historical analysis"""
    st.markdown("### 📈 Trends & History")
    
    history_df = db_manager.get_validation_history()
    
    if history_df.empty:
        st.markdown('<div class="info-box"><strong>ℹ Info:</strong> No historical data available.</div>', unsafe_allow_html=True)
        return
    
    history_df['upload_time'] = pd.to_datetime(history_df['upload_time'])
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=history_df['upload_time'],
        y=history_df['total_exceptions'],
        mode='lines+markers',
        name='Exceptions',
        line=dict(color='#667eea')
    ))
    fig.add_trace(go.Scatter(
        x=history_df['upload_time'],
        y=history_df['total_records'],
        mode='lines+markers',
        name='Total Records',
        line=dict(color='#48bb78')
    ))
    fig.update_layout(
        title="Validation Trends Over Time",
        xaxis_title="Upload Date",
        yaxis_title="Count",
        showlegend=True,
        margin=dict(l=20, r=20, t=50, b=20)
    )
    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("##### 📜 Validation History")
    st.dataframe(
        history_df[['id', 'filename', 'upload_time', 'total_records', 'total_exceptions', 'file_size']],
        use_container_width=True
    )

def show_exception_details_page():
    """Display detailed exceptions for a specific validation run"""
    st.header("🔍 Exception Details")
    db_manager = DatabaseManager()
    
    try:
        validation_runs = db_manager.get_validation_history()
        if validation_runs.empty:
            st.markdown('<div class="info-box"><strong>ℹ️ Info:</strong> No validation runs found.</div>', unsafe_allow_html=True)
            return
        
        run_options = [(row['id'], row['filename'], row['upload_time']) for _, row in validation_runs.iterrows()]
        selected_run = st.selectbox(
            "Select Validation Run",
            run_options,
            format_func=lambda x: f"Run {x[0]}: {x[1]} ({x[2]})"
        )
        
        if selected_run:
            run_id = selected_run[0]
            exceptions_df = db_manager.get_exceptions_by_run(run_id)
            
            st.markdown("### Debug: Exceptions DataFrame")
            st.write(f"Empty: {exceptions_df is None or exceptions_df.empty}")
            st.write(f"Columns: {exceptions_df.columns.tolist()}")
            st.write(f"Shape: {exceptions_df.shape}")
            if not exceptions_df.empty:
                st.write(f"Preview: {exceptions_df.head().to_dict()}")
            else:
                st.markdown('<div class="warning-box"><strong>⚠ Warning:</strong> No exceptions found for this run. Check database or upload a new file.</div>', unsafe_allow_html=True)
            
            st.markdown("### Exceptions Found")
            if not exceptions_df.empty:
                st.dataframe(exceptions_df, use_container_width=True)
                
                # Severity distribution
                st.markdown("#### 📊 Severity Distribution")
                severity_counts = exceptions_df['Severity'].value_counts().reset_index()
                severity_counts.columns = ['Severity', 'Count']
                fig = px.bar(
                    severity_counts,
                    x='Severity',
                    y='Count',
                    title='Exception Severity Distribution',
                    color='Severity',
                    color_continuous_scale='Reds'
                )
                fig.update_layout(
                    xaxis_title="Severity Score",
                    yaxis_title="Number of Exceptions",
                    showlegend=False
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.markdown('<div class="info-box"><strong>ℹ️ Success!</strong> No exceptions found for this run.</div>', unsafe_allow_html=True)
                return
            
            conn = sqlite3.connect(db_manager.db_path)
            dept_stats_df = pd.read_sql_query(
                "SELECT department, total_records, exception_records, exception_rate FROM department_summary WHERE run_id = ?",
                conn,
                params=(run_id,)
            )
            conn.close()
            
            dept_stats = {
                row['department']: {
                    'total_records': row['total_records'],
                    'exception_records': row['exception_records'],
                    'exception_rate': row['exception_rate']
                }
                for _, row in dept_stats_df.iterrows()
            }
            
            st.write(f"Department Stats: {dept_stats}")
            
            try:
                excel_data = create_excel_report(exceptions_df, dept_stats, f"run_{run_id}.xlsx")
                if excel_data is None:
                    st.markdown('<div class="error-box"><strong>❌ Error:</strong> Failed to generate the Exceptions Report. Check dashboard.log for details.</div>', unsafe_allow_html=True)
                    logging.error(f"show_exception_details_page: excel_data is None for run_id {run_id}")
                else:
                    st.download_button(
                        label="📥 Download Exceptions Report",
                        data=excel_data,
                        file_name=f"Exceptions_Report_Run_{run_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error generating report: {e}</div>', unsafe_allow_html=True)
                logging.error(f"show_exception_details_page: Error in create_excel_report for run_id {run_id}: {e}")
    
    except Exception as e:
        st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error loading exception details: {e}</div>', unsafe_allow_html=True)
        logging.error(f"show_exception_details_page: General error: {e}")

def show_user_location_page():
    """Display exceptions by user and location"""
    st.markdown("### 👤📍 User & Location Analysis")
    
    history_df = db_manager.get_validation_history()
    
    if history_df.empty:
        st.markdown('<div class="info-box"><strong>ℹ Info:</strong> No validation runs found.</div>', unsafe_allow_html=True)
        return
    
    run_id = st.selectbox(
        "Select Validation Run",
        options=history_df['id'],
        format_func=lambda x: f"Run {x}: {history_df[history_df['id'] == x]['filename'].iloc[0]} ({history_df[history_df['id'] == x]['upload_time'].iloc[0]})",
        key="user_location_run_select"
    )
    
    exceptions_df = db_manager.get_exceptions_by_run(run_id)
    
    if exceptions_df.empty:
        st.markdown('<div class="success-box"><strong>✅ Success!</strong> No exceptions found for this run.</div>', unsafe_allow_html=True)
        return
    
    st.markdown("#### 📊 Exceptions by User and Location")
    
    user_location_summary = exceptions_df.groupby(['Created user', 'Location.Name']).agg({
        'Severity': 'sum',
        'Exception Reasons': 'count'
    }).reset_index()
    user_location_summary = user_location_summary.rename(columns={
        'Created user': 'User',
        'Location.Name': 'Location',
        'Exception Reasons': 'Exception Count',
        'Severity': 'Total Severity'
    })
    
    col1, col2, col3 = st.columns(3)
    display_metric("Total Exceptions", f"{len(exceptions_df):,}", container=col1)
    display_metric("Unique Users", f"{exceptions_df['Created user'].nunique()}", container=col2)
    display_metric("Average Severity", f"{exceptions_df['Severity'].mean():.2f}", container=col3)
    
    st.markdown("##### 📋 User-Location Exception Summary")
    st.dataframe(user_location_summary, use_container_width=True)
    
    fig = px.bar(
        user_location_summary,
        x='User',
        y='Exception Count',
        color='Location',
        title="Exceptions by User and Location",
        labels={'Exception Count': 'Number of Exceptions', 'User': 'User'},
        color_discrete_sequence=px.colors.qualitative.Set3
    )
    fig.update_layout(
        xaxis_title="User",
        yaxis_title="Number of Exceptions",
        showlegend=True,
        margin=dict(l=20, r=20, t=50, b=20),
        xaxis_tickangle=45
    )
    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("##### 📊 Severity by User")
    fig_severity = px.bar(
        user_location_summary,
        x='User',
        y='Total Severity',
        color='Location',
        title="Total Severity by User and Location",
        labels={'Total Severity': 'Severity Score', 'User': 'User'},
        color_discrete_sequence=px.colors.qualitative.Set3
    )
    fig_severity.update_layout(
        xaxis_title="User",
        yaxis_title="Severity Score",
        showlegend=True,
        margin=dict(l=20, r=20, t=50, b=20),
        xaxis_tickangle=45
    )
    st.plotly_chart(fig_severity, use_container_width=True)
    
    st.markdown("##### 📊 Common Error Types")
    error_types = exceptions_df['Exception Reasons'].str.split('; ', expand=True).stack().value_counts().reset_index()
    error_types.columns = ['Error Type', 'Count']
    st.dataframe(error_types, use_container_width=True)
    
    fig_errors = px.bar(
        error_types.head(10),
        x='Error Type',
        y='Count',
        title="Top 10 Error Types",
        labels={'Count': 'Number of Occurrences', 'Error Type': 'Error Type'},
        color='Count',
        color_continuous_scale='Blues'
    )
    fig_errors.update_layout(
        xaxis_title="Error Type",
        yaxis_title="Number of Occurrences",
        showlegend=False,
        margin=dict(l=20, r=20, t=50, b=20),
        xaxisangle=45
    )
    st.plotly_chart(fig_errors, use_container_width=True)
    
    st.markdown("##### 📊 User Risk Scores")
    user_risk_df = exceptions_df.groupby('Created user').agg({
        'Exception Reasons': 'count',
        'Severity': 'mean'
    }).reset_index()
    user_risk_df['avg_exception_rate'] = (user_risk_df['Exception Reasons'] / user_risk_df['Exception Reasons'].sum() * 100).round(2)
    user_risk_df['risk_score'] = (user_risk_df['avg_exception_rate'] * 0.5 + user_risk_df['Severity'] * 0.5).round(2)
    user_risk_df = user_risk_df.rename(columns={'Created user': 'User', 'Exception Reasons': 'Total Exceptions', 'Severity': 'Avg Severity'})
    st.dataframe(user_risk_df[['User', 'Total Exceptions', 'Avg Severity', 'avg_exception_rate', 'risk_score']], use_container_width=True)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        user_location_summary.to_excel(writer, sheet_name='User_Location_Summary', index=False)
        error_types.to_excel(writer, sheet_name='Error_Types', index=False)
        user_risk_df.to_excel(writer, sheet_name='User_Risk_Scores', index=False)
        
        workbook = writer.book
        for sheet_name in workbook.sheet_names:
            worksheet = workbook[sheet_name]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="667ee", end_color="667ee", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    st.download_button(
        label="📥 Download User-Location Report",
        data=output,
        file_name=f"User_Location_Report_Run_{run_id}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def show_user_performance_page():
    """Display user performance history with enhanced visualizations"""
    st.markdown("### 👤📊 User Performance")
    
    history_df = db_manager.get_validation_history()
    
    if history_df.empty:
        st.markdown('<div class="info-box"><strong>ℹ Info:</strong> No validation runs found.</div>', unsafe_allow_html=True)
        return
    
    # Get unique users from exceptions table
    conn = sqlite3.connect(db_manager.db_path)
    users_df = pd.read_sql_query('''
        SELECT DISTINCT created_user AS user
        FROM exceptions
        WHERE created_user IS NOT NULL AND created_user != ''
        UNION
        SELECT DISTINCT user
        FROM user_performance
        WHERE user IS NOT NULL AND user != ''
    ''', conn)
    conn.close()
    
    if users_df.empty:
        st.markdown(
            '<div class="info-box"><strong>ℹ Info:</strong> No user performance data available. Try uploading a file with valid Created user values.</div>',
            unsafe_allow_html=True
        )
        return
    
    user = st.selectbox("Select User", options=users_df['user'].sort_values(), key="user_select")
    
    # Fetch user performance data
    conn = sqlite3.connect(db_manager.db_path)
    user_perf_df = pd.read_sql_query('''
        SELECT up.run_id, vr.upload_time, up.user, up.total_records, up.exception_records, up.exception_rate
        FROM user_performance up
        JOIN validation_runs vr ON up.run_id = vr.id
        WHERE up.user = ?
        ORDER BY vr.upload_time DESC
    ''', conn, params=(user,))
    
    # Fetch exception details
    exceptions_df = pd.read_sql_query('''
        SELECT run_id, exception_reason, severity, vr.upload_time
        FROM exceptions e
        JOIN validation_runs vr ON e.run_id = vr.id
        WHERE created_user = ?
    ''', conn, params=(user,))
    conn.close()
    
    if user_perf_df.empty and exceptions_df.empty:
        st.markdown(f'<div class="info-box"><strong>ℹ Info:</strong> No data for user {user}.</div>', unsafe_allow_html=True)
        return
    
    # Display metrics
    st.markdown("#### 📈 User Metrics")
    col1, col2, col3 = st.columns(3)
    display_metric("Total Exceptions", f"{user_perf_df['exception_records'].sum():,}" if not user_perf_df.empty else "0", container=col1)
    display_metric("Average Exception Rate", f"{user_perf_df['exception_rate'].mean():.2f}%" if not user_perf_df.empty else "0.00%", container=col2)
    display_metric("Validation Runs", f"{len(user_perf_df)}" if not user_perf_df.empty else "0", container=col3)
    
    # Mistake breakdown
    if not exceptions_df.empty:
        st.markdown("#### 📊 Mistake Breakdown")
        
        # Process exception reasons
        mistake_counts = exceptions_df['exception_reason'].str.split('; ', expand=True).stack().value_counts().reset_index()
        mistake_counts.columns = ['Mistake Type', 'Count']
        
        # Pie chart
        fig_pie = px.pie(
            mistake_counts,
            names='Mistake Type',
            values='Count',
            title=f"Distribution of Mistake Types for {user}",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        fig_pie.update_traces(textposition='inside', textinfo='percent+label')
        fig_pie.update_layout(showlegend=True, margin=dict(l=20, r=20, t=50, b=20))
        st.plotly_chart(fig_pie, use_container_width=True)
        
        # Bar chart
        fig_bar = px.bar(
            mistake_counts,
            x='Mistake Type',
            y='Count',
            title=f"Mistake Counts for {user}",
            labels={'Count': 'Number of Occurrences', 'Mistake Type': 'Mistake Type'},
            color='Count',
            color_continuous_scale='Blues'
        )
        fig_bar.update_layout(
            xaxis_title="Mistake Type",
            yaxis_title="Number of Occurrences",
            xaxis_tickangle=45,
            margin=dict(l=20, r=20, t=50, b=20)
        )
        st.plotly_chart(fig_bar, use_container_width=True)
        
        # Detailed table
        st.markdown("##### 📋 Detailed Mistake History")
        mistake_details = exceptions_df[['run_id', 'upload_time', 'exception_reason', 'severity']].copy()
        mistake_details['severity'] = mistake_details['severity'].map({1: 'Low', 2: 'Medium', 3: 'High'})
        st.dataframe(mistake_details, use_container_width=True)
    
    # Exception rate trend
    if not user_perf_df.empty:
        st.markdown("#### 📈 Exception Rate Trend")
        fig_trend = go.Figure()
        fig_trend.add_trace(go.Scatter(
            x=user_perf_df['upload_time'],
            y=user_perf_df['exception_rate'],
            mode='lines+markers',
            name='Exception Rate',
            line=dict(color='#667eea')
        ))
        fig_trend.update_layout(
            title=f"Exception Rate Trend for {user}",
            xaxis_title="Run Date",
            yaxis_title="Exception Rate (%)",
            margin=dict(l=20, r=20, t=50, b=20)
        )
        st.plotly_chart(fig_trend, use_container_width=True)
        
        st.markdown("##### 📋 User Performance History")
        st.dataframe(user_perf_df[['run_id', 'upload_time', 'total_records', 'exception_records', 'exception_rate']], use_container_width=True)
    
    # Training recommendations
    if not exceptions_df.empty:
        st.markdown("#### 📚 Training Recommendations")
        user_errors = exceptions_df['exception_reason'].str.split('; ', expand=True).stack().value_counts()
        validator = DataValidator()
        for error_type, _ in user_errors.head(3).items():
            if error_type in validator.training_map:
                st.markdown(f"- **{error_type}**: {validator.training_map[error_type]}")


def show_settings_page():
    """Display settings page"""
    st.markdown("### ⚙️ Settings")
    
    st.markdown("#### 🛠 Database Management")
    if st.button("🗑 Clear Database"):
        try:
            conn = sqlite3.connect(db_manager.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM validation_runs")
            cursor.execute("DELETE FROM exceptions")
            cursor.execute("DELETE FROM department_summary")
            cursor.execute("DELETE FROM user_performance")
            conn.commit()
            conn.close()
            st.markdown('<div class="success-box"><strong>✅ Success!</strong> Database cleared successfully.</div>', unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Failed to clear database: {str(e)}</div>', unsafe_allow_html=True)
    
    st.markdown("#### ℹ About")
    st.markdown

def main():
    st.markdown("<h1>🎯 Data Validation Dashboard</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #718096; font-size: 1.2rem; margin-bottom: 2rem;'>Upload your expense reports or enter data manually for validation and insights.</p>", unsafe_allow_html=True)
    
    st.sidebar.markdown(f"**Dashboard Version**: 2.3 (Updated June 03, 2025)")
    
    with st.sidebar:
        st.markdown("### 📋 Navigation")
        page = st.selectbox("Choose Page", [
            "🏠 Upload & Validate",
            "📊 Dashboard Analytics",
            "📈 Trends & History",
            "📋 Exception Details",
            "👤📍 User & Location Analysis",
            "👤📊 User Performance",
            "⚙️ Settings"
        ])
    
    if page == "🏠 Upload & Validate":
        show_upload_page()
    elif page == "📊 Dashboard Analytics":
        show_analytics_page()
    elif page == "📈 Trends & History":
        show_trends_page()
    elif page == "📋 Exception Details":
        show_exception_details_page()
    elif page == "👤📍 User & Location Analysis":
        show_user_location_page()
    elif page == "👤📊 User Performance":
        show_user_performance_page()
    elif page == "⚙️ Settings":
        show_settings_page()

if __name__ == "__main__":
    main()
