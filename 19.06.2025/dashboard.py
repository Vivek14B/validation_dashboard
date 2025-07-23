import streamlit as st
import pandas as pd
import sqlite3
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import logging
import os
# Add these with your other imports near the top of the file
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import io
import json
import concurrent.futures # Added for parallel processing

# Helper function to serialize objects not recognized by default json.dumps
def json_serializer_default(obj):
    if isinstance(obj, datetime):
        if hasattr(obj, 'tzinfo') and obj.tzinfo is not None and obj.tzinfo.utcoffset(obj) is not None:
            if isinstance(obj, pd.Timestamp) and pd.isna(obj):
                 return None
            return obj.astimezone(pd.timezone('UTC')).isoformat()
        else:
            if isinstance(obj, pd.Timestamp) and pd.isna(obj):
                 return None
            return obj.isoformat()
    elif isinstance(obj, np.datetime64):
        if np.isnat(obj):
            return None
        return pd.Timestamp(obj).isoformat()
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        if np.isnan(obj) or np.isinf(obj):
            return str(obj)
        return float(obj)
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, pd.Series):
        return obj.to_dict()
    elif isinstance(obj, (pd.Interval, pd.Period, pd.Categorical)):
        return str(obj)
    if hasattr(obj, 'isoformat') and callable(getattr(obj, 'isoformat')):
        try:
            return obj.isoformat()
        except Exception:
            pass
    try:
        return str(obj)
    except Exception:
        return f"Unserializable object: {type(obj).__name__}"

# OPTIMIZATION: This helper function will be run in parallel by each CPU core.
# It takes a chunk of the DataFrame and applies the validation logic.
def _validate_chunk(validator_instance, df_chunk):
    """
    Helper function to be executed in parallel.
    Validates a small DataFrame chunk.
    """
    exceptions = []
    if df_chunk.empty:
        return exceptions

    for index, row in df_chunk.iterrows():
        dept = str(row.get('Department.Name', ''))
        if not dept:
            continue
        reasons, severity = validator_instance.validate_row(dept, row)
        if reasons:
            record = row.to_dict()
            record['Exception Reasons'] = "; ".join(reasons)
            record['Severity'] = severity
            exceptions.append(record)
    return exceptions


# Page configuration
st.set_page_config(
    page_title="Data Validation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    body { /* Apply Inter to the whole body for wider reach */
        font-family: 'Inter', sans-serif;
    }
    .main {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    .main > div { /* This targets the main content blocks */
        background: white;
        border-radius: 20px;
        margin: 1rem;
        padding: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    h1 {
        color: #2d3748;
        font-weight: 700;
        text-align: center;
        margin-bottom: 1rem; /* Adjusted for underline */
        font-size: 2.5rem;
        position: relative;
        padding-bottom: 0.5rem;
    }
    h1::after { /* Graphical underline for main title */
        content: '';
        display: block;
        width: 100px;
        height: 4px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        margin: 0.5rem auto 1rem auto;
        border-radius: 2px;
    }
    h2, h3 {
        color: #4a5568;
        font-weight: 600;
    }
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin: 0.5rem;
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.3);
        transition: transform 0.3s ease;
    }
    .metric-container:hover {
        transform: translateY(-5px);
    }
    .metric-title {
        font-size: 0.9rem;
        font-weight: 500;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    .upload-section {
        border: 3px dashed #667eea;
        border-radius: 20px;
        padding: 3rem;
        text-align: center;
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        margin: 2rem 0;
        transition: all 0.3s ease;
    }
    .upload-section:hover {
        border-color: #764ba2;
        background: linear-gradient(135deg, #edf2f7 0%, #e2e8f0 100%);
    }
    .upload-title {
        color: #2d3748;
        font-size: 1.5rem;
        font-weight: 600;
        margin-bottom: 1rem;
    }
    .upload-subtitle {
        color: #718096;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }
    .success-box {
        background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(72, 187, 120, 0.3);
    }
    .error-box {
        background: linear-gradient(135deg, #f56565 0%, #e53e3e 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(245, 101, 101, 0.3);
    }
    .warning-box {
        background: linear-gradient(135deg, #ed8936 0%, #dd6b20 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(237, 137, 54, 0.3);
    }
    .info-box {
        background: linear-gradient(135deg, #4299e1 0%, #3182ce 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(66, 153, 225, 0.3);
    }
    .css-1d391kg {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }
    .css-1d391kg .css-1v0mbm {
        color: white;
    }
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 3px 10px rgba(102, 126, 234, 0.3);
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
    }
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
        border-radius: 10px;
        font-weight: 600;
    }
    .stProgress .st-bo {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

DB_TIMEOUT = 30

# Font dictionary for Plotly charts
PLOTLY_FONT = dict(family="Inter, sans-serif", size=12, color="#2d3748")
PLOTLY_TITLE_FONT = dict(family="Inter, sans-serif", size=16, color="#2d3748")


class DatabaseManager:
    def __init__(self, db_path="validation_dashboard.db"):
        self.db_path = db_path
        self.init_database()

    def init_database(self):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS validation_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT NOT NULL,
                    upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP, total_records INTEGER,
                    total_exceptions INTEGER, status TEXT DEFAULT 'completed',
                    file_size INTEGER, excel_report_data BLOB)''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS exceptions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, department TEXT,
                    sub_department TEXT, created_user TEXT, modified_user TEXT,
                    exception_reason TEXT, severity INTEGER, net_amount REAL, location TEXT,
                    crop TEXT, activity TEXT, function_name TEXT, vertical_name TEXT,
                    region_name TEXT, zone_name TEXT, business_unit TEXT,
                    account2_code TEXT, sub_ledger_code TEXT,
                    original_row_data TEXT,
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            
            cursor.execute("PRAGMA table_info(validation_runs)")
            run_columns = [info[1] for info in cursor.fetchall()]
            if 'excel_report_data' not in run_columns:
                cursor.execute('ALTER TABLE validation_runs ADD COLUMN excel_report_data BLOB')
                logging.info("Added excel_report_data column to validation_runs table.")

            cursor.execute("PRAGMA table_info(exceptions)")
            columns = [info[1] for info in cursor.fetchall()]
            if 'function_name' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN function_name TEXT')
                logging.info("Added function_name column to exceptions table.")
            if 'original_row_data' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN original_row_data TEXT')
                logging.info("Added original_row_data column to exceptions table.")
            if 'account2_code' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN account2_code TEXT')
                logging.info("Added account2_code column to exceptions table.")
            if 'sub_ledger_code' not in columns:
                cursor.execute('ALTER TABLE exceptions ADD COLUMN sub_ledger_code TEXT')
                logging.info("Added sub_ledger_code column to exceptions table.")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS department_summary (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, department TEXT,
                    total_records INTEGER, exception_records INTEGER, exception_rate REAL,
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS user_performance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id INTEGER, user TEXT,
                    total_records INTEGER, exception_records INTEGER, exception_rate REAL,
                    FOREIGN KEY (run_id) REFERENCES validation_runs (id))''')
            conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Database initialization error: {e}", exc_info=True)
            if conn: conn.rollback()
        finally:
            if conn: conn.close()

    def save_validation_run(self, filename, total_records, total_exceptions, file_size, upload_time=None):
        conn = None; run_id = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            
            final_upload_time = upload_time if upload_time else datetime.now()

            cursor.execute('''INSERT INTO validation_runs (filename, upload_time, total_records, total_exceptions, file_size)
                              VALUES (?, ?, ?, ?, ?)''', 
                           (filename, final_upload_time, total_records, total_exceptions, file_size))
            run_id = cursor.lastrowid
            conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Error in save_validation_run: {e}", exc_info=True)
            if conn: conn.rollback()
            raise
        finally:
            if conn: conn.close()
        return run_id
    
    def save_excel_report(self, run_id, excel_data):
        if not excel_data:
            return
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            cursor.execute('''UPDATE validation_runs SET excel_report_data = ? WHERE id = ?''',
                           (sqlite3.Binary(excel_data.getvalue()), run_id))
            conn.commit()
            logging.info(f"Successfully saved Excel report for run_id {run_id} to database.")
        except sqlite3.Error as e:
            logging.error(f"Error saving Excel report for run_id {run_id}: {e}", exc_info=True)
            if conn: conn.rollback()
        finally:
            if conn: conn.close()
            
    def get_archived_report(self, run_id):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            cursor.execute("SELECT excel_report_data, filename FROM validation_runs WHERE id = ?", (run_id,))
            result = cursor.fetchone()
            if result and result[0] is not None:
                return result[0], result[1]
            return None, None
        except sqlite3.Error as e:
            logging.error(f"Error fetching archived report for run_id {run_id}: {e}", exc_info=True)
            return None, None
        finally:
            if conn: conn.close()

    def delete_run(self, run_id):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            conn.execute("BEGIN TRANSACTION")
            cursor.execute("DELETE FROM exceptions WHERE run_id = ?", (run_id,))
            cursor.execute("DELETE FROM department_summary WHERE run_id = ?", (run_id,))
            cursor.execute("DELETE FROM user_performance WHERE run_id = ?", (run_id,))
            cursor.execute("DELETE FROM validation_runs WHERE id = ?", (run_id,))
            conn.commit()
            logging.info(f"Successfully deleted run ID: {run_id} and all associated data.")
            return True
        except sqlite3.Error as e:
            logging.error(f"Error deleting run ID {run_id}: {e}", exc_info=True)
            if conn: conn.rollback()
            return False
        finally:
            if conn: conn.close()

    def save_exceptions(self, run_id, exceptions_df):
        conn = None
        if exceptions_df.empty: return
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()

            data_to_insert = []
            for _, row_series in exceptions_df.iterrows():
                original_row_data_dict = row_series.to_dict()
                try:
                    serialized_row_data = json.dumps(original_row_data_dict, default=json_serializer_default)
                except Exception as json_e:
                    logging.error(f"Could not serialize row for run_id {run_id}. Row data: {original_row_data_dict}. Error: {json_e}", exc_info=True)
                    try:
                        serialized_row_data = json.dumps({"error": "Failed to serialize original row", "details": str(original_row_data_dict)})
                    except:
                        serialized_row_data = '{"error": "Critical serialization failure for original row"}'

                data_to_insert.append(
                    (run_id,
                     row_series.get('Department.Name', ''),
                     row_series.get('Sub Department.Name', ''),
                     row_series.get('Created user', ''),
                     row_series.get('Modified user', ''),
                     str(row_series.get('Exception Reasons', '')),
                     row_series.get('Severity', 0),
                     row_series.get('Net amount', 0.0),
                     row_series.get('Location.Name', ''),
                     row_series.get('Crop.Name', ''),
                     row_series.get('Activity.Name', ''),
                     row_series.get('Function.Name', ''),
                     row_series.get('FC-Vertical.Name', ''),
                     row_series.get('Region.Name', ''),
                     row_series.get('Zone.Name', ''),
                     row_series.get('Business Unit.Name', ''),
                     row_series.get('Account2.Code', ''),
                     row_series.get('Sub Ledger.Code', ''),
                     serialized_row_data
                    )
                )

            if data_to_insert:
                cursor.executemany('''INSERT INTO exceptions (
                    run_id, department, sub_department, created_user, modified_user, exception_reason,
                    severity, net_amount, location, crop, activity, function_name, vertical_name,
                    region_name, zone_name, business_unit, account2_code, sub_ledger_code, original_row_data)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', data_to_insert)
                conn.commit()
        except sqlite3.Error as e:
            logging.error(f"Error in save_exceptions (SQLite): {e}", exc_info=True)
            if conn: conn.rollback()
            st.error(f"Database error while saving exceptions: {e}")
            raise
        except Exception as ex:
            logging.error(f"Unexpected error in save_exceptions: {ex}", exc_info=True)
            st.error(f"An unexpected error occurred while saving exceptions: {ex}")
            if conn and getattr(conn, 'in_transaction', False): conn.rollback()
        finally:
            if conn: conn.close()

    def save_user_performance(self, run_id, df, exceptions_df):
        conn = None
        try:
            if 'Created user' not in df.columns:
                st.markdown('<div class="error-box"><strong>❌ Error:</strong> Created user column missing in input for user performance.</div>', unsafe_allow_html=True)
                logging.error("Created user column missing in input df for save_user_performance.")
                return

            df['Created user'] = df['Created user'].dropna().astype(str).str.strip()
            valid_users = df['Created user'][df['Created user'] != ''].unique()

            if len(valid_users) == 0:
                st.markdown('<div class="warning-box"><strong>⚠ Warning:</strong> No valid Created user values for user performance.</div>', unsafe_allow_html=True)
                logging.warning("No valid Created user values in input df for save_user_performance.")
                return

            if exceptions_df.empty or 'Created user' not in exceptions_df.columns:
                user_stats = pd.DataFrame({'Created user': valid_users, 'exception_records': 0})
            else:
                exceptions_df_copy = exceptions_df.copy()
                exceptions_df_copy['Created user'] = exceptions_df_copy['Created user'].astype(str).str.strip()
                user_stats = exceptions_df_copy.groupby('Created user').size().reset_index(name='exception_records')
                user_stats = user_stats[user_stats['Created user'].isin(valid_users)]
                users_with_no_exceptions_list = [u for u in valid_users if u not in user_stats['Created user'].tolist()]
                if users_with_no_exceptions_list:
                    users_with_no_exceptions_df = pd.DataFrame({'Created user': users_with_no_exceptions_list, 'exception_records': 0})
                    user_stats = pd.concat([user_stats, users_with_no_exceptions_df], ignore_index=True)

            total_records_user_df = df[df['Created user'].isin(valid_users)].groupby('Created user').size().reset_index(name='total_records')

            user_stats['Created user'] = user_stats['Created user'].astype(str)
            total_records_user_df['Created user'] = total_records_user_df['Created user'].astype(str)
            user_stats = pd.merge(user_stats, total_records_user_df, on='Created user', how='left').fillna({'total_records': 0, 'exception_records': 0})

            user_stats['exception_rate'] = 0.0
            mask = user_stats['total_records'] > 0
            user_stats.loc[mask, 'exception_rate'] = (user_stats.loc[mask, 'exception_records'] / user_stats.loc[mask, 'total_records'] * 100).round(2)
            user_stats['exception_rate'] = user_stats['exception_rate'].fillna(0)

            if user_stats.empty: return

            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            cursor = conn.cursor()
            data_to_insert_perf = [
                (run_id, row['Created user'], int(row['total_records']), int(row['exception_records']), row['exception_rate'])
                for _, row in user_stats.iterrows()
            ]
            if data_to_insert_perf:
                cursor.executemany('''INSERT INTO user_performance (run_id, user, total_records, exception_records, exception_rate)
                                      VALUES (?,?,?,?,?)''', data_to_insert_perf)
                conn.commit()
        except sqlite3.Error as e:
            logging.error(f"SQLite error in save_user_performance: {e}", exc_info=True)
            if conn: conn.rollback()
            raise
        except Exception as ex_gen:
            logging.error(f"General error in save_user_performance: {ex_gen}", exc_info=True)
            st.error(f"An unexpected error occurred while saving user performance: {ex_gen}")
        finally:
            if conn: conn.close()

    def get_validation_history(self):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            df = pd.read_sql_query('SELECT * FROM validation_runs ORDER BY upload_time DESC', conn)
            if not df.empty and 'upload_time' in df.columns:
                # Use format='mixed' to handle different datetime string formats
                df['upload_time'] = pd.to_datetime(df['upload_time'], format='mixed')
            return df
        except sqlite3.Error as e:
            logging.error(f"Error in get_validation_history: {e}", exc_info=True); return pd.DataFrame()
        finally:
            if conn: conn.close()

    def get_exceptions_by_run(self, run_id):
        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=DB_TIMEOUT)
            raw_exceptions_df = pd.read_sql_query('SELECT * FROM exceptions WHERE run_id = ?', conn, params=(run_id,))

            if raw_exceptions_df.empty:
                return pd.DataFrame()

            all_records = []
            for _, db_row in raw_exceptions_df.iterrows():
                record_data = db_row.to_dict()

                if 'original_row_data' in record_data and pd.notna(record_data['original_row_data']):
                    try:
                        json_data = json.loads(record_data['original_row_data'])
                        record_data.update(json_data)
                    except json.JSONDecodeError:
                        logging.warning(f"Failed to parse original_row_data for exception id {db_row.get('id')}")

                record_data['id'] = db_row['id']
                record_data['run_id'] = db_row['run_id']
                record_data['Exception Reasons'] = db_row['exception_reason']
                record_data['Severity'] = db_row['severity']
                record_data['Department.Name'] = db_row['department']
                # --- FIX: Ensure 'Sub Department.Name' is correctly populated ---
                record_data['Sub Department.Name'] = db_row['sub_department']
                record_data['Created user'] = db_row['created_user']
                record_data['Account2.Code'] = db_row['account2_code']
                record_data['Sub Ledger.Code'] = db_row['sub_ledger_code']

                all_records.append(record_data)

            final_df = pd.DataFrame(all_records)

            first_cols = ['id', 'run_id']
            last_cols = ['Exception Reasons', 'Severity']
            middle_cols = sorted([col for col in final_df.columns if col not in first_cols + last_cols and col != 'original_row_data'])
            ordered_cols = first_cols + middle_cols + last_cols
            final_ordered_columns = [col for col in ordered_cols if col in final_df.columns]
            for col in final_df.columns:
                if col not in final_ordered_columns and col != 'original_row_data':
                    final_ordered_columns.append(col)

            return final_df.reindex(columns=final_ordered_columns, fill_value=pd.NA)

        except sqlite3.Error as e:
            st.error(f"Database error retrieving exceptions for run {run_id}: {e}")
            logging.error(f"get_exceptions_by_run SQLite error for run_id {run_id}: {e}", exc_info=True)
            return pd.DataFrame()
        except Exception as ex:
            st.error(f"Unexpected error retrieving exceptions for run {run_id}: {ex}")
            logging.error(f"get_exceptions_by_run general error for run_id {run_id}: {ex}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()


@st.cache_resource
def get_database_manager():
    return DatabaseManager()

db_manager = get_database_manager()

@st.cache_data
def load_ledger_validation_mapping(base_ref_path="reference_data"):
    """Loads the combined ledger/sub-ledger mapping for VALIDATION purposes."""
    mapping_file = os.path.join(base_ref_path, "Ledgersubledger mapping.xlsx")
    LEDGER_CODE_COL = "Account2.Code"
    SUBLEDGER_CODE_COL = "Sub Ledger.Code"
    
    try:
        if not os.path.exists(mapping_file):
            st.error(f"VALIDATION Error: Ledger mapping file not found at '{mapping_file}'. Ledger combination validation will not work.")
            return None

        mapping_df = pd.read_excel(mapping_file)
        required_cols = [LEDGER_CODE_COL, SUBLEDGER_CODE_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.error(f"VALIDATION Error: Mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}")
            return None

        for col in required_cols:
            mapping_df[col] = mapping_df[col].astype(str).str.strip()

        return mapping_df

    except Exception as e:
        st.error(f"Failed to read or process validation mapping file {mapping_file}: {str(e)}")
        return None

@st.cache_data
def load_account_name_mapping(base_ref_path="reference_data"):
    """Loads the account code-to-name mapping for DISPLAY purposes."""
    mapping_file = os.path.join(base_ref_path, "account_mapping.xlsx")
    CODE_COL = "Account2.Code"
    NAME_COL = "Account2.Name"

    try:
        if not os.path.exists(mapping_file):
            st.warning(f"DISPLAY Warning: Account name mapping file not found at '{mapping_file}'. Ledger names may not display correctly.")
            return None
        
        mapping_df = pd.read_excel(mapping_file)
        required_cols = [CODE_COL, NAME_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.warning(f"DISPLAY Warning: Account name mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}. Names may not display.")
            return None
        
        mapping_df[CODE_COL] = mapping_df[CODE_COL].astype(str).str.strip()
        mapping_df[NAME_COL] = mapping_df[NAME_COL].astype(str).str.strip()

        return mapping_df[required_cols].drop_duplicates(subset=[CODE_COL]).reset_index(drop=True)

    except Exception as e:
        st.error(f"Failed to read or process account name mapping file {mapping_file}: {str(e)}")
        return None

@st.cache_data
def load_subledger_name_mapping(base_ref_path="reference_data"):
    """Loads the sub-ledger code-to-name mapping for DISPLAY purposes."""
    mapping_file = os.path.join(base_ref_path, "subledger_mapping.xlsx")
    CODE_COL = "Sub Ledger.Code"
    NAME_COL = "SubLedger.Name"

    try:
        if not os.path.exists(mapping_file):
            st.warning(f"DISPLAY Warning: Sub-ledger name mapping file not found at '{mapping_file}'. Sub-ledger names may not display correctly.")
            return None
        
        mapping_df = pd.read_excel(mapping_file)
        required_cols = [CODE_COL, NAME_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.warning(f"DISPLAY Warning: Sub-ledger name mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}. Names may not display.")
            return None

        mapping_df[CODE_COL] = mapping_df[CODE_COL].astype(str).str.strip()
        mapping_df[NAME_COL] = mapping_df[NAME_COL].astype(str).str.strip()

        return mapping_df[required_cols].drop_duplicates(subset=[CODE_COL]).reset_index(drop=True)

    except Exception as e:
        st.error(f"Failed to read or process sub-ledger name mapping file {mapping_file}: {str(e)}")
        return None


class DataValidator:
    def __init__(self, base_ref_path="reference_data"):
        self.base_ref_path = base_ref_path
        self.no_crop_check = {
            "Finance & Account", "Human Resource", "Administration",
            "Information Technology", "Legal", "Accounts Receivable & MIS", "Management"
            }
        self.no_activity_check = self.no_crop_check.copy()
        self.no_activity_check.update({"Production", "Processing", "Parent Seed"})
        self.ref_files = self._load_reference_data()

        self.valid_ledger_keys = set()
        self.LEDGER_CODE_COL = "Account2.Code"
        self.SUBLEDGER_CODE_COL = "Sub Ledger.Code"

        mapping_df = load_ledger_validation_mapping(self.base_ref_path)
        if mapping_df is not None:
             self.valid_ledger_keys = set(mapping_df[self.LEDGER_CODE_COL] + "_" + mapping_df[self.SUBLEDGER_CODE_COL])
             logging.info(f"Successfully loaded {len(self.valid_ledger_keys)} ledger-subledger mapping combinations into validator.")
        else:
             logging.error("Could not load ledger validation mapping into validator. Ledger validation will be skipped.")

        self.training_map = {
            'Incorrect Ledger/Sub-Ledger Combination': 'Review the Ledger-Subledger mapping file for valid combinations.',
            'Incorrect Location Name': 'Review Location Name Guidelines: Ensure locations are valid (e.g., Bandamailaram).',
            'Incorrect Activity Name': 'Complete Activity Name Training: Use only approved activities for your department.',
            'Incorrect Crop Name': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'FC-Vertical Name cannot be blank': 'Ensure Valid FC-Vertical Name.',
            'Crop Name cannot be blank': 'Ensure Crop Name is provided when required.',
            'Incorrect Crop Name starting with ZZ': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'Incorrect Crop Name for FC-field crop Vertical': 'Ensure Valid Crop Name for FC-field crop.',
            'Incorrect Crop Name for VC-Veg Crop Vertical': 'Ensure Valid Crop Name for VC-Veg Crop.',
            'Incorrect Crop Name for Fruit Crop Vertical': 'Ensure Valid Crop Name for Fruit Crop.',
            'Incorrect Crop Name for Common vertical': 'Ensure Valid Crop Name for Common vertical.',
            'Incorrect Sub Department Name': 'Verify Sub-Department Standards.',
            'Incorrect Function Name': 'Check Function Name Guidelines.',
            'Incorrect FC-Vertical Name': 'Ensure Valid FC-Vertical Name.',
            'Need to Update Processing Location': 'Use Approved Processing Locations.',
            'Incorrect Activity Name for Lab QC': 'Use Lab QA Approved Activities.',
            'Incorrect Activity Name for Field QA': 'Use Field QA Approved Activities.',
            'Incorrect Activity Name for Bio Tech Services': 'Use Bio Tech Approved Activities.',
            'Sub Department should be blank': 'Ensure Sub-Department is blank for this department.',
            'Activity Name cannot be blank or start with ZZ': 'Complete Activity Name Training: Use only approved activities.',
            'Incorrect Activity Name for Biotech - Markers': 'Use Approved Activities for Biotech - Markers.',
            'Incorrect Activity Name for Biotech - Tissue Culture': 'Use Approved Activities for Biotech - Tissue Culture.',
            'Incorrect Activity Name for Biotech - Mutation': 'Use Approved Activities for Biotech - Mutation.',
            'Incorrect Activity Name for Entomology': 'Use Approved Activities for Entomology.',
            'Incorrect Activity Name for Pathology': 'Use Approved Activities for Pathology.',
            'Incorrect Activity Name for Bioinformatics': 'Use Approved Activities for Bioinformatics.',
            'Incorrect Activity Name for Biochemistry': 'Use Approved Activities for Biochemistry.',
            'Incorrect Activity Name for Common': 'Use Approved Activities for Common sub-department in Breeding Support.',
            'Need to update Zone can not left Blank': 'Ensure Zone is specified.',
            'Incorrect Zone Name for FC-field crop Vertical': 'Use Approved Zone Names for FC-field crop.',
            'Incorrect Zone Name for VC-Veg Crop Vertical': 'Use Approved Zone Names for VC-Veg Crop.',
            'Need to update Zone Name can not left Blank': 'Ensure Zone is specified for Common vertical in Production.',
            'Need to update Business Unit can not left Blank': 'Ensure Business Unit is specified.',
            'Incorrect Business Unit Name for FC-field crop Vertical': 'Use Approved Business Unit Names for FC-field crop.',
            'Incorrect Business Unit Name for VC-Veg Crop Vertical': 'Use Approved Business Unit Names for VC-Veg Crop.',
            'Need to update Region Name can not left Blank': 'Ensure Region is specified.',
            'Incorrect Region Name for FC-field crop Vertical': 'Use Approved Region Names for FC-field crop.',
            'Incorrect Region Name for VC-Veg Crop Vertical': 'Use Approved Region Names for VC-Veg Crop.',
            'Region, Zone, BU need to check for Root Stock': 'Ensure Region, Zone, and BU are blank for Root Stock vertical in Sales/Marketing.',
            'Incorrect Activity Name for Sales': 'Use Approved Activities for Sales.',
            'Incorrect Activity Name for Marketing': 'Use Approved Activities for Marketing.'
        }

    def _load_reference_data(self):
        loaded_ref_files = {}
        ref_file_mappings = {
            "FC_Crop": ("FC-field crop.xlsx", "Crop.Name"),
            "VC_Crop": ("VC-Veg Crop.xlsx", "Crop.Name"),
            "SBFC_Region": ("SBFC-Region.xlsx", "Region.Name"),
            "SBVC_Region": ("SBVC-Region.xlsx", "Region.Name"),
            "SaleFC_Zone": ("SaleFC-Zone.xlsx", "Zone.Name"),
            "SaleVC_Zone": ("SaleVC-Zone.xlsx", "Zone.Name"),
            "FC_BU": ("FC-BU.xlsx", "Business Unit.Name"),
            "VC_BU": ("VC-BU.xlsx", "Business Unit.Name"),
            "Fruit_Crop": ("Fruit Crop.xlsx", "Crop.Name"),
            "Common_Crop": ("Common crop.xlsx", "Crop.Name"),
            "ProductionFC_Zone": ("ProductionFC-Zone.xlsx", "Zone.Name"),
            "ProductionVC_Zone": ("ProductionVC-Zone.xlsx", "Zone.Name"),
            "SalesActivity": ("SalesActivity.xlsx", "Activity.Name"),
            "MarketingActivity": ("MarketingActivity.xlsx", "Activity.Name"),
            "RS_BU": ("RS-BU.xlsx", "Business Unit.Name"),
            "SaleRS_Zone": ("SaleRS-Zone.xlsx", "Zone.Name"),
            "SBRS_Region": ("SBRS-Region.xlsx", "Region.Name"),
            "Root Stock_Crop": ("Root Stock Crop.xlsx", "Crop.Name"),
            "Region_Excluded_Accounts": ("Region.Name excluded.xlsx", "Account.Code"),
            "Zone_Excluded_Accounts": ("Zone.Name excluded.xlsx", "Account.Code"),
        }

        if not os.path.isdir(self.base_ref_path):
            st.error(f"Reference data directory not found: '{self.base_ref_path}'. Please create it and add reference Excel files. Validations will be highly inaccurate.")
            logging.critical(f"Reference data directory not found: '{self.base_ref_path}'. Cannot load reference data.")
            return {key: [] for key in ref_file_mappings.keys()}

        all_files_loaded_successfully = True
        for key, (filename, col_name) in ref_file_mappings.items():
            try:
                file_path = os.path.join(self.base_ref_path, filename)
                if os.path.exists(file_path):
                    df_ref = pd.read_excel(file_path, engine='openpyxl')
                    if col_name in df_ref.columns:
                        if key in ["Region_Excluded_Accounts", "Zone_Excluded_Accounts"]:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        else:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        logging.info(f"Loaded {len(loaded_ref_files[key])} items for '{key}' from '{filename}'.")
                    else:
                        st.warning(f"Column '{col_name}' not found in reference file '{filename}'. '{key}' will be empty.")
                        logging.warning(f"Column '{col_name}' not found in '{file_path}' for key '{key}'.")
                        loaded_ref_files[key] = []
                        all_files_loaded_successfully = False
                else:
                    st.warning(f"Reference file '{filename}' not found in '{self.base_ref_path}'. Validations for '{key}' may be inaccurate.")
                    logging.warning(f"Reference file '{filename}' not found for key '{key}' at path '{file_path}'.")
                    loaded_ref_files[key] = []
                    all_files_loaded_successfully = False
            except Exception as e:
                st.error(f"Error loading reference file '{filename}' for '{key}': {e}")
                logging.error(f"Error loading reference file '{filename}' for key '{key}': {e}", exc_info=True)
                loaded_ref_files[key] = []
                all_files_loaded_successfully = False

        if not all_files_loaded_successfully:
            st.error("One or more reference data files could not be loaded correctly. Please check logs and file paths. Validation accuracy will be affected.")
        elif not any(loaded_ref_files.values()):
            st.error("All reference data lists are empty after attempting to load files. This indicates a problem with file contents or loading logic. Validations will be highly inaccurate.")
            logging.critical("All loaded reference file lists are empty. Check file contents and parsing logic.")
        else:
            logging.info("Reference data loading process completed.")
        return loaded_ref_files

    def is_not_blank(self, value):
        if pd.isna(value) or value is None:
            return False
        val = str(value).strip().replace("\u00A0", "").replace("\u200B", "")
        return val != "" and val.upper() not in ["N/A", "NULL", "NONE", "NA", "0", "-"]

    def is_blank(self, value):
        return not self.is_not_blank(value)

    def validate_row(self, dept, row):
        reasons = []
        # Extract values from row
        sub_dept = str(row.get("Sub Department.Name", "") or "").strip().replace("\u00A0", "").replace("\u200B", "")
        func = str(row.get("Function.Name", "") or "").strip()
        vertical = str(row.get("FC-Vertical.Name", "") or "").strip()
        loc = str(row.get("Location.Name", "") or "").strip()
        crop = str(row.get("Crop.Name", "") or "").strip()
        act = str(row.get("Activity.Name", "") or "").strip()
        region = row.get("Region.Name", "")
        zone = row.get("Zone.Name", "")
        bu = row.get("Business Unit.Name", "")
        account_code = str(row.get("Account.Code", "") or "").strip()
        ledger_code = str(row.get(self.LEDGER_CODE_COL, "") or "").strip()
        subledger_code = str(row.get(self.SUBLEDGER_CODE_COL, "") or "").strip()

        # <<< INTEGRATED LEDGER/SUB-LEDGER CHECK >>>
        # *** Only check for an invalid combination if codes have actually been entered. ***
        if self.is_not_blank(ledger_code) or self.is_not_blank(subledger_code):
            combination_key = f"{ledger_code}_{subledger_code}"
            if combination_key not in self.valid_ledger_keys:
                reasons.append("Incorrect Ledger/Sub-Ledger Combination")

        # --- Existing Generic and Department-Specific Checks ---
        if self.is_blank(loc) or loc.startswith("ZZ"):
            reasons.append("Incorrect Location Name")
        if dept not in self.no_activity_check and dept not in ["Breeding", "Trialing & PD", "Sales", "Marketing", "Breeding Support"]:
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Incorrect Activity Name")

        # Crop and Vertical validation
        if dept not in self.no_crop_check:
            if self.is_blank(vertical):
                reasons.append("FC-Vertical Name cannot be blank")
            if self.is_blank(crop):
                reasons.append("Crop Name cannot be blank")
            elif crop.startswith("ZZ"):
                reasons.append("Incorrect Crop Name starting with ZZ")
            elif vertical == "FC-field crop" and crop not in self.ref_files.get("FC_Crop", []):
                reasons.append("Incorrect Crop Name for FC-field crop Vertical")
            elif vertical == "VC-Veg Crop" and crop not in self.ref_files.get("VC_Crop", []):
                reasons.append("Incorrect Crop Name for VC-Veg Crop Vertical")
            elif vertical == "Fruit Crop" and crop not in self.ref_files.get("Fruit_Crop", []):
                reasons.append("Incorrect Crop Name for Fruit Crop Vertical")
            elif vertical == "Common" and crop not in self.ref_files.get("Common_Crop", []):
                reasons.append("Incorrect Crop Name for Common vertical")
            elif vertical == "Root Stock" and crop not in self.ref_files.get("Root Stock_Crop", []):
                reasons.append("Incorrect Crop Name for Root Stock Crop Vertical")

        # Account Code exclusion checks
        if account_code in self.ref_files.get("Region_Excluded_Accounts", []) and self.is_not_blank(region):
            reasons.append("Region Name should be blank for this Account Code")
        if account_code in self.ref_files.get("Zone_Excluded_Accounts", []) and self.is_not_blank(zone):
            reasons.append("Zone Name should be blank for this Account Code")

        # Department-specific checks
        if dept == "Parent Seed":
            if sub_dept not in ["Breeder Seed Production", "Foundation Seed Production", "Processing FS"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Production":
            if sub_dept not in ["Commercial Seed Production", "Seed Production Research"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            # Zone validation for Commercial Seed Production sub-department
            if sub_dept == "Commercial Seed Production":
                if vertical == "FC-field crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("ProductionFC_Zone", []):
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("ProductionVC_Zone", []):
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                elif vertical == "Common" and self.is_blank(zone):
                    reasons.append("Need to update Zone Name can not left Blank")

        elif dept == "Processing":
            if sub_dept not in ["Processing", "Warehousing", "Project & Maintenance"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            if loc not in ["Bandamailaram", "Deorjhal", "Boriya"]:
                reasons.append("Need to Update Processing Location")

        elif dept == "Quality Assurance":
            if sub_dept not in ["Field QA", "Lab QC", "Bio Tech Services"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            # Sub-department-specific activity checks
            if sub_dept == "Lab QC" and act not in ["Lab Operations QA", "All Activity"]:
                reasons.append("Incorrect Activity Name for Lab QC")
            if sub_dept == "Field QA" and act not in ["Field Operations QA", "All Activity", "GOT"]:
                reasons.append("Incorrect Activity Name for Field QA")
            if sub_dept == "Bio Tech Services" and act not in ["Molecular", "All Activity"]:
                reasons.append("Incorrect Activity Name for Bio Tech Services")

        elif dept == "Seed Tech":
            if sub_dept not in ["Aging Test", "Pelleting", "Priming", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "In Licensing & Procurement":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if vertical in ["", "N/A", "Common"]:
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Breeding":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if vertical in ["", "N/A"]:
                reasons.append("Incorrect FC-Vertical Name")
            if dept not in self.no_activity_check and act not in ["Breeding", "All Activity", "Trialing", "Pre Breeding", "Germplasm Maintainance", "Experimental Seed Production"]:
                reasons.append("Incorrect Activity Name")

        elif dept == "Breeding Support":
            if sub_dept not in ["Pathology", "Biotech - Tissue Culture", "Biotech - Mutation", "Biotech - Markers", "Bioinformatics", "Biochemistry", "Entomology", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            # Activity validation: Check for blank or ZZ first, then sub-department-specific checks
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Activity Name cannot be blank or start with ZZ")
            else:
                # Sub-department-specific activity checks
                if sub_dept == "Biotech - Markers" and act not in ["Molecular", "Grain Quality", "Seed Treatment", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Markers")
                elif sub_dept == "Biotech - Tissue Culture" and act not in ["Tissue Culture", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Tissue Culture")
                elif sub_dept == "Biotech - Mutation" and act not in ["Mutation", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Mutation")
                elif sub_dept == "Entomology" and act not in ["Entomology", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Entomology")
                elif sub_dept == "Pathology" and act not in ["Pathalogy", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Pathology")
                elif sub_dept == "Bioinformatics" and act not in ["Bioinformatics", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Bioinformatics")
                elif sub_dept == "Biochemistry" and act not in ["Biochemistry", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biochemistry")
                elif sub_dept == "Common" and act not in ["All Activity"]:
                    reasons.append("Incorrect Activity Name for Common")

        elif dept == "Trialing & PD":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            if dept not in self.no_activity_check and act not in ["CT", "All Activity", "Trialing", "RST", "Disease"]:
                reasons.append("Incorrect Activity Name")

        elif dept == "Sales":
            valid_subs = ["Sales Brand", "Sales Export", "Sales Institutional & Govt"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            # Activity validation for Sales
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files.get("SalesActivity",[]):
                reasons.append("Incorrect Activity Name for Sales")
            # Business Unit, Zone, and Region validation for Sales Brand sub-department
            if sub_dept == "Sales Brand":
                if vertical == "FC-field crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("FC_BU", []):
                        reasons.append("Incorrect Business Unit Name for FC-field crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleFC_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBFC_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for FC-field crop Vertical")

                elif vertical == "VC-Veg Crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("VC_BU", []):
                        reasons.append("Incorrect Business Unit Name for VC-Veg Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleVC_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBVC_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for VC-Veg Crop Vertical")

                elif vertical == "Root Stock":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("RS_BU", []):
                        reasons.append("Incorrect Business Unit Name for Root Stock Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleRS_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for Root Stock Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBRS_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for Root Stock Crop Vertical")

        elif dept == "Marketing":
            valid_subs = ["Business Development", "Digital Marketing", "Product Management"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")
            elif vertical == "Root Stock" and any(self.is_not_blank(x) for x in [region, zone, bu]):
                reasons.append("Region, Zone, BU need to check for Root Stock")
            # Activity validation for Marketing
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files.get("MarketingActivity",[]):
                reasons.append("Incorrect Activity Name for Marketing")

        elif dept == "Finance & Account":
            if sub_dept not in ["Accounts", "Finance", "Analytics, Internal Control & Budget", "Purchase ops", "Secretarial", "Document Management System", "Automation", "Group Company", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Human Resource":
            if sub_dept not in ["Compliances", "HR Ops", "Recruitment", "Team Welfare", "Training", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Administration":
            if sub_dept not in ["Events", "Maintenance", "Travel Desk","Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Information Technology":
            if sub_dept not in ["ERP Support", "Infra & Hardware", "Application Development"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Legal":
            if sub_dept not in ["Compliances", "Litigation","Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Accounts Receivable & MIS":
            if sub_dept not in ["Branch and C&F Ops", "Commercial & AR Management", "Common", "Order Processing", "Transport & Logistic"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Management":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Management":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical):
                reasons.append("Incorrect FC-Vertical Name")

        unique_reasons = sorted(list(set(reasons)))
        severity = len(unique_reasons) * 2
        return unique_reasons, severity

    def validate_dataframe(self, df):
        """
        Validates the entire DataFrame by splitting it into chunks and processing them in parallel.
        This method utilizes multiple CPU cores to significantly speed up validation for large files.
        """
        exceptions = []
        
        null_like_values = [pd.NA, "N/A", "NaN", "null", "NONE", "", " ", "-", "\u00A0", None, 0, "0"]
        if 'Sub Department.Name' in df.columns:
            df['Sub Department.Name'] = df['Sub Department.Name'].replace(null_like_values, "").astype(str).str.strip()
        else:
            df['Sub Department.Name'] = ""

        input_columns = df.columns.tolist()
        
        if 'Department.Name' not in df.columns:
            st.error("Critical Error: 'Department.Name' column is missing from the input file.")
            logging.critical("Critical Error: 'Department.Name' column is missing in validate_dataframe.")
            return pd.DataFrame(columns=input_columns + ['Exception Reasons', 'Severity']), {}

        # OPTIMIZATION: Use the number of available CPUs for parallel processing. Fallback to 4 if undetected.
        num_workers = os.cpu_count() or 4
        logging.info(f"Starting parallel validation with {num_workers} workers.")
        
        # OPTIMIZATION: Split the DataFrame into chunks for each worker.
        df_chunks = np.array_split(df, num_workers)
        
        # OPTIMIZATION: Use ProcessPoolExecutor to run validation in parallel.
        with concurrent.futures.ProcessPoolExecutor(max_workers=num_workers) as executor:
            # Submit each chunk to the _validate_chunk function
            futures = [executor.submit(_validate_chunk, self, chunk) for chunk in df_chunks]
            
            # Collect the results as they are completed
            for future in concurrent.futures.as_completed(futures):
                try:
                    exceptions.extend(future.result())
                except Exception as e:
                    logging.error(f"A validation chunk failed: {e}", exc_info=True)
                    st.error(f"An error occurred during parallel processing: {e}")

        # --- Post-processing after parallel execution ---
        
        # Create the final exceptions DataFrame from the collected results
        exceptions_df_output = pd.DataFrame()
        output_columns_with_exceptions = input_columns + ['Exception Reasons', 'Severity']

        if exceptions:
            exceptions_df_output = pd.DataFrame(exceptions)
            # Ensure all original columns are present in the output
            for col in output_columns_with_exceptions:
                if col not in exceptions_df_output.columns:
                    exceptions_df_output[col] = pd.NA
            exceptions_df_output = exceptions_df_output.reindex(columns=output_columns_with_exceptions, fill_value=pd.NA)
        else:
            exceptions_df_output = pd.DataFrame(columns=output_columns_with_exceptions)
        
        # Calculate department statistics after all exceptions have been found
        department_stats = {}
        if 'Department.Name' in df.columns:
            # Get total records per department from the original dataframe
            total_records_by_dept = df.groupby('Department.Name').size().to_dict()
            
            # Get exception records per department from the exceptions dataframe
            exception_records_by_dept = {}
            if not exceptions_df_output.empty and 'Department.Name' in exceptions_df_output.columns:
                exception_records_by_dept = exceptions_df_output.groupby('Department.Name').size().to_dict()

            for dept, total_records in total_records_by_dept.items():
                exception_records = exception_records_by_dept.get(dept, 0)
                department_stats[dept] = {
                    'total_records': total_records,
                    'exception_records': exception_records,
                    'exception_rate': (exception_records / total_records * 100) if total_records > 0 else 0
                }

        return exceptions_df_output, department_stats


def display_metric(title, value, delta=None, container=None):
    target_container = container if container else st
    delta_html = f'<div class="metric-delta" style="font-size:0.8rem; color:#e2e8f0;">{delta}</div>' if delta else ""
    target_container.markdown(f"""
    <div class="metric-container">
        <div class="metric-title">{title}</div>
        <div class="metric-value">{value}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

def create_excel_report(exceptions_df, dept_stats, filename_for_logging="ExcelReport"):
    output = io.BytesIO()
    try:
        if exceptions_df is None or exceptions_df.empty:
            logging.info(f"create_excel_report: exceptions_df is empty for {filename_for_logging}. Creating empty Exceptions sheet.")
            exceptions_df_prepared = exceptions_df.copy() if exceptions_df is not None else pd.DataFrame()
        else:
            exceptions_df_prepared = exceptions_df.copy()

        if 'Exception Reasons' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Exception Reasons'] = exceptions_df_prepared['Exception Reasons'].astype(str).replace('<NA>', '').replace('nan', '').replace('None','')
        if 'Severity' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Severity'] = pd.to_numeric(exceptions_df_prepared['Severity'], errors='coerce').fillna(0)
            if exceptions_df_prepared['Severity'].notna().all():
                 try:
                    exceptions_df_prepared['Severity'] = exceptions_df_prepared['Severity'].astype(int)
                 except ValueError:
                    pass

        numeric_cols_to_preserve = ['Net amount', 'Severity', 'id', 'run_id']
        if dept_stats:
             numeric_cols_to_preserve.extend(['Total Records', 'Exception Records', 'Exception Rate (%)'])

        for col in exceptions_df_prepared.columns:
            if col in numeric_cols_to_preserve:
                if col == 'Net amount':
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce').fillna(0.0)
                elif col in ['id', 'run_id']:
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce')
            else:
                if not pd.api.types.is_numeric_dtype(exceptions_df_prepared[col]) and \
                   not pd.api.types.is_string_dtype(exceptions_df_prepared[col]):
                    exceptions_df_prepared[col] = exceptions_df_prepared[col].astype(str).replace('<NA>', '').replace('nan', '').replace('None','').replace('NaT','')

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            exceptions_df_prepared.to_excel(writer, sheet_name='Exceptions', index=False)
            if not dept_stats:
                 dept_summary_df = pd.DataFrame(columns=['Department', 'Total Records', 'Exception Records', 'Exception Rate (%)'])
            else:
                dept_summary_df = pd.DataFrame([{'Department': dept, 'Total Records': stats.get('total_records', 0), 'Exception Records': stats.get('exception_records', 0), 'Exception Rate (%)': round(stats.get('exception_rate', 0), 2)} for dept, stats in dept_stats.items()])
            dept_summary_df.to_excel(writer, sheet_name='Department Summary', index=False)
            workbook = writer.book
            header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            thin_border_side = Side(style='thin'); cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.max_row == 0 : continue
                df_ref_for_headers = exceptions_df_prepared if sheet_name == 'Exceptions' else dept_summary_df
                if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                    if not df_ref_for_headers.columns.empty:
                        for c_idx, h_val in enumerate(df_ref_for_headers.columns, 1):
                            cell = ws.cell(row=1, column=c_idx, value=str(h_val)); cell.font=header_font; cell.fill=header_fill; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=cell_border
                    else: continue
                for cell in ws[1]:
                    if cell.value is not None: cell.value = str(cell.value); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = cell_border
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx); cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cell.border = cell_border
                        header_cell_value = str(ws.cell(row=1, column=col_idx).value or "")
                        if header_cell_value == "Net amount": cell.number_format = '#,##0.00'
                        elif header_cell_value == "Exception Rate (%)": cell.number_format = '0.00"%"'
                        elif header_cell_value == "Severity" or "Records" in header_cell_value or "ID" in header_cell_value or "id" in header_cell_value : cell.number_format = '0'
                for col_idx_letter_enum, column_cells_obj in enumerate(ws.columns, 1):
                    current_col_letter = get_column_letter(col_idx_letter_enum); header_val_str = str(ws.cell(row=1, column=col_idx_letter_enum).value or ""); max_length = len(header_val_str)
                    for cell_in_col_obj in column_cells_obj:
                        if cell_in_col_obj.row == 1: continue
                        try:
                            if cell_in_col_obj.value is not None: cell_str_val = str(cell_in_col_obj.value); cell_len = max(len(s) for s in cell_str_val.split('\n')) if '\n' in cell_str_val else len(cell_str_val); max_length = max(max_length, cell_len)
                        except: pass
                    adjusted_width = min(max_length + 5, 60); ws.column_dimensions[current_col_letter].width = adjusted_width
        output.seek(0)
        if output.getbuffer().nbytes == 0:
            logging.error(f"create_excel_report: Excel output buffer is empty for {filename_for_logging}")
            st.markdown('<div class="error-box"><strong>❌ Error:</strong> Failed to generate Excel report (output empty).</div>', unsafe_allow_html=True); return None
        return output
    except Exception as e:
        logging.exception(f"create_excel_report: Error generating Excel report for {filename_for_logging}: {e}")
        st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error generating Excel report: {e}</div>', unsafe_allow_html=True); return None

def display_interactive_exceptions(df, key_prefix="df"):
    """
    Displays an interactive dataframe of exceptions where a user can select a row
    to view the original data record.

    Args:
        df (pd.DataFrame): The DataFrame containing exceptions.
        key_prefix (str): A unique prefix for Streamlit widget keys to avoid conflicts.
    """
    if df.empty:
        st.success("No exceptions to display for the current selection.")
        return

    # Columns to display in the main table
    display_cols = [col for col in [
        'id', 'run_id', 'Department.Name', 'Created user', 'Location.Name',
        'Account2.Code', 'Sub Ledger.Code', 'Exception Reasons', 'Severity'
    ] if col in df.columns]

    st.info("💡 Click on any row to see the complete original data for that record.")

    # Use st.data_editor to get row selections
    # We use a selection key to check for events
    selection_key = f"{key_prefix}_selection"
    if selection_key not in st.session_state:
        st.session_state[selection_key] = None

    edited_df = st.data_editor(
        df[display_cols],
        key=f"{key_prefix}_editor",
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        disabled=df.columns,
        on_change=lambda: st.session_state.update({selection_key: st.session_state[f"{key_prefix}_editor"].get("selection")})
    )

    # Check if a selection has been made
    selection = st.session_state[selection_key]
    if selection and selection.get("rows"):
        try:
            selected_index = selection["rows"][0]
            # Get the full data for the selected row from the original dataframe
            selected_record = df.iloc[selected_index]
            exception_id = selected_record.get('id', 'N/A')

            with st.expander(f"**👁️ Viewing Original Data for Exception ID: {exception_id}**", expanded=True):
                if 'original_row_data' in selected_record and pd.notna(selected_record['original_row_data']):
                    try:
                        original_data = json.loads(selected_record['original_row_data'])
                        
                        # Convert the dictionary to a DataFrame for clean display
                        original_df = pd.DataFrame([original_data])
                        
                        # Tidy up the display: show non-empty columns only
                        original_df = original_df.dropna(axis=1, how='all')
                        st.markdown("#### Original Record Details")
                        st.dataframe(original_df.T.rename(columns={0: 'Value'}), use_container_width=True)

                    except json.JSONDecodeError:
                        st.error("Could not parse the original row data. It might be corrupted.")
                        st.text(selected_record['original_row_data'])
                else:
                    st.warning("Original row data is not available for this exception record.")
        except IndexError:
            st.warning("Could not retrieve the selected row. Please try again.")
        finally:
            # Reset selection to allow re-selection of the same row
            st.session_state[selection_key] = None


def process_uploaded_file(uploaded_file, selected_date=None):
    try:
        df = pd.DataFrame()
        with st.spinner(f"📖 Reading file: {uploaded_file.name}..."):
            try:
                df = pd.read_excel(uploaded_file, engine='openpyxl')
                df.columns = df.columns.str.strip()
            except Exception as e:
                st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Could not read Excel file "{uploaded_file.name}". Details: {str(e)}</div>', unsafe_allow_html=True)
                logging.error(f"Error reading Excel file {uploaded_file.name}: {e}", exc_info=True)
                return

        if df.empty:
            st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> The uploaded file "{uploaded_file.name}" is empty or could not be parsed.</div>', unsafe_allow_html=True)
            return

        required_columns_for_processing = ['Department.Name', 'Created user', 'Account2.Code', 'Sub Ledger.Code']
        validator_expected_columns = [
            'Department.Name', 'Sub Department.Name', 'Created user', 'Modified user',
            'Net amount', 'Location.Name', 'Crop.Name', 'Activity.Name', 'Function.Name',
            'FC-Vertical.Name', 'Region.Name', 'Zone.Name', 'Business Unit.Name',
            'Account.Code', 'Account2.Code', 'Sub Ledger.Code'
        ]

        missing_core_cols = [col for col in required_columns_for_processing if col not in df.columns]
        if missing_core_cols:
            st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Missing essential columns for processing: {", ".join(missing_core_cols)}. Cannot proceed with file "{uploaded_file.name}".</div>', unsafe_allow_html=True)
            return
        
        missing_validator_cols_list = [col for col in validator_expected_columns if col not in df.columns]
        if missing_validator_cols_list:
             st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> For file "{uploaded_file.name}", the following columns are missing and might affect validation accuracy: {", ".join(missing_validator_cols_list)}. Missing columns will be treated as blank during validation.</div>', unsafe_allow_html=True)
             for col in missing_validator_cols_list:
                 df[col] = ""

        summary_tab, exceptions_tab, data_tab = st.tabs(["📊 Validation Summary", "📋 Exception Records", "📖 Original Uploaded Data"])

        with summary_tab:
            st.markdown("#### 📊 File Information")
            col_info1, col_info2, col_info3, col_info4 = st.columns(4)
            display_metric("Total Records", f"{len(df):,}", container=col_info1)
            display_metric("Total Columns", len(df.columns), container=col_info2)
            display_metric("File Size", f"{uploaded_file.size / 1024:.1f} KB", container=col_info3)
            display_metric("Departments", df['Department.Name'].nunique() if 'Department.Name' in df.columns else "N/A", container=col_info4)
        
        exceptions_df_from_validation = pd.DataFrame()
        department_statistics = {}

        with st.spinner(f"🔍 Validating data for {uploaded_file.name}... This may take a moment for large files."):
            validator_instance = DataValidator(base_ref_path="reference_data")
            ref_data_loaded_check = any(isinstance(lst, list) and len(lst) > 0 for lst in validator_instance.ref_files.values())
            ledger_keys_loaded = len(validator_instance.valid_ledger_keys) > 0

            if not ref_data_loaded_check or not ledger_keys_loaded:
                error_msg = f"CRITICAL ERROR for file '{uploaded_file.name}': One or more reference data files could not be loaded or are all empty. This includes the essential 'Ledgersubledger mapping.xlsx'. Validation results will be highly inaccurate. Please check the 'reference_data' directory. See logs for details."
                with summary_tab:
                    st.error(error_msg)
                logging.critical(f"Reference data files appear to be uninitialized or empty in DataValidator for file '{uploaded_file.name}'. Aborting full validation display logic.")
                run_id_ref_error = db_manager.save_validation_run(
                    filename=uploaded_file.name, total_records=len(df),
                    total_exceptions=0, file_size=uploaded_file.size,
                    upload_time=selected_date
                )
                with summary_tab:
                    st.markdown(f"A validation run (ID: {run_id_ref_error}) was logged for '{uploaded_file.name}', but no exceptions were processed due to missing/empty reference data.")
                return

            exceptions_df_from_validation, department_statistics = validator_instance.validate_dataframe(df.copy())

        current_run_id = db_manager.save_validation_run(
            filename=uploaded_file.name,
            total_records=len(df),
            total_exceptions=len(exceptions_df_from_validation),
            file_size=uploaded_file.size,
            upload_time=selected_date
        )

        if not exceptions_df_from_validation.empty:
            db_manager.save_exceptions(current_run_id, exceptions_df_from_validation)
        
        exceptions_for_perf_calc = exceptions_df_from_validation
        if not exceptions_df_from_validation.empty and 'Created user' not in exceptions_df_from_validation.columns:
            logging.warning("'Created user' column missing in exceptions_df for user performance calculation. Adding it as blank.")
            exceptions_for_perf_calc = exceptions_df_from_validation.copy()
            exceptions_for_perf_calc['Created user'] = ""
        db_manager.save_user_performance(current_run_id, df, exceptions_for_perf_calc)
        
        if department_statistics:
            conn_dept_save = None
            try:
                conn_dept_save = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                cursor_dept_save = conn_dept_save.cursor()
                for dept_key_save, stats_val_save in department_statistics.items():
                    cursor_dept_save.execute('''
                        INSERT INTO department_summary (run_id, department, total_records, exception_records, exception_rate)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (current_run_id, dept_key_save, stats_val_save['total_records'], stats_val_save['exception_records'], stats_val_save['exception_rate']))
                conn_dept_save.commit()
            except sqlite3.Error as e_dept_save:
                logging.error(f"Error saving department_summary for run {current_run_id}: {e_dept_save}", exc_info=True)
                if conn_dept_save: conn_dept_save.rollback()
                st.error(f"Failed to save department summary for '{uploaded_file.name}': {e_dept_save}")
            finally:
                if conn_dept_save: conn_dept_save.close()

        with summary_tab:
            st.markdown("#### 🛠 Validation Results")
            if exceptions_df_from_validation.empty:
                st.markdown(f'<div class="success-box"><strong>✅ Perfect!</strong> No validation issues found in "{uploaded_file.name}"!</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> Found {len(exceptions_df_from_validation)} records with validation issues in "{uploaded_file.name}".</div>', unsafe_allow_html=True)
                
                col_res1, col_res2, col_res3 = st.columns(3)
                display_metric("Total Exceptions", f"{len(exceptions_df_from_validation):,}", container=col_res1)
                current_exception_rate = (len(exceptions_df_from_validation)/len(df)*100) if len(df) > 0 else 0
                display_metric("Exception Rate", f"{current_exception_rate:.2f}%", container=col_res2)
                
                avg_sev = 0.0
                if not exceptions_df_from_validation.empty and 'Severity' in exceptions_df_from_validation.columns and not exceptions_df_from_validation['Severity'].dropna().empty:
                    avg_sev = exceptions_df_from_validation['Severity'].mean()
                display_metric("Average Severity", f"{avg_sev:.2f}", container=col_res3)
        
        excel_report_data = create_excel_report(exceptions_df_from_validation, department_statistics, uploaded_file.name)
        
        if excel_report_data:
            db_manager.save_excel_report(current_run_id, excel_report_data)
            excel_report_data.seek(0)

        with exceptions_tab:
            if exceptions_df_from_validation.empty:
                 st.success("No exceptions to display.")
            else:
                st.markdown("##### 📋 Exception Records")
                exceptions_display_df = exceptions_df_from_validation.copy()
                
                exceptions_display_df.reset_index(inplace=True)
                exceptions_display_df.rename(columns={'index': 'id'}, inplace=True)
                exceptions_display_df['id'] = exceptions_display_df['id'] + 1
                exceptions_display_df['run_id'] = current_run_id

                def row_to_json(row):
                    return json.dumps(row.to_dict(), default=json_serializer_default)

                exceptions_display_df['original_row_data'] = exceptions_display_df.apply(row_to_json, axis=1)

                display_interactive_exceptions(exceptions_display_df, key_prefix="upload_view")
                
                if excel_report_data:
                    st.download_button(
                        label=f"📥 Download Validation Report for {uploaded_file.name}",
                        data=excel_report_data,
                        file_name=f"Validation_Report_{uploaded_file.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error(f"Could not generate the Excel report for download for '{uploaded_file.name}'.")
        
            st.markdown(f"#### 📖 Full Uploaded Dataset: `{uploaded_file.name}`")
        with data_tab:
            st.dataframe(df, use_container_width=True)

    except Exception as e_process:
        st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Failed to process file "{uploaded_file.name if uploaded_file else "N/A"}": {str(e_process)}. Check logs for more details.</div>', unsafe_allow_html=True)
        logging.exception(f"Unhandled error processing uploaded file {uploaded_file.name if uploaded_file else 'N/A'}: {e_process}")


def show_upload_page():
    st.markdown("### 📁 File Upload & Validation")

    # --- FINAL, ROBUST & STABLE FILE UPLOADER ---
    # We use a native Streamlit container to ensure no conflicts.
    # The uploader itself is the drop zone.

    with st.container(border=True):
        st.write("📤 **Drag & Drop Your Excel Files Here**")
        st.caption("or click 'Browse files' to select them from your computer")
        uploaded_files_list = st.file_uploader(
            "File Uploader", # The label is now minimal and mostly for internal use
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            label_visibility="collapsed" # We hide the simple label
        )

    custom_upload_date = st.date_input(
        "Override upload date (optional)",
        value=None,
        help="If you select a date, it will be used as the upload time for all files in this batch. If left empty, the current time will be used."
    )

    if uploaded_files_list:
        st.success(f"✅ {len(uploaded_files_list)} file(s) selected successfully! Processing...")
        for individual_uploaded_file in uploaded_files_list:
            with st.expander(f"⚙️ Processing: {individual_uploaded_file.name}", expanded=True):
                process_uploaded_file(individual_uploaded_file, selected_date=custom_upload_date)

    st.markdown("---")
    st.markdown("### 📝 Manual Data Entry & Validation")
    with st.form("manual_data_entry_form"):
        st.markdown("##### Enter Record Details (fields with * are mandatory):")

        manual_custom_date = st.date_input("Override record date (optional)", value=None, key="manual_date_override")

        cols_m1_row1, cols_m1_row2, cols_m1_row3 = st.columns(3)
        department_manual_input = cols_m1_row1.selectbox("Department*", [""] + sorted(["Parent Seed", "Production", "Processing", "Quality Assurance", "Finance & Account", "Human Resource", "Administration", "Information Technology", "Legal", "Accounts Receivable & MIS", "Seed Tech", "In Licensing & Procurement", "Breeding", "Breeding Support", "Trialing & PD", "Sales", "Marketing", "Management"]), key="manual_input_department", help="Select the department.")
        location_manual_input = cols_m1_row2.text_input("Location Name*", key="manual_input_location", help="Enter location name (e.g., Bandamailaram).")
        created_user_manual_input = cols_m1_row3.text_input("Created User*", key="manual_input_created_user", help="Enter the user ID of the creator.")

        cols_m2_row1, cols_m2_row2, cols_m2_row3 = st.columns(3)
        activity_manual_input = cols_m2_row1.text_input("Activity Name", key="manual_input_activity", help="Enter activity name if applicable.")
        net_amount_manual_input = cols_m2_row2.number_input("Net Amount", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="manual_input_net_amount", help="Enter the net amount.")
        sub_dept_manual_input = cols_m2_row3.text_input("Sub Department Name", value="", key="manual_input_sub_department", help="Enter sub-department if applicable.")

        with st.expander("Optional Fields for Manual Entry"):
            cols_m3_row1, cols_m3_row2, cols_m3_row3 = st.columns(3)
            modified_user_manual_input = cols_m3_row1.text_input("Modified User", value="", key="manual_input_modified_user")
            crop_manual_input = cols_m3_row2.text_input("Crop Name", value="", key="manual_input_crop")
            function_manual_input = cols_m3_row3.text_input("Function Name", value="", key="manual_input_function")
            cols_m4_row1, cols_m4_row2, cols_m4_row3, cols_m4_row4 = st.columns(4)
            vertical_manual_input = cols_m4_row1.text_input("FC-Vertical Name", value="", key="manual_input_vertical")
            region_manual_input = cols_m4_row2.text_input("Region Name", value="", key="manual_input_region")
            zone_manual_input = cols_m4_row3.text_input("Zone Name", value="", key="manual_input_zone")
            business_unit_manual_input = cols_m4_row4.text_input("Business Unit Name", value="", key="manual_input_business_unit")

        submitted_manual_record = st.form_submit_button("Validate & Submit Manual Record")
        if submitted_manual_record:
            if not department_manual_input or not location_manual_input or not created_user_manual_input:
                st.error("Department, Location Name, and Created User are mandatory fields for manual entry.")
            else:
                manual_validator = DataValidator(base_ref_path="reference_data")
                manual_ref_data_loaded = any(isinstance(lst, list) and len(lst) > 0 for lst in manual_validator.ref_files.values())
                if not manual_ref_data_loaded:
                    st.error("CRITICAL: Reference data not loaded. Manual validation cannot be performed accurately.")
                    logging.critical("Reference data not loaded for manual validation.")
                else:
                    manual_row_data = pd.Series({'Department.Name': department_manual_input, 'Location.Name': location_manual_input, 'Activity.Name': activity_manual_input, 'Created user': created_user_manual_input, 'Net amount': net_amount_manual_input, 'Sub Department.Name': sub_dept_manual_input, 'Modified user': modified_user_manual_input, 'Crop.Name': crop_manual_input, 'Function.Name': function_manual_input, 'FC-Vertical.Name': vertical_manual_input, 'Region.Name': region_manual_input, 'Zone.Name': zone_manual_input, 'Business Unit.Name': business_unit_manual_input})
                    manual_reasons, manual_severity = manual_validator.validate_row(department_manual_input, manual_row_data)
                    manual_df_for_db = pd.DataFrame([manual_row_data])
                    if manual_reasons:
                        st.warning(f"Validation Issues for manual entry: {'; '.join(manual_reasons)} (Severity: {manual_severity})")
                        manual_df_for_db['Exception Reasons'] = "; ".join(manual_reasons)
                        manual_df_for_db['Severity'] = manual_severity
                        manual_entry_run_id = db_manager.save_validation_run(
                            filename=f"Manual_Entry_Error_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                            total_records=1, total_exceptions=1, file_size=0,
                            upload_time=manual_custom_date
                        )
                        db_manager.save_exceptions(manual_entry_run_id, manual_df_for_db)
                        db_manager.save_user_performance(manual_entry_run_id, pd.DataFrame([manual_row_data]), manual_df_for_db)
                        st.info(f"Manual record submitted with noted validation issues (Run ID: {manual_entry_run_id}).")
                    else:
                        valid_manual_entry_run_id = db_manager.save_validation_run(
                            filename=f"Manual_Entry_OK_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                            total_records=1, total_exceptions=0, file_size=0,
                            upload_time=manual_custom_date
                        )
                        empty_exceptions_for_valid_manual = pd.DataFrame(columns=['Created user', 'Exception Reasons', 'Severity'])
                        db_manager.save_user_performance(valid_manual_entry_run_id, pd.DataFrame([manual_row_data]), empty_exceptions_for_valid_manual)
                        st.success(f"Manual record validated successfully and run logged (Run ID: {valid_manual_entry_run_id}).")

def show_analytics_page(start_date, end_date):
    st.markdown("### 📊 Dashboard Analytics")
    
    # The validation_history is now filtered by the main function based on sidebar selection
    validation_history = db_manager.get_validation_history()
    
    # Filter based on the selected date range
    if start_date and end_date:
        validation_history = validation_history[
            (validation_history['upload_time'].dt.date >= start_date) & 
            (validation_history['upload_time'].dt.date <= end_date)
        ]

    if validation_history.empty:
        st.info("No validation runs found for the selected period. Adjust the filter or upload a file to see analytics.")
        return

    st.markdown("#### 📈 Overall Statistics (for selected period)")
    stat_col1, stat_col2, stat_col3 = st.columns(3)
    total_runs = len(validation_history)
    display_metric("Total Validation Runs", f"{total_runs:,}", container=stat_col1)
    
    total_recs_processed = validation_history['total_records'].sum()
    display_metric("Total Records Processed", f"{total_recs_processed:,}", container=stat_col2)
    
    total_excs_found = validation_history['total_exceptions'].sum()
    display_metric("Total Exceptions Found", f"{total_excs_found:,}", container=stat_col3)

    if total_recs_processed > 0:
        st.markdown("#### 🔍 Data Quality Snapshot (for selected period)")
        labels = ['Records with Exceptions', 'Records without Exceptions']
        values = [total_excs_found, total_recs_processed - total_excs_found]
        colors = ['#FF6B6B', '#6BCB77']
        fig_overall_quality = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.4, marker_colors=colors,
                                                     hoverinfo='label+percent+value', textinfo='value+label',
                                                     insidetextorientation='radial', pull=[0.05, 0])])
        fig_overall_quality.update_layout(
            annotations=[dict(text='Quality', x=0.5, y=0.5, font_size=20, showarrow=False, font=PLOTLY_FONT)],
            legend_title_text='Record Status', margin=dict(t=30, b=30, l=10, r=10), font=PLOTLY_FONT,
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=400)
        st.plotly_chart(fig_overall_quality, use_container_width=True)

    run_ids_in_period = validation_history['id'].tolist()
    if run_ids_in_period:
        conn_analytics = None
        dept_summary_df = pd.DataFrame()
        try:
            conn_analytics = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
            query_placeholder = ','.join('?' for _ in run_ids_in_period)
            dept_summary_df = pd.read_sql_query(f"SELECT department, total_records, exception_records FROM department_summary WHERE run_id IN ({query_placeholder})", conn_analytics, params=run_ids_in_period)
        except sqlite3.Error as e_analytics_dept:
            logging.error(f"Error fetching department summary for analytics: {e_analytics_dept}", exc_info=True)
        finally:
            if conn_analytics: conn_analytics.close()

        if not dept_summary_df.empty:
            st.markdown(f"#### 🏭 Department Analysis (for selected period)")
            
            # Aggregate data for the period
            agg_dept_summary = dept_summary_df.groupby('department').agg(
                total_records=('total_records', 'sum'),
                exception_records=('exception_records', 'sum')
            ).reset_index()
            agg_dept_summary['exception_rate'] = (agg_dept_summary['exception_records'] / agg_dept_summary['total_records'] * 100).fillna(0)

            agg_dept_summary_sorted = agg_dept_summary.sort_values(by='exception_rate', ascending=False)
            
            fig_dept_analysis = px.bar(agg_dept_summary_sorted, x='department', y='exception_rate',
                                       labels={'exception_rate': 'Exception Rate (%)', 'department': 'Department'},
                                       color='exception_rate', color_continuous_scale='Sunsetdark', text_auto='.2f',
                                       hover_name='department', custom_data=['total_records', 'exception_records'])
            fig_dept_analysis.update_traces(hovertemplate="<b>%{hovertext}</b><br><br>" +
                                                        "Exception Rate: %{y:.2f}%<br>" +
                                                        "Total Records: %{customdata[0]:,}<br>" +
                                                        "Exception Records: %{customdata[1]:,}<extra></extra>")
            fig_dept_analysis.update_layout(title_text="Exception Rate by Department", title_x=0.5,
                                            title_font=PLOTLY_TITLE_FONT, xaxis_title="Department",
                                            yaxis_title="Exception Rate (%)", margin=dict(l=40, r=20, t=60, b=150),
                                            xaxis_tickangle=-45, font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)',
                                            plot_bgcolor='rgba(0,0,0,0)', yaxis=dict(gridcolor='#e9ecef'),
                                            xaxis=dict(showgrid=False))
            st.plotly_chart(fig_dept_analysis, use_container_width=True)
            
            st.markdown("##### 📋 Department Summary Table (for selected period)")
            st.dataframe(agg_dept_summary.style.format({"exception_rate": "{:.2f}%", "total_records":"{:,}","exception_records":"{:,}"}), use_container_width=True, hide_index=True)
        else:
            st.info(f"No department summary found for the selected period.")
    else:
        st.info("No run history available to analyze for the selected period.")


def show_trends_page(start_date, end_date):
    st.markdown("### 📈 Trends & History")
    trends_history_df = db_manager.get_validation_history()

    if start_date and end_date:
        trends_history_df = trends_history_df[
            (trends_history_df['upload_time'].dt.date >= start_date) & 
            (trends_history_df['upload_time'].dt.date <= end_date)
        ]

    if trends_history_df.empty:
        st.info("No historical data available for the selected period. Adjust the filter to see trends.")
        return

    trends_history_df = trends_history_df.sort_values(by='upload_time', ascending=True)
    
    fig_trends = go.Figure()
    theme_colors = {'exceptions': '#FF6B6B', 'records': '#6A89CC'}

    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'],
                                    y=trends_history_df['total_exceptions'],
                                    mode='lines+markers', name='Total Exceptions',
                                    line=dict(color=theme_colors['exceptions'], width=2.5, shape='spline'),
                                    marker=dict(symbol="circle", size=8, line=dict(width=1,color=theme_colors['exceptions'])),
                                    hovertemplate="<b>Total Exceptions</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>"))
    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'],
                                    y=trends_history_df['total_records'],
                                    mode='lines+markers', name='Total Records Processed',
                                    line=dict(color=theme_colors['records'], width=2.5, shape='spline'),
                                    marker=dict(symbol="square", size=8, line=dict(width=1,color=theme_colors['records'])),
                                    hovertemplate="<b>Total Records</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>",
                                    fill='tozeroy',
                                    fillcolor='rgba(106, 137, 204, 0.2)'
                                    ))
    fig_trends.update_layout(
        title_text="Validation Trends Over Time",
        title_x=0.5,
        title_font=PLOTLY_TITLE_FONT,
        xaxis_title="Upload Date",
        yaxis_title="Count",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    bgcolor='rgba(255,255,255,0.7)', bordercolor='rgba(0,0,0,0.05)', borderwidth=1),
        margin=dict(l=50, r=20, t=70, b=40),
        hovermode="x unified",
        font=PLOTLY_FONT,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da'),
        xaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da', rangeslider=dict(visible=True, thickness=0.05), type="date")
    )
    st.plotly_chart(fig_trends, use_container_width=True)
    st.markdown("##### 📜 Validation History Log")
    display_history_log_df = trends_history_df[['id', 'filename', 'upload_time', 'total_records', 'total_exceptions', 'file_size']].copy()
    display_history_log_df.columns = ['Run ID', 'Filename', 'Upload Time', 'Total Records', 'Total Exceptions', 'File Size (Bytes)']
    display_history_log_df['Upload Time'] = display_history_log_df['Upload Time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    st.dataframe(display_history_log_df.sort_values(by='Upload Time', ascending=False), use_container_width=True, hide_index=True)


def show_exception_details_page(start_date, end_date):
    st.markdown("### 🔍 Exception Details Viewer")
    try:
        validation_runs_list = db_manager.get_validation_history()
        
        if start_date and end_date:
            validation_runs_list = validation_runs_list[
                (validation_runs_list['upload_time'].dt.date >= start_date) & 
                (validation_runs_list['upload_time'].dt.date <= end_date)
            ]

        if validation_runs_list.empty:
            st.info("No validation runs found for the selected period.")
            return

        run_options_for_details = {f"Run {run_row['id']}: {run_row['filename']} ({pd.to_datetime(run_row['upload_time']).strftime('%Y-%m-%d %H:%M')})": run_row['id'] for _, run_row in validation_runs_list.iterrows()}
        selected_run_display_text = st.selectbox("Select Validation Run to View Details", options=list(run_options_for_details.keys()), index=0, key="exception_details_run_select")
        
        if selected_run_display_text:
            selected_run_id = run_options_for_details[selected_run_display_text]
            exceptions_for_run_df = db_manager.get_exceptions_by_run(selected_run_id)

            st.markdown(f"#### Exceptions Found for Run ID: {selected_run_id}")
            display_interactive_exceptions(exceptions_for_run_df, key_prefix="details_view")

            if exceptions_for_run_df.empty:
                return
            
            if 'Severity' in exceptions_for_run_df.columns and not exceptions_for_run_df['Severity'].dropna().empty:
                st.markdown("#### 📊 Severity Distribution")
                severity_value_counts = exceptions_for_run_df['Severity'].value_counts().reset_index()
                severity_value_counts.columns = ['Severity', 'Count']
                fig_severity_dist = px.bar(severity_value_counts.sort_values(by='Severity'),
                                           x='Severity', y='Count',
                                           title=f'Exception Severity Distribution (Run ID {selected_run_id})',
                                           color='Severity',
                                           color_continuous_scale=px.colors.sequential.OrRd,
                                           text_auto=True,
                                           hover_name='Severity',
                                           labels={'Count': 'Number of Exceptions'})
                fig_severity_dist.update_traces(hovertemplate="Severity: %{x}<br>Count: %{y}<extra></extra>")
                fig_severity_dist.update_layout(
                    title_x=0.5, title_font=PLOTLY_TITLE_FONT,
                    xaxis_title="Severity Score", yaxis_title="Number of Exceptions",
                    font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False),
                    margin=dict(l=40, r=20, t=60, b=40)
                )
                st.plotly_chart(fig_severity_dist, use_container_width=True)
            else:
                st.info("Severity data not available or no exceptions with severity for this run.")
            
            conn_details_dept = None; dept_stats_for_excel_df = pd.DataFrame()
            try:
                conn_details_dept = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                dept_stats_for_excel_df = pd.read_sql_query("SELECT department, total_records, exception_records, exception_rate FROM department_summary WHERE run_id = ?", conn_details_dept, params=(selected_run_id,))
            except sqlite3.Error as e_details_dept_excel:
                logging.error(f"Error fetching department stats for Excel report (run {selected_run_id}): {e_details_dept_excel}", exc_info=True)
            finally:
                if conn_details_dept: conn_details_dept.close()
            
            dept_stats_dict_for_excel = {row_stat['department']: {'total_records': row_stat['total_records'], 'exception_records': row_stat['exception_records'], 'exception_rate': row_stat['exception_rate']} for _, row_stat in dept_stats_for_excel_df.iterrows()}
            
            try:
                original_filename_for_report = validation_runs_list[validation_runs_list['id'] == selected_run_id]['filename'].iloc[0]
                excel_report_name = f"Exceptions_Report_Run_{selected_run_id}_{original_filename_for_report}.xlsx"
                excel_binary_data = create_excel_report(exceptions_for_run_df, dept_stats_dict_for_excel, excel_report_name)
                
                if excel_binary_data is None:
                    st.error("Failed to generate the Exceptions Report. Please check the application logs for more details.")
                    logging.error(f"Excel report generation returned None for run_id {selected_run_id}")
                else:
                    st.download_button(label="📥 Download Exceptions Report (Excel)", data=excel_binary_data, file_name=excel_report_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e_generate_report:
                st.error(f"Error generating or providing download for Excel report: {e_generate_report}")
                logging.exception(f"Error in Excel report generation/download setup for run_id {selected_run_id}: {e_generate_report}")
    except Exception as e_page_load:
        st.error(f"An error occurred while loading the exception details page: {e_page_load}")
        logging.exception(f"General error on exception details page: {e_page_load}")


def show_user_location_page(start_date, end_date):
    st.markdown("### 👤📍 User & Location Analysis")
    
    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[
            (history_df['upload_time'].dt.date >= start_date) & 
            (history_df['upload_time'].dt.date <= end_date)
        ]

    if history_df.empty: 
        st.info("No validation runs found for the selected period. Adjust filter to enable User & Location analysis.")
        return

    # Add "All Runs" option
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="user_location_run_selector")
    
    if not selected_run_display:
        st.info("Please select a validation run to view analysis.")
        return
    
    selected_run_id = run_options_with_all[selected_run_display]

    # Load data based on selection
    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        if selected_run_id == "all":
            run_ids_to_load = history_df['id'].tolist()
            all_exceptions = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
            ul_exceptions_df_master = pd.concat([df for df in all_exceptions if not df.empty], ignore_index=True)
        else:
            ul_exceptions_df_master = db_manager.get_exceptions_by_run(selected_run_id)

    if ul_exceptions_df_master.empty:
        st.success(f"No exceptions found for the selected scope. Nothing to analyze.")
        return

    st.markdown("#### Filters")
    user_list = ["All Users"] + sorted(ul_exceptions_df_master['Created user'].dropna().unique().tolist())
    selected_user = st.selectbox("Filter by User", options=user_list, key="user_location_user_filter")

    if selected_user == "All Users":
        ul_exceptions_df = ul_exceptions_df_master.copy()
    else:
        ul_exceptions_df = ul_exceptions_df_master[ul_exceptions_df_master['Created user'] == selected_user].copy()
    
    if ul_exceptions_df.empty:
        st.warning(f"No exceptions found for user '{selected_user}' in this scope.")
        return

    st.markdown(f"---")
    st.markdown(f"#### 📊 Exceptions by User and Location (Scope: {selected_run_display})")
    ul_required_cols = ['Created user', 'Location.Name', 'Severity', 'Exception Reasons']
    for col_ul in ul_required_cols:
        if col_ul not in ul_exceptions_df.columns:
            st.warning(f"Missing required column '{col_ul}' in exceptions data. Analysis might be incomplete.")
            ul_exceptions_df[col_ul] = pd.NA
    
    ul_summary_df = ul_exceptions_df.groupby(['Created user', 'Location.Name'], dropna=False).agg(
        Total_Severity_Score=('Severity', 'sum'),
        Number_of_Exceptions=('Exception Reasons', lambda x: x.notna().sum())
    ).reset_index().rename(columns={'Created user': 'User', 'Location.Name': 'Location'})
    
    ul_metric_col1, ul_metric_col2, ul_metric_col3 = st.columns(3)
    display_metric("Total Exceptions", f"{ul_exceptions_df['Exception Reasons'].notna().sum():,}", container=ul_metric_col1)
    display_metric("Unique Users with Exceptions", f"{ul_exceptions_df['Created user'].nunique()}", container=ul_metric_col2)
    ul_avg_severity = ul_exceptions_df['Severity'].mean() if not ul_exceptions_df['Severity'].dropna().empty else 0.0
    display_metric("Average Severity", f"{ul_avg_severity:.2f}", container=ul_metric_col3)
    
    st.markdown("##### 📋 User-Location Exception Summary Table")
    st.dataframe(ul_summary_df.sort_values(by="Number_of_Exceptions", ascending=False), use_container_width=True, hide_index=True)
    
    if not ul_summary_df.empty:
        fig_ul_exc_count_chart = px.bar(ul_summary_df, x='User', y='Number_of_Exceptions', color='Location',
                                        barmode='group', title="Exceptions by User and Location",
                                        labels={'Number_of_Exceptions': 'Number of Exceptions'},
                                        color_discrete_sequence=px.colors.qualitative.Pastel)
        fig_ul_exc_count_chart.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=120)
        )
        st.plotly_chart(fig_ul_exc_count_chart, use_container_width=True)
        
        fig_ul_severity_chart = px.bar(ul_summary_df, x='User', y='Total_Severity_Score', color='Location',
                                       barmode='stack', title="Total Severity by User and Location (Stacked)",
                                       labels={'Total_Severity_Score': 'Total Severity Score'},
                                       color_discrete_sequence=px.colors.qualitative.Antique)
        fig_ul_severity_chart.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=120)
        )
        st.plotly_chart(fig_ul_severity_chart, use_container_width=True)
        
    ul_error_types_df = pd.DataFrame(columns=['Error Type', 'Count'])
    if 'Exception Reasons' in ul_exceptions_df.columns and ul_exceptions_df['Exception Reasons'].notna().any():
        st.markdown("##### 📊 Top 10 Common Error Types for this Scope")
        ul_error_types_df = ul_exceptions_df['Exception Reasons'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        ul_error_types_df.columns = ['Error Type', 'Count']
        st.dataframe(ul_error_types_df.head(10), use_container_width=True, hide_index=True)
        
        fig_ul_top_errors = px.bar(ul_error_types_df.head(10), x='Error Type', y='Count',
                                   title="Top 10 Error Types by Occurrence",
                                   color='Count', color_continuous_scale=px.colors.sequential.Tealgrn, text_auto=True)
        fig_ul_top_errors.update_layout(
            title_x=0.5, title_font=PLOTLY_TITLE_FONT,
            font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45),
            margin=dict(l=40, r=20, t=60, b=150)
        )
        st.plotly_chart(fig_ul_top_errors, use_container_width=True)
        
    ul_user_risk_df = pd.DataFrame(columns=['User', 'Total Exceptions by User', 'Average Severity Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)'])
    if 'Created user' in ul_exceptions_df.columns and ul_exceptions_df['Created user'].notna().any():
        st.markdown("##### 🎯 User Risk Analysis for this Scope")
        ul_user_risk_df_agg = ul_exceptions_df.groupby('Created user', dropna=False).agg(Total_Exceptions_by_User=('Exception Reasons', lambda x: x.notna().sum()), Average_Severity_Score=('Severity', 'mean')).reset_index().rename(columns={'Created user': 'User'})
        ul_total_exceptions_in_run = ul_user_risk_df_agg['Total_Exceptions_by_User'].sum()
        ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] = (ul_user_risk_df_agg['Total_Exceptions_by_User'] / ul_total_exceptions_in_run * 100) if ul_total_exceptions_in_run > 0 else 0.0
        ul_user_risk_df_agg['Average_Severity_Score'] = ul_user_risk_df_agg['Average_Severity_Score'].fillna(0)
        ul_max_possible_severity = 10
        ul_user_risk_df_agg['Normalized_Severity'] = (ul_user_risk_df_agg['Average_Severity_Score'] / ul_max_possible_severity).clip(0,1) * 100
        ul_user_risk_df_agg['Risk_Score (0-100)'] = (ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] * 0.6 + ul_user_risk_df_agg['Normalized_Severity'] * 0.4).round(2)
        ul_user_risk_df = ul_user_risk_df_agg[['User', 'Total_Exceptions_by_User', 'Average_Severity_Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)']]
        st.dataframe(ul_user_risk_df.sort_values(by="Risk_Score (0-100)", ascending=False), use_container_width=True, hide_index=True)
        
    ul_excel_output = io.BytesIO()
    with pd.ExcelWriter(ul_excel_output, engine='openpyxl') as ul_writer:
        ul_summary_df.to_excel(ul_writer, sheet_name='User_Location_Summary', index=False)
        ul_error_types_df.to_excel(ul_writer, sheet_name='Common_Error_Types', index=False)
        ul_user_risk_df.to_excel(ul_writer, sheet_name='User_Risk_Analysis', index=False)
    ul_excel_output.seek(0)
    if ul_excel_output.getbuffer().nbytes > 0:
        file_name_suffix = f"Run_{selected_run_id}" if selected_run_id != 'all' else 'All_Runs'
        st.download_button(label="📥 Download User-Location Analysis Report", data=ul_excel_output, file_name=f"User_Location_Report_{file_name_suffix}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.warning("Could not generate the User-Location Analysis report for download (no data to write).")


def show_user_performance_page(start_date, end_date):
    st.markdown("### 👤📊 User Performance Dashboard")

    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty:
        st.info("No validation runs found for the selected period.")
        return

    # Add Run Selector
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run Scope", options=list(run_options_with_all.keys()), key="perf_run_selector")

    if not selected_run_display:
        return

    selected_run_id = run_options_with_all[selected_run_display]
    
    # Determine which run IDs are in the current scope
    if selected_run_id == "all":
        run_ids_in_scope = history_df['id'].tolist()
    else:
        run_ids_in_scope = [selected_run_id]

    if not run_ids_in_scope:
        st.info("No runs match the selected criteria.")
        return

    # Fetch unique users from the runs within the selected scope
    conn = None
    try:
        conn = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
        placeholders = ','.join('?' for _ in run_ids_in_scope)
        query = f"SELECT DISTINCT user FROM user_performance WHERE run_id IN ({placeholders}) AND user IS NOT NULL AND user != ''"
        users_df = pd.read_sql_query(query, conn, params=run_ids_in_scope)
        unique_sorted_users = sorted(users_df['user'].unique())
    except sqlite3.Error as e:
        st.error(f"Failed to fetch users for the selected scope: {e}")
        return
    finally:
        if conn: conn.close()

    if not unique_sorted_users:
        st.info("No user performance data available for the selected scope.")
        return

    # Add User Selector
    selected_user_for_perf = st.selectbox(
        "Select User to View",
        options=["All Users (Summary)"] + unique_sorted_users,
        key="user_performance_selector"
    )

    # --- "All Users" View ---
    if selected_user_for_perf == "All Users (Summary)":
        st.markdown(f"#### Performance Summary for **All Users** (Scope: {selected_run_display})")
        
        conn = None
        try:
            conn = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
            perf_query = f"SELECT * FROM user_performance WHERE run_id IN ({placeholders})"
            all_perf_df = pd.read_sql_query(perf_query, conn, params=run_ids_in_scope)
            exc_query = f"SELECT created_user, exception_reason FROM exceptions WHERE run_id IN ({placeholders})"
            all_exc_df = pd.read_sql_query(exc_query, conn, params=run_ids_in_scope)
        except sqlite3.Error as e:
            st.error(f"Database error fetching summary data: {e}")
            return
        finally:
            if conn: conn.close()
        
        if all_perf_df.empty:
            st.info("No performance records to summarize for this scope.")
            return

        total_exceptions = all_perf_df['exception_records'].sum()
        total_records = all_perf_df['total_records'].sum()
        overall_avg_exception_rate = (total_exceptions / total_records * 100) if total_records > 0 else 0

        # Changed from 3 columns to 4
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        display_metric("Total Exceptions (All Users)", f"{int(total_exceptions):,}", container=kpi_col1)
        display_metric("Overall Exception Rate", f"{overall_avg_exception_rate:.2f}%", container=kpi_col2)
        display_metric("Total Records Processed", f"{int(total_records):,}", container=kpi_col3)
        # Added Total Runs metric
        display_metric("Total Validation Runs", f"{len(run_ids_in_scope)}", container=kpi_col4)

        st.markdown("##### 📋 Performance by User")
        summary_by_user = all_perf_df.groupby('user').agg(
            total_records=('total_records', 'sum'),
            exception_records=('exception_records', 'sum')
        ).reset_index()
        summary_by_user['exception_rate'] = (summary_by_user['exception_records'] / summary_by_user['total_records'] * 100).fillna(0)
        st.dataframe(summary_by_user.sort_values('exception_rate', ascending=False), use_container_width=True, hide_index=True)

        fig_mistakes_summary = None
        if not all_exc_df.empty:
            st.markdown("##### 🛠️ Common Mistake Analysis (All Users in Scope)")
            all_mistakes_df = all_exc_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
            all_mistakes_df.columns = ['Mistake Type', 'Count']
            
            fig_mistakes_summary = px.bar(
                all_mistakes_df.head(15), 
                x='Mistake Type', 
                y='Count', 
                title="Top 15 Mistake Types Across All Users",
                template="plotly_white" 
            )
            st.plotly_chart(fig_mistakes_summary, use_container_width=True)

        st.markdown("---")
        with st.expander("📧 Email Overall Summary Report"):
            with st.form("email_summary_form"):
                to_input = st.text_input("To (separate multiple emails with a comma)")
                cc_input = st.text_input("CC (separate multiple emails with a comma)")
                subject = st.text_input("Subject", value="Overall User Performance Summary")
                
                summary_table_html = summary_by_user.sort_values('exception_rate', ascending=False).to_html(index=False, border=0, classes="dataframe", float_format='{:.2f}'.format)
                
                email_body_html = f"""
                <html style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;">
                <head>
                    <style>
                        .dataframe {{ border-collapse: collapse; width: 100%; margin-top: 20px; font-size: 14px; text-align: left; }}
                        .dataframe th {{ background-color: #007bff; color: white; padding: 10px; border: 1px solid #dddddd; }}
                        .dataframe td {{ padding: 8px; border: 1px solid #dddddd; }}
                        .dataframe tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    </style>
                </head>
                <body style="background-color: #f4f4f4; margin: 0; padding: 20px;">
                    <table width="100%" border="0" cellspacing="0" cellpadding="0">
                        <tr><td align="center">
                            <table width="600" border="0" cellspacing="0" cellpadding="20" style="background-color: #ffffff; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1);">
                                <tr><td>
                                    <h2 style="color: #333333; border-bottom: 2px solid #007bff; padding-bottom: 10px;">Overall User Performance Summary</h2>
                                    <table width="100%" border="0" cellspacing="0" cellpadding="10" style="margin-top: 20px; background-color: #eef7ff; border-radius: 5px;">
                                        <tr>
                                            <td align="center"><strong>Total Exceptions:</strong><br><span style="font-size: 24px; color: #d9534f;">{int(total_exceptions):,}</span></td>
                                            <td align="center"><strong>Overall Exception Rate:</strong><br><span style="font-size: 24px; color: #f0ad4e;">{overall_avg_exception_rate:.2f}%</span></td>
                                            <td align="center"><strong>Total Records:</strong><br><span style="font-size: 24px; color: #5bc0de;">{int(total_records):,}</span></td>
                                        </tr>
                                    </table>
                                    <h3 style="color: #333333; margin-top: 30px;">Performance by User</h3>
                                    {summary_table_html}
                                    <h3 style="color: #333333; margin-top: 30px;">Top 15 Common Mistake Types</h3>
                                    <img src="cid:mistake_summary_chart" width="560" style="max-width: 100%; border-radius: 5px;">
                                    <hr style="border: none; border-top: 1px solid #dddddd; margin: 30px 0;">
                                    <p style="color: #888888; font-size: 12px; text-align: center;"><i>This is an automated report from the Data Validation Dashboard.</i></p>
                                </td></tr>
                            </table>
                        </td></tr>
                    </table>
                </body>
                </html>
                """
                
                submitted = st.form_submit_button("Send Summary Email")
                if submitted:
                    to_recipients_list = [email.strip() for email in to_input.split(',') if email.strip()]
                    cc_recipients_list = [email.strip() for email in cc_input.split(',') if email.strip()]

                    if not to_recipients_list:
                        st.error("Please provide at least one recipient in the 'To' field.")
                    elif fig_mistakes_summary is None:
                        st.error("Cannot send email because the summary chart could not be generated.")
                    else:
                        with st.spinner("Preparing and sending email..."):
                            fig_summary_email = fig_mistakes_summary
                            
                            fig_summary_email.update_layout(
                                title_font={"size": 18},
                                xaxis_tickangle=-45
                            )
                            
                            mistakes_summary_bytes = fig_summary_email.to_image(format="png", scale=2)
                            images_to_embed = {"mistake_summary_chart": mistakes_summary_bytes}
                            
                            send_performance_email(
                                to_recipients=to_recipients_list, 
                                subject=subject, 
                                html_body=email_body_html, 
                                cc_recipients=cc_recipients_list, 
                                images=images_to_embed
                            )

    # --- Single User View ---
    else:
        st.markdown(f"#### Performance Details for **{selected_user_for_perf}** (Scope: {selected_run_display})")
        conn = None
        try:
            conn = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
            params = [selected_user_for_perf] + run_ids_in_scope
            placeholders_sql = ','.join('?' for _ in run_ids_in_scope)
            perf_query = f"SELECT up.*, vr.upload_time, vr.filename FROM user_performance up JOIN validation_runs vr ON up.run_id = vr.id WHERE up.user = ? AND up.run_id IN ({placeholders_sql}) ORDER BY vr.upload_time ASC"
            user_perf_df = pd.read_sql_query(perf_query, conn, params=params)
            exc_query = f"SELECT vr.upload_time, e.exception_reason FROM exceptions e JOIN validation_runs vr ON e.run_id = vr.id WHERE e.created_user = ? AND e.run_id IN ({placeholders_sql})"
            user_exc_df = pd.read_sql_query(exc_query, conn, params=params)
        except sqlite3.Error as e:
            st.error(f"Error fetching performance details: {e}")
            return
        finally:
            if conn: conn.close()
        
        if user_perf_df.empty:
            st.info(f"No performance data found for user '{selected_user_for_perf}' in the selected scope."); return

        # Changed from 3 columns to 4
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        total_exc = user_perf_df['exception_records'].sum()
        total_recs_by_user = user_perf_df['total_records'].sum()
        avg_rate = (total_exc / total_recs_by_user * 100) if total_recs_by_user > 0 else 0
        runs_involved = user_perf_df['run_id'].nunique()
        display_metric("Total Exceptions by User", f"{int(total_exc):,}", container=kpi_col1)
        display_metric("User Exception Rate", f"{avg_rate:.2f}%", container=kpi_col2)
        display_metric("Total Records by User", f"{int(total_recs_by_user):,}", container=kpi_col3)
        # Added Runs Involved metric
        display_metric("Runs Involved In User", f"{runs_involved}", container=kpi_col4)

        fig_mistakes = None
        fig_trend = None
        mistake_counts = pd.DataFrame() # Initialize empty dataframe

        if not user_exc_df.empty:
            st.markdown("##### 🛠️ Common Mistake Analysis")
            mistake_counts = user_exc_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
            mistake_counts.columns = ['Mistake Type', 'Count']
            fig_mistakes = px.pie(mistake_counts.head(10), names='Mistake Type', values='Count', 
                                  title="Top Mistake Types Distribution",
                                  color_discrete_sequence=px.colors.qualitative.Pastel)
            fig_mistakes.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig_mistakes, use_container_width=True)

        if not user_perf_df.empty:
            st.markdown("##### 📉 User Exception Rate Trend")
            # FIX: Use format='mixed' to handle different timestamp formats from the database
            user_perf_df['upload_time'] = pd.to_datetime(user_perf_df['upload_time'], format='mixed')
            fig_trend = px.line(user_perf_df, x='upload_time', y='exception_rate', markers=True, 
                                title="User Exception Rate Trend", labels={'upload_time': 'Date', 'exception_rate': 'Exception Rate (%)'})
            fig_trend.update_traces(line=dict(color='#007bff', width=3))
            st.plotly_chart(fig_trend, use_container_width=True)

        # Training Recommendations Section
        if not mistake_counts.empty:
            st.markdown("#### 📚 Training Recommendations")
            st.info("Based on the top 3 most frequent errors for this user, here are some suggested areas for training and review.")
            
            # Create an instance of the validator to access the training map
            validator = DataValidator(base_ref_path="reference_data")
            
            top_mistakes_for_training = mistake_counts.head(3)

            for index, row in top_mistakes_for_training.iterrows():
                error_type = row['Mistake Type']
                if error_type in validator.training_map:
                    recommendation = validator.training_map[error_type]
                    st.markdown(f"- **For '{error_type}':** {recommendation}")
                else:
                    # Fallback for any error messages not in the map
                    st.markdown(f"- **For '{error_type}':** No specific training suggestion available. Please review general data entry guidelines for this error type.")

        st.markdown("---")
        with st.expander("📧 Email Performance Report"):
            with st.form("email_performance_form"):
                to_input = st.text_input("To (separate multiple emails with a comma)", value=selected_user_for_perf if '@' in selected_user_for_perf else "")
                cc_input = st.text_input("CC (separate multiple emails with a comma)")
                subject = st.text_input("Subject", value=f"Performance Report for {selected_user_for_perf}")
                
                email_body_html = f"""
                <html style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;">
                <body style="background-color: #f4f4f4; margin: 0; padding: 20px;">
                    <table width="100%" border="0" cellspacing="0" cellpadding="0">
                        <tr><td align="center">
                            <table width="600" border="0" cellspacing="0" cellpadding="20" style="background-color: #ffffff; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1);">
                                <tr><td>
                                    <h2 style="color: #333333; border-bottom: 2px solid #007bff; padding-bottom: 10px;">Performance Summary for {selected_user_for_perf}</h2>
                                    <p style="color: #555555; font-size: 16px;">Here is a summary of the performance based on the selected data validation runs.</p>
                                    <table width="100%" border="0" cellspacing="0" cellpadding="10" style="margin-top: 20px; background-color: #eef7ff; border-radius: 5px;">
                                        <tr>
                                            <td align="center"><strong>Total Exceptions:</strong><br><span style="font-size: 24px; color: #d9534f;">{int(total_exc):,}</span></td>
                                            <td align="center"><strong>Avg. Exception Rate:</strong><br><span style="font-size: 24px; color: #f0ad4e;">{avg_rate:.2f}%</span></td>
                                            <td align="center"><strong>Runs Involved:</strong><br><span style="font-size: 24px; color: #5bc0de;">{runs_involved}</span></td>
                                        </tr>
                                    </table>
                                    <h3 style="color: #333333; margin-top: 30px;">Common Mistake Analysis</h3>
                                    <p style="color: #555555;">The chart below shows the distribution of the most common mistake types.</p>
                                    <img src="cid:mistake_analysis_chart" width="560" style="max-width: 100%; border-radius: 5px;">
                                    <h3 style="color: #333333; margin-top: 30px;">User Exception Rate Trend</h3>
                                    <p style="color: #555555;">This chart shows the trend of the user's exception rate over time.</p>
                                    <img src="cid:exception_rate_trend" width="560" style="max-width: 100%; border-radius: 5px;">
                                    <hr style="border: none; border-top: 1px solid #dddddd; margin: 30px 0;">
                                    <p style="color: #888888; font-size: 12px; text-align: center;"><i>This is an automated report from the Data Validation Dashboard.</i></p>
                                </td></tr>
                            </table>
                        </td></tr>
                    </table>
                </body>
                </html>
                """
                
                submitted = st.form_submit_button("Send Email")
                if submitted:
                    to_recipients_list = [email.strip() for email in to_input.split(',') if email.strip()]
                    cc_recipients_list = [email.strip() for email in cc_input.split(',') if email.strip()]

                    if not to_recipients_list:
                        st.error("Please provide at least one recipient in the 'To' field.")
                    elif fig_mistakes is None or fig_trend is None:
                        st.error("Cannot send email because one or more charts could not be generated.")
                    else:
                        with st.spinner("Preparing and sending email..."):
                            fig_mistakes_email = go.Figure(fig_mistakes)
                            fig_trend_email = go.Figure(fig_trend)
                            
                            email_layout_update = {"paper_bgcolor": "white", "plot_bgcolor": "white", "font": {"color": "#333333"}, "title_font": {"color": "#111111", "size": 18}}
                            fig_mistakes_email.update_layout(**email_layout_update)
                            fig_trend_email.update_layout(**email_layout_update)
                            fig_trend_email.update_xaxes(showgrid=True, gridwidth=1, gridcolor='#e0e0e0')
                            fig_trend_email.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#e0e0e0')
                            
                            mistakes_img_bytes = fig_mistakes_email.to_image(format="png", scale=2)
                            trend_img_bytes = fig_trend_email.to_image(format="png", scale=2)
                            
                            images_to_embed = {"mistake_analysis_chart": mistakes_img_bytes, "exception_rate_trend": trend_img_bytes}
                            
                            send_performance_email(
                                to_recipients=to_recipients_list, 
                                subject=subject, 
                                html_body=email_body_html, 
                                cc_recipients=cc_recipients_list, 
                                images=images_to_embed
                            )

def show_ledger_summary_page(start_date, end_date):
    st.markdown("### 🧾 Ledger & Sub-Ledger Exception Summary")

    account_names_df = load_account_name_mapping()
    subledger_names_df = load_subledger_name_mapping()
    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display this page because account or sub-ledger name mapping files could not be loaded.")
        return

    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty:
        st.info("No validation runs found for the selected period.")
        return

    # Add "All Runs" option
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="ledger_summary_run_selector")

    if not selected_run_display:
        return
    selected_run_id = run_options_with_all[selected_run_display]
    
    # Load data based on selection
    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        if selected_run_id == "all":
            run_ids_to_load = history_df['id'].tolist()
            all_exceptions = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
            exceptions_df = pd.concat([df for df in all_exceptions if not df.empty], ignore_index=True)
            total_records_in_scope = history_df['total_records'].sum()
        else:
            exceptions_df = db_manager.get_exceptions_by_run(selected_run_id)
            total_records_in_scope = history_df.loc[history_df['id'] == selected_run_id, 'total_records'].iloc[0]

    ledger_exception_reason = "Incorrect Ledger/Sub-Ledger Combination"
    ledger_errors_df = exceptions_df[exceptions_df['Exception Reasons'].str.contains(ledger_exception_reason, na=False)].copy()

    if ledger_errors_df.empty:
        st.success(f"No 'Incorrect Ledger/Sub-Ledger Combination' exceptions found for the selected scope.")
        return

    # Merge names
    if 'Account2.Code' in ledger_errors_df.columns:
        ledger_errors_df['merge_key_account'] = ledger_errors_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(ledger_errors_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in merged_df.columns:
            merged_df['Account2.Name'] = merged_df['Account2.Name_map'].fillna(merged_df['Account2.Name'])
            merged_df.drop(columns=['Account2.Name_map'], inplace=True)
    else: merged_df = ledger_errors_df
    
    if 'Sub Ledger.Code' in merged_df.columns:
        merged_df['merge_key_subledger'] = merged_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(merged_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in merged_df.columns:
            merged_df['SubLedger.Name'] = merged_df['SubLedger.Name_map'].fillna(merged_df['SubLedger.Name'])
            merged_df.drop(columns=['SubLedger.Name_map'], inplace=True)

    merged_df.drop(columns=[col for col in ['merge_key_account', 'merge_key_subledger'] if col in merged_df.columns], inplace=True)
    merged_df['Account2.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['SubLedger.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['Department.Name'].fillna('N/A - Department Missing', inplace=True)

    st.markdown("---")
    st.markdown("#### 📈 Metrics for Ledger/Sub-Ledger Exceptions")
    total_ledger_exceptions = len(merged_df)
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    display_metric("Total Transactions in Scope", f"{int(total_records_in_scope):,}", container=metric_col1)
    display_metric("Ledger/Sub-Ledger Exceptions", f"{total_ledger_exceptions:,}", container=metric_col2)
    exception_percentage = (total_ledger_exceptions / total_records_in_scope * 100) if total_records_in_scope > 0 else 0
    display_metric("% of Total", f"{exception_percentage:.2f}%", "of transactions in this scope have this error", container=metric_col3)
    
    st.markdown("#### 📊 Exception Count by Ledger and Department")
    summary_df = merged_df.groupby(['Account2.Name', 'SubLedger.Name', 'Department.Name']).size().reset_index(name='Count').sort_values(by='Count', ascending=False)
    summary_df.columns = ['Ledger Name', 'Sub-Ledger Name', 'Department', 'Count']
    chart_type = st.radio("Select Chart Type", ["Bar Chart", "Pie Chart"], horizontal=True, key="ledger_summary_chart_type")
    
    if chart_type == "Bar Chart":
        fig = px.bar(summary_df, x='Ledger Name', y='Count', color='Department', barmode='group', hover_data=['Sub-Ledger Name', 'Department'], title='Ledger/Sub-Ledger Exception Counts by Department')
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)
    if chart_type == "Pie Chart":
        pie_summary = summary_df.groupby('Ledger Name')['Count'].sum().reset_index().sort_values(by='Count', ascending=False)
        fig = px.pie(pie_summary.head(10), names='Ledger Name', values='Count', title='Top 10 Distribution of Exceptions by Ledger Name')
        st.plotly_chart(fig, use_container_width=True)
        
    st.markdown("#### 📋 Detailed Summary Table")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)


def show_user_ledger_exceptions_page(start_date, end_date):
    st.markdown("### 👤🧾 User-wise Ledger Exception Details")

    account_names_df = load_account_name_mapping()
    subledger_names_df = load_subledger_name_mapping()
    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display page: account or sub-ledger name mapping files could not be loaded.")
        return

    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty:
        st.info("No validation runs found for the selected period.")
        return

    # Add "All Runs" option
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="user_ledger_run_selector")

    if not selected_run_display:
        return
    selected_run_id = run_options_with_all[selected_run_display]

    # Load data based on selection
    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        if selected_run_id == "all":
            run_ids_to_load = history_df['id'].tolist()
            all_exceptions = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
            exceptions_df = pd.concat([df for df in all_exceptions if not df.empty], ignore_index=True)
        else:
            exceptions_df = db_manager.get_exceptions_by_run(selected_run_id)

    ledger_exception_reason = "Incorrect Ledger/Sub-Ledger Combination"
    ledger_errors_df = exceptions_df[exceptions_df['Exception Reasons'].str.contains(ledger_exception_reason, na=False)].copy()

    if ledger_errors_df.empty:
        st.success(f"No 'Incorrect Ledger/Sub-Ledger Combination' exceptions found for the selected scope.")
        return

    st.markdown("---")
    st.markdown("#### Filters")

    user_list = sorted(ledger_errors_df['Created user'].dropna().unique().tolist())
    if not user_list:
        st.warning("No users found with this type of exception in the selected scope.")
        return

    selected_user = st.selectbox("Select User", options=user_list, key="user_ledger_user_selector")

    if not selected_user:
        st.info("Please select a user to see their specific exceptions.")
        return

    user_specific_errors_df = ledger_errors_df[ledger_errors_df['Created user'] == selected_user].copy()

    if user_specific_errors_df.empty:
        st.warning(f"No ledger/sub-ledger exceptions found for user '{selected_user}' in this scope.")
        return

    # Merge names
    if 'Account2.Code' in user_specific_errors_df.columns:
        user_specific_errors_df['merge_key_account'] = user_specific_errors_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(user_specific_errors_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in merged_df.columns:
            merged_df['Account2.Name'] = merged_df['Account2.Name_map'].fillna(merged_df['Account2.Name'])
            merged_df.drop(columns=['Account2.Name_map'], inplace=True)
    else: merged_df = user_specific_errors_df

    if 'Sub Ledger.Code' in merged_df.columns:
        merged_df['merge_key_subledger'] = merged_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(merged_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in merged_df.columns:
            merged_df['SubLedger.Name'] = merged_df['SubLedger.Name_map'].fillna(merged_df['SubLedger.Name'])
            merged_df.drop(columns=['SubLedger.Name_map'], inplace=True)

    merged_df.drop(columns=[col for col in ['merge_key_account', 'merge_key_subledger'] if col in merged_df.columns], inplace=True)
    merged_df['Account2.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['SubLedger.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['Department.Name'].fillna('N/A - Department Missing', inplace=True)

    st.markdown("---")
    st.markdown(f"#### 📊 Visual Summary for **{selected_user}**")

    plot_df = merged_df.groupby(['Account2.Name', 'SubLedger.Name']).size().reset_index(name='Count')
    plot_df['Combination'] = plot_df['Account2.Name'] + " / " + plot_df['SubLedger.Name']
    
    fig = px.bar(plot_df.sort_values('Count', ascending=False).head(15), x='Count', y='Combination', orientation='h',
                 title=f"Top 15 Ledger/Sub-Ledger Exception Counts for {selected_user}",
                 labels={'Combination': 'Ledger / Sub-Ledger Combination', 'Count': 'Number of Exceptions'},
                 text='Count', color='Count', color_continuous_scale=px.colors.sequential.Plasma)
    fig.update_layout(yaxis={'categoryorder':'total ascending'}, title_x=0.5, title_font=PLOTLY_TITLE_FONT, font=PLOTLY_FONT,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(l=40, r=20, t=60, b=40))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown(f"#### 📋 Detailed Exceptions by **{selected_user}**")
    display_interactive_exceptions(merged_df, key_prefix="user_ledger_view")


def show_location_expenses_page(start_date, end_date):
    st.markdown("### 📍 Location Expenses")
    st.markdown("Analyze expenses aggregated by Ledger and Sub-Ledger across all locations. Select a single validation run or 'All Runs' for an aggregate view over the selected time period.")

    # 1. Select a Run or "All Runs"
    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[
            (history_df['upload_time'].dt.date >= start_date) & 
            (history_df['upload_time'].dt.date <= end_date)
        ]
        
    if history_df.empty:
        st.info("No validation runs found for the selected period. Upload a file to see analytics.")
        return

    # Add "All Runs" option to the beginning of the list
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}

    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="location_exp_run_selector")

    if not selected_run_display:
        return
    
    selected_run_id = run_options_with_all[selected_run_display]

    # 2. Load Data and Mappings
    with st.spinner(f"Loading expense data for '{selected_run_display}'..."):
        all_dfs = []
        if selected_run_id == "all":
            run_ids_to_load = history_df['id'].tolist()
        else:
            run_ids_to_load = [selected_run_id]

        for run_id in run_ids_to_load:
            df = db_manager.get_exceptions_by_run(run_id)
            if not df.empty:
                all_dfs.append(df)
        
        if not all_dfs:
            st.success(f"No expense data (from exception records) found for the selected scope.")
            return
            
        expense_data_df = pd.concat(all_dfs, ignore_index=True)
        account_names_df = load_account_name_mapping()
        subledger_names_df = load_subledger_name_mapping()

    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display page because account or sub-ledger name mapping files are missing from the 'reference_data' directory.")
        return

    # 3. Merge to get Names and prepare data
    if 'Net amount' in expense_data_df.columns:
        expense_data_df['Net amount'] = pd.to_numeric(expense_data_df['Net amount'], errors='coerce').fillna(0)
    else:
        st.error("The 'Net amount' column is missing from the data.")
        return
        
    # Merge Account Names
    if 'Account2.Code' in expense_data_df.columns and not expense_data_df.empty:
        expense_data_df['merge_key_account'] = expense_data_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        expense_data_df = pd.merge(expense_data_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in expense_data_df.columns:
            expense_data_df['Account2.Name'] = expense_data_df['Account2.Name'].fillna(expense_data_df['Account2.Name_map'])
            expense_data_df.drop(columns=[col for col in ['Account2.Name_map', 'Account2.Code_map', 'merge_key_account'] if col in expense_data_df.columns], inplace=True)

    # Merge Sub-Ledger Names
    if 'Sub Ledger.Code' in expense_data_df.columns and not expense_data_df.empty:
        expense_data_df['merge_key_subledger'] = expense_data_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        expense_data_df = pd.merge(expense_data_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in expense_data_df.columns:
            expense_data_df['SubLedger.Name'] = expense_data_df['SubLedger.Name'].fillna(expense_data_df['SubLedger.Name_map'])
            expense_data_df.drop(columns=[col for col in ['SubLedger.Name_map', 'Sub Ledger.Code_map', 'merge_key_subledger'] if col in expense_data_df.columns], inplace=True)

    # Fill NA for key fields to ensure they appear in filters and pivots
    for col, fill_val in {
        'Account2.Name': 'N/A - Name Not Mapped',
        'SubLedger.Name': 'N/A - Name Not Mapped',
        'Department.Name': 'N/A - Department Missing',
        'Sub Department.Name': 'N/A',
        'Location.Name': 'N/A - Location Missing'
    }.items():
        if col not in expense_data_df.columns: expense_data_df[col] = fill_val
        else: expense_data_df[col].fillna(fill_val, inplace=True)

    # 4. Filters
    st.markdown("---")
    st.markdown("#### 🔍 Filters")
    
    filter_col1, filter_col2 = st.columns(2)
    
    with filter_col1:
        departments = ["All"] + sorted(expense_data_df['Department.Name'].unique().tolist())
        selected_dept = st.selectbox("Filter by Department", options=departments, key="loc_exp_dept_filter")

    # Filter dataframe based on department selection to populate sub-department options
    df_for_sub_dept_options = expense_data_df[expense_data_df['Department.Name'] == selected_dept] if selected_dept != "All" else expense_data_df

    with filter_col2:
        sub_dept_options = ["All"] + sorted([sd for sd in df_for_sub_dept_options['Sub Department.Name'].unique() if sd != 'N/A'])
        if 'N/A' in df_for_sub_dept_options['Sub Department.Name'].unique():
             sub_dept_options.append('N/A')
        selected_sub_dept = st.selectbox("Filter by Sub-Department", options=sub_dept_options, key="loc_exp_sub_dept_filter")

    # Apply final sub-department filter
    final_filtered_df = df_for_sub_dept_options[df_for_sub_dept_options['Sub Department.Name'] == selected_sub_dept] if selected_sub_dept != "All" else df_for_sub_dept_options
    
    if final_filtered_df.empty:
        st.warning("No data matches the current filter criteria.")
        return

    # 5. Create and Display Pivot Table
    st.markdown("---")
    st.markdown("#### 💰 Ledger Expense Summary by Location")
    st.markdown("This table shows the total expense amount for each Ledger/Sub-Ledger combination, broken down by Location.")
    
    try:
        # Dynamically set the index for the pivot table
        pivot_index = ['Account2.Name', 'SubLedger.Name']
        if selected_dept == "All":
            pivot_index.insert(0, 'Department.Name')
            if selected_sub_dept == "All":
                pivot_index.insert(1, 'Sub Department.Name')
        
        pivot_table = pd.pivot_table(
            final_filtered_df,
            values='Net amount',
            index=pivot_index,
            columns='Location.Name',
            aggfunc=np.sum,
            fill_value=0,
            margins=True, # Add row and column totals (Grand Total)
            margins_name="Grand Total"
        )
        
        st.dataframe(pivot_table.style.format("{:,.2f}"), use_container_width=True)

        # 6. Download Button
        @st.cache_data
        def convert_df_to_excel(df_to_convert):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_to_convert.to_excel(writer, sheet_name='Ledger_Expense_by_Location')
            return output.getvalue()

        excel_data = convert_df_to_excel(pivot_table)
        
        st.download_button(
            label="📥 Download Summary as Excel",
            data=excel_data,
            file_name=f"Ledger_Expense_by_Location_{selected_run_display.replace(':', '')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_loc_exp"
        )

    except Exception as e:
        st.error(f"Could not generate the pivot table. Error: {e}")
        logging.error(f"Error creating pivot table on location expense page: {e}", exc_info=True)


def show_report_archive_page():
    st.markdown("### 🗂️ Report Archive")
    st.markdown("Access originally generated validation reports or create new aggregate reports from past runs.")
    
    history_df = db_manager.get_validation_history()
    if history_df.empty:
        st.warning("No validation runs found in the archive.")
        return

    st.markdown("---")
    st.markdown("#### 📥 Run-wise Download")
    st.info("Select a single past validation run to download the exact Excel report that was generated at that time.")

    run_options_single = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    selected_run_display = st.selectbox("Select a Validation Run to Download", options=list(run_options_single.keys()), key="archive_run_select")

    if selected_run_display:
        selected_run_id = run_options_single[selected_run_display]
        report_data, filename = db_manager.get_archived_report(selected_run_id)
        if report_data:
            st.download_button(
                label=f"📥 Download Archived Report for Run {selected_run_id}",
                data=report_data,
                file_name=f"Archived_Report_Run_{selected_run_id}_{filename}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_btn_{selected_run_id}"
            )
        else:
            st.error("No archived report found for this run. This can happen if the run occurred before the archive feature was added, or if the report was empty and not saved.")

    st.markdown("---")
    st.markdown("#### 📦 Aggregate-wise Download")
    st.info("Select multiple validation runs to combine all their exceptions into a single, newly-generated summary report.")

    run_options_multi = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    selected_multi_runs = st.multiselect("Select runs to aggregate", options=list(run_options_multi.keys()), key="archive_multi_run_select")

    if st.button("Generate & Download Aggregate Report", key="agg_report_btn"):
        if not selected_multi_runs:
            st.warning("Please select at least two runs to aggregate.")
        else:
            with st.spinner("Aggregating data and generating report..."):
                run_ids_to_agg = [run_options_multi[key] for key in selected_multi_runs]
                
                all_exceptions_list = []
                all_dept_stats_list = []
                
                conn = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                try:
                    for run_id in run_ids_to_agg:
                        exceptions_df = db_manager.get_exceptions_by_run(run_id)
                        if not exceptions_df.empty:
                            all_exceptions_list.append(exceptions_df)
                        
                        dept_summary_df = pd.read_sql_query("SELECT * FROM department_summary WHERE run_id = ?", conn, params=(run_id,))
                        if not dept_summary_df.empty:
                            all_dept_stats_list.append(dept_summary_df)
                finally:
                    conn.close()

                if not all_exceptions_list:
                    st.error("None of the selected runs contained any exception records to aggregate.")
                else:
                    combined_exceptions_df = pd.concat(all_exceptions_list, ignore_index=True)
                    
                    final_dept_stats_dict = {}
                    if all_dept_stats_list:
                        combined_dept_stats_df = pd.concat(all_dept_stats_list, ignore_index=True)
                        agg_dept_stats = combined_dept_stats_df.groupby('department').agg(
                            total_records=('total_records', 'sum'),
                            exception_records=('exception_records', 'sum')
                        ).reset_index()
                        agg_dept_stats['exception_rate'] = (agg_dept_stats['exception_records'] / agg_dept_stats['total_records'] * 100).fillna(0)
                        final_dept_stats_dict = {row['department']: {'total_records': row['total_records'], 'exception_records': row['exception_records'], 'exception_rate': row['exception_rate']} for _, row in agg_dept_stats.iterrows()}
                    
                    agg_report_data = create_excel_report(combined_exceptions_df, final_dept_stats_dict, "AggregateReport")
                    
                    if agg_report_data:
                        st.download_button(
                            label="✔️ Click to Download Aggregate Report",
                            data=agg_report_data,
                            file_name=f"Aggregate_Report_Runs_{'_'.join(map(str, run_ids_to_agg))}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error("Failed to generate the aggregate report.")

def show_data_management_page():
    st.markdown("### 🗑️ Data Management")
    st.warning("🚨 **Caution:** Actions on this page are permanent and cannot be undone.")

    st.markdown("---")
    st.markdown("#### ❌ Delete a Specific Validation Run")
    st.markdown("Select a validation run to permanently delete it and all its associated data (exceptions, summaries, etc.).")

    history_df = db_manager.get_validation_history()

    if history_df.empty:
        st.info("No validation runs available to delete.")
        return

    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    
    if 'run_to_delete_display' not in st.session_state:
        st.session_state.run_to_delete_display = None

    selected_run_display = st.selectbox(
        "Select a run to delete", 
        options=[None] + list(run_options.keys()), 
        key="delete_run_select"
    )

    if selected_run_display:
        selected_run_id = run_options[selected_run_display]
        run_details = history_df[history_df['id'] == selected_run_id].iloc[0]

        with st.expander("⚠️ Review Run Details Before Deleting", expanded=True):
            st.markdown(f"- **Run ID:** `{run_details['id']}`")
            st.markdown(f"- **Filename:** `{run_details['filename']}`")
            st.markdown(f"- **Upload Time:** `{pd.to_datetime(run_details['upload_time']).strftime('%Y-%m-%d %H:%M:%S')}`")
            st.markdown(f"- **Total Records:** `{run_details['total_records']}`")
            st.markdown(f"- **Total Exceptions:** `{run_details['total_exceptions']}`")

            st.error("This action will permanently delete the run entry, all its exceptions, department summaries, and user performance records.")
            
            confirm_delete = st.checkbox("I understand this is permanent and want to delete this run.", key=f"confirm_delete_{selected_run_id}")

            if st.button(f"Permanently Delete Run {selected_run_id}", disabled=not confirm_delete):
                with st.spinner(f"Deleting run {selected_run_id}..."):
                    success = db_manager.delete_run(selected_run_id)
                    if success:
                        st.success(f"Successfully deleted Run ID {selected_run_id}.")
                        st.rerun()
                    else:
                        st.error(f"Failed to delete Run ID {selected_run_id}. Check logs for details.")
    else:
        st.info("Select a run from the dropdown list to manage it.")


def show_settings_page():
    st.markdown("### ⚙️ Application Settings")
    st.markdown("#### 🛠 Database Management")
    st.warning("🚨 **Caution:** Clearing the database is irreversible and will delete all stored validation history, exceptions, and performance data. This action cannot be undone.")
    if 'confirm_db_clear_pressed_once' not in st.session_state: st.session_state.confirm_db_clear_pressed_once = False
    if st.button("🗑️ Clear Entire Database", type="primary", key="clear_db_initial_button"): st.session_state.confirm_db_clear_pressed_once = True
    if st.session_state.confirm_db_clear_pressed_once:
        st.error("ARE YOU ABSOLUTELY SURE? This will delete all data.")
        if st.button("✅ Yes, I am sure. Delete all data.", key="clear_db_confirm_button"):
            conn_settings_clear = None
            try:
                conn_settings_clear = sqlite3.connect(db_manager.db_path, timeout=DB_TIMEOUT)
                cursor_settings_clear = conn_settings_clear.cursor()
                tables_to_be_cleared = ["validation_runs", "exceptions", "department_summary", "user_performance"]
                for table_name in tables_to_be_cleared: cursor_settings_clear.execute(f"DELETE FROM {table_name}"); logging.info(f"Cleared table: {table_name}")
                conn_settings_clear.commit(); st.success("Database cleared successfully. Please refresh the application to see the changes across all pages."); logging.info("Database cleared successfully by user from settings page.")
                st.session_state.confirm_db_clear_pressed_once = False
            except sqlite3.Error as e_settings_clear: st.error(f"Failed to clear database: {str(e_settings_clear)}"); logging.exception("Failed to clear database from settings page.");
            finally:
                if conn_settings_clear: conn_settings_clear.close()
        elif st.button("❌ No, cancel.", key="clear_db_cancel_button"): st.session_state.confirm_db_clear_pressed_once = False; st.info("Database clearing cancelled.")
    st.markdown("---")
    st.markdown("#### ℹ️ About This Dashboard")
    dashboard_version = "3.7.0" 
    st.markdown(f"""**Data Validation Dashboard - Version {dashboard_version}**\n\nThis application is designed to help users validate data from Excel files against a predefined set of business rules. It provides insights into data quality, tracks validation history, and analyzes user performance.\n\n- **Upload & Validate:** Upload your Excel files with an optional date override, or enter data manually for instant validation.\n- **Dashboard Analytics:** Get an overview of validation statistics and department-level performance for selected time periods.\n- **Trends & History:** Track validation metrics over time, with date filtering.\n- **Exception Details:** Dive deep into specific validation runs and their exceptions with date filtering. Now with interactive drill-down to see original data!\n- **Location Expenses**: Analyze expenses by location with departmental filters and date filtering.\n- **Ledger/Sub-Ledger Summary**: Analyze ledger-related exceptions within a specific run and time period.\n- **User-wise Ledger Exceptions**: View detailed ledger exceptions by user within a specific run and time period.\n- **User & Location Analysis:** Analyze exceptions based on users and locations within a specific run and time period.\n- **User Performance:** Monitor individual user performance and get training recommendations, now with date filtering.\n- **Report Archive**: Access a permanent library of all past validation reports.\n- **Data Management**: Select and delete specific validation runs from the database.\n\nBuilt with Streamlit, Pandas, Plotly, and SQLite.""")
    st.markdown(f"**SQLite Database Path:** `{os.path.abspath(db_manager.db_path)}`")

# SNIPPET 1: Replace your existing send_performance_email function

def send_performance_email(to_recipients, subject, html_body, cc_recipients=None, images=None):
    """
    Connects to the SMTP server and sends a multipart HTML email with embedded images.

    Args:
        to_recipients (list): A list of email addresses for the 'To' field.
        subject (str): The subject line of the email.
        html_body (str): The HTML content of the email body.
        cc_recipients (list, optional): A list of email addresses for the 'CC' field. Defaults to None.
        images (dict, optional): A dictionary of images to embed. Defaults to None.
    """
    try:
        # Fetch credentials from Streamlit secrets
        sender_email = st.secrets["email_credentials"]["sender_email"]
        password = st.secrets["email_credentials"]["sender_password"]
        smtp_server = st.secrets["email_credentials"]["smtp_server"]
        smtp_port = st.secrets["email_credentials"]["smtp_port"]

        if not to_recipients:
            st.error("No recipients specified in the 'To' field.")
            return
            
        if cc_recipients is None:
            cc_recipients = []

        # Combine all recipients for the mail server
        all_recipients_list = to_recipients + cc_recipients

        # Create the email message
        message = MIMEMultipart("related")
        message["From"] = sender_email
        message["To"] = ", ".join(to_recipients)
        if cc_recipients:
            message["Cc"] = ", ".join(cc_recipients)
        message["Subject"] = subject

        # Attach the HTML body and images
        message.attach(MIMEText(html_body, "html"))
        if images:
            for cid, img_data in images.items():
                img = MIMEImage(img_data, _subtype="png")
                img.add_header('Content-ID', f'<{cid}>')
                message.attach(img)

        # Send the email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, all_recipients_list, message.as_string())
        
        st.success(f"Performance report successfully sent to: {', '.join(all_recipients_list)}")
        logging.info(f"Successfully sent performance email to: {', '.join(all_recipients_list)}")

    except KeyError:
        st.error("Email credentials are not configured in st.secrets. Please check your secrets.toml file.")
        logging.error("Email credentials missing in st.secrets.")
    except Exception as e:
        st.error(f"Failed to send email: {e}")
        logging.error(f"Failed to send email: {e}", exc_info=True)

def add_date_filters_to_sidebar(history_df, key_suffix=""):
    """Adds date filter widgets to the sidebar and returns start and end dates."""
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🗓️ Date Filter")

    if history_df.empty:
        st.sidebar.warning("No historical data to filter.")
        return None, None

    history_df['upload_time'] = pd.to_datetime(history_df['upload_time'])
    min_date = history_df['upload_time'].min().date()
    max_date = history_df['upload_time'].max().date()

    filter_type = st.sidebar.radio(
        "Filter by:",
        ("All Time", "Month", "Custom Range"),
        key=f"date_filter_type_{key_suffix}"
    )

    start_date, end_date = None, None

    if filter_type == "Month":
        history_df['month_year'] = history_df['upload_time'].dt.to_period('M')
        month_options = sorted(history_df['month_year'].unique(), reverse=True)
        selected_month = st.sidebar.selectbox(
            "Select Month",
            options=month_options,
            format_func=lambda x: x.strftime('%B %Y'),
            key=f"month_select_{key_suffix}"
        )
        if selected_month:
            start_date = selected_month.start_time.date()
            end_date = selected_month.end_time.date()

    elif filter_type == "Custom Range":
        start_date_input = st.sidebar.date_input("Start date", min_date, min_value=min_date, max_value=max_date, key=f"start_date_{key_suffix}")
        end_date_input = st.sidebar.date_input("End date", max_date, min_value=min_date, max_value=max_date, key=f"end_date_{key_suffix}")
        if start_date_input and end_date_input:
            start_date = start_date_input
            end_date = end_date_input
    else:  # All Time
        start_date = min_date
        end_date = max_date

    return start_date, end_date


def main():
    st.markdown("<h1>🎯 Data Validation Dashboard</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align:center;color:#718096;font-size:1.2rem;margin-bottom:2rem;'>Upload expense reports or enter data manually for validation and insights.</p>", unsafe_allow_html=True)
    
    st.sidebar.image("https://vnrseeds.com/wp-content/uploads/2022/11/logo.png", width=150)
    st.sidebar.markdown(f"**Version**: 3.7.0")
    st.sidebar.markdown("---")
    
    with st.sidebar:
        page_navigation_options = [
            "🏠 Upload & Validate",
            "📊 Dashboard Analytics",
            "📈 Trends & History",
            "📋 Exception Details",
            "📍 Location Expenses",
            "🧾 Ledger/Sub-Ledger Summary",
            "👤🧾 User-wise Ledger Exceptions",
            "👤📍 User & Location Analysis",
            "👤📊 User Performance",
            "🗂️ Report Archive",
            "🗑️ Data Management",
            "⚙️ Settings"
        ]
        selected_page = st.radio("Main Navigation:", page_navigation_options, label_visibility="collapsed")

    # Centralized date filter logic
    start_date, end_date = None, None
    pages_with_filter = [
        "📊 Dashboard Analytics", "📈 Trends & History", "📋 Exception Details", 
        "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", 
        "👤📍 User & Location Analysis", "👤📊 User Performance"
    ]

    if selected_page in pages_with_filter:
        history_df = db_manager.get_validation_history()
        start_date, end_date = add_date_filters_to_sidebar(history_df, key_suffix=selected_page.replace(" ", "_"))

    # Page routing
    if selected_page == "🏠 Upload & Validate": show_upload_page()
    elif selected_page == "📊 Dashboard Analytics": show_analytics_page(start_date, end_date)
    elif selected_page == "📈 Trends & History": show_trends_page(start_date, end_date)
    elif selected_page == "📋 Exception Details": show_exception_details_page(start_date, end_date)
    elif selected_page == "📍 Location Expenses": show_location_expenses_page(start_date, end_date)
    elif selected_page == "🧾 Ledger/Sub-Ledger Summary": show_ledger_summary_page(start_date, end_date)
    elif selected_page == "👤🧾 User-wise Ledger Exceptions": show_user_ledger_exceptions_page(start_date, end_date)
    elif selected_page == "👤📍 User & Location Analysis": show_user_location_page(start_date, end_date)
    elif selected_page == "👤📊 User Performance": show_user_performance_page(start_date, end_date)
    elif selected_page == "🗂️ Report Archive": show_report_archive_page()
    elif selected_page == "🗑️ Data Management": show_data_management_page()
    elif selected_page == "⚙️ Settings": show_settings_page()


if __name__ == "__main__":
    log_file_path = "dashboard.log"
    try:
        log_dir = os.path.dirname(os.path.abspath(log_file_path)) or '.'
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        if not os.access(log_dir, os.W_OK):
            raise IOError(f"Log file directory '{log_dir}' is not writable.")
        
        with open(log_file_path, "a") as f:
            f.write(f"--- Log session started at {datetime.now()} ---\n")
    except (IOError, OSError) as e:
        print(f"Warning: Log file '{log_file_path}' is not writable or cannot be created. Logging to console only. Error: {e}")
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[logging.StreamHandler()]
        )
    else:
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[
                logging.FileHandler(log_file_path, mode='a', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
    
    logging.info("Dashboard application started.")
    main()