import bcrypt
import streamlit as st
import pandas as pd
import mysql.connector # UPDATED: Replaced sqlite3 with mysql.connector
from mysql.connector import errorcode # UPDATED: Added for specific MySQL error handling
import io
import sqlite3  # <--- MAKE SURE THIS LINE EXISTS
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import logging
import os
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication # <-- ADD THIS LINE
import json
import math
import concurrent.futures
import time
import hashlib


# Helper function to serialize objects not recognized by default json.dumps
# Helper function to serialize objects not recognized by default json.dumps
# Helper function to serialize objects not recognized by default json.dumps
# Helper function to serialize objects not recognized by default json.dumps
def json_serializer_default(obj):
    # Final production version of the serializer function.

    # Handle all standard "Not a Number" or "Not a Time" types
    if pd.isna(obj):
        return None
    
    # Explicitly handle 'Infinity' and '-Infinity' values for floats
    if isinstance(obj, float) and math.isinf(obj):
        return None

    # Handle specific data types
    if isinstance(obj, (datetime, np.datetime64)):
        return obj.isoformat()
    if isinstance(obj, (np.integer, int)):
        return int(obj)
    if isinstance(obj, (np.floating, float)):
        return float(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    
    # Fallback for any other complex types
    try:
        return str(obj)
    except Exception:
        return f"Unserializable object: {type(obj).__name__}"

# OPTIMIZATION: This helper function will be run in parallel by each CPU core.
def _validate_chunk(validator_instance, df_chunk):
    """
    Helper function to be executed in parallel.
    Validates a small DataFrame chunk.
    """
    exceptions = []
    if df_chunk.empty:
        return exceptions

    for index, row in df_chunk.iterrows():
        dept = str(row.get('Department.Name', ''))
        if not dept:
            continue
        reasons, severity = validator_instance.validate_row(dept, row)
        if reasons:
            record = row.to_dict()
            record['Exception Reasons'] = "; ".join(reasons)
            record['Severity'] = severity
            # === FIX: CAPTURE THE ORIGINAL INDEX ===
            record['original_index'] = index
            # === END OF FIX ===
            exceptions.append(record)
    return exceptions


# Page configuration
st.set_page_config(
    page_title="Data Validation Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS (No changes here, but included for completeness)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    body {
        font-family: 'Inter', sans-serif;
    }
    .main {
        font-family: 'Inter', sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    .main > div {
        background: white;
        border-radius: 20px;
        margin: 1rem;
        padding: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
    }
    h1 {
        color: #2d3748;
        font-weight: 700;
        text-align: center;
        margin-bottom: 1rem;
        font-size: 2.5rem;
        position: relative;
        padding-bottom: 0.5rem;
    }
    h1::after {
        content: '';
        display: block;
        width: 100px;
        height: 4px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        margin: 0.5rem auto 1rem auto;
        border-radius: 2px;
    }
    h2, h3 {
        color: #4a5568;
        font-weight: 600;
    }
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin: 0.5rem;
        box-shadow: 0 5px 15px rgba(102, 126, 234, 0.3);
        transition: transform 0.3s ease;
    }
    .metric-container:hover {
        transform: translateY(-5px);
    }
    .metric-title {
        font-size: 0.9rem;
        font-weight: 500;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    /* ... rest of your CSS ... */
    .stProgress .st-bo {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

# REMOVED: DB_TIMEOUT is less relevant for MySQL connector which has its own config
# DB_TIMEOUT = 30

# Font dictionary for Plotly charts
PLOTLY_FONT = dict(family="Inter, sans-serif", size=12, color="#2d3748")
PLOTLY_TITLE_FONT = dict(family="Inter, sans-serif", size=16, color="#2d3748")


class DatabaseManager:
    def __init__(self, db_creds=st.secrets["mysql"]):
        self.db_creds = db_creds
        self.init_database()

    def _get_connection(self):
        """Establishes and returns a new connection to the MySQL database."""
        try:
            # --- MODIFIED --- Added autocommit=True to ensure writes are saved immediately.
            conn = mysql.connector.connect(**self.db_creds, autocommit=True)
            return conn
        except mysql.connector.Error as err:
            if err.errno == errorcode.ER_ACCESS_DENIED_ERROR: st.error("FATAL: MySQL access denied. Please check 'user' and 'password' in secrets.toml.")
            elif err.errno == errorcode.ER_BAD_DB_ERROR: st.error(f"FATAL: The database '{self.db_creds.get('database')}' does not exist.")
            else: st.error(f"FATAL: Could not connect to MySQL. Error: {err}")
            logging.critical(f"Database connection failed: {err}", exc_info=True)
            st.stop()

    def init_database(self):
        """Initializes and updates all tables for the application."""
        conn = self._get_connection()
        try:
            table_options = "ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
            with conn.cursor() as cursor:
                # --- EXISTING TABLES (No changes here) ---
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `users` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT, `username` VARCHAR(255) UNIQUE NOT NULL, `full_name` VARCHAR(255),
                        `email` VARCHAR(255), `mobile_number` VARCHAR(20), `hashed_password` TEXT NOT NULL,
                        `role` VARCHAR(255) NOT NULL, `reports_to` TEXT, `mapped_to_management` TEXT,
                        `can_upload` BOOLEAN DEFAULT NULL, `disabled_pages` TEXT, `disabled` BOOLEAN DEFAULT FALSE,
                        `can_accept_corrections` BOOLEAN DEFAULT NULL,
                        `receive_auto_notifications` BOOLEAN DEFAULT TRUE
                    ) {table_options}''')
                
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `role_permissions` (
                        `role` VARCHAR(255) PRIMARY KEY, `can_upload` BOOLEAN DEFAULT TRUE, `disabled_pages` TEXT
                    ) {table_options}''')
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `clarifications` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT, `username` VARCHAR(255) NOT NULL, `run_ids` TEXT NOT NULL,
                        `clarification_text` TEXT NOT NULL, `submitted_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        `status` VARCHAR(50) DEFAULT 'Submitted', `acknowledged_by` VARCHAR(255),
                        `acknowledged_at` TIMESTAMP NULL
                    ) {table_options}''')
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `clarification_waivers` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT, `username` VARCHAR(255) UNIQUE NOT NULL,
                        `waived_until` DATE NOT NULL, `waived_by` VARCHAR(255) NOT NULL,
                        `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    ) {table_options}''')
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `notifications` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT, `username` VARCHAR(255) NOT NULL,
                        `notification_type` VARCHAR(255) NOT NULL, `message` TEXT NOT NULL,
                        `is_read` BOOLEAN DEFAULT FALSE, `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    ) {table_options}''')
                cursor.execute(f'''CREATE TABLE IF NOT EXISTS `validation_runs` (`id` INT PRIMARY KEY AUTO_INCREMENT, `filename` TEXT NOT NULL, `upload_time` TIMESTAMP DEFAULT CURRENT_TIMESTAMP, `total_records` INT, `total_exceptions` INT, `status` TEXT, `file_size` BIGINT, `excel_report_data` LONGBLOB) {table_options}''')
                # --- MODIFIED --- Added narration, correction_status, and is_accepted to the exceptions table
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `exceptions` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT, `run_id` INT, `department` TEXT, 
                        `sub_department` TEXT, `created_user` TEXT, `modified_user` TEXT, 
                        `exception_reason` TEXT, `severity` INT, `net_amount` FLOAT, 
                        `location` TEXT, `crop` TEXT, `activity` TEXT, `function_name` TEXT, 
                        `vertical_name` TEXT, `region_name` TEXT, `zone_name` TEXT, 
                        `business_unit` TEXT, `account2_code` TEXT, `sub_ledger_code` TEXT, 
                        `original_row_data` LONGTEXT, 
                        `narration` TEXT,
                        `correction_status` ENUM('Pending', 'Yes', 'No') NOT NULL DEFAULT 'Pending',
                        `is_accepted` BOOLEAN NOT NULL DEFAULT FALSE,
                        FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE
                    ) {table_options}''')
                cursor.execute(f'''CREATE TABLE IF NOT EXISTS `department_summary` (`id` INT PRIMARY KEY AUTO_INCREMENT, `run_id` INT, `department` TEXT, `total_records` INT, `exception_records` INT, `exception_rate` FLOAT, FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE) {table_options}''')
                cursor.execute(f'''CREATE TABLE IF NOT EXISTS `user_performance` (`id` INT PRIMARY KEY AUTO_INCREMENT, `run_id` INT, `user` TEXT, `total_records` INT, `exception_records` INT, `exception_rate` FLOAT, FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE) {table_options}''')
                cursor.execute(f'''CREATE TABLE IF NOT EXISTS `correction_status` (`id` INT PRIMARY KEY AUTO_INCREMENT, `run_id` INT NOT NULL, `username` VARCHAR(255) NOT NULL, `status` ENUM('Yes', 'No', 'Pending') NOT NULL, `update_time` TIMESTAMP DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE, UNIQUE KEY `unique_run_user` (`run_id`, `username`)) {table_options}''')
            
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `transaction_fingerprints` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `run_id` INT NOT NULL,
                        `fingerprint_hash` VARCHAR(255) UNIQUE NOT NULL,
                        `created_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE
                    ) {table_options}''')
                # --- NEW --- Tables for Suspicious Transaction System
                # --- NEW --- Tables for Suspicious Transaction System
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `suspicious_rule_options` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `rule_column` VARCHAR(255) NOT NULL,
                        `option_value` VARCHAR(255) NOT NULL,
                        UNIQUE KEY `unique_option` (`rule_column`, `option_value`)
                    ) {table_options}''')
                
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `entry_clarifications` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `username` VARCHAR(255) NOT NULL,
                        `trigger_details` TEXT NOT NULL,
                        `user_clarification` TEXT,
                        `management_reply` TEXT,
                        `status` ENUM('Pending User Action', 'Submitted', 'Replied', 'Accepted') NOT NULL DEFAULT 'Pending User Action',
                        `created_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        `submitted_at` TIMESTAMP NULL,
                        `replied_by` VARCHAR(255),
                        `replied_at` TIMESTAMP NULL,
                        `accepted_by` VARCHAR(255),
                        `accepted_at` TIMESTAMP NULL
                    ) {table_options}''')
                
                cursor.execute('''
                    INSERT INTO global_settings (setting_key, setting_value) 
                    VALUES ('automatic_notifications_enabled', 'true') 
                    ON DUPLICATE KEY UPDATE setting_value=setting_value
                ''')

                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `suspicious_rules` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `sub_department_name` VARCHAR(255) NOT NULL,
                        `rule_column` VARCHAR(255) NOT NULL,
                        `rule_values` JSON,
                        UNIQUE KEY `unique_rule` (`sub_department_name`, `rule_column`)
                    ) {table_options}''')
                
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `suspicious_transactions_log` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `run_id` INT NOT NULL,
                        `original_row_data` JSON,
                        `created_user` VARCHAR(255),
                        `status` VARCHAR(50) DEFAULT 'Pending Admin Review',
                        `admin_comment` TEXT,
                        `reviewed_by` VARCHAR(255),
                        `reviewed_at` TIMESTAMP NULL,
                        `user_corrected_at` TIMESTAMP NULL,
                        FOREIGN KEY (`run_id`) REFERENCES `validation_runs`(`id`) ON DELETE CASCADE
                    ) {table_options}''')
                
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `correction_logs` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `exception_id` INT NOT NULL,
                        `action` ENUM('Corrected') NOT NULL,
                        `action_by` VARCHAR(255) NOT NULL,
                        `action_at` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (`exception_id`) REFERENCES `exceptions`(`id`) ON DELETE CASCADE
                    ) {table_options}''')
                
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS `accepted_exception_fingerprints` (
                        `id` INT PRIMARY KEY AUTO_INCREMENT,
                        `data_hash` VARCHAR(255) NOT NULL,
                        `reason_hash` VARCHAR(255) NOT NULL,
                        `combined_hash` VARCHAR(255) UNIQUE NOT NULL, -- Combined hash for quick lookup
                        `accepted_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    ) {table_options}''')

                # --- Populate default roles and Super User (No changes here) ---
                default_roles = ["User", "Manager", "Management", "Super User"]
                for role in default_roles:
                    cursor.execute("INSERT INTO `role_permissions` (role) VALUES (%s) ON DUPLICATE KEY UPDATE role=role", (role,))

                cursor.execute("SELECT COUNT(*) FROM `users`")
                if cursor.fetchone()[0] == 0:
                    super_user = os.environ.get("SUPER_USER_USERNAME")
                    super_pass = os.environ.get("SUPER_USER_PASSWORD")
                    if super_user and super_pass:
                        hashed_password = bcrypt.hashpw(super_pass.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                        cursor.execute("INSERT INTO `users` (username, full_name, hashed_password, role) VALUES (%s, %s, %s, %s)",(super_user, 'Super User Account', hashed_password, 'Super User'))
                        conn.commit()
                        logging.info(f"Successfully created initial 'Super User' account for '{super_user}'.")
        except mysql.connector.Error as err:
            logging.error(f"Database initialization error (MySQL): {err}", exc_info=True); conn.rollback()
        finally:
            if conn: conn.close()

    # --- User Management Methods ---
    def add_user(self, username, password, role, full_name=None, email=None, mobile_number=None, reports_to=None):
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "INSERT INTO `users` (username, full_name, email, mobile_number, hashed_password, role, reports_to) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                val = (username, full_name, email, mobile_number, hashed_password, role, reports_to)
                cursor.execute(sql, val)
            conn.commit(); return True
        except mysql.connector.Error as err:
            if err.errno == errorcode.ER_DUP_ENTRY: return "Username already exists."
            logging.error(f"Error in add_user: {err}", exc_info=True); return f"An error occurred: {err}"
        finally:
            if conn: conn.close()

    # REPLACE your existing get_user function with this one
    def get_user(self, username):
        conn = self._get_connection()
        try:
            # Added `disabled` to the SELECT query.
            with conn.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT username, full_name, hashed_password, role, disabled FROM `users` WHERE username = %s", (username,))
                return cursor.fetchone()
        except mysql.connector.Error as err:
            logging.error(f"Error in get_user: {err}", exc_info=True)
        finally:
            if conn: conn.close()

    # REPLACE your old get_all_users function with this one
    def get_all_users(self):
        conn = self._get_connection()
        try:
            # --- MODIFIED --- Added the new 'receive_auto_notifications' column to the SELECT query
            query = "SELECT id, username, full_name, email, mobile_number, role, reports_to, mapped_to_management, can_upload, disabled_pages, disabled, can_accept_corrections, receive_auto_notifications FROM `users`"
            return pd.read_sql_query(query, conn)
        except mysql.connector.Error as err:
            logging.error(f"Error in get_all_users: {err}", exc_info=True); return pd.DataFrame()
        finally:
            if conn: conn.close()

    def get_user_profile(self, username):
        """Gets all profile data for a single user."""
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT * FROM `users` WHERE username = %s", (username,)); return cursor.fetchone()
        except mysql.connector.Error as err:
            logging.error(f"Error in get_user_profile: {err}", exc_info=True)
        finally:
            if conn: conn.close()
    
    # ... (Other user management methods like get_user, get_managed_users, etc. go here) ...
    
    # --- Permission, Clarification, Waiver, and Notification Methods ---
    
    def get_pending_correction_runs_for_user(self, username, consecutive_limit=3):
        """
        Finds if a user has N or more consecutive unresolved correction statuses
        from their most recent runs.
        """
        conn = self._get_connection()
        try:
            # Get the user's most recent runs where they had exceptions
            query = """
                SELECT cs.run_id, cs.status
                FROM correction_status cs
                JOIN (
                    SELECT DISTINCT up.run_id
                    FROM user_performance up
                    WHERE up.user = %s AND up.exception_records > 0
                    ORDER BY up.run_id DESC
                    LIMIT 10
                ) AS recent_runs ON cs.run_id = recent_runs.run_id
                WHERE cs.username = %s
                ORDER BY cs.run_id DESC;
            """
            df = pd.read_sql_query(query, conn, params=(username, username))

            if len(df) < consecutive_limit:
                return [] # Not enough runs to meet the trigger

            consecutive_count = 0
            failed_run_ids = []
            # Iterate from the most recent run backwards
            for index, row in df.iterrows():
                if row['status'] in ['Pending', 'No']:
                    consecutive_count += 1
                    failed_run_ids.append(row['run_id'])
                else:
                    # If we find a 'Yes', the consecutive chain is broken
                    break
            
            if consecutive_count >= consecutive_limit:
                # Check if a clarification for this exact issue already exists and is not 'Acknowledged'
                run_ids_str = ",".join(map(str, sorted(failed_run_ids)))
                with conn.cursor() as cursor:
                    cursor.execute(
                        "SELECT 1 FROM clarifications WHERE username = %s AND run_ids = %s AND status != 'Acknowledged'",
                        (username, run_ids_str)
                    )
                    if cursor.fetchone():
                        return [] # A clarification is already pending for this issue
                return sorted(failed_run_ids)
            
            return []
        except Exception as e:
            logging.error(f"Error in get_pending_correction_runs_for_user for {username}: {e}", exc_info=True)
            return []
        finally:
            if conn: conn.close()
            
    def submit_clarification(self, username, run_ids, text):
        """Saves a user's clarification to the database."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                run_ids_str = ",".join(map(str, sorted(run_ids)))
                sql = "INSERT INTO clarifications (username, run_ids, clarification_text) VALUES (%s, %s, %s)"
                cursor.execute(sql, (username, run_ids_str, text))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error in submit_clarification: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
        
    def get_clarifications(self, user_role, username=None, managed_users=None, management_map=None):
        """Gets clarifications based on role hierarchy."""
        conn = self._get_connection()
        try:
            query = "SELECT * FROM clarifications WHERE status = 'Submitted' "
            params = []
            if user_role == 'Manager':
                if not (managed_users and username): return pd.DataFrame()
                team = managed_users + [username]
                placeholders = ','.join(['%s'] * len(team))
                query += f" AND username IN ({placeholders})"
                params.extend(team)
            elif user_role == 'Management':
                if not management_map: return pd.DataFrame()
                placeholders = ','.join(['%s'] * len(management_map))
                query += f" AND username IN (SELECT username FROM users WHERE reports_to IN ({placeholders}))"
                params.extend(management_map)

            query += " ORDER BY submitted_at DESC"
            return pd.read_sql_query(query, conn, params=params)
        finally:
            if conn: conn.close()

    def acknowledge_clarification(self, clarification_id, acknowledged_by):
        """Updates a clarification's status to Acknowledged."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE clarifications SET status = 'Acknowledged', acknowledged_by = %s, acknowledged_at = %s WHERE id = %s"
                cursor.execute(sql, (acknowledged_by, datetime.now(), clarification_id))
            conn.commit(); return True
        except mysql.connector.Error as err:
            logging.error(f"Error in acknowledge_clarification: {err}", exc_info=True); return False
        finally:
            if conn: conn.close()

    def check_waiver_status(self, username):
        """Checks if a user has an active clarification waiver."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT waived_until FROM clarification_waivers WHERE username = %s", (username,))
                result = cursor.fetchone()
                if result and result[0] >= datetime.now().date():
                    return True # Waiver is active
            return False
        finally:
            if conn: conn.close()
                 

    def get_managed_users(self, manager_username):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT username FROM `users` WHERE reports_to = %s", (manager_username,)); return [row[0] for row in cursor.fetchall()]
        except mysql.connector.Error as err:
            logging.error(f"Error in get_managed_users: {err}", exc_info=True); return []
        finally:
            if conn: conn.close()

    def get_management_users(self):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT username FROM `users` WHERE role = 'Management'"); return [row[0] for row in cursor.fetchall()]
        except mysql.connector.Error as err:
            logging.error(f"Error in get_management_users: {err}", exc_info=True); return []
        finally:
            if conn: conn.close()
    
    def get_user_permissions(self, username):
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT role, can_upload, disabled_pages FROM users WHERE username = %s", (username,))
                user_data = cursor.fetchone()
                if not user_data: return {"can_upload": False, "disabled_pages": []}
                
                cursor.execute("SELECT can_upload, disabled_pages FROM role_permissions WHERE role = %s", (user_data['role'],))
                role_data = cursor.fetchone()
                if not role_data: return {"can_upload": False, "disabled_pages": []}
                
                final_can_upload = user_data['can_upload'] if user_data['can_upload'] is not None else role_data['can_upload']
                final_disabled_pages_str = user_data['disabled_pages'] if user_data['disabled_pages'] is not None else role_data['disabled_pages']
                
                return {
                    "can_upload": bool(final_can_upload),
                    "disabled_pages": final_disabled_pages_str.split(',') if final_disabled_pages_str else []
                }
        except mysql.connector.Error as err:
            logging.error(f"Error in get_user_permissions for {username}: {err}", exc_info=True)
            return {"can_upload": False, "disabled_pages": []}
        finally:
            if conn: conn.close()

    def get_all_permissions(self):
        conn = self._get_connection()
        permissions = {'roles': {}, 'users': {}}
        try:
            with conn.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT * FROM role_permissions")
                for row in cursor.fetchall(): permissions['roles'][row['role']] = row
                cursor.execute("SELECT username, role, can_upload, disabled_pages FROM users")
                for row in cursor.fetchall(): permissions['users'][row['username']] = row
            return permissions
        except mysql.connector.Error as err:
            logging.error(f"Error in get_all_permissions: {err}", exc_info=True); return permissions
        finally:
            if conn: conn.close()
    
    def update_role_permissions(self, role, can_upload, disabled_pages_list):
        conn = self._get_connection()
        try:
            disabled_pages_str = ",".join(disabled_pages_list) if disabled_pages_list else ""
            with conn.cursor() as cursor:
                cursor.execute("UPDATE role_permissions SET can_upload = %s, disabled_pages = %s WHERE role = %s", (can_upload, disabled_pages_str, role))
            conn.commit(); return True
        except mysql.connector.Error as err:
            logging.error(f"Error in update_role_permissions: {err}", exc_info=True); return False
        finally:
            if conn: conn.close()

    def update_user_permissions(self, username, can_upload, disabled_pages_list):
        conn = self._get_connection()
        try:
            disabled_pages_str = ",".join(disabled_pages_list) if disabled_pages_list else ""
            with conn.cursor() as cursor:
                cursor.execute("UPDATE users SET can_upload = %s, disabled_pages = %s WHERE username = %s", (can_upload, disabled_pages_str, username))
            conn.commit(); return True
        except mysql.connector.Error as err:
            logging.error(f"Error in update_user_permissions: {err}", exc_info=True); return False
        finally:
            if conn: conn.close()

    # ADD THIS NEW FUNCTION INSIDE your DatabaseManager class
    def get_users_by_role(self, role_name):
        """Fetches a list of usernames for a given role."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT username FROM users WHERE role = %s", (role_name,))
                return [row[0] for row in cursor.fetchall()]
        except mysql.connector.Error as err:
            logging.error(f"Error fetching users by role '{role_name}': {err}", exc_info=True)
            return []
        finally:
            if conn: conn.close()            

    def update_user_role(self, username, new_role):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("UPDATE `users` SET role = %s WHERE username = %s", (new_role, username))
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error in update_user_role: {err}", exc_info=True)
        finally:
            if conn: conn.close()

    # ADD THESE NEW METHODS INSIDE THE DatabaseManager CLASS

    def update_user_profile(self, username, full_name, email, mobile_number):
        """Updates a user's profile information."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE users SET full_name = %s, email = %s, mobile_number = %s WHERE username = %s"
                cursor.execute(sql, (full_name, email, mobile_number, username))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error updating profile for {username}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def set_user_disabled_status(self, username, disabled):
        """Sets a user's disabled status (True or False)."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE users SET disabled = %s WHERE username = %s"
                cursor.execute(sql, (disabled, username))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error setting disabled status for {username}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def delete_user(self, username):
        """Permanently deletes a user from the database."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Ensure the Super User from environment variables cannot be deleted
                super_user_env = os.environ.get("SUPER_USER_USERNAME")
                if username == super_user_env:
                    return "Cannot delete the primary Super User account."
                
                sql = "DELETE FROM users WHERE username = %s"
                cursor.execute(sql, (username,))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error deleting user {username}: {err}", exc_info=True)
            return f"Database error: {err}"
        finally:
            if conn: conn.close()

    def update_user_mapping(self, username, manager_username):
        manager = manager_username if manager_username and manager_username != "None" else None
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("UPDATE `users` SET reports_to = %s WHERE username = %s", (manager, username))
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error in update_user_mapping: {err}", exc_info=True)
        finally:
            if conn: conn.close()
            
    def update_manager_to_management_mapping(self, manager_username, management_username):
        management_user = management_username if management_username and management_username != "None" else None
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("UPDATE `users` SET mapped_to_management = %s WHERE username = %s AND role = 'Manager'", (management_user, manager_username))
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error in update_manager_to_management_mapping: {err}", exc_info=True)
        finally:
            if conn: conn.close()

    def update_user_password(self, username, new_password):
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("UPDATE `users` SET hashed_password = %s WHERE username = %s", (hashed_password, username))
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error in update_user_password: {err}", exc_info=True)
        finally:
            if conn: conn.close()

    # --- Data Saving and Retrieval Methods ---
    # The following methods are the final, stable versions and do not need further changes.
    def save_validation_run(self, filename, total_records, total_exceptions, file_size, upload_time=None):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                final_upload_time = upload_time if upload_time else datetime.now()
                cursor.execute('''INSERT INTO `validation_runs` (filename, upload_time, total_records, total_exceptions, file_size) VALUES (%s, %s, %s, %s, %s)''', (filename, final_upload_time, total_records, total_exceptions, file_size))
                run_id = cursor.lastrowid
                conn.commit()
                return run_id
        except mysql.connector.Error as err:
            logging.error(f"Error in save_validation_run: {err}", exc_info=True); conn.rollback(); raise
        finally:
            if conn: conn.close()

    def save_excel_report(self, run_id, excel_data):
        if not excel_data: return
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute('''UPDATE `validation_runs` SET excel_report_data = %s WHERE id = %s''', (excel_data.getvalue(), run_id))
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error saving Excel report for run_id {run_id}: {err}", exc_info=True); conn.rollback()
        finally:
            if conn: conn.close()

    def delete_run(self, run_id):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM `validation_runs` WHERE id = %s", (run_id,)); conn.commit(); return True
        except mysql.connector.Error as err:
            logging.error(f"Error deleting run ID {run_id}: {err}", exc_info=True); conn.rollback(); return False
        finally:
            if conn: conn.close()

    def save_exceptions(self, run_id, exceptions_df):
        conn = self._get_connection()
        if exceptions_df.empty: return
        try:
            with conn.cursor() as cursor:
                data_to_insert = []
                for _, row_series in exceptions_df.iterrows():
                    original_row_data_dict = row_series.to_dict()
                    # Apply json_serializer_default and ensure sorted keys for consistent hashing
                    serialized_row_data = json.dumps(original_row_data_dict, default=json_serializer_default, sort_keys=True) # ADDED sort_keys=True
                    data_to_insert.append((
                        run_id, row_series.get('Department.Name', ''), row_series.get('Sub Department.Name', ''),
                        row_series.get('Created user', ''), row_series.get('Modified user', ''),
                        str(row_series.get('Exception Reasons', '')), row_series.get('Severity', 0),
                        row_series.get('Net amount', 0.0), row_series.get('Location.Name', ''),
                        row_series.get('Crop.Name', ''), row_series.get('Activity.Name', ''),
                        row_series.get('Function.Name', ''), row_series.get('FC-Vertical.Name', ''),
                        row_series.get('Region.Name', ''), row_series.get('Zone.Name', ''),
                        row_series.get('Business Unit.Name', ''), row_series.get('Account2.Code', ''),
                        row_series.get('Sub Ledger.Code', ''),
                        row_series.get('Narration', ''), # This is new
                        serialized_row_data
                    ))
                if data_to_insert:
                    cursor.executemany('''
                        INSERT INTO `exceptions` (
                            run_id, department, sub_department, created_user, modified_user,
                            exception_reason, severity, net_amount, location, crop, activity,
                            function_name, vertical_name, region_name, zone_name, business_unit,
                            account2_code, sub_ledger_code, narration, original_row_data
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ''', data_to_insert)
            conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error in save_exceptions (MySQL): {err}", exc_info=True); conn.rollback(); raise
        finally:
            if conn: conn.close()

    def save_transaction_fingerprints(self, run_id, fingerprints_to_save):
        """Saves a list of unique transaction fingerprints to the new historical table."""
        if not fingerprints_to_save:
            return
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Prepare data for insertion: (run_id, fingerprint)
                data_to_insert = [(run_id, fp) for fp in fingerprints_to_save]
                
                # Use INSERT IGNORE to prevent errors if a duplicate fingerprint is somehow processed.
                # The UNIQUE constraint on the column is the primary guard.
                query = "INSERT IGNORE INTO `transaction_fingerprints` (run_id, fingerprint_hash) VALUES (%s, %s)"
                cursor.executemany(query, data_to_insert)
            conn.commit()  
        except mysql.connector.Error as err:
            logging.error(f"Error saving transaction fingerprints for run {run_id}: {err}", exc_info=True)
            conn.rollback()
        finally:
            if conn and conn.is_connected():
                conn.close()
    
    def save_department_summary(self, run_id, department_statistics):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                data_to_insert = [(run_id, dept, stats['total_records'], stats['exception_records'], stats['exception_rate']) for dept, stats in department_statistics.items()]
                if data_to_insert:
                    cursor.executemany('''INSERT INTO `department_summary` (run_id, department, total_records, exception_records, exception_rate) VALUES (%s, %s, %s, %s, %s)''', data_to_insert)
                conn.commit()
        except mysql.connector.Error as err:
            logging.error(f"Error saving department_summary for run {run_id}: {err}", exc_info=True); conn.rollback()
        finally:
            if conn: conn.close()

    def save_user_performance(self, run_id, df, exceptions_df):
        try:
            if 'Created user' not in df.columns:
                logging.error(f"Run ID {run_id}: 'Created user' column not in source file. Cannot save user performance.")
                return
    
            # --- FIX: Standardize user names to prevent case-sensitivity and whitespace issues ---
            df_copy = df.copy()
            exceptions_df_copy = exceptions_df.copy()
            df_copy['Created user'] = df_copy['Created user'].fillna('').astype(str).str.strip()
            exceptions_df_copy['Created user'] = exceptions_df_copy['Created user'].fillna('').astype(str).str.strip()
    
            # --- FIX: Group by standardized user names ---
            total_records_by_user = df_copy[df_copy['Created user'] != ''].groupby('Created user').size().reset_index(name='total_records')
    
            if exceptions_df_copy.empty or 'Created user' not in exceptions_df_copy.columns:
                exception_records_by_user = pd.DataFrame(columns=['Created user', 'exception_records'])
            else:
                exception_records_by_user = exceptions_df_copy[exceptions_df_copy['Created user'] != ''].groupby('Created user').size().reset_index(name='exception_records')
    
            # --- FIX: Use a robust 'outer' merge to combine the data ---
            if total_records_by_user.empty and exception_records_by_user.empty:
                logging.warning(f"Run ID {run_id}: No user data found to save for performance.")
                return
            
            user_stats = pd.merge(total_records_by_user, exception_records_by_user, on='Created user', how='outer')
            
            user_stats['total_records'] = user_stats['total_records'].fillna(0).astype(int)
            user_stats['exception_records'] = user_stats['exception_records'].fillna(0).astype(int)
    
            user_stats['exception_rate'] = (user_stats['exception_records'] / user_stats['total_records'].replace(0, pd.NA) * 100).fillna(0).round(2)
    
            if user_stats.empty:
                logging.error(f"Run ID {run_id}: Final user_stats DataFrame is empty. Aborting save.")
                return
                
            conn = self._get_connection()
            try:
                with conn.cursor() as cursor:
                    data_to_insert_perf = [
                        (run_id, row['Created user'], row['total_records'], row['exception_records'], row['exception_rate'])
                        for _, row in user_stats.iterrows()
                    ]
                    if data_to_insert_perf:
                        cursor.executemany(
                            '''INSERT INTO `user_performance` (run_id, `user`, total_records, exception_records, exception_rate) VALUES (%s,%s,%s,%s,%s)''',
                            data_to_insert_perf
                        )
                        conn.commit()
            except mysql.connector.Error as err:
                 logging.error(f"DB ERROR in save_user_performance for run ID {run_id}: {err}", exc_info=True)
                 conn.rollback()
            finally:
                if conn: conn.close()
    
        except Exception as e:
            logging.error(f"LOGIC ERROR in save_user_performance for run ID {run_id}: {e}", exc_info=True)

    def get_validation_history(self, user_role=None, username=None, managed_users=None):
        conn = self._get_connection()
        try:
            query = 'SELECT * FROM `validation_runs`'
            params = ()
            
            # --- UPDATED HIERARCHICAL LOGIC ---
            if user_role == 'User' and username:
                # User sees only runs where they have performance records
                query = '''
                    SELECT vr.* FROM `validation_runs` vr 
                    JOIN (SELECT DISTINCT run_id FROM `user_performance` WHERE `user` = %s) up 
                    ON vr.id = up.run_id
                '''
                params = (username,)
            elif user_role == 'Manager' and managed_users is not None and username is not None:
                # Manager sees runs involving themselves or their team
                team_members = managed_users + [username]
                if not team_members: return pd.DataFrame() # Return empty if team is empty
                placeholders = ','.join(['%s'] * len(team_members))
                query = f'''
                    SELECT vr.* FROM `validation_runs` vr 
                    JOIN (SELECT DISTINCT run_id FROM `user_performance` WHERE `user` IN ({placeholders})) up 
                    ON vr.id = up.run_id
                '''
                params = tuple(team_members)
            # For 'Management' and 'Super User', the original query fetching all runs is used.
            
            query += ' ORDER BY upload_time DESC'
            df = pd.read_sql_query(query, conn, params=params)
            if not df.empty and 'upload_time' in df.columns: 
                df['upload_time'] = pd.to_datetime(df['upload_time'], format='mixed')
            return df
        except mysql.connector.Error as err:
            logging.error(f"Error in get_validation_history: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()
    
    # ... All other methods from the class should be here ...

    def get_archived_report(self, run_id, user_role=None, username=None, managed_users=None):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                has_access = False
                # Management and Super User have universal access
                if user_role in ['Management', 'Super User']:
                    has_access = True
                elif user_role == 'User' and username:
                    # Check if the user participated in this specific run
                    cursor.execute("SELECT 1 FROM `user_performance` WHERE run_id = %s AND `user` = %s", (run_id, username))
                    if cursor.fetchone(): has_access = True
                elif user_role == 'Manager' and username and managed_users is not None:
                    # Check if the manager or anyone on their team participated
                    team_members = managed_users + [username]
                    if team_members:
                        placeholders = ','.join(['%s'] * len(team_members))
                        query = f"SELECT 1 FROM `user_performance` WHERE run_id = %s AND `user` IN ({placeholders})"
                        params = (run_id,) + tuple(team_members)
                        cursor.execute(query, params)
                        if cursor.fetchone(): has_access = True

                if not has_access:
                    logging.warning(f"ACCESS DENIED: User '{username}' (Role: {user_role}) tried to access report for run_id {run_id}.")
                    return None, None

                # If access is confirmed, fetch the report
                cursor.execute("SELECT excel_report_data, filename FROM `validation_runs` WHERE id = %s", (run_id,))
                result = cursor.fetchone()
                return result if result and result[0] is not None else (None, None)
        except mysql.connector.Error as err:
            logging.error(f"Error fetching archived report for run_id {run_id}: {err}", exc_info=True)
            return None, None
        finally:
            if conn.is_connected():
                conn.close()

    def get_exceptions_by_run(self, run_id):
        conn = self._get_connection()
        try:
            query = "SELECT id, run_id, exception_reason, severity, original_row_data FROM `exceptions` WHERE run_id = %s AND is_accepted = FALSE"
            raw_exceptions_df = pd.read_sql_query(query, conn, params=(run_id,))

            if raw_exceptions_df.empty:
                return pd.DataFrame()

            processed_records = []
            for _, db_row in raw_exceptions_df.iterrows():
                try:
                    # JSON type from DB can be directly loaded if it's a string, or used if it's already a dict
                    record = db_row['original_row_data'] if isinstance(db_row['original_row_data'], dict) else json.loads(db_row['original_row_data'])
                except (json.JSONDecodeError, TypeError):
                    record = {}

                record['id'] = db_row['id']
                record['run_id'] = db_row['run_id']
                record['Exception Reasons'] = db_row['exception_reason']
                record['Severity'] = db_row['severity']
                processed_records.append(record)

            return pd.DataFrame(processed_records) if processed_records else pd.DataFrame()

        except Exception as e:
            st.error(f"An error occurred while retrieving exception details for run {run_id}: {e}")
            logging.error(f"Error in get_exceptions_by_run for run_id {run_id}: {e}", exc_info=True)
            return pd.DataFrame()
        finally:
            conn.close()

    def add_or_update_correction_status(self, run_id, username, status):
        """Inserts or updates a user's correction status for a specific run using MySQL's syntax."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                query = """
                    INSERT INTO `correction_status` (run_id, username, status, update_time)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                    status = VALUES(status),
                    update_time = VALUES(update_time);
                """
                cursor.execute(query, (run_id, username, status, datetime.now()))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error updating correction status for run {run_id}, user {username}: {err}", exc_info=True)
            conn.rollback()
            return False
        finally:
            conn.close()

    def get_correction_status_for_run(self, run_id):
        """Fetches all correction statuses for a given run ID."""
        conn = self._get_connection()
        try:
            query = "SELECT username, status FROM `correction_status` WHERE run_id = %s"
            df = pd.read_sql_query(query, conn, params=(run_id,))
            return pd.Series(df.status.values, index=df.username).to_dict()
        except mysql.connector.Error as err:
            logging.error(f"Error fetching correction status for run {run_id}: {err}", exc_info=True)
            return {}
        finally:
            conn.close()

    # --- NEW --- Methods for Correction Entries Dashboard
    
    def get_correction_entries(self, user_role, username=None, managed_users=None, narration_filter=None):
        """
        Fetches exception entries for the Correction Entries dashboard.
        - Only fetches entries that have not been accepted (is_accepted = FALSE).
        - Applies hierarchical logic for users, managers, and admins.
        - Optionally filters by a search term in the narration column.
        """
        conn = self._get_connection()
        try:
            # Base query always excludes accepted entries
            query = "SELECT * FROM `exceptions` WHERE is_accepted = FALSE"
            params = []

            # Hierarchical access control
            if user_role == 'User' and username:
                query += " AND created_user = %s"
                params.append(username)
            elif user_role == 'Manager' and managed_users is not None and username is not None:
                team_members = managed_users + [username]
                if not team_members: return pd.DataFrame()
                placeholders = ','.join(['%s'] * len(team_members))
                query += f" AND created_user IN ({placeholders})"
                params.extend(team_members)
            # For Management/Super User, no additional user filter is needed.

            # Narration text filter
            if narration_filter:
                query += " AND narration LIKE %s"
                params.append(f"%{narration_filter}%")

            query += " ORDER BY id DESC"
            df = pd.read_sql_query(query, conn, params=tuple(params))
            return df
        except mysql.connector.Error as err:
            logging.error(f"Error in get_correction_entries: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()


    def batch_update_exception_status(self, exception_ids, new_status, action_by, user_role):
        """
        Updates the correction_status for a list of exception IDs.
        If status is 'Yes', it also logs the actions.
        """
        if not exception_ids:
            return False
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                placeholders = ','.join(['%s'] * len(exception_ids))
                # This query now ONLY updates the status column.
                sql = f"UPDATE exceptions SET correction_status = %s WHERE id IN ({placeholders})"
                params = [new_status] + exception_ids
                cursor.execute(sql, tuple(params))
            conn.commit()
            
            # Log the actions if corrected, regardless of role
            if new_status == 'Yes':
                self.batch_log_correction_action(exception_ids, action_by)

            return True
        except mysql.connector.Error as err:
            logging.error(f"Error in batch updating exception status: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def accept_correction_entry(self, exception_id):
        """Marks a correction entry as accepted, hiding it from future views."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE exceptions SET is_accepted = TRUE WHERE id = %s"
                cursor.execute(sql, (exception_id,))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error accepting correction entry for id {exception_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def set_manager_acceptance_permission(self, manager_username, can_accept):
        """Updates the can_accept_corrections permission for a manager."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # The permission value can be True, False, or None (to inherit)
                sql = "UPDATE users SET can_accept_corrections = %s WHERE username = %s"
                cursor.execute(sql, (can_accept, manager_username))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error setting manager acceptance permission for {manager_username}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    # --- NEW --- Methods for Correction Entries Dashboard
    
    def get_correction_entries(self, user_role, username=None, managed_users=None, narration_filter=None, filter_user=None, filter_run_id=None):
        """
        Fetches exception entries for the Correction Entries dashboard.
        - Hides entries that are accepted OR have status 'Yes'.
        - Applies hierarchical logic for users, managers, and admins.
        - Optionally filters by narration, user, and run_id.
        """
        conn = self._get_connection()
        try:
            # --- MODIFIED --- Query now also excludes entries where correction_status is 'Yes'
            query = "SELECT * FROM `exceptions` WHERE is_accepted = FALSE AND correction_status != 'Yes'"
            params = []

            # Hierarchical access control
            if user_role == 'User' and username:
                query += " AND created_user = %s"
                params.append(username)
            elif user_role == 'Manager' and managed_users is not None and username is not None:
                team_members = managed_users + [username]
                if not team_members: return pd.DataFrame()
                placeholders = ','.join(['%s'] * len(team_members))
                query += f" AND created_user IN ({placeholders})"
                params.extend(team_members)

            # --- NEW --- Add optional filters from the UI
            if narration_filter:
                query += " AND narration LIKE %s"
                params.append(f"%{narration_filter}%")
            if filter_user:
                query += " AND created_user = %s"
                params.append(filter_user)
            if filter_run_id:
                query += " AND run_id = %s"
                params.append(filter_run_id)

            query += " ORDER BY id DESC"
            df = pd.read_sql_query(query, conn, params=tuple(params))
            return df
        except mysql.connector.Error as err:
            logging.error(f"Error in get_correction_entries: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()


    def update_exception_status(self, exception_id, new_status, action_by, user_role):
        """
        Updates the correction_status for a single exception entry.
        If status is 'Yes', it also logs the action.
        """
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # This query now ONLY updates the status column.
                sql = "UPDATE exceptions SET correction_status = %s WHERE id = %s"
                cursor.execute(sql, (new_status, exception_id))
            conn.commit()

            # Log the action if corrected, regardless of role
            if new_status == 'Yes':
                self.log_correction_action(exception_id, action_by)
            
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error updating exception status for id {exception_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def accept_correction_entry(self, exception_id):
        """Marks a correction entry as accepted, hiding it from future views."""
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor: # Use dictionary=True to fetch row data easily
                # First, retrieve the necessary data before updating
                cursor.execute("SELECT original_row_data, exception_reason FROM exceptions WHERE id = %s", (exception_id,))
                exception_data = cursor.fetchone()

                if exception_data:
                    original_row_data_str = exception_data['original_row_data']
                    exception_reason_str = exception_data['exception_reason']

                    # Update the exception status
                    sql = "UPDATE exceptions SET is_accepted = TRUE WHERE id = %s"
                    cursor.execute(sql, (exception_id,))
                    conn.commit()

                    # Save the fingerprint of this accepted exception
                    self.save_accepted_exception_fingerprint(original_row_data_str, exception_reason_str)
                    return True
                else:
                    logging.warning(f"Attempted to accept non-existent exception ID: {exception_id}")
                    return False
        except mysql.connector.Error as err:
            logging.error(f"Error accepting correction entry for id {exception_id}: {err}", exc_info=True)
            conn.rollback()
            return False
        finally:
            if conn: conn.close()

    def log_correction_action(self, exception_id, action_by, action='Corrected'):
        """Logs an action (like 'Corrected') for an exception."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "INSERT INTO correction_logs (exception_id, action_by, action) VALUES (%s, %s, %s)"
                cursor.execute(sql, (exception_id, action_by, action))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error logging correction action for exception {exception_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
            
    def batch_log_correction_action(self, exception_ids, action_by, action='Corrected'):
        """Logs an action for a list of exception IDs."""
        if not exception_ids:
            return False
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                data_to_insert = [(exc_id, action_by, action) for exc_id in exception_ids]
                sql = "INSERT INTO correction_logs (exception_id, action_by, action) VALUES (%s, %s, %s)"
                cursor.executemany(sql, data_to_insert)
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error batch logging correction actions: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def save_accepted_exception_fingerprint(self, original_row_data_json_string, exception_reason_string):
        """
        Saves a unique fingerprint for an accepted exception (original data + reason).
        This helps prevent re-flagging the exact same accepted issue in future uploads.
        """
        import hashlib # Ensure hashlib is imported at the top of the file if not already

        data_hash = hashlib.sha256(original_row_data_json_string.encode('utf-8')).hexdigest()
        reason_hash = hashlib.sha256(exception_reason_string.encode('utf-8')).hexdigest()
        combined_hash = hashlib.sha256(f"{data_hash}_{reason_hash}".encode('utf-8')).hexdigest()

        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Use INSERT IGNORE to avoid errors if for some reason a duplicate is attempted
                sql = """
                    INSERT IGNORE INTO `accepted_exception_fingerprints` 
                    (data_hash, reason_hash, combined_hash) VALUES (%s, %s, %s)
                """
                cursor.execute(sql, (data_hash, reason_hash, combined_hash))
            conn.commit()
            logging.info(f"Saved accepted exception fingerprint: {combined_hash}")
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error saving accepted exception fingerprint: {err}", exc_info=True)
            return False
        finally:
            if conn and conn.is_connected():
                conn.close()

    # ADD THIS BLOCK OF METHODS INSIDE YOUR DatabaseManager CLASS

    def create_notification(self, username, notif_type, message):
        """Creates a new notification for a user. [WITH ENHANCED LOGGING]"""
        logging.info(f"Attempting to create notification for user: '{username}' with type: '{notif_type}'")
        conn = self._get_connection()
        if not conn:
            logging.error("Failed to get a database connection in create_notification.")
            return False
            
        try:
            with conn.cursor() as cursor:
                sql = "INSERT INTO notifications (username, notification_type, message) VALUES (%s, %s, %s)"
                params = (username, notif_type, message)
                logging.info(f"Executing SQL: {sql} with params: {params}")
                cursor.execute(sql, params)
            conn.commit()
            logging.info(f"Successfully committed notification for '{username}' to the database.")
            return True
        except mysql.connector.Error as err:
            logging.error(f"DATABASE ERROR caught while creating notification for {username}: {err}", exc_info=True)
            return False
        finally:
            if conn and conn.is_connected():
                conn.close()
                logging.info("Database connection closed in create_notification.")

    def get_notifications_for_user(self, username):
        """Fetches all unread notifications for a user."""
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor:
                sql = "SELECT * FROM notifications WHERE username = %s AND is_read = FALSE ORDER BY created_at DESC"
                cursor.execute(sql, (username,))
                return cursor.fetchall()
        except mysql.connector.Error as err:
            logging.error(f"Error fetching notifications for {username}: {err}", exc_info=True)
            return []
        finally:
            if conn: conn.close()
    
    def mark_notifications_as_read(self, notification_ids):
        """Marks a list of notification IDs as read."""
        if not notification_ids: return False
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Use a format string for the IN clause
                format_strings = ','.join(['%s'] * len(notification_ids))
                sql = f"UPDATE notifications SET is_read = TRUE WHERE id IN ({format_strings})"
                cursor.execute(sql, tuple(notification_ids))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error marking notifications as read: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    # ADD THIS NEW FUNCTION INSIDE your DatabaseManager class
    
    def get_notification_counts(self, run_ids, user_list=None):
        """Fetches notification counts for a given list of users and runs."""
        if not run_ids: return pd.DataFrame()
        conn = self._get_connection()
        try:
            placeholders = ','.join(['%s'] * len(run_ids))
            # We filter notifications to be within the date range of the selected runs
            query = f"""
                SELECT username, notification_type, COUNT(*) as count
                FROM notifications
                WHERE created_at >= (SELECT MIN(upload_time) FROM validation_runs WHERE id IN ({placeholders}))
                  AND created_at <= (SELECT MAX(upload_time) + INTERVAL 1 DAY FROM validation_runs WHERE id IN ({placeholders}))
            """
            params = list(run_ids) * 2 # Params are needed for both subqueries

            if user_list:
                user_placeholders = ','.join(['%s'] * len(user_list))
                query += f" AND username IN ({user_placeholders})"
                params.extend(user_list)
            
            query += " GROUP BY username, notification_type"
            df = pd.read_sql_query(query, conn, params=params)
            return df
        except mysql.connector.Error as err:
            logging.error(f"Error in get_notification_counts: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()
    
    # ADD THIS BLOCK OF METHODS INSIDE YOUR DatabaseManager CLASS

    def grant_waiver(self, username, waived_until, waived_by):
        """Grants or updates a clarification waiver for a user."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # This 'ON DUPLICATE KEY UPDATE' will insert a new waiver or update the date if one already exists.
                sql = """
                    INSERT INTO clarification_waivers (username, waived_until, waived_by)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                    waived_until = VALUES(waived_until),
                    waived_by = VALUES(waived_by);
                """
                cursor.execute(sql, (username, waived_until, waived_by))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error in grant_waiver for {username}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def get_all_waivers(self):
        """Fetches all active clarification waivers."""
        conn = self._get_connection()
        try:
            # Only get waivers that have not expired yet
            query = "SELECT * FROM clarification_waivers WHERE waived_until >= CURDATE() ORDER BY waived_until DESC"
            return pd.read_sql_query(query, conn)
        except mysql.connector.Error as err:
            logging.error(f"Error in get_all_waivers: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()

    def revoke_waiver(self, waiver_id):
        """Removes a specific waiver."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM clarification_waivers WHERE id = %s", (waiver_id,))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error in revoke_waiver for id {waiver_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def get_correction_summary(self, run_ids, accessible_users=None):
        """Fetches a summary of correction statuses for given runs and users."""
        if not run_ids: return pd.DataFrame()
        conn = self._get_connection()
        try:
            placeholders = ','.join(['%s'] * len(run_ids))
            
            # --- FIX: The query now ONLY selects users who have exception records. ---
            perf_query = f"SELECT DISTINCT run_id, `user` FROM `user_performance` WHERE run_id IN ({placeholders}) AND exception_records > 0"
            params = list(run_ids)
            
            if accessible_users:
                user_placeholders = ','.join(['%s'] * len(accessible_users))
                perf_query += f" AND `user` IN ({user_placeholders})"
                params.extend(accessible_users)
            
            all_participants_df = pd.read_sql_query(perf_query, conn, params=params)
    
            if all_participants_df.empty: 
                return pd.DataFrame(), pd.DataFrame() # Return two empty dataframes
    
            status_query = f"SELECT run_id, username, status, update_time FROM `correction_status` WHERE run_id IN ({placeholders})"
            status_df = pd.read_sql_query(status_query, conn, params=tuple(run_ids))
    
            summary_df = pd.merge(all_participants_df, status_df, how='left', left_on=['run_id', 'user'], right_on=['run_id', 'username'])
    
            run_dates_query = f"SELECT id, upload_time FROM `validation_runs` WHERE id IN ({placeholders})"
            run_dates = pd.read_sql_query(run_dates_query, conn, params=tuple(run_ids))
            run_dates['upload_date'] = pd.to_datetime(run_dates['upload_time'], format='mixed').dt.date
            today = datetime.now().date()
            todays_run_ids = set(run_dates[run_dates['upload_date'] == today]['id'])
    
            def determine_status(row):
                if pd.notna(row['status']):
                    return row['status']
                return 'Pending' if row['run_id'] in todays_run_ids else 'No'
    
            summary_df['status'] = summary_df.apply(determine_status, axis=1)
            
            # FOR DEBUGGING: Return the intermediate dataframe along with the final one
            return summary_df[['run_id', 'user', 'status']], all_participants_df
    
        except Exception as e:
            logging.error(f"Error fetching correction summary: {e}", exc_info=True)
            return pd.DataFrame(), pd.DataFrame()
        finally:
            if conn.is_connected():
                conn.close()

    def get_correction_analytics_data(self, target_user=None):
        """
        Fetches the counts of correction statuses for the analytics pie chart.
        - Fetches for all users if target_user is None.
        - Fetches for a specific user if a username is provided.
        - Excludes entries that have already been accepted.
        """
        conn = self._get_connection()
        try:
            query = "SELECT correction_status, COUNT(*) as count FROM `exceptions` WHERE is_accepted = FALSE"
            params = []

            if target_user:
                query += " AND created_user = %s"
                params.append(target_user)
            
            query += " GROUP BY correction_status"

            df = pd.read_sql_query(query, conn, params=tuple(params))
            
            # Initialize a dictionary with all possible statuses to ensure they exist
            status_counts = {'Yes': 0, 'No': 0, 'Pending': 0}
            # Update the dictionary with the actual counts from the database
            if not df.empty:
                status_counts.update(pd.Series(df['count'].values, index=df.correction_status).to_dict())
            
            return status_counts

        except mysql.connector.Error as err:
            logging.error(f"Error getting correction analytics data: {err}", exc_info=True)
            return {'Yes': 0, 'No': 0, 'Pending': 0} # Return zero counts on error
        finally:
            if conn: conn.close()
    # --- NEW --- Methods for Suspicious Transaction System
    
    @st.cache_data(ttl=3600)
    def load_suspense_immunity_list(_self):
        """Loads and caches the list of account/sub-ledger combinations exempt from suspicious checks."""
        immunity_file = "reference_data/do not check suspense.xlsx"
        try:
            if not os.path.exists(immunity_file):
                logging.warning(f"Immunity file '{immunity_file}' not found. No transactions will be exempt.")
                return set()
            
            df = pd.read_excel(immunity_file)
            required_cols = ["Account2.Code", "Sub Ledger.Code"]
            if not all(col in df.columns for col in required_cols):
                logging.error(f"Immunity file '{immunity_file}' is missing required columns: {required_cols}")
                return set()

            df.dropna(subset=required_cols, inplace=True)
            immunity_set = set(
                df["Account2.Code"].astype(str).str.strip() + "_" + df["Sub Ledger.Code"].astype(str).str.strip()
            )
            logging.info(f"Loaded {len(immunity_set)} immunity combinations from '{immunity_file}'.")
            return immunity_set
        except Exception as e:
            st.error(f"Failed to load suspense immunity file: {e}")
            logging.error(f"Error loading suspense immunity file: {e}", exc_info=True)
            return set()

    def get_rule_options(self, rule_column):
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor:
                cursor.execute("SELECT id, option_value FROM suspicious_rule_options WHERE rule_column = %s ORDER BY option_value", (rule_column,))
                return cursor.fetchall()
        finally:
            if conn: conn.close()

    def add_rule_option(self, rule_column, option_value):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("INSERT INTO suspicious_rule_options (rule_column, option_value) VALUES (%s, %s)", (rule_column, option_value))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            if err.errno == errorcode.ER_DUP_ENTRY:
                return "This option already exists for this column."
            logging.error(f"Error adding rule option: {err}", exc_info=True)
            return str(err)
        finally:
            if conn: conn.close()

    def delete_rule_option(self, option_id):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM suspicious_rule_options WHERE id = %s", (option_id,))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error deleting rule option: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
            
    def get_all_suspicious_rules(self):
        conn = self._get_connection()
        try:
            query = "SELECT sub_department_name, rule_column, rule_values FROM suspicious_rules"
            df = pd.read_sql_query(query, conn)
            if not df.empty and 'rule_values' in df.columns:
                 df['rule_values'] = df['rule_values'].apply(lambda x: json.loads(x) if isinstance(x, str) else x if x is not None else [])
            return df
        finally:
            if conn: conn.close()

    def save_suspicious_rule(self, sub_department, rule_column, rule_values):
        conn = self._get_connection()
        try:
            rule_values_json = json.dumps(rule_values) if rule_values else '[]'
            with conn.cursor() as cursor:
                query = """
                    INSERT INTO suspicious_rules (sub_department_name, rule_column, rule_values)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                    rule_values = VALUES(rule_values);
                """
                cursor.execute(query, (sub_department, rule_column, rule_values_json))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error saving suspicious rule: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
            
    def log_suspicious_transaction(self, run_id, original_row_data, created_user):
        # --- FINAL PRODUCTION FIX ---
        # Pre-process the dictionary to explicitly convert any pandas/numpy NaN values to None
        # before serialization. This is the most direct way to solve the JSON error.
        sanitized_data = {k: None if pd.isna(v) else v for k, v in original_row_data.items()}
        
        conn = self._get_connection()
        try:
            # Use the sanitized_data dictionary for serialization
            row_json = json.dumps(sanitized_data, default=json_serializer_default)
            
            with conn.cursor() as cursor:
                query = "INSERT INTO suspicious_transactions_log (run_id, original_row_data, created_user) VALUES (%s, %s, %s)"
                cursor.execute(query, (run_id, row_json, created_user))
            
            logging.info(f"Successfully logged suspicious transaction for user {created_user} in run {run_id}.")

        except mysql.connector.Error as err:
            st.error(f"Database Error: Could not log suspicious transaction. Details: {err}")
        except Exception as e:
            st.error(f"An unexpected error occurred while logging the transaction: {e}")
        finally:
            if conn and conn.is_connected():
                conn.close()
            
    def get_suspicious_transactions_for_admin(self):
        conn = self._get_connection()
        try:
            query = "SELECT * FROM suspicious_transactions_log WHERE status = 'Pending Admin Review' ORDER BY id DESC"
            df = pd.read_sql_query(query, conn)
            return self._process_log_df(df)
        finally:
            if conn: conn.close()

    def get_notification_settings(self):
        """Fetches both global and per-user notification settings."""
        conn = self._get_connection()
        settings = {'global': {}, 'users': {}}
        try:
            with conn.cursor(dictionary=True) as cursor:
                # Fetch global settings
                cursor.execute("SELECT setting_key, setting_value FROM global_settings")
                for row in cursor.fetchall():
                    settings['global'][row['setting_key']] = row['setting_value']
                
                # Fetch per-user settings
                cursor.execute("SELECT username, receive_auto_notifications FROM users")
                for row in cursor.fetchall():
                    settings['users'][row['username']] = row['receive_auto_notifications']
            return settings
        except mysql.connector.Error as err:
            logging.error(f"Error getting notification settings: {err}", exc_info=True)
            return settings
        finally:
            if conn: conn.close()

    def update_notification_setting(self, setting_type, key, value):
        """Updates a global or per-user notification setting."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                if setting_type == 'global':
                    sql = "UPDATE global_settings SET setting_value = %s WHERE setting_key = %s"
                    params = (str(value).lower(), key)
                elif setting_type == 'user':
                    sql = "UPDATE users SET receive_auto_notifications = %s WHERE username = %s"
                    params = (value, key)
                else:
                    return False
                
                cursor.execute(sql, params)
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error updating notification setting for {key}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    # --- NEW --- Methods for the Enhanced Clarification Workflow
    
    def create_entry_clarification(self, username, trigger_details):
        """Creates a new record in the entry_clarifications table."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "INSERT INTO entry_clarifications (username, trigger_details) VALUES (%s, %s)"
                cursor.execute(sql, (username, trigger_details))
            conn.commit()
            return cursor.lastrowid
        except mysql.connector.Error as err:
            logging.error(f"Error creating entry clarification for {username}: {err}", exc_info=True)
            return None
        finally:
            if conn: conn.close()
            
    def get_entry_clarifications(self, user_role, username=None, managed_users=None):
        """Fetches entry clarifications with hierarchical access."""
        conn = self._get_connection()
        try:
            # Base query orders by status to show active items first
            query = "SELECT * FROM `entry_clarifications`"
            params = []
            
            # Add role-based filtering
            if user_role == 'User' and username:
                query += " WHERE username = %s"
                params.append(username)
            elif user_role == 'Manager' and managed_users is not None and username is not None:
                team_members = managed_users + [username]
                if not team_members: return pd.DataFrame()
                placeholders = ','.join(['%s'] * len(team_members))
                query += f" WHERE username IN ({placeholders})"
                params.extend(team_members)
            
            # Order by status priority and then by creation date
            query += " ORDER BY FIELD(status, 'Pending User Action', 'Submitted', 'Replied', 'Accepted'), created_at DESC"
            
            df = pd.read_sql_query(query, conn, params=tuple(params))
            return df
        except mysql.connector.Error as err:
            logging.error(f"Error in get_entry_clarifications: {err}", exc_info=True)
            return pd.DataFrame()
        finally:
            if conn: conn.close()

    def submit_user_clarification(self, clarification_id, clarification_text):
        """Saves the user's clarification text and updates status to 'Submitted'."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE entry_clarifications SET user_clarification = %s, status = 'Submitted', submitted_at = %s WHERE id = %s"
                cursor.execute(sql, (clarification_text, datetime.now(), clarification_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error submitting user clarification for id {clarification_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
            
    def reply_to_clarification(self, clarification_id, reply_text, replied_by):
        """Saves management's reply and updates status to 'Replied'."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE entry_clarifications SET management_reply = %s, status = 'Replied', replied_by = %s, replied_at = %s WHERE id = %s"
                cursor.execute(sql, (reply_text, replied_by, datetime.now(), clarification_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error replying to clarification for id {clarification_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def batch_accept_correction_entries(self, exception_ids):
        """Marks a list of correction entries as accepted."""
        if not exception_ids:
            return False
        conn = self._get_connection()
        try:
            with conn.cursor(dictionary=True) as cursor:
                # First, retrieve all necessary data for the given exception_ids
                placeholders = ','.join(['%s'] * len(exception_ids))
                cursor.execute(f"SELECT id, original_row_data, exception_reason FROM exceptions WHERE id IN ({placeholders})", tuple(exception_ids))
                exceptions_data = cursor.fetchall()

                # Update the exception status in the main exceptions table
                update_sql = f"UPDATE exceptions SET is_accepted = TRUE WHERE id IN ({placeholders})"
                cursor.execute(update_sql, tuple(exception_ids))
                conn.commit()

                # Now, save the fingerprint for each accepted exception
                for exc_data in exceptions_data:
                    self.save_accepted_exception_fingerprint(exc_data['original_row_data'], exc_data['exception_reason'])
                
                return True
        except mysql.connector.Error as err:
            logging.error(f"Error in batch accepting correction entries: {err}", exc_info=True)
            conn.rollback()
            return False
        finally:
            if conn: conn.close()

    def accept_entry_clarification(self, clarification_id, accepted_by):
        """Marks a clarification as 'Accepted'."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                sql = "UPDATE entry_clarifications SET status = 'Accepted', accepted_by = %s, accepted_at = %s WHERE id = %s"
                cursor.execute(sql, (accepted_by, datetime.now(), clarification_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error accepting clarification for id {clarification_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def has_open_clarification(self, username):
        """Checks if a user has an active, non-accepted clarification request."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Checks for any clarification that isn't yet 'Accepted'
                sql = "SELECT 1 FROM entry_clarifications WHERE username = %s AND status != 'Accepted' LIMIT 1"
                cursor.execute(sql, (username,))
                if cursor.fetchone():
                    return True # User has an open request
            return False
        except mysql.connector.Error as err:
            logging.error(f"Error checking for open clarifications for {username}: {err}", exc_info=True)
            return False # Assume no open clarification on error
        finally:
            if conn: conn.close()
            
    def get_suspicious_transactions_for_user(self, username):
        conn = self._get_connection()
        try:
            query = "SELECT * FROM suspicious_transactions_log WHERE status = 'Rejected' AND created_user = %s ORDER BY reviewed_at DESC"
            df = pd.read_sql_query(query, conn, params=(username,))
            return self._process_log_df(df)
        finally:
            if conn: conn.close()


    # START of new function to be inserted
    def get_rejected_transactions(self):
        """Fetches all transactions that are pending user correction or have been corrected."""
        conn = self._get_connection()
        try:
            # Fetches items that have been rejected or that the user has already corrected.
            query = "SELECT * FROM suspicious_transactions_log WHERE status IN ('Rejected', 'User Corrected') ORDER BY reviewed_at DESC"
            df = pd.read_sql_query(query, conn)
            return self._process_log_df(df)
        finally:
            if conn: conn.close()
# END of new function

    # START of new function to be inserted
    # START of new missing function to be inserted
    def call_back_rejected_transaction(self, log_id):
        """Resets a 'Rejected' transaction back to 'Pending Admin Review'."""
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                # Reset the status and clear the rejection details
                query = """
                    UPDATE suspicious_transactions_log 
                    SET status = 'Pending Admin Review', 
                        reviewed_by = NULL, 
                        reviewed_at = NULL, 
                        admin_comment = NULL 
                    WHERE id = %s
                """
                cursor.execute(query, (log_id,))
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error calling back transaction ID {log_id}: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()
# END of new missing function

    # START of new function to be inserted

    def get_historical_fingerprints(self):
        """
        Fetches all historical transaction fingerprints from the dedicated log table.
        This is the final, most efficient version.
        """
        conn = self._get_connection()
        try:
            # Query the new, dedicated table which is much faster.
            query = "SELECT fingerprint_hash FROM transaction_fingerprints"
            df_fingerprints = pd.read_sql_query(query, conn)

            if df_fingerprints.empty:
                return set()

            # Return the fingerprints as a set for fast lookups
            return set(df_fingerprints['fingerprint_hash'])

        except Exception as e:
            st.error(f"Failed to fetch historical fingerprints for duplicate check: {e}")
            logging.error(f"Error in get_historical_fingerprints: {e}", exc_info=True)
            return set() # Return an empty set on failure
        finally:
            if conn and conn.is_connected():
                conn.close()
            
    def get_accepted_exception_fingerprints(self):
        """
        Fetches all combined_hash values for exceptions that have been marked as 'accepted'.
        Used to prevent re-flagging the exact same accepted issue in future uploads.
        """
        conn = self._get_connection()
        try:
            query = "SELECT combined_hash FROM accepted_exception_fingerprints"
            df_fingerprints = pd.read_sql_query(query, conn)

            if df_fingerprints.empty:
                return set()

            return set(df_fingerprints['combined_hash'])

        except mysql.connector.Error as e:
            logging.error(f"Error fetching accepted exception fingerprints: {e}", exc_info=True)
            return set() # Return an empty set on failure
        finally:
            if conn and conn.is_connected():
                conn.close()
# END of new function

    def _process_log_df(self, df):
        if df.empty:
            return pd.DataFrame()
        
        if 'original_row_data' in df.columns:
            try:
                # Handle case where data might already be a dict or needs loading from JSON string
                expanded_data = df['original_row_data'].apply(lambda x: x if isinstance(x, dict) else json.loads(x) if pd.notna(x) else {})
                expanded_df = pd.json_normalize(expanded_data)
                
                # Combine original df (without the json column) with the new expanded columns
                original_cols = df.columns.drop('original_row_data')
                df = pd.concat([df[original_cols].reset_index(drop=True), expanded_df.reset_index(drop=True)], axis=1)
            except Exception as e:
                logging.error(f"Error processing JSON log data: {e}", exc_info=True)
                # Return the original dataframe but without the problematic column to avoid crashing
                return df.drop(columns=['original_row_data'])
        return df

    def accept_suspicious_transaction(self, log_id, admin_username):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                query = "UPDATE suspicious_transactions_log SET status = 'Accepted', reviewed_by = %s, reviewed_at = %s WHERE id = %s"
                cursor.execute(query, (admin_username, datetime.now(), log_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error accepting suspicious transaction: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def reject_suspicious_transaction(self, log_id, admin_username, comment):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                query = "UPDATE suspicious_transactions_log SET status = 'Rejected', reviewed_by = %s, reviewed_at = %s, admin_comment = %s WHERE id = %s"
                cursor.execute(query, (admin_username, datetime.now(), comment, log_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error rejecting suspicious transaction: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()

    def confirm_user_correction(self, log_id):
        conn = self._get_connection()
        try:
            with conn.cursor() as cursor:
                query = "UPDATE suspicious_transactions_log SET status = 'User Corrected', user_corrected_at = %s WHERE id = %s"
                cursor.execute(query, (datetime.now(), log_id))
            conn.commit()
            return True
        except mysql.connector.Error as err:
            logging.error(f"Error confirming user correction: {err}", exc_info=True)
            return False
        finally:
            if conn: conn.close()


def check_and_trigger_notifications():
    """
    Checks all users for the 3-consecutive-run unresolved entry condition
    and triggers the clarification workflow if necessary.
    """
    logging.info("Running automatic check for unresolved entries...")
    db_manager = get_database_manager()

    # 1. Check if the global service is enabled
    settings = db_manager.get_notification_settings()
    if settings.get('global', {}).get('automatic_notifications_enabled') != 'true':
        logging.info("Automatic notification service is globally disabled. Skipping check.")
        return

    # 2. Get all users who have notifications enabled
    all_users_df = db_manager.get_all_users()
    if all_users_df.empty:
        return
    
    # Filter for users who are set to receive notifications (NULL is treated as True)
    eligible_users = all_users_df[all_users_df['receive_auto_notifications'] != False]['username'].tolist()

    conn = db_manager._get_connection()
    try:
        for user in eligible_users:
            # 3. Check if user already has an open clarification request
            if db_manager.has_open_clarification(user):
                logging.info(f"Skipping user {user}, they already have an open clarification request.")
                continue

            # 4. Check the "3 consecutive runs" condition
            # Get the user's most recent runs where they had exceptions
            perf_query = """
                SELECT DISTINCT run_id FROM user_performance 
                WHERE user = %s AND exception_records > 0 
                ORDER BY run_id DESC LIMIT 10
            """
            user_runs_df = pd.read_sql_query(perf_query, conn, params=(user,))
            
            if len(user_runs_df) < 3:
                continue # Not enough runs to meet the condition

            consecutive_unresolved_count = 0
            failed_run_ids = []

            for run_id in user_runs_df['run_id']:
                with conn.cursor() as cursor:
                    # Check if this run has any unresolved entries for this user
                    exc_query = """
                        SELECT 1 FROM exceptions 
                        WHERE run_id = %s AND created_user = %s 
                        AND correction_status IN ('Pending', 'No') 
                        AND is_accepted = FALSE LIMIT 1
                    """
                    cursor.execute(exc_query, (run_id, user))
                    if cursor.fetchone():
                        # This run is unresolved
                        consecutive_unresolved_count += 1
                        failed_run_ids.append(run_id)
                    else:
                        # The chain of unresolved runs is broken
                        break
            
            # 5. If condition is met, trigger the actions
            if consecutive_unresolved_count >= 3:
                logging.warning(f"Condition met for user {user}. Triggering clarification request.")
                trigger_details = f"For unresolved entries in runs: {', '.join(map(str, sorted(failed_run_ids)))}"
                
                # Create the clarification record
                clarification_id = db_manager.create_entry_clarification(user, trigger_details)
                
                if clarification_id:
                    # Send a notification to the user's dashboard
                    message = f"You have unresolved entries for {consecutive_unresolved_count} consecutive runs. Please provide a clarification in the Clarification Center."
                    db_manager.create_notification(user, 'Clarification Required', message)

    except Exception as e:
        logging.error(f"An error occurred during check_and_trigger_notifications: {e}", exc_info=True)
    finally:
        if conn and conn.is_connected():
            conn.close()


@st.cache_resource
def get_database_manager():
    return DatabaseManager()

db_manager = get_database_manager()

# ==============================================================================
# ADD THE FULLY UPDATED check_password() FUNCTION HERE
# ==============================================================================
def run_user_session_checks(username):
    """
    Runs checks when a user session starts, e.g., for pending clarifications.
    Stores results in the session state for the UI to use.
    """
    # Do not run these checks for the Super User
    if st.session_state.get("role") == "Super User":
        st.session_state['clarification_required'] = False
        return

    # First, check if the user has a waiver from the Super User
    has_waiver = db_manager.check_waiver_status(username)
    if has_waiver:
        st.session_state['clarification_required'] = False
        # Optional: Show an info message that a waiver is active
        # st.sidebar.info("A clarification waiver is currently active for your account.")
        return # Do not proceed with other checks if waiver is active

    # If no waiver, check for consecutive pending corrections
    # We use get() to avoid an error if the key doesn't exist yet
    if st.session_state.get('clarification_required') is None:
        pending_runs = db_manager.get_pending_correction_runs_for_user(username, consecutive_limit=3)
        if pending_runs:
            st.session_state['clarification_required'] = True
            st.session_state['clarification_run_ids'] = pending_runs
        else:
            st.session_state['clarification_required'] = False


# Fully updated check_password function using cookies

def check_password():
    """
    Returns `True` if the user is authenticated, `False` otherwise.
    This function uses st.session_state for modern, reliable session management.
    """
    db_manager = get_database_manager()

    # --- 1. Check if the user is already authenticated in the current session ---
    if st.session_state.get("authentication_status"):
        return True

    # --- 2. If not authenticated, display the login form ---
    st.markdown("### Please Log In")
    username_input = st.text_input("Username", key="login_username")
    password_input = st.text_input("Password", type="password", key="login_password")

    if st.button("Log in", key="login_button"):
        user_data = db_manager.get_user(username_input)

        # NEW: Check for disabled status first
        if user_data and user_data.get('disabled'):
            st.error("❌ Your account has been disabled. Please contact an administrator.")
            st.session_state["authentication_status"] = False
        
        # This logic correctly handles both "user not found" and "password incorrect"
        if user_data and bcrypt.checkpw(password_input.encode('utf-8'), user_data['hashed_password'].encode('utf-8')):
            # --- On successful login, set all necessary session state variables ---
            st.session_state["authentication_status"] = True
            st.session_state["username_actual"] = user_data['username']
            st.session_state["full_name"] = user_data['full_name']
            st.session_state["role"] = user_data['role']
            st.session_state["managed_users"] = db_manager.get_managed_users(user_data['username']) if user_data['role'] == "Manager" else []
            
            # Rerun the script to immediately reflect the logged-in state
            st.rerun()
        else:
            st.error("😕 User not known or password incorrect.")
            st.session_state["authentication_status"] = False

    # If the code reaches this point, the user is not yet authenticated.
    return False

@st.cache_data
def load_ledger_validation_mapping(base_ref_path="reference_data"):
    """Loads the combined ledger/sub-ledger mapping for VALIDATION purposes."""
    mapping_file = os.path.join(base_ref_path, "Ledgersubledger mapping.xlsx")
    LEDGER_CODE_COL = "Account2.Code"
    SUBLEDGER_CODE_COL = "Sub Ledger.Code"
    
    try:
        if not os.path.exists(mapping_file):
            st.error(f"VALIDATION Error: Ledger mapping file not found at '{mapping_file}'. Ledger combination validation will not work.")
            return None

        mapping_df = pd.read_excel(mapping_file)
        required_cols = [LEDGER_CODE_COL, SUBLEDGER_CODE_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.error(f"VALIDATION Error: Mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}")
            return None

        for col in required_cols:
            mapping_df[col] = mapping_df[col].astype(str).str.strip()

        return mapping_df

    except Exception as e:
        st.error(f"Failed to read or process validation mapping file {mapping_file}: {str(e)}")
        return None

@st.cache_data
def load_account_name_mapping(base_ref_path="reference_data"):
    """Loads the account code-to-name mapping for DISPLAY purposes."""
    mapping_file = os.path.join(base_ref_path, "account_mapping.xlsx")
    CODE_COL = "Account2.Code"
    NAME_COL = "Account2.Name"

    try:
        if not os.path.exists(mapping_file):
            st.warning(f"DISPLAY Warning: Account name mapping file not found at '{mapping_file}'. Ledger names may not display correctly.")
            return None
        
        mapping_df = pd.read_excel(mapping_file)
        required_cols = [CODE_COL, NAME_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.warning(f"DISPLAY Warning: Account name mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}. Names may not display.")
            return None
        
        mapping_df[CODE_COL] = mapping_df[CODE_COL].astype(str).str.strip()
        mapping_df[NAME_COL] = mapping_df[NAME_COL].astype(str).str.strip()

        return mapping_df[required_cols].drop_duplicates(subset=[CODE_COL]).reset_index(drop=True)

    except Exception as e:
        st.error(f"Failed to read or process account name mapping file {mapping_file}: {str(e)}")
        return None
    
@st.cache_data
def load_sub_departments(base_ref_path="reference_data"):
    """Loads the canonical list of Sub Department names."""
    sub_dept_file = os.path.join(base_ref_path, "SubDepartment.xlsx")
    try:
        if not os.path.exists(sub_dept_file):
            st.error(f"Sub Department reference file not found at '{sub_dept_file}'. The rule control page will not function correctly.")
            return []
        
        df = pd.read_excel(sub_dept_file)
        # Assuming the column name is 'Sub Department.Name' as in the main data
        if "Sub Department.Name" not in df.columns:
            st.error(f"The file '{sub_dept_file}' must contain a column named 'Sub Department.Name'.")
            return []
        
        sub_depts = df["Sub Department.Name"].dropna().astype(str).str.strip().unique().tolist()
        return sorted(sub_depts)
        
    except Exception as e:
        st.error(f"Failed to read or process sub-department file {sub_dept_file}: {str(e)}")
        return []

@st.cache_data
def load_subledger_name_mapping(base_ref_path="reference_data"):
    """Loads the sub-ledger code-to-name mapping for DISPLAY purposes."""
    mapping_file = os.path.join(base_ref_path, "subledger_mapping.xlsx")
    CODE_COL = "Sub Ledger.Code"
    NAME_COL = "SubLedger.Name"

    try:
        if not os.path.exists(mapping_file):
            st.warning(f"DISPLAY Warning: Sub-ledger name mapping file not found at '{mapping_file}'. Sub-ledger names may not display correctly.")
            return None
        
        mapping_df = pd.read_excel(mapping_file)
        required_cols = [CODE_COL, NAME_COL]
        if not all(col in mapping_df.columns for col in required_cols):
            st.warning(f"DISPLAY Warning: Sub-ledger name mapping file '{mapping_file}' is missing required columns. It needs: {required_cols}. Names may not display.")
            return None

        mapping_df[CODE_COL] = mapping_df[CODE_COL].astype(str).str.strip()
        mapping_df[NAME_COL] = mapping_df[NAME_COL].astype(str).str.strip()

        return mapping_df[required_cols].drop_duplicates(subset=[CODE_COL]).reset_index(drop=True)

    except Exception as e:
        st.error(f"Failed to read or process sub-ledger name mapping file {mapping_file}: {str(e)}")
        return None


class DataValidator:
    def __init__(self, base_ref_path="reference_data" , accepted_exception_fingerprints_set=None):
        self.base_ref_path = base_ref_path

        if 'hashlib' not in globals():
            import hashlib

        self.no_crop_check = {
            "Finance & Account", "Human Resource", "Administration",
            "Information Technology", "Legal", "Accounts Receivable & MIS", "Management"
            }
        self.no_activity_check = self.no_crop_check.copy()
        self.no_activity_check.update({"Production", "Processing", "Parent Seed"})
        self.ref_files = self._load_reference_data()
        
        self.valid_ledger_keys = set()
        self.LEDGER_CODE_COL = "Account2.Code"
        self.SUBLEDGER_CODE_COL = "Sub Ledger.Code"

        mapping_df = load_ledger_validation_mapping(self.base_ref_path)
        if mapping_df is not None:
             self.valid_ledger_keys = set(mapping_df[self.LEDGER_CODE_COL] + "_" + mapping_df[self.SUBLEDGER_CODE_COL])
             logging.info(f"Successfully loaded {len(self.valid_ledger_keys)} ledger-subledger mapping combinations into validator.")
        else:
             logging.error("Could not load ledger validation mapping into validator. Ledger validation will be skipped.")
        
        self.accepted_exception_fingerprints = accepted_exception_fingerprints_set if accepted_exception_fingerprints_set is not None else set()
        if not self.accepted_exception_fingerprints: # Check if it's empty, not if db_manager exists
            logging.warning("No accepted exception fingerprints provided to DataValidator. Accepted exception filtering will be skipped.")
        else:
            logging.info(f"Loaded {len(self.accepted_exception_fingerprints)} accepted exception fingerprints into validator.")

        self.training_map = {
            'Incorrect Ledger/Sub-Ledger Combination': 'Review the Ledger-Subledger mapping file for valid combinations.',
            'Incorrect Location Name': 'Review Location Name Guidelines: Ensure locations are valid (e.g., Bandamailaram).',
            'Incorrect Activity Name': 'Complete Activity Name Training: Use only approved activities for your department.',
            'Incorrect Crop Name': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'FC-Vertical Name cannot be blank': 'Ensure Valid FC-Vertical Name.',
            'Crop Name cannot be blank': 'Ensure Crop Name is provided when required.',
            'Incorrect Crop Name starting with ZZ': 'Check Crop Name Standards: Ensure crops are valid for your department.',
            'Incorrect Crop Name for FC-field crop Vertical': 'Ensure Valid Crop Name for FC-field crop.',
            'Incorrect Crop Name for VC-Veg Crop Vertical': 'Ensure Valid Crop Name for VC-Veg Crop.',
            'Incorrect Crop Name for Fruit Crop Vertical': 'Ensure Valid Crop Name for Fruit Crop.',
            'Incorrect Crop Name for Common vertical': 'Ensure Valid Crop Name for Common vertical.',
            'Incorrect Sub Department Name': 'Verify Sub-Department Standards.',
            'Incorrect Function Name': 'Check Function Name Guidelines.',
            'Incorrect FC-Vertical Name': 'Ensure Valid FC-Vertical Name.',
            'Need to Update Processing Location': 'Use Approved Processing Locations.',
            'Incorrect Activity Name for Lab QC': 'Use Lab QA Approved Activities.',
            'Incorrect Activity Name for Field QA': 'Use Field QA Approved Activities.',
            'Incorrect Activity Name for Bio Tech Services': 'Use Bio Tech Approved Activities.',
            'Sub Department should be blank': 'Ensure Sub-Department is blank for this department.',
            'Activity Name cannot be blank or start with ZZ': 'Complete Activity Name Training: Use only approved activities.',
            'Incorrect Activity Name for Biotech - Markers': 'Use Approved Activities for Biotech - Markers.',
            'Incorrect Activity Name for Biotech - Tissue Culture': 'Use Approved Activities for Biotech - Tissue Culture.',
            'Incorrect Activity Name for Biotech - Mutation': 'Use Approved Activities for Biotech - Mutation.',
            'Incorrect Activity Name for Entomology': 'Use Approved Activities for Entomology.',
            'Incorrect Activity Name for Pathology': 'Use Approved Activities for Pathology.',
            'Incorrect Activity Name for Bioinformatics': 'Use Approved Activities for Bioinformatics.',
            'Incorrect Activity Name for Biochemistry': 'Use Approved Activities for Biochemistry.',
            'Incorrect Activity Name for Common': 'Use Approved Activities for Common sub-department in Breeding Support.',
            'Need to update Zone can not left Blank': 'Ensure Zone is specified.',
            'Incorrect Zone Name for FC-field crop Vertical': 'Use Approved Zone Names for FC-field crop.',
            'Incorrect Zone Name for VC-Veg Crop Vertical': 'Use Approved Zone Names for VC-Veg Crop.',
            'Need to update Zone Name can not left Blank': 'Ensure Zone is specified for Common vertical in Production.',
            'Need to update Business Unit can not left Blank': 'Ensure Business Unit is specified.',
            'Incorrect Business Unit Name for FC-field crop Vertical': 'Use Approved Business Unit Names for FC-field crop.',
            'Incorrect Business Unit Name for VC-Veg Crop Vertical': 'Use Approved Business Unit Names for VC-Veg Crop.',
            'Need to update Region Name can not left Blank': 'Ensure Region is specified.',
            'Incorrect Region Name for FC-field crop Vertical': 'Use Approved Region Names for FC-field crop.',
            'Incorrect Region Name for VC-Veg Crop Vertical': 'Use Approved Region Names for VC-Veg Crop.',
            'Region, Zone, BU need to check for Root Stock': 'Ensure Region, Zone, and BU are blank for Root Stock vertical in Sales/Marketing.',
            'Incorrect Activity Name for Sales': 'Use Approved Activities for Sales.',
            'Incorrect Activity Name for Marketing': 'Use Approved Activities for Marketing.'
        }

    def _load_reference_data(self):
        loaded_ref_files = {}
        ref_file_mappings = {
            "FC_Crop": ("FC-field crop.xlsx", "Crop.Name"),
            "VC_Crop": ("VC-Veg Crop.xlsx", "Crop.Name"),
            "SBFC_Region": ("SBFC-Region.xlsx", "Region.Name"),
            "SBVC_Region": ("SBVC-Region.xlsx", "Region.Name"),
            "SaleFC_Zone": ("SaleFC-Zone.xlsx", "Zone.Name"),
            "SaleVC_Zone": ("SaleVC-Zone.xlsx", "Zone.Name"),
            "FC_BU": ("FC-BU.xlsx", "Business Unit.Name"),
            "VC_BU": ("VC-BU.xlsx", "Business Unit.Name"),
            "Fruit_Crop": ("Fruit Crop.xlsx", "Crop.Name"),
            "Common_Crop": ("Common crop.xlsx", "Crop.Name"),
            "ProductionFC_Zone": ("ProductionFC-Zone.xlsx", "Zone.Name"),
            "ProductionVC_Zone": ("ProductionVC-Zone.xlsx", "Zone.Name"),
            "SalesActivity": ("SalesActivity.xlsx", "Activity.Name"),
            "MarketingActivity": ("MarketingActivity.xlsx", "Activity.Name"),
            "RS_BU": ("RS-BU.xlsx", "Business Unit.Name"),
            "SaleRS_Zone": ("SaleRS-Zone.xlsx", "Zone.Name"),
            "SBRS_Region": ("SBRS-Region.xlsx", "Region.Name"),
            "Root Stock_Crop": ("Root Stock Crop.xlsx", "Crop.Name"),
            "Region_Excluded_Accounts": ("Region.Name excluded.xlsx", "Account.Code"),
            "Zone_Excluded_Accounts": ("Zone.Name excluded.xlsx", "Account.Code"),
        }

        if not os.path.isdir(self.base_ref_path):
            st.error(f"Reference data directory not found: '{self.base_ref_path}'. Please create it and add reference Excel files. Validations will be highly inaccurate.")
            logging.critical(f"Reference data directory not found: '{self.base_ref_path}'. Cannot load reference data.")
            return {key: [] for key in ref_file_mappings.keys()}

        all_files_loaded_successfully = True
        for key, (filename, col_name) in ref_file_mappings.items():
            try:
                file_path = os.path.join(self.base_ref_path, filename)
                if os.path.exists(file_path):
                    df_ref = pd.read_excel(file_path, engine='openpyxl')
                    if col_name in df_ref.columns:
                        if key in ["Region_Excluded_Accounts", "Zone_Excluded_Accounts"]:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        else:
                             loaded_ref_files[key] = df_ref[col_name].dropna().astype(str).str.strip().unique().tolist()
                        logging.info(f"Loaded {len(loaded_ref_files[key])} items for '{key}' from '{filename}'.")
                    else:
                        st.warning(f"Column '{col_name}' not found in reference file '{filename}'. '{key}' will be empty.")
                        logging.warning(f"Column '{col_name}' not found in '{file_path}' for key '{key}'.")
                        loaded_ref_files[key] = []
                        all_files_loaded_successfully = False
                else:
                    st.warning(f"Reference file '{filename}' not found in '{self.base_ref_path}'. Validations for '{key}' may be inaccurate.")
                    logging.warning(f"Reference file '{filename}' not found for key '{key}' at path '{file_path}'.")
                    loaded_ref_files[key] = []
                    all_files_loaded_successfully = False
            except Exception as e:
                st.error(f"Error loading reference file '{filename}' for '{key}': {e}")
                logging.error(f"Error loading reference file '{filename}' for key '{key}': {e}", exc_info=True)
                loaded_ref_files[key] = []
                all_files_loaded_successfully = False

        if not all_files_loaded_successfully:
            st.error("One or more reference data files could not be loaded correctly. Please check logs and file paths. Validation accuracy will be affected.")
        elif not any(loaded_ref_files.values()):
            st.error("All reference data lists are empty after attempting to load files. This indicates a problem with file contents or loading logic. Validations will be highly inaccurate.")
            logging.critical("All loaded reference file lists are empty. Check file contents and parsing logic.")
        else:
            logging.info("Reference data loading process completed.")
        return loaded_ref_files

    def is_not_blank(self, value):
        if pd.isna(value) or value is None:
            return False
        val = str(value).strip().replace("\u00A0", "").replace("\u200B", "")
        return val != "" and val.upper() not in ["N/A", "NULL", "NONE", "NA", "0", "-"]

    def is_blank(self, value):
        return not self.is_not_blank(value)

    def validate_row(self, dept, row):
        reasons = []
        # Extract values from row
        sub_dept = str(row.get("Sub Department.Name", "") or "").strip().replace("\u00A0", "").replace("\u200B", "")
        func = str(row.get("Function.Name", "") or "").strip()
        vertical = str(row.get("FC-Vertical.Name", "") or "").strip()
        loc = str(row.get("Location.Name", "") or "").strip()
        crop = str(row.get("Crop.Name", "") or "").strip()
        act = str(row.get("Activity.Name", "") or "").strip()
        region = row.get("Region.Name", "")
        zone = row.get("Zone.Name", "")
        bu = row.get("Business Unit.Name", "")
        account_code = str(row.get("Account.Code", "") or "").strip()
        ledger_code = str(row.get(self.LEDGER_CODE_COL, "") or "").strip()
        subledger_code = str(row.get(self.SUBLEDGER_CODE_COL, "") or "").strip()

        try:
            # Ensure sorted keys for consistent hashing
            original_row_data_str = json.dumps(row.to_dict(), default=json_serializer_default, sort_keys=True) # ADDED sort_keys=True
            data_hash = hashlib.sha256(original_row_data_str.encode('utf-8')).hexdigest()
        except Exception as e:
            logging.error(f"Error generating data hash for row: {e}", exc_info=True)
            data_hash = "" # Fallback to empty hash if serialization fails

        # <<< INTEGRATED LEDGER/SUB-LEDGER CHECK >>>
        # *** Only check for an invalid combination if codes have actually been entered. ***
        if self.is_not_blank(ledger_code) or self.is_not_blank(subledger_code):
            combination_key = f"{ledger_code}_{subledger_code}"
            if combination_key not in self.valid_ledger_keys:
                reasons.append("Incorrect Ledger/Sub-Ledger Combination")

        # --- Existing Generic and Department-Specific Checks ---
        if self.is_blank(loc) or loc.startswith("ZZ"):
            reasons.append("Incorrect Location Name")
        if dept not in self.no_activity_check and dept not in ["Breeding", "Trialing & PD", "Sales", "Marketing", "Breeding Support"]:
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Incorrect Activity Name")

        # Crop and Vertical validation
        if dept not in self.no_crop_check:
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("FC-Vertical Name cannot be blank")
            if self.is_blank(crop):
                reasons.append("Crop Name cannot be blank")
            elif crop.startswith("ZZ"):
                reasons.append("Incorrect Crop Name starting with ZZ")
            elif vertical == "FC-field crop" and crop not in self.ref_files.get("FC_Crop", []):
                reasons.append("Incorrect Crop Name for FC-field crop Vertical")
            elif vertical == "VC-Veg Crop" and crop not in self.ref_files.get("VC_Crop", []):
                reasons.append("Incorrect Crop Name for VC-Veg Crop Vertical")
            elif vertical == "Fruit Crop" and crop not in self.ref_files.get("Fruit_Crop", []):
                reasons.append("Incorrect Crop Name for Fruit Crop Vertical")
            elif vertical == "Common" and crop not in self.ref_files.get("Common_Crop", []):
                reasons.append("Incorrect Crop Name for Common vertical")
            elif vertical == "Root Stock" and crop not in self.ref_files.get("Root Stock_Crop", []):
                reasons.append("Incorrect Crop Name for Root Stock Crop Vertical")

        # Account Code exclusion checks
        if account_code in self.ref_files.get("Region_Excluded_Accounts", []) and self.is_not_blank(region):
            reasons.append("Region Name should be blank for this Account Code")
        if account_code in self.ref_files.get("Zone_Excluded_Accounts", []) and self.is_not_blank(zone):
            reasons.append("Zone Name should be blank for this Account Code")
        

        # Department-specific checks
        if dept == "Parent Seed":
            if sub_dept not in ["Breeder Seed Production", "Foundation Seed Production", "Processing FS"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Production":
            if sub_dept not in ["Commercial Seed Production", "Seed Production Research"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            # Zone validation for Commercial Seed Production sub-department
            if sub_dept == "Commercial Seed Production":
                if vertical == "FC-field crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("ProductionFC_Zone", []):
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                elif vertical == "VC-Veg Crop":
                    if self.is_blank(zone):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("ProductionVC_Zone", []):
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")                
                    elif vertical == "Common" and not self.is_blank(zone):
                        reasons.append("Need to update Vertical Accordingly")


        elif dept == "Processing":
            if sub_dept not in ["Processing", "Warehousing", "Project & Maintenance"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            if loc not in ["Bandamailaram", "Deorjhal", "Boriya"]:
                reasons.append("Need to Update Processing Location")

        elif dept == "Quality Assurance":
            if sub_dept not in ["Field QA", "Lab QC", "Bio Tech Services"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            # Sub-department-specific activity checks
            if sub_dept == "Lab QC" and act not in ["Lab Operations QA", "All Activity"]:
                reasons.append("Incorrect Activity Name for Lab QC")
            if sub_dept == "Field QA" and act not in ["Field Operations QA", "All Activity", "GOT"]:
                reasons.append("Incorrect Activity Name for Field QA")
            if sub_dept == "Bio Tech Services" and act not in ["Molecular", "All Activity"]:
                reasons.append("Incorrect Activity Name for Bio Tech Services")

        elif dept == "Seed Tech":
            if sub_dept not in ["Aging Test", "Pelleting", "Priming", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "In Licensing & Procurement":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Supply Chain":
                reasons.append("Incorrect Function Name")
            if vertical in ["", "N/A", "Common", "ZZ"]:
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Breeding":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if vertical in ["", "N/A", "ZZ"]:
                reasons.append("Incorrect FC-Vertical Name")
            if dept not in self.no_activity_check and act not in ["Breeding", "All Activity", "Trialing", "Pre Breeding", "Germplasm Maintainance", "Experimental Seed Production"]:
                reasons.append("Incorrect Activity Name")

        elif dept == "Breeding Support":
            if sub_dept not in ["Pathology", "Biotech - Tissue Culture", "Biotech - Mutation", "Biotech - Markers", "Bioinformatics", "Biochemistry", "Entomology", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            # Activity validation: Check for blank or ZZ first, then sub-department-specific checks
            if self.is_blank(act) or act.startswith("ZZ"):
                reasons.append("Activity Name cannot be blank or start with ZZ")
            else:
                # Sub-department-specific activity checks
                if sub_dept == "Biotech - Markers" and act not in ["Molecular", "Grain Quality", "Seed Treatment", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Markers")
                elif sub_dept == "Biotech - Tissue Culture" and act not in ["Tissue Culture", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Tissue Culture")
                elif sub_dept == "Biotech - Mutation" and act not in ["Mutation", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biotech - Mutation")
                elif sub_dept == "Entomology" and act not in ["Entomology", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Entomology")
                elif sub_dept == "Pathology" and act not in ["Pathalogy", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Pathology")
                elif sub_dept == "Bioinformatics" and act not in ["Bioinformatics", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Bioinformatics")
                elif sub_dept == "Biochemistry" and act not in ["Biochemistry", "All Activity"]:
                    reasons.append("Incorrect Activity Name for Biochemistry")
                elif sub_dept == "Common" and act not in ["All Activity"]:
                    reasons.append("Incorrect Activity Name for Common")

        elif dept == "Trialing & PD":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Research and Development":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            if dept not in self.no_activity_check and act not in ["CT", "All Activity", "Trialing", "RST", "OFD", "Disease"]:
                reasons.append("Incorrect Activity Name")

        elif dept == "Sales":
            valid_subs = ["Sales Brand", "Sales Export", "Sales Institutional & Govt"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            # Activity validation for Sales
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files.get("SalesActivity",[]):
                reasons.append("Incorrect Activity Name for Sales")
            # Business Unit, Zone, and Region validation for Sales Brand sub-department
            if sub_dept == "Sales Brand":
                if vertical == "FC-field crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("FC_BU", []):
                        reasons.append("Incorrect Business Unit Name for FC-field crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleFC_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for FC-field crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBFC_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for FC-field crop Vertical")

                elif vertical == "VC-Veg Crop":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("VC_BU", []):
                        reasons.append("Incorrect Business Unit Name for VC-Veg Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleVC_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for VC-Veg Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBVC_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for VC-Veg Crop Vertical")

                elif vertical == "Root Stock":
                    if self.is_blank(bu):
                        reasons.append("Need to update Business Unit can not left Blank")
                    elif bu not in self.ref_files.get("RS_BU", []):
                        reasons.append("Incorrect Business Unit Name for Root Stock Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(zone) and account_code not in self.ref_files.get("Zone_Excluded_Accounts", []):
                        reasons.append("Need to update Zone can not left Blank")
                    elif zone not in self.ref_files.get("SaleRS_Zone", []) and self.is_not_blank(zone):
                        reasons.append("Incorrect Zone Name for Root Stock Crop Vertical")
                    # MODIFIED: Added check to ensure account_code is not in the exclusion list
                    if self.is_blank(region) and account_code not in self.ref_files.get("Region_Excluded_Accounts", []):
                        reasons.append("Need to update Region Name can not left Blank")
                    elif region not in self.ref_files.get("SBRS_Region", []) and self.is_not_blank(region):
                        reasons.append("Incorrect Region Name for Root Stock Crop Vertical")

        elif dept == "Marketing":
            valid_subs = ["Business Development", "Digital Marketing", "Product Management"]
            if sub_dept not in valid_subs:
                reasons.append("Incorrect Sub Department Name")
            if func != "Sales and Marketing":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")
            elif vertical == "Root Stock" and any(self.is_not_blank(x) for x in [region, zone, bu]):
                reasons.append("Region, Zone, BU need to check for Root Stock")
            # Activity validation for Marketing
            if self.is_blank(act) or act.startswith("ZZ") or act not in self.ref_files.get("MarketingActivity",[]):
                reasons.append("Incorrect Activity Name for Marketing")

        elif dept == "Finance & Account":
            if sub_dept not in ["Accounts", "Finance", "Analytics, Internal Control & Budget", "Purchase ops", "Secretarial", "Document Management System", "Automation", "Group Company"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Human Resource":
            if sub_dept not in ["Compliances", "HR Ops", "Recruitment", "Team Welfare", "Training", "Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Administration":
            if sub_dept not in ["Events", "Maintenance", "Travel Desk","Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Information Technology":
            if sub_dept not in ["ERP Support", "Infra & Hardware", "Application Development"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Legal":
            if sub_dept not in ["Compliances", "Litigation","Common"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Accounts Receivable & MIS":
            if sub_dept not in ["Branch and C&F Ops", "Commercial & AR Management", "Common", "Order Processing", "Transport & Logistic"]:
                reasons.append("Incorrect Sub Department Name")
            if func != "Support Functions":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        elif dept == "Management":
            if self.is_not_blank(sub_dept):
                reasons.append("Sub Department should be blank")
            if func != "Management":
                reasons.append("Incorrect Function Name")
            if self.is_blank(vertical) or vertical.startswith("ZZ"):
                reasons.append("Incorrect FC-Vertical Name")

        unique_reasons = sorted(list(set(reasons))) # Keep this sorted for consistent reason_hash
        filtered_reasons = []
        if unique_reasons:
            for reason_str in unique_reasons:
                reason_hash = hashlib.sha256(reason_str.encode('utf-8')).hexdigest()
                combined_hash = hashlib.sha256(f"{data_hash}_{reason_hash}".encode('utf-8')).hexdigest()

                if combined_hash in self.accepted_exception_fingerprints:
                    logging.info(f"Ignoring previously accepted exception: {combined_hash} for row {row.get('Document No.')}")
                else:
                    filtered_reasons.append(reason_str)
        unique_reasons = filtered_reasons # Assign the filtered list back

        severity = len(unique_reasons) * 2
        return unique_reasons, severity

    def validate_dataframe(self, df):
        """
        Validates the entire DataFrame by splitting it into chunks and processing them in parallel.
        This method utilizes multiple CPU cores to significantly speed up validation for large files.
        """
        exceptions = []
        
        null_like_values = [pd.NA, "N/A", "NaN", "null", "NONE", "", " ", "-", "\u00A0", None, 0, "0"]
        if 'Sub Department.Name' in df.columns:
            df['Sub Department.Name'] = df['Sub Department.Name'].replace(null_like_values, "").astype(str).str.strip()
        else:
            df['Sub Department.Name'] = ""

        input_columns = df.columns.tolist()
        
        if 'Department.Name' not in df.columns:
            st.error("Critical Error: 'Department.Name' column is missing from the input file.")
            logging.critical("Critical Error: 'Department.Name' column is missing in validate_dataframe.")
            return pd.DataFrame(columns=input_columns + ['Exception Reasons', 'Severity']), {}

        # OPTIMIZATION: Use the number of available CPUs for parallel processing. Fallback to 4 if undetected.
        num_workers = os.cpu_count() or 4
        logging.info(f"Starting parallel validation with {num_workers} workers.")
        
        # OPTIMIZATION: Split the DataFrame into chunks for each worker.
        df_chunks = np.array_split(df, num_workers)

        global db_manager 
        accepted_fingerprints_for_validator = db_manager.get_accepted_exception_fingerprints()

        # OPTIMIZATION: Use ProcessPoolExecutor to run validation in parallel.
        with concurrent.futures.ProcessPoolExecutor(max_workers=num_workers) as executor:
            # Submit each chunk to the _validate_chunk function
            futures = [
                executor.submit(
                    _validate_chunk, 
                    # NEW: Create a new DataValidator instance for each worker
                    DataValidator(base_ref_path=self.base_ref_path, accepted_exception_fingerprints_set=accepted_fingerprints_for_validator), 
                    chunk
                ) 
                for chunk in df_chunks
            ]
            
            # Collect the results as they are completed
            for future in concurrent.futures.as_completed(futures):
                try:
                    exceptions.extend(future.result())
                except Exception as e:
                    logging.error(f"A validation chunk failed: {e}", exc_info=True)
                    st.error(f"An error occurred during parallel processing: {e}")

        # --- Post-processing after parallel execution ---
        
        # Create the final exceptions DataFrame from the collected results
        exceptions_df_output = pd.DataFrame()
        output_columns_with_exceptions = input_columns + ['Exception Reasons', 'Severity']

        if exceptions:
            exceptions_df_output = pd.DataFrame(exceptions)
            # === FIX: SET THE INDEX TO THE ORIGINAL INDEX WE CAPTURED ===
            if 'original_index' in exceptions_df_output.columns:
                exceptions_df_output.set_index('original_index', inplace=True)
            # === END OF FIX ===
            
            # Ensure all original columns are present in the output
            for col in output_columns_with_exceptions:
                if col not in exceptions_df_output.columns:
                    exceptions_df_output[col] = pd.NA
            exceptions_df_output = exceptions_df_output.reindex(columns=output_columns_with_exceptions, fill_value=pd.NA)
        else:
            exceptions_df_output = pd.DataFrame(columns=output_columns_with_exceptions)
        
        # Calculate department statistics after all exceptions have been found
        department_stats = {}
        if 'Department.Name' in df.columns:
            # Get total records per department from the original dataframe
            total_records_by_dept = df.groupby('Department.Name').size().to_dict()
            
            # Get exception records per department from the exceptions dataframe
            exception_records_by_dept = {}
            if not exceptions_df_output.empty and 'Department.Name' in exceptions_df_output.columns:
                exception_records_by_dept = exceptions_df_output.groupby('Department.Name').size().to_dict()

            for dept, total_records in total_records_by_dept.items():
                exception_records = exception_records_by_dept.get(dept, 0)
                department_stats[dept] = {
                    'total_records': total_records,
                    'exception_records': exception_records,
                    'exception_rate': (exception_records / total_records * 100) if total_records > 0 else 0
                }

        return exceptions_df_output, department_stats


def display_metric(title, value, delta=None, container=None):
    target_container = container if container else st
    delta_html = f'<div class="metric-delta" style="font-size:0.8rem; color:#e2e8f0;">{delta}</div>' if delta else ""
    target_container.markdown(f"""
    <div class="metric-container">
        <div class="metric-title">{title}</div>
        <div class="metric-value">{value}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

def create_excel_report(exceptions_df, dept_stats, filename_for_logging="ExcelReport"):
    output = io.BytesIO()
    try:
        if exceptions_df is None or exceptions_df.empty:
            logging.info(f"create_excel_report: exceptions_df is empty for {filename_for_logging}. Creating empty Exceptions sheet.")
            exceptions_df_prepared = exceptions_df.copy() if exceptions_df is not None else pd.DataFrame()
        else:
            exceptions_df_prepared = exceptions_df.copy()

        if 'Exception Reasons' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Exception Reasons'] = exceptions_df_prepared['Exception Reasons'].astype(str).replace('<NA>', '').replace('nan', '').replace('None','')
        if 'Severity' in exceptions_df_prepared.columns:
            exceptions_df_prepared['Severity'] = pd.to_numeric(exceptions_df_prepared['Severity'], errors='coerce').fillna(0)
            if exceptions_df_prepared['Severity'].notna().all():
                 try:
                    exceptions_df_prepared['Severity'] = exceptions_df_prepared['Severity'].astype(int)
                 except ValueError:
                    pass

        numeric_cols_to_preserve = ['Net amount', 'Severity', 'id', 'run_id']
        if dept_stats:
             numeric_cols_to_preserve.extend(['Total Records', 'Exception Records', 'Exception Rate (%)'])

        for col in exceptions_df_prepared.columns:
            if col in numeric_cols_to_preserve:
                if col == 'Net amount':
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce').fillna(0.0)
                elif col in ['id', 'run_id']:
                    exceptions_df_prepared[col] = pd.to_numeric(exceptions_df_prepared[col], errors='coerce')
            else:
                if not pd.api.types.is_numeric_dtype(exceptions_df_prepared[col]) and \
                   not pd.api.types.is_string_dtype(exceptions_df_prepared[col]):
                    exceptions_df_prepared[col] = exceptions_df_prepared[col].astype(str).replace('<NA>', '').replace('nan', '').replace('None','').replace('NaT','')

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            exceptions_df_prepared.to_excel(writer, sheet_name='Exceptions', index=False)
            if not dept_stats:
                 dept_summary_df = pd.DataFrame(columns=['Department', 'Total Records', 'Exception Records', 'Exception Rate (%)'])
            else:
                dept_summary_df = pd.DataFrame([{'Department': dept, 'Total Records': stats.get('total_records', 0), 'Exception Records': stats.get('exception_records', 0), 'Exception Rate (%)': round(stats.get('exception_rate', 0), 2)} for dept, stats in dept_stats.items()])
            dept_summary_df.to_excel(writer, sheet_name='Department Summary', index=False)
            workbook = writer.book
            header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            thin_border_side = Side(style='thin'); cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.max_row == 0 : continue
                df_ref_for_headers = exceptions_df_prepared if sheet_name == 'Exceptions' else dept_summary_df
                if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                    if not df_ref_for_headers.columns.empty:
                        for c_idx, h_val in enumerate(df_ref_for_headers.columns, 1):
                            cell = ws.cell(row=1, column=c_idx, value=str(h_val)); cell.font=header_font; cell.fill=header_fill; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=cell_border
                    else: continue
                for cell in ws[1]:
                    if cell.value is not None: cell.value = str(cell.value); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = cell_border
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx); cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cell.border = cell_border
                        header_cell_value = str(ws.cell(row=1, column=col_idx).value or "")
                        if header_cell_value == "Net amount": cell.number_format = '#,##0.00'
                        elif header_cell_value == "Exception Rate (%)": cell.number_format = '0.00"%"'
                        elif header_cell_value == "Severity" or "Records" in header_cell_value or "ID" in header_cell_value or "id" in header_cell_value : cell.number_format = '0'
                for col_idx_letter_enum, column_cells_obj in enumerate(ws.columns, 1):
                    current_col_letter = get_column_letter(col_idx_letter_enum); header_val_str = str(ws.cell(row=1, column=col_idx_letter_enum).value or ""); max_length = len(header_val_str)
                    for cell_in_col_obj in column_cells_obj:
                        if cell_in_col_obj.row == 1: continue
                        try:
                            if cell_in_col_obj.value is not None: cell_str_val = str(cell_in_col_obj.value); cell_len = max(len(s) for s in cell_str_val.split('\n')) if '\n' in cell_str_val else len(cell_str_val); max_length = max(max_length, cell_len)
                        except: pass
                    adjusted_width = min(max_length + 5, 60); ws.column_dimensions[current_col_letter].width = adjusted_width
        output.seek(0)
        if output.getbuffer().nbytes == 0:
            logging.error(f"create_excel_report: Excel output buffer is empty for {filename_for_logging}")
            st.markdown('<div class="error-box"><strong>❌ Error:</strong> Failed to generate Excel report (output empty).</div>', unsafe_allow_html=True); return None
        return output
    except Exception as e:
        logging.exception(f"create_excel_report: Error generating Excel report for {filename_for_logging}: {e}")
        st.markdown(f'<div class="error-box"><strong>❌ Error:</strong> Error generating Excel report: {e}</div>', unsafe_allow_html=True); return None

def display_interactive_exceptions(df, key_prefix="df"):
    if df.empty:
        st.success("No exceptions to display for the current selection.")
        return

    display_cols = [col for col in [
        'id', 'run_id', 'Department.Name', 'Created user', 'Location.Name',
        'Account2.Code', 'Sub Ledger.Code', 'Exception Reasons', 'Severity'
    ] if col in df.columns]

    st.info("💡 Click on any row to see the complete original data for that record.")

    selection_key = f"{key_prefix}_selection"
    if selection_key not in st.session_state:
        st.session_state[selection_key] = None

    edited_df = st.data_editor(
        df[display_cols],
        key=f"{key_prefix}_editor",
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        disabled=df.columns,
        on_change=lambda: st.session_state.update({selection_key: st.session_state[f"{key_prefix}_editor"].get("selection")})
    )

    selection = st.session_state[selection_key]
    if selection and selection.get("rows"):
        try:
            selected_index = selection["rows"][0]
            selected_record = df.iloc[selected_index]
            exception_id = selected_record.get('id', 'N/A')

            with st.expander(f"**👁️ Viewing Original Data for Exception ID: {exception_id}**", expanded=True):
                # In MySQL version, this can be a dict or a string
                original_data_raw = selected_record.get('original_row_data')
                if pd.notna(original_data_raw):
                    try:
                        original_data = original_data_raw if isinstance(original_data_raw, dict) else json.loads(original_data_raw)
                        original_df = pd.DataFrame([original_data]).dropna(axis=1, how='all')
                        st.markdown("#### Original Record Details")
                        st.dataframe(original_df.T.rename(columns={0: 'Value'}), use_container_width=True)
                    except (json.JSONDecodeError, TypeError):
                        st.error("Could not parse the original row data. It might be corrupted.")
                        st.text(original_data_raw)
                else:
                    st.warning("Original row data is not available for this exception record.")
        except IndexError:
            st.warning("Could not retrieve the selected row. Please try again.")
        finally:
            st.session_state[selection_key] = None

def process_uploaded_file(uploaded_file, selected_date=None):
    try:
        # --- Get User Context (from your code) ---
        user_role = st.session_state.get("role")
        username = st.session_state.get("username_actual")
        managed_users = st.session_state.get("managed_users", [])

        df_original = pd.DataFrame()
        with st.spinner(f"📖 Reading file: {uploaded_file.name}..."):
            try:
                df_original = pd.read_excel(uploaded_file, engine='openpyxl', skiprows=5, skipfooter=1)
                df_original.columns = df_original.columns.str.strip()
            except Exception as e:
                st.markdown(f'<div class="error-box"><strong>❌ Error!</strong> Could not read Excel file "{uploaded_file.name}". Details: {str(e)}</div>', unsafe_allow_html=True)
                return

        if df_original.empty:
            st.markdown(f'<div class="warning-box"><strong>⚠ Warning!</strong> The uploaded file "{uploaded_file.name}" is empty or could not be parsed.</div>', unsafe_allow_html=True)
            return
            
        if 'Created user' not in df_original.columns:
            st.error("CRITICAL ERROR: The uploaded file must contain a 'Created user' column to be processed.")
            return

        # --- Role-Based Filtering (from your code) ---
        df_to_process = pd.DataFrame()
        df_original['Created user'] = df_original['Created user'].astype(str)

        if user_role == 'User':
            df_to_process = df_original[df_original['Created user'].str.lower() == username.lower()].copy()
            filter_message = f"As a **User**, this file has been automatically filtered to process records created by you."
        elif user_role == 'Manager':
            accessible_users = [username.lower()] + [u.lower() for u in st.session_state.get("managed_users", [])]
            df_to_process = df_original[df_original['Created user'].str.lower().isin(accessible_users)].copy()
            filter_message = f"As a **Manager**, this file has been filtered for you and your team."
        else: # Management and Super User
            df_to_process = df_original.copy()
            filter_message = "As **Management/Super User**, all records in the file will be processed."
        
        st.info(f"""
        {filter_message}\n
        - **{len(df_original)}** records found in the original file.
        - **{len(df_to_process)}** records to be initially processed.
        """)

        if df_to_process.empty:
            st.warning("No records in the uploaded file match your user profile or team. Nothing to process.")
            return

        # --- Check for essential columns (from your code) ---
        required_columns_for_processing = ['Department.Name', 'Account2.Code', 'Sub Ledger.Code']
        missing_core_cols = [col for col in required_columns_for_processing if col not in df_to_process.columns]
        if missing_core_cols:
            st.error(f'Error! Missing essential columns for processing: {", ".join(missing_core_cols)}. Cannot proceed.')
            return
        
        # --- NEW DUPLICATE CHECK LOGIC ---
        with st.spinner("🔍 Checking for duplicate transactions from previous runs..."):
            historical_fingerprints = db_manager.get_historical_fingerprints()
            
            # NEW: Fetch accepted exception fingerprints here
            accepted_exception_fingerprints_from_db = db_manager.get_accepted_exception_fingerprints()
            
            # MODIFIED: Pass accepted_exception_fingerprints_from_db instead of db_manager
            validator = DataValidator(base_ref_path="reference_data", accepted_exception_fingerprints_set=accepted_exception_fingerprints_from_db)
            
            # 1. Run validation on ALL incoming data first
            all_exceptions_df, _ = validator.validate_dataframe(df_to_process.copy())
            
            # 2. Separate the incoming data into two groups
            exception_indices = all_exceptions_df.index
            exceptions_to_process = df_to_process.loc[exception_indices].copy()
            clean_df_from_upload = df_to_process.drop(index=exception_indices).copy()

            # 3. Filter the CLEAN rows to remove historical duplicates
            new_clean_rows = []
            ignored_clean_count = 0
            for index, row in clean_df_from_upload.iterrows():
                # ... inside the loop `for index, row in clean_df_from_upload.iterrows():` ...
                doc_no = str(row.get("Document No.", "")).strip().lower()
                location = str(row.get("Location.Name", "")).strip().lower()
                activity = str(row.get("Activity.Name", "")).strip().lower()
                crop = str(row.get("Crop.Name", "")).strip().lower()
                
                # --- NORMALIZED FINGERPRINT ---
                try:
                    net_amount_val = float(row.get("Net amount", 0.0))
                    net_amount = f"{net_amount_val:.2f}"
                except (ValueError, TypeError):
                    net_amount = "0.00"
                
                fingerprint = f"{doc_no}|{location}|{activity}|{crop}|{net_amount}"

                if fingerprint in historical_fingerprints:
                    ignored_clean_count += 1
                else:
                    new_clean_rows.append(row)
            
            if new_clean_rows:
                final_clean_df = pd.DataFrame(new_clean_rows, columns=clean_df_from_upload.columns)
            else:
                final_clean_df = pd.DataFrame(columns=clean_df_from_upload.columns)

            # 4. Re-assemble the final dataframe for processing
            final_df_to_process = pd.concat([exceptions_to_process, final_clean_df], ignore_index=True)
        
        st.success(f"Duplicate check complete. Ignored **{ignored_clean_count}** clean rows that were duplicates of past transactions.")
        st.info(f"Processing **{len(final_df_to_process)}** unique transactions (**{len(exceptions_to_process)}** with exceptions, **{len(final_clean_df)}** new clean rows).")
        # --- END OF NEW DUPLICATE CHECK LOGIC ---

        # --- Save the main validation run ---
        current_run_id = db_manager.save_validation_run(
            filename=uploaded_file.name,
            total_records=len(final_df_to_process),
            total_exceptions=len(exceptions_to_process),
            file_size=uploaded_file.size,
            upload_time=selected_date
        )

        # --- Suspicious Transaction Check (from your code, now runs on de-duplicated data) ---
        processing_log = [] 
        with st.spinner(f"🕵️‍♀️ Checking for suspicious transactions..."):
            immunity_list = db_manager.load_suspense_immunity_list()
            suspicious_rules_df = db_manager.get_all_suspicious_rules()
            flagged_count = 0

            if not suspicious_rules_df.empty:
                rules_dict = {}
                for _, rule in suspicious_rules_df.iterrows():
                    if rule['rule_values']:
                        key = (rule['sub_department_name'], rule['rule_column'])
                        rules_dict[key] = [str(v).lower() for v in rule['rule_values']]

                for index, row in final_df_to_process.iterrows(): # MODIFIED: Runs on de-duplicated data
                    user = row.get('Created user', 'Unknown User')
                    
                                        # ... inside the loop `for index, row in final_df_to_process.iterrows():` ...
                    doc_no_s = str(row.get("Document No.", "")).strip().lower()
                    location_s = str(row.get("Location.Name", "")).strip().lower()
                    activity_s = str(row.get("Activity.Name", "")).strip().lower()
                    crop_s = str(row.get("Crop.Name", "")).strip().lower()
                    
                    # --- NORMALIZED FINGERPRINT ---
                    try:
                        net_amount_val_s = float(row.get("Net amount", 0.0))
                        net_amount_s = f"{net_amount_val_s:.2f}"
                    except (ValueError, TypeError):
                        net_amount_s = "0.00"
                    
                    fingerprint_s = f"{doc_no_s}|{location_s}|{activity_s}|{crop_s}|{net_amount_s}"
                    if fingerprint_s in historical_fingerprints:
                        continue
                        
                    account_code = str(row.get("Account2.Code", "")).strip()
                    sub_ledger_code = str(row.get("Sub Ledger.Code", "")).strip()
                    if f"{account_code}_{sub_ledger_code}" in immunity_list:
                        continue

                    sub_dept = str(row.get('Sub Department.Name', '')).strip()
                    if not sub_dept: continue
                    
                    for (rule_sub_dept, rule_col), rule_vals_lower in rules_dict.items():
                        if sub_dept == rule_sub_dept:
                            row_val_lower = str(row.get(rule_col, '')).strip().lower()
                            log_entry = f"Row for **{user}**: Checking Sub-Dept `'{sub_dept}'`. Comparing value `'{row_val_lower}'` in column `'{rule_col}'` against rule `'{rule_vals_lower}'`."
                            
                            if row_val_lower in rule_vals_lower:
                                db_manager.log_suspicious_transaction(current_run_id, row.to_dict(), user)
                                flagged_count += 1
                                processing_log.append(log_entry + " -> **MATCH FOUND**")
                                break
                            else:
                                processing_log.append(log_entry + " -> No Match")
        
        if flagged_count > 0:
            st.success(f"✅ Flagged **{flagged_count}** new suspicious transaction(s) for manual admin review.")
        else:
            st.info("ℹ️ No transactions matched the custom suspicious rules.")
            
        with st.expander("🔍 View Suspicious Rule Check Log"):
            if not processing_log:
                st.write("No applicable rules were found for the sub-departments in this file.")
            else:
                for entry in processing_log:
                    st.markdown(entry, unsafe_allow_html=True)
        
        # --- Existing DataValidator Logic ---
        summary_tab, exceptions_tab, data_tab = st.tabs(["📊 Validation Summary", "📋 Exception Records", "📖 Processed Data"])
        
        exceptions_df_from_validation = all_exceptions_df
        # Re-calculate department statistics on the final de-duplicated dataframe
        _, department_statistics = validator.validate_dataframe(final_df_to_process)

        if not exceptions_df_from_validation.empty:
            db_manager.save_exceptions(current_run_id, exceptions_df_from_validation)
        
        db_manager.save_user_performance(current_run_id, final_df_to_process, exceptions_df_from_validation)
        if department_statistics:
            db_manager.save_department_summary(current_run_id, department_statistics)

        # --- Ghost User Detection (from your code, on de-duplicated data) ---
        try:
            all_users_in_db_df = db_manager.get_all_users()
            known_users = set(all_users_in_db_df['username'].str.lower()) if not all_users_in_db_df.empty else set()
            uploaded_users = set(final_df_to_process['Created user'].dropna().astype(str).str.lower())
            ghost_users = uploaded_users - known_users
            if ghost_users:
                ghost_users_str = ", ".join(sorted(list(ghost_users)))
                with summary_tab:
                     st.warning(f"👻 **Ghost Users Found:** The following users from the file do not exist in the system: `{ghost_users_str}`.")
                super_users = db_manager.get_users_by_role('Super User')
                if super_users:
                    message = f"In file '{uploaded_file.name}', these usernames were found but do not exist: **{ghost_users_str}**. Please add them if they are valid users."
                    for su in super_users:
                        db_manager.create_notification(username=su, notif_type='Ghost User Detected', message=message)
        except Exception as e_ghost:
            logging.error(f"Error during ghost user detection: {e_ghost}", exc_info=True)

        # --- Create and Save Excel Report (from your code) ---
        excel_report_data = create_excel_report(exceptions_df_from_validation, department_statistics, uploaded_file.name)
        if excel_report_data:
            db_manager.save_excel_report(current_run_id, excel_report_data)
            excel_report_data.seek(0)

        # --- Display Results in UI (from your code, using new counts) ---
        with summary_tab:
            st.markdown("#### 📊 File Information (Post-Deduplication)")
            col_info1, col_info2, col_info3 = st.columns(3)
            display_metric("Unique Records Processed", f"{len(final_df_to_process):,}", container=col_info1)
            display_metric("Total Columns", len(final_df_to_process.columns), container=col_info2)
            display_metric("File Size", f"{uploaded_file.size / 1024:.1f} KB", container=col_info3)

            st.markdown("#### 🛠 Standard Validation Results")
            if exceptions_df_from_validation.empty:
                st.success(f'**Perfect!** No standard validation issues found in "{uploaded_file.name}"!')
            else:
                st.warning(f'**Warning!** Found {len(exceptions_df_from_validation)} records with standard validation issues.')
                col_res1, col_res2, col_res3 = st.columns(3)
                display_metric("Total Exceptions", f"{len(exceptions_df_from_validation):,}", container=col_res1)
                current_exception_rate = (len(exceptions_df_from_validation)/len(final_df_to_process)*100) if len(final_df_to_process) > 0 else 0
                display_metric("Exception Rate", f"{current_exception_rate:.2f}%", container=col_res2)
                avg_sev = exceptions_df_from_validation['Severity'].mean() if 'Severity' in exceptions_df_from_validation.columns else 0.0
                display_metric("Average Severity", f"{avg_sev:.2f}", container=col_res3)

        with exceptions_tab:
            if exceptions_df_from_validation.empty:
                 st.success("No standard exceptions to display.")
            else:
                st.markdown("##### 📋 Standard Exception Records")
                display_interactive_exceptions(exceptions_df_from_validation, key_prefix="upload_view")
                if excel_report_data:
                    st.download_button(
                        label=f"📥 Download Standard Validation Report",
                        data=excel_report_data,
                        file_name=f"Validation_Report_{uploaded_file.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        with data_tab:
            st.markdown(f"#### 📖 Final De-duplicated Dataset for Processing")
            st.dataframe(final_df_to_process, use_container_width=True)

        if not final_df_to_process.empty:
            with st.spinner("Saving transaction history for future duplicate checks..."):
                processed_fingerprints = set()
                for _, row in final_df_to_process.iterrows():
                    try:
                        doc_no = str(row.get("Document No.", "")).strip().lower()
                        location = str(row.get("Location.Name", "")).strip().lower()
                        activity = str(row.get("Activity.Name", "")).strip().lower()
                        crop = str(row.get("Crop.Name", "")).strip().lower()
                        net_amount_val = float(row.get("Net amount", 0.0))
                        net_amount = f"{net_amount_val:.2f}"
                        fingerprint = f"{doc_no}|{location}|{activity}|{crop}|{net_amount}"
                        processed_fingerprints.add(fingerprint)
                    except (ValueError, TypeError):
                        continue
                db_manager.save_transaction_fingerprints(current_run_id, list(processed_fingerprints))
                st.success("Transaction history saved.")
                check_and_trigger_notifications()

    except Exception as e_process:
        st.error(f'An unhandled error occurred while processing "{uploaded_file.name}": {str(e_process)}')
        logging.exception(f"Unhandled error processing uploaded file {uploaded_file.name}: {e_process}")

def show_upload_page():
    st.markdown("### 🏠 Upload & Validate")

    # --- NEW PERMISSION CHECK ---
    # This value is calculated in main() and stored in the session state.
    # It correctly checks the hierarchical permissions (role and user-specific override).
    if not st.session_state.get('can_upload', False):
        st.warning("🔒 Your account does not have permission to upload files. Please contact a Super User for access.")
        return  # Stop rendering the rest of the page if permission is denied
    # --- END OF PERMISSION CHECK ---

    # The rest of the page is only visible if the user has upload permission
    with st.container(border=True):
        st.write("📤 **Drag & Drop Your Excel Files Here**")
        st.caption("or click 'Browse files' to select them from your computer")
        uploaded_files_list = st.file_uploader(
            "File Uploader",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
    custom_upload_date = st.date_input(
        "Override upload date (optional)",
        value=None,
        help="If you select a date, it will be used as the upload time for all files in this batch."
    )
    if uploaded_files_list:
        st.success(f"✅ {len(uploaded_files_list)} file(s) selected! Processing...")
        for individual_uploaded_file in uploaded_files_list:
            with st.expander(f"⚙️ Processing: {individual_uploaded_file.name}", expanded=True):
                process_uploaded_file(individual_uploaded_file, selected_date=custom_upload_date)
    
    st.markdown("---")
    st.markdown("### 📝 Manual Data Entry & Validation")
    with st.form("manual_data_entry_form"):
        st.markdown("##### Enter Record Details (fields with * are mandatory):")
        manual_custom_date = st.date_input("Override record date (optional)", value=None, key="manual_date_override")
        cols_m1_row1, cols_m1_row2, cols_m1_row3 = st.columns(3)
        department_manual_input = cols_m1_row1.selectbox("Department*", [""] + sorted(["Parent Seed", "Production", "Processing", "Quality Assurance", "Finance & Account", "Human Resource", "Administration", "Information Technology", "Legal", "Accounts Receivable & MIS", "Seed Tech", "In Licensing & Procurement", "Breeding", "Breeding Support", "Trialing & PD", "Sales", "Marketing", "Management"]), key="manual_input_department", help="Select the department.")
        location_manual_input = cols_m1_row2.text_input("Location Name*", key="manual_input_location", help="Enter location name (e.g., Bandamailaram).")
        created_user_manual_input = cols_m1_row3.text_input("Created User*", key="manual_input_created_user", help="Enter the user ID of the creator.")
        
        # You can expand this form with more fields as needed
        activity_manual_input = "" 
        net_amount_manual_input = 0.0
        sub_dept_manual_input = ""
        modified_user_manual_input = ""
        crop_manual_input = ""
        function_manual_input = ""
        vertical_manual_input = ""
        region_manual_input = ""
        zone_manual_input = ""
        business_unit_manual_input = ""

        submitted_manual_record = st.form_submit_button("Validate & Submit Manual Record")
        if submitted_manual_record:
            if not department_manual_input or not location_manual_input or not created_user_manual_input:
                st.error("Department, Location Name, and Created User are mandatory fields for manual entry.")
            else:
                manual_validator = DataValidator(base_ref_path="reference_data")
                manual_ref_data_loaded = any(isinstance(lst, list) and len(lst) > 0 for lst in manual_validator.ref_files.values())
                if not manual_ref_data_loaded:
                    st.error("CRITICAL: Reference data not loaded. Manual validation cannot be performed accurately.")
                else:
                    manual_row_data = pd.Series({'Department.Name': department_manual_input, 'Location.Name': location_manual_input, 'Activity.Name': activity_manual_input, 'Created user': created_user_manual_input, 'Net amount': net_amount_manual_input, 'Sub Department.Name': sub_dept_manual_input, 'Modified user': modified_user_manual_input, 'Crop.Name': crop_manual_input, 'Function.Name': function_manual_input, 'FC-Vertical.Name': vertical_manual_input, 'Region.Name': region_manual_input, 'Zone.Name': zone_manual_input, 'Business Unit.Name': business_unit_manual_input})
                    manual_reasons, manual_severity = manual_validator.validate_row(department_manual_input, manual_row_data)
                    manual_df_for_db = pd.DataFrame([manual_row_data])
                    if manual_reasons:
                        st.warning(f"Validation Issues for manual entry: {'; '.join(manual_reasons)} (Severity: {manual_severity})")
                        manual_df_for_db['Exception Reasons'] = "; ".join(manual_reasons)
                        manual_df_for_db['Severity'] = manual_severity
                        manual_entry_run_id = db_manager.save_validation_run(filename=f"Manual_Entry_Error_{datetime.now().strftime('%Y%m%d_%H%M%S')}", total_records=1, total_exceptions=1, file_size=0, upload_time=manual_custom_date)
                        db_manager.save_exceptions(manual_entry_run_id, manual_df_for_db)
                        db_manager.save_user_performance(manual_entry_run_id, pd.DataFrame([manual_row_data]), manual_df_for_db)
                        st.info(f"Manual record submitted with noted validation issues (Run ID: {manual_entry_run_id}).")
                    else:
                        valid_manual_entry_run_id = db_manager.save_validation_run(filename=f"Manual_Entry_OK_{datetime.now().strftime('%Y%m%d_%H%M%S')}", total_records=1, total_exceptions=0, file_size=0, upload_time=manual_custom_date)
                        empty_exceptions_for_valid_manual = pd.DataFrame(columns=['Created user', 'Exception Reasons', 'Severity'])
                        db_manager.save_user_performance(valid_manual_entry_run_id, pd.DataFrame([manual_row_data]), empty_exceptions_for_valid_manual)
                        st.success(f"Manual record validated successfully and run logged (Run ID: {valid_manual_entry_run_id}).")

def show_exception_details_page(start_date, end_date):
    """
    Displays a detailed, filterable view of all exceptions within the user's scope
    and the selected date range.
    """
    st.markdown("### 📋 All Exception Details")
    st.info("This page provides a detailed, interactive view of all individual exception records you have access to. Use the filters to narrow down the results.")

    # 1. Get user context from session state
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    # 2. Get all accessible validation runs within the filtered date range
    history_df = db_manager.get_validation_history(user_role, username, managed_users)
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty:
        st.warning("No validation runs found for your account in the selected date range.")
        return

    # 3. Load all exception data for the accessible runs
    with st.spinner("Loading exception data..."):
        run_ids_to_load = history_df['id'].tolist()
        all_exceptions_list = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
        # Concatenate all dataframes, ignoring any that might be empty
        unfiltered_df = pd.concat([df for df in all_exceptions_list if not df.empty], ignore_index=True)

    if unfiltered_df.empty:
        st.success("No exception records were found in the selected date range.")
        return

    # 4. Apply hierarchical security filtering to the master dataframe
    scoped_df = pd.DataFrame()
    if user_role == 'User':
        scoped_df = unfiltered_df[unfiltered_df['Created user'].str.lower() == username.lower()]
    elif user_role == 'Manager':
        team_to_filter = [username.lower()] + [u.lower() for u in managed_users]
        scoped_df = unfiltered_df[unfiltered_df['Created user'].str.lower().isin(team_to_filter)]
    else: # Management and Super User see all data
        scoped_df = unfiltered_df.copy()

    if scoped_df.empty:
        st.warning("No exception records fall within your specific access scope for this period.")
        return

    st.markdown("---")
    st.markdown("#### 🔍 Filters")

    # 5. Create UI filters based on the *scoped* data
    filter_col1, filter_col2 = st.columns(2)

    with filter_col1:
        # The user list is now correctly populated based on the user's role and access
        user_options = ["All Users"] + sorted(scoped_df['Created user'].dropna().unique().tolist())
        selected_user = st.selectbox("Filter by User", options=user_options)

    with filter_col2:
        dept_options = ["All Departments"] + sorted(scoped_df['Department.Name'].dropna().unique().tolist())
        selected_dept = st.selectbox("Filter by Department", options=dept_options)

    # 6. Apply filters to create the final dataframe for display
    final_df = scoped_df.copy()
    if selected_user != "All Users":
        final_df = final_df[final_df['Created user'] == selected_user]
    if selected_dept != "All Departments":
        final_df = final_df[final_df['Department.Name'] == selected_dept]

    st.markdown("---")

    # 7. Display the results
    if final_df.empty:
        st.warning("No records match the selected filter criteria.")
    else:
        st.markdown(f"####  Displaying **{len(final_df)}** Exception Records")
        # Reuse the existing helper function to display the interactive table
        display_interactive_exceptions(final_df, key_prefix="details_page")

def show_analytics_page(start_date, end_date):
    st.markdown("### 📊 Dashboard Analytics")
    
    validation_history = db_manager.get_validation_history()
    if start_date and end_date:
        validation_history = validation_history[(validation_history['upload_time'].dt.date >= start_date) & (validation_history['upload_time'].dt.date <= end_date)]

    if validation_history.empty:
        st.info("No validation runs found for the selected period.")
        return

    st.markdown("#### 📈 Overall Statistics (for selected period)")
    stat_col1, stat_col2, stat_col3 = st.columns(3)
    display_metric("Total Validation Runs", f"{len(validation_history):,}", container=stat_col1)
    total_recs_processed = validation_history['total_records'].sum()
    display_metric("Total Records Processed", f"{total_recs_processed:,}", container=stat_col2)
    total_excs_found = validation_history['total_exceptions'].sum()
    display_metric("Total Exceptions Found", f"{total_excs_found:,}", container=stat_col3)

    if total_recs_processed > 0:
        st.markdown("#### 🔍 Data Quality Snapshot (for selected period)")
        # ... (Pie chart logic is unchanged) ...
        labels = ['Records with Exceptions', 'Records without Exceptions']
        values = [total_excs_found, total_recs_processed - total_excs_found]
        colors = ['#FF6B6B', '#6BCB77']
        fig_overall_quality = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.4, marker_colors=colors, hoverinfo='label+percent+value', textinfo='value+label', insidetextorientation='radial', pull=[0.05, 0])])
        fig_overall_quality.update_layout(annotations=[dict(text='Quality', x=0.5, y=0.5, font_size=20, showarrow=False, font=PLOTLY_FONT)], legend_title_text='Record Status', margin=dict(t=30, b=30, l=10, r=10), font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=400)
        st.plotly_chart(fig_overall_quality, use_container_width=True)

    run_ids_in_period = validation_history['id'].tolist()
    if run_ids_in_period:
        # UPDATED: Replaced direct sqlite3 connection with a proper MySQL connection
        conn_analytics = db_manager._get_connection()
        dept_summary_df = pd.DataFrame()
        try:
            # Use %s placeholders for MySQL
            query_placeholder = ','.join(['%s'] * len(run_ids_in_period))
            query = f"SELECT department, total_records, exception_records FROM department_summary WHERE run_id IN ({query_placeholder})"
            dept_summary_df = pd.read_sql_query(query, conn_analytics, params=run_ids_in_period)
        except mysql.connector.Error as e_analytics_dept:
            logging.error(f"Error fetching department summary for analytics: {e_analytics_dept}", exc_info=True)
        finally:
            if conn_analytics and conn_analytics.is_connected():
                conn_analytics.close()

        if not dept_summary_df.empty:
            st.markdown(f"#### 🏭 Department Analysis (for selected period)")
            # ... (Department analysis logic is unchanged) ...
            agg_dept_summary = dept_summary_df.groupby('department').agg(total_records=('total_records', 'sum'), exception_records=('exception_records', 'sum')).reset_index()
            agg_dept_summary['exception_rate'] = (agg_dept_summary['exception_records'] / agg_dept_summary['total_records'] * 100).fillna(0)
            agg_dept_summary_sorted = agg_dept_summary.sort_values(by='exception_rate', ascending=False)
            fig_dept_analysis = px.bar(agg_dept_summary_sorted, x='department', y='exception_rate', labels={'exception_rate': 'Exception Rate (%)', 'department': 'Department'}, color='exception_rate', color_continuous_scale='Sunsetdark', text_auto='.2f', hover_name='department', custom_data=['total_records', 'exception_records'])
            fig_dept_analysis.update_traces(hovertemplate="<b>%{hovertext}</b><br><br>" + "Exception Rate: %{y:.2f}%<br>" + "Total Records: %{customdata[0]:,}<br>" + "Exception Records: %{customdata[1]:,}<extra></extra>")
            fig_dept_analysis.update_layout(title_text="Exception Rate by Department", title_x=0.5, title_font=PLOTLY_TITLE_FONT, xaxis_title="Department", yaxis_title="Exception Rate (%)", margin=dict(l=40, r=20, t=60, b=150), xaxis_tickangle=-45, font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False))
            st.plotly_chart(fig_dept_analysis, use_container_width=True)
            st.markdown("##### 📋 Department Summary Table (for selected period)")
            st.dataframe(agg_dept_summary.style.format({"exception_rate": "{:.2f}%", "total_records":"{:,}","exception_records":"{:,}"}), use_container_width=True, hide_index=True)


def show_trends_page(start_date, end_date):
    st.markdown("### 📈 Trends & History")
    trends_history_df = db_manager.get_validation_history()

    if start_date and end_date:
        trends_history_df = trends_history_df[
            (trends_history_df['upload_time'].dt.date >= start_date) & 
            (trends_history_df['upload_time'].dt.date <= end_date)
        ]

    if trends_history_df.empty:
        st.info("No historical data available for the selected period. Adjust the filter to see trends.")
        return

    trends_history_df = trends_history_df.sort_values(by='upload_time', ascending=True)
    
    fig_trends = go.Figure()
    theme_colors = {'exceptions': '#FF6B6B', 'records': '#6A89CC'}

    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'],
                                    y=trends_history_df['total_exceptions'],
                                    mode='lines+markers', name='Total Exceptions',
                                    line=dict(color=theme_colors['exceptions'], width=2.5, shape='spline'),
                                    marker=dict(symbol="circle", size=8, line=dict(width=1,color=theme_colors['exceptions'])),
                                    hovertemplate="<b>Total Exceptions</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>"))
    fig_trends.add_trace(go.Scatter(x=trends_history_df['upload_time'],
                                    y=trends_history_df['total_records'],
                                    mode='lines+markers', name='Total Records Processed',
                                    line=dict(color=theme_colors['records'], width=2.5, shape='spline'),
                                    marker=dict(symbol="square", size=8, line=dict(width=1,color=theme_colors['records'])),
                                    hovertemplate="<b>Total Records</b><br>Date: %{x|%Y-%m-%d %H:%M}<br>Count: %{y:,}<extra></extra>",
                                    fill='tozeroy',
                                    fillcolor='rgba(106, 137, 204, 0.2)'
                                    ))
    fig_trends.update_layout(
        title_text="Validation Trends Over Time",
        title_x=0.5,
        title_font=PLOTLY_TITLE_FONT,
        xaxis_title="Upload Date",
        yaxis_title="Count",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    bgcolor='rgba(255,255,255,0.7)', bordercolor='rgba(0,0,0,0.05)', borderwidth=1),
        margin=dict(l=50, r=20, t=70, b=40),
        hovermode="x unified",
        font=PLOTLY_FONT,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da'),
        xaxis=dict(gridcolor='#e9ecef', zerolinecolor='#ced4da', rangeslider=dict(visible=True, thickness=0.05), type="date")
    )
    st.plotly_chart(fig_trends, use_container_width=True)
    st.markdown("##### 📜 Validation History Log")
    display_history_log_df = trends_history_df[['id', 'filename', 'upload_time', 'total_records', 'total_exceptions', 'file_size']].copy()
    display_history_log_df.columns = ['Run ID', 'Filename', 'Upload Time', 'Total Records', 'Total Exceptions', 'File Size (Bytes)']
    display_history_log_df['Upload Time'] = display_history_log_df['Upload Time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    st.dataframe(display_history_log_df.sort_values(by='Upload Time', ascending=False), use_container_width=True, hide_index=True)


def show_user_location_page(start_date, end_date):
    st.markdown("### 👤📍 User & Location Analysis")

    # --- 1. Get User Context & Initial Data ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])
    
    # This call is already correctly scoped based on previous updates
    history_df = db_manager.get_validation_history(user_role, username, managed_users)
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty: 
        st.info("No validation runs found for the selected period. Adjust filter to enable User & Location analysis.")
        return

    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="user_location_run_selector")
    
    if not selected_run_display: return
    
    selected_run_id = run_options_with_all[selected_run_display]

    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        run_ids_to_load = history_df['id'].tolist() if selected_run_id == "all" else [selected_run_id]
        all_exceptions = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
        ul_exceptions_df_master = pd.concat([df for df in all_exceptions if not df.empty], ignore_index=True)

    if ul_exceptions_df_master.empty:
        st.success(f"No exceptions found for the selected scope. Nothing to analyze.")
        return

    # --- NEW: HIERARCHICAL DATA SCOPING BLOCK ---
    # This block ensures the data is filtered *before* being used in the dashboard.
    scoped_exceptions_df = pd.DataFrame()
    if user_role == 'User':
        scoped_exceptions_df = ul_exceptions_df_master[ul_exceptions_df_master['Created user'].str.lower() == username.lower()].copy()
    elif user_role == 'Manager':
        team_to_filter = [u.lower() for u in managed_users] + [username.lower()]
        scoped_exceptions_df = ul_exceptions_df_master[ul_exceptions_df_master['Created user'].str.lower().isin(team_to_filter)].copy()
    else: # Management and Super User see all data within the selected runs
        scoped_exceptions_df = ul_exceptions_df_master.copy()

    if scoped_exceptions_df.empty:
        st.warning("No exception data found for your specific access scope (user or team).")
        return
    # --- END OF NEW BLOCK ---

    st.markdown("#### Filters")
    # The user list for the filter is now created from the *scoped* data.
    user_list = ["All Users"] + sorted(scoped_exceptions_df['Created user'].dropna().unique().tolist())
    selected_user = st.selectbox("Filter by User", options=user_list, key="user_location_user_filter")

    # The final DataFrame for analysis is now correctly filtered from the start.
    ul_exceptions_df = scoped_exceptions_df.copy()
    if selected_user != "All Users":
        ul_exceptions_df = scoped_exceptions_df[scoped_exceptions_df['Created user'] == selected_user].copy()
    
    if ul_exceptions_df.empty:
        st.warning(f"No exceptions found for user '{selected_user}' in this scope.")
        return

    st.markdown(f"---")
    st.markdown(f"#### 📊 Exceptions by User and Location (Scope: {selected_run_display})")
    
    # The rest of the function remains the same as it now operates on correctly filtered data.
    ul_required_cols = ['Created user', 'Location.Name', 'Severity', 'Exception Reasons']
    for col_ul in ul_required_cols:
        if col_ul not in ul_exceptions_df.columns:
            st.warning(f"Missing required column '{col_ul}' in exceptions data. Analysis might be incomplete.")
            ul_exceptions_df[col_ul] = pd.NA
    
    ul_summary_df = ul_exceptions_df.groupby(['Created user', 'Location.Name'], dropna=False).agg(
        Total_Severity_Score=('Severity', 'sum'),
        Number_of_Exceptions=('Exception Reasons', lambda x: x.notna().sum())
    ).reset_index().rename(columns={'Created user': 'User', 'Location.Name': 'Location'})
    
    ul_metric_col1, ul_metric_col2, ul_metric_col3 = st.columns(3)
    display_metric("Total Exceptions", f"{ul_exceptions_df['Exception Reasons'].notna().sum():,}", container=ul_metric_col1)
    display_metric("Unique Users with Exceptions", f"{ul_exceptions_df['Created user'].nunique()}", container=ul_metric_col2)
    ul_avg_severity = ul_exceptions_df['Severity'].mean() if not ul_exceptions_df['Severity'].dropna().empty else 0.0
    display_metric("Average Severity", f"{ul_avg_severity:.2f}", container=ul_metric_col3)
    
    st.markdown("##### 📋 User-Location Exception Summary Table")
    st.dataframe(ul_summary_df.sort_values(by="Number_of_Exceptions", ascending=False), use_container_width=True, hide_index=True)
    
    if not ul_summary_df.empty:
        fig_ul_exc_count_chart = px.bar(ul_summary_df, x='User', y='Number_of_Exceptions', color='Location', barmode='group', title="Exceptions by User and Location", labels={'Number_of_Exceptions': 'Number of Exceptions'}, color_discrete_sequence=px.colors.qualitative.Pastel)
        fig_ul_exc_count_chart.update_layout(title_x=0.5, title_font=PLOTLY_TITLE_FONT, font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45), margin=dict(l=40, r=20, t=60, b=120))
        st.plotly_chart(fig_ul_exc_count_chart, use_container_width=True)
        
        fig_ul_severity_chart = px.bar(ul_summary_df, x='User', y='Total_Severity_Score', color='Location', barmode='stack', title="Total Severity by User and Location (Stacked)", labels={'Total_Severity_Score': 'Total Severity Score'}, color_discrete_sequence=px.colors.qualitative.Antique)
        fig_ul_severity_chart.update_layout(title_x=0.5, title_font=PLOTLY_TITLE_FONT, font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45), margin=dict(l=40, r=20, t=60, b=120))
        st.plotly_chart(fig_ul_severity_chart, use_container_width=True)
        
    ul_error_types_df = pd.DataFrame(columns=['Error Type', 'Count'])
    if 'Exception Reasons' in ul_exceptions_df.columns and ul_exceptions_df['Exception Reasons'].notna().any():
        st.markdown("##### 📊 Top 10 Common Error Types for this Scope")
        ul_error_types_df = ul_exceptions_df['Exception Reasons'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        ul_error_types_df.columns = ['Error Type', 'Count']
        st.dataframe(ul_error_types_df.head(10), use_container_width=True, hide_index=True)
        
        fig_ul_top_errors = px.bar(ul_error_types_df.head(10), x='Error Type', y='Count', title="Top 10 Error Types by Occurrence", color='Count', color_continuous_scale=px.colors.sequential.Tealgrn, text_auto=True)
        fig_ul_top_errors.update_layout(title_x=0.5, title_font=PLOTLY_TITLE_FONT, font=PLOTLY_FONT, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', yaxis=dict(gridcolor='#e9ecef'), xaxis=dict(showgrid=False, tickangle=-45), margin=dict(l=40, r=20, t=60, b=150))
        st.plotly_chart(fig_ul_top_errors, use_container_width=True)
        
    ul_user_risk_df = pd.DataFrame(columns=['User', 'Total Exceptions by User', 'Average Severity Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)'])
    if 'Created user' in ul_exceptions_df.columns and ul_exceptions_df['Created user'].notna().any():
        st.markdown("##### 🎯 User Risk Analysis for this Scope")
        ul_user_risk_df_agg = ul_exceptions_df.groupby('Created user', dropna=False).agg(Total_Exceptions_by_User=('Exception Reasons', lambda x: x.notna().sum()), Average_Severity_Score=('Severity', 'mean')).reset_index().rename(columns={'Created user': 'User'})
        ul_total_exceptions_in_run = ul_user_risk_df_agg['Total_Exceptions_by_User'].sum()
        ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] = (ul_user_risk_df_agg['Total_Exceptions_by_User'] / ul_total_exceptions_in_run * 100) if ul_total_exceptions_in_run > 0 else 0.0
        ul_user_risk_df_agg['Average_Severity_Score'] = ul_user_risk_df_agg['Average_Severity_Score'].fillna(0)
        ul_max_possible_severity = 10
        ul_user_risk_df_agg['Normalized_Severity'] = (ul_user_risk_df_agg['Average_Severity_Score'] / ul_max_possible_severity).clip(0,1) * 100
        ul_user_risk_df_agg['Risk_Score (0-100)'] = (ul_user_risk_df_agg['Contribution_to_Exceptions (%)'] * 0.6 + ul_user_risk_df_agg['Normalized_Severity'] * 0.4).round(2)
        ul_user_risk_df = ul_user_risk_df_agg[['User', 'Total_Exceptions_by_User', 'Average_Severity_Score', 'Contribution_to_Exceptions (%)', 'Risk_Score (0-100)']]
        st.dataframe(ul_user_risk_df.sort_values(by="Risk_Score (0-100)", ascending=False), use_container_width=True, hide_index=True)
        
    ul_excel_output = io.BytesIO()
    with pd.ExcelWriter(ul_excel_output, engine='openpyxl') as ul_writer:
        ul_summary_df.to_excel(ul_writer, sheet_name='User_Location_Summary', index=False)
        ul_error_types_df.to_excel(ul_writer, sheet_name='Common_Error_Types', index=False)
        ul_user_risk_df.to_excel(ul_writer, sheet_name='User_Risk_Analysis', index=False)
    ul_excel_output.seek(0)
    if ul_excel_output.getbuffer().nbytes > 0:
        file_name_suffix = f"Run_{selected_run_id}" if selected_run_id != 'all' else 'All_Runs'
        st.download_button(label="📥 Download User-Location Analysis Report", data=ul_excel_output, file_name=f"User_Location_Report_{file_name_suffix}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ADD THIS ENTIRE NEW FUNCTION TO YOUR SCRIPT

def show_clarification_center_page():
    st.markdown("### 📢 Clarification Center")
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    tab1_title = "Submitted Clarifications"
    tab2_title = "Manage Waivers (Super User Only)"

    # Only Super User can see the waiver management tab
    if user_role == 'Super User':
        tab1, tab2 = st.tabs([tab1_title, tab2_title])
    else:
        tab1 = st.tabs([tab1_title])[0]

    with tab1:
        st.subheader("Review Submitted Clarifications")
        st.caption("Review and acknowledge clarifications submitted by users for unresolved correction statuses.")

        clarifications_df = db_manager.get_clarifications(user_role, username, managed_users)

        if clarifications_df.empty:
            st.info("There are no pending clarifications to review in your scope.")
        else:
            for _, row in clarifications_df.iterrows():
                with st.expander(f"Clarification from **{row['username']}** submitted on {row['submitted_at'].strftime('%d-%b-%Y %H:%M')}"):
                    st.markdown(f"**Associated Run IDs:** `{row['run_ids']}`")
                    st.markdown("**User's Explanation:**")
                    st.info(row['clarification_text'])
                    
                    if st.button("Acknowledge & Dismiss", key=f"ack_{row['id']}"):
                        success = db_manager.acknowledge_clarification(row['id'], username)
                        if success:
                            st.success(f"Clarification ID {row['id']} has been acknowledged.")
                            # Create a notification for the user whose clarification was acknowledged
                            db_manager.create_notification(
                                username=row['username'],
                                notif_type="Clarification Acknowledged",
                                message=f"Your clarification regarding runs {row['run_ids']} has been reviewed and acknowledged by {username}."
                            )
                            st.rerun()
                        else:
                            st.error("Failed to acknowledge the clarification.")
    
    # This tab will only be created if the user is a Super User
    if user_role == 'Super User':
        with tab2:
            st.subheader("Grant or Revoke Clarification Waivers")
            st.caption("Grant a waiver to a user to temporarily prevent the clarification form from appearing for them.")

            all_users_df = db_manager.get_all_users()
            users_to_manage = all_users_df[all_users_df['role'] != 'Super User']['username'].tolist()

            with st.form("waiver_form"):
                st.markdown("##### Grant a New Waiver")
                selected_user = st.selectbox("Select a user to grant a waiver", options=users_to_manage)
                waiver_end_date = st.date_input("Waive until (end of day)", value=datetime.now().date() + pd.Timedelta(days=30))
                
                submitted = st.form_submit_button("Grant Waiver")
                if submitted:
                    if selected_user and waiver_end_date:
                        success = db_manager.grant_waiver(selected_user, waiver_end_date, username)
                        if success:
                            st.success(f"Waiver granted for {selected_user} until {waiver_end_date.strftime('%Y-%m-%d')}.")
                            db_manager.create_notification(
                                username=selected_user,
                                notif_type="Clarification Waiver Granted",
                                message=f"A clarification waiver has been granted to your account until {waiver_end_date.strftime('%Y-%m-%d')} by {username}."
                            )
                        else:
                            st.error("Failed to grant waiver.")
            
            st.markdown("---")
            st.markdown("##### Active Waivers")
            active_waivers_df = db_manager.get_all_waivers()
            if active_waivers_df.empty:
                st.info("No active waivers.")
            else:
                for _, waiver in active_waivers_df.iterrows():
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("User", waiver['username'])
                    col2.metric("Waived Until", waiver['waived_until'].strftime('%d-%b-%Y'))
                    col3.metric("Granted By", waiver['waived_by'])
                    with col4:
                        st.write("") # for vertical alignment
                        st.write("")
                        if st.button("Revoke", key=f"revoke_{waiver['id']}"):
                            db_manager.revoke_waiver(waiver['id'])
                            st.success(f"Waiver for {waiver['username']} has been revoked.")
                            st.rerun()

def show_user_performance_page(start_date, end_date):
    st.markdown("### 👤📊 User Performance Dashboard")

    # --- 1. Get User Context & Initial Data ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])
    
    all_users_df = db_manager.get_all_users()
    full_name_map = pd.Series(all_users_df.full_name.values,index=all_users_df.username).to_dict() if not all_users_df.empty else {}

    history_df = db_manager.get_validation_history(user_role, username, managed_users)
    if start_date and end_date:
        history_df = history_df[(pd.to_datetime(history_df['upload_time']).dt.date >= start_date) & 
                                (pd.to_datetime(history_df['upload_time']).dt.date <= end_date)]

    if history_df.empty:
        st.info("No validation runs found for your accessible scope and selected date range.")
        return

    # --- 2. Run Scope Filter ---
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs in Period": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run Scope", options=list(run_options_with_all.keys()), key="perf_run_selector")

    if not selected_run_display: return
    run_ids_in_scope = history_df['id'].tolist() if selected_run_display == "All Runs in Period" else [run_options_with_all[selected_run_display]]
    if not run_ids_in_scope: st.info("No runs match the selected criteria."); return

    # --- 3. HIERARCHICAL VIEW LOGIC ---
    
    if user_role == 'User':
        display_name = f"{full_name_map.get(username, username)} ({username})"
        st.markdown(f"#### Your Performance Details")
        display_single_user_performance(username, run_ids_in_scope, full_name_map)

    elif user_role == 'Manager':
        team_list = managed_users + [username]
        team_with_exceptions = get_filtered_users_with_exceptions(run_ids_in_scope, team_list)
        
        if not team_with_exceptions:
            st.info("No users in your team had exceptions in the selected scope.")
            return

        options_map = {"All My Team (Summary)": "summary"}
        for user in sorted(team_with_exceptions):
            options_map[f"{full_name_map.get(user, user)} ({user})"] = user
        
        selected_view_display = st.selectbox("Select View", options=list(options_map.keys()))
        selected_view_user = options_map[selected_view_display]

        if selected_view_user == "summary":
            st.markdown(f"#### Performance Summary for **Your Team**")
            display_all_users_performance_summary(run_ids_in_scope, team_list, full_name_map)
        else:
            st.markdown(f"#### Performance Details for **{selected_view_display}**")
            display_single_user_performance(selected_view_user, run_ids_in_scope, full_name_map)

    elif user_role in ['Management', 'Super User']:
        # MODIFIED: Added "View by Specific User" option
        view_type = st.radio(
            "Select View Type",
            ["View All Users (Summary)", "View by Specific Manager", "View by Specific User"],
            horizontal=True,
            key="mgmt_view_type"
        )
        
        if view_type == "View All Users (Summary)":
            st.markdown(f"#### Global Performance Summary for **All Users**")
            display_all_users_performance_summary(run_ids_in_scope, None, full_name_map)

        elif view_type == "View by Specific Manager":
            all_managers = all_users_df[all_users_df['role'] == 'Manager']['username'].tolist()
            if not all_managers:
                st.warning("No users with the 'Manager' role exist to filter by."); return
            
            manager_options_map = {"": ""}
            for manager in sorted(all_managers):
                manager_options_map[f"{full_name_map.get(manager, manager)} ({manager})"] = manager
            
            selected_manager_display = st.selectbox("Select a Manager to view their team", options=list(manager_options_map.keys()))
            
            if selected_manager_display:
                selected_manager_username = manager_options_map[selected_manager_display]
                manager_team_users = db_manager.get_managed_users(selected_manager_username)
                team_list = manager_team_users + [selected_manager_username]
                team_with_exceptions = get_filtered_users_with_exceptions(run_ids_in_scope, team_list)
                
                if not team_with_exceptions:
                    st.info(f"No users in {selected_manager_display}'s team had exceptions in the selected scope.")
                else:
                    drill_down_options_map = {f"Summary for {selected_manager_display}'s Team": "summary"}
                    for user in sorted(team_with_exceptions):
                        drill_down_options_map[f"{full_name_map.get(user, user)} ({user})"] = user

                    selected_drill_down_display = st.selectbox("Select Team View or Individual User", options=list(drill_down_options_map.keys()))
                    selected_drill_down_user = drill_down_options_map[selected_drill_down_display]
                    
                    if selected_drill_down_user == "summary":
                        st.markdown(f"#### Performance Summary for **{selected_manager_display}'s Team**")
                        display_all_users_performance_summary(run_ids_in_scope, team_list, full_name_map)
                    else:
                        st.markdown(f"#### Performance Details for **{selected_drill_down_display}**")
                        display_single_user_performance(selected_drill_down_user, run_ids_in_scope, full_name_map)
        
        # NEW: Logic to handle the direct user-wise filter
        elif view_type == "View by Specific User":
            st.markdown("#### Direct User Performance Lookup")
            # Get all users who have data in the selected scope to populate the dropdown
            all_users_with_data = get_filtered_users_with_exceptions(run_ids_in_scope, None)
            
            if not all_users_with_data:
                st.info("No users with exceptions found in the selected scope.")
            else:
                user_options_map = {"": ""} # Add a blank default option
                for user in sorted(all_users_with_data):
                    user_options_map[f"{full_name_map.get(user, user)} ({user})"] = user

                selected_user_display = st.selectbox(
                    "Select a User to View Their Performance",
                    options=list(user_options_map.keys()),
                    key="specific_user_select"
                )

                if selected_user_display:
                    selected_username = user_options_map[selected_user_display]
                    st.markdown(f"#### Performance Details for **{selected_user_display}**")
                    display_single_user_performance(selected_username, run_ids_in_scope, full_name_map)

def get_filtered_users_with_exceptions(run_ids, user_list=None):
    """Queries the database to get a list of users who have exceptions."""
    conn = db_manager._get_connection()
    try:
        placeholders = ','.join(['%s'] * len(run_ids))
        base_query = f"SELECT DISTINCT `user` FROM `user_performance` WHERE run_id IN ({placeholders}) AND `user` IS NOT NULL AND `user` != '' AND exception_records > 0"
        query_params = list(run_ids)
        if user_list:
            user_placeholders = ','.join(['%s'] * len(user_list))
            base_query += f" AND `user` IN ({user_placeholders})"
            query_params.extend(user_list)
        users_df = pd.read_sql_query(base_query, conn, params=tuple(query_params))
        return sorted(users_df['user'].unique())
    finally:
        if conn and conn.is_connected(): conn.close()

#
# --- FULLY COMPLETED HELPER FUNCTIONS FOR USER PERFORMANCE PAGE ---
#

def display_single_user_performance(user, run_ids, full_name_map):
    """Displays the detailed performance page for a single user, with all charts and forms."""
    conn = db_manager._get_connection()
    try:
        params = [user] + run_ids
        placeholders_sql = ','.join(['%s'] * len(run_ids))
        perf_query = f"SELECT up.*, vr.upload_time, vr.filename FROM `user_performance` up JOIN `validation_runs` vr ON up.run_id = vr.id WHERE up.`user` = %s AND up.run_id IN ({placeholders_sql}) ORDER BY vr.upload_time ASC"
        user_perf_df = pd.read_sql_query(perf_query, conn, params=params)
        exc_query = f"SELECT vr.upload_time, e.exception_reason FROM `exceptions` e JOIN `validation_runs` vr ON e.run_id = vr.id WHERE e.created_user = %s AND e.run_id IN ({placeholders_sql})"
        user_exc_df = pd.read_sql_query(exc_query, conn, params=params)
    except mysql.connector.Error as e:
        st.error(f"Error fetching performance details: {e}"); return
    finally:
        if conn and conn.is_connected(): conn.close()
    
    if user_perf_df.empty:
        st.info(f"No performance data found for user '{user}' in the selected scope."); return

    kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
    total_exc = user_perf_df['exception_records'].sum()
    total_recs_by_user = user_perf_df['total_records'].sum()
    avg_rate = (total_exc / total_recs_by_user * 100) if total_recs_by_user > 0 else 0
    runs_involved = user_perf_df['run_id'].nunique()
    display_metric("Total Exceptions", f"{int(total_exc):,}", container=kpi_col1)
    display_metric("User Exception Rate", f"{avg_rate:.2f}%", container=kpi_col2)
    display_metric("Total Records by User", f"{int(total_recs_by_user):,}", container=kpi_col3)
    display_metric("Runs Involved", f"{runs_involved}", container=kpi_col4)

    fig_mistakes, fig_trend, mistake_counts = None, None, pd.DataFrame()
    if not user_exc_df.empty:
        st.markdown("##### 🛠️ Common Mistake Analysis")
        mistake_counts = user_exc_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        mistake_counts.columns = ['Mistake Type', 'Count']
        fig_mistakes = px.pie(mistake_counts.head(10), names='Mistake Type', values='Count', 
                              title="Top Mistake Types Distribution",
                              color_discrete_sequence=px.colors.qualitative.Pastel)
        fig_mistakes.update_traces(textposition='inside', textinfo='percent+label')
        st.plotly_chart(fig_mistakes, use_container_width=True)

    if not user_perf_df.empty:
        st.markdown("##### 📉 User Exception Rate Trend")
        user_perf_df['upload_time'] = pd.to_datetime(user_perf_df['upload_time'], format='mixed')
        fig_trend = px.line(user_perf_df, x='upload_time', y='exception_rate', markers=True, 
                            title="User Exception Rate Trend", labels={'upload_time': 'Date', 'exception_rate': 'Exception Rate (%)'})
        fig_trend.update_traces(line=dict(color='#007bff', width=3))
        st.plotly_chart(fig_trend, use_container_width=True)

    if not mistake_counts.empty:
        st.markdown("#### 📚 Training Recommendations")
        st.info("Based on the top 3 most frequent errors for this user, here are some suggested areas for training and review.")
        validator = DataValidator(base_ref_path="reference_data")
        for _, row in mistake_counts.head(3).iterrows():
            error_type = row['Mistake Type']
            recommendation = validator.training_map.get(error_type, "No specific training suggestion available. Please review general data entry guidelines.")
            st.markdown(f"- **For '{error_type}':** {recommendation}")
    
    st.markdown("---")
    with st.expander("📧 Email Performance Report"):
        with st.form(f"email_single_user_form_{user.replace('.', '_')}"):
            to_input = st.text_input("To", value=user if '@' in user else "")
            cc_input = st.text_input("CC")
            subject = st.text_input("Subject", value=f"Performance Report for {user}")
            
            email_body_html = f"""
            <html style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;">
            <body style="background-color: #f4f4f4; margin: 0; padding: 20px;">
                <h2 style="color: #333333;">Performance Summary for {full_name_map.get(user, user)}</h2>
                <p>Total Exceptions: {int(total_exc):,}</p>
                <p>Average Exception Rate: {avg_rate:.2f}%</p>
                <h3 style="color: #333333;">Common Mistake Analysis</h3>
                <img src="cid:mistake_analysis_chart" width="560">
                <h3 style="color: #333333;">Exception Rate Trend</h3>
                <img src="cid:exception_rate_trend" width="560">
            </body></html>"""
            
            submitted = st.form_submit_button("Send Email")
            if submitted:
                to_recipients_list = [email.strip() for email in to_input.split(',') if email.strip()]
                cc_recipients_list = [email.strip() for email in cc_input.split(',') if email.strip()] if cc_input else []
                if not to_recipients_list:
                    st.error("Please provide at least one recipient.")
                elif fig_mistakes is None or fig_trend is None:
                    st.error("Cannot send email as one or more charts could not be generated.")
                else:
                    with st.spinner("Preparing and sending email..."):
                        mistakes_img_bytes = fig_mistakes.to_image(format="png", scale=2)
                        trend_img_bytes = fig_trend.to_image(format="png", scale=2)
                        images_to_embed = {"mistake_analysis_chart": mistakes_img_bytes, "exception_rate_trend": trend_img_bytes}
                        send_performance_email(to_recipients=to_recipients_list, subject=subject, html_body=email_body_html, cc_recipients=cc_recipients_list, images=images_to_embed)

def display_all_users_performance_summary(run_ids, user_list, full_name_map):
    """Displays the summary performance for a list of users, with all charts and forms."""
    conn = db_manager._get_connection()
    try:
        placeholders = ','.join(['%s'] * len(run_ids))
        perf_query = f"SELECT * FROM user_performance WHERE run_id IN ({placeholders})"
        all_perf_df = pd.read_sql_query(perf_query, conn, params=tuple(run_ids))
        exc_query = f"SELECT created_user, exception_reason FROM exceptions WHERE run_id IN ({placeholders})"
        all_exc_df = pd.read_sql_query(exc_query, conn, params=tuple(run_ids))
        notif_df = db_manager.get_notification_counts(run_ids, user_list)
    except mysql.connector.Error as e:
        st.error(f"Database error fetching summary data: {e}"); return
    finally:
        if conn and conn.is_connected(): conn.close()
    
    if user_list is not None:
        all_perf_df = all_perf_df[all_perf_df['user'].isin(user_list)]
        all_exc_df = all_exc_df[all_exc_df['created_user'].isin(user_list)]
    
    if all_perf_df.empty:
        st.info("No performance records to summarize for this scope."); return

    total_exceptions = all_perf_df['exception_records'].sum()
    total_records = all_perf_df['total_records'].sum()
    overall_avg_exception_rate = (total_exceptions / total_records * 100) if total_records > 0 else 0
    kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
    display_metric("Total Exceptions", f"{int(total_exceptions):,}", container=kpi_col1)
    display_metric("Overall Exception Rate", f"{overall_avg_exception_rate:.2f}%", container=kpi_col2)
    display_metric("Total Records Processed", f"{int(total_records):,}", container=kpi_col3)
    display_metric("Total Validation Runs", f"{len(run_ids)}", container=kpi_col4)

    st.markdown("##### 📋 Performance by User")
    summary_by_user = all_perf_df.groupby('user').agg(total_records=('total_records', 'sum'), exception_records=('exception_records', 'sum')).reset_index()
    summary_by_user['exception_rate'] = (summary_by_user['exception_records'] / summary_by_user['total_records'] * 100).fillna(0)
    summary_by_user['Full Name'] = summary_by_user['user'].map(full_name_map).fillna("N/A")
    st.dataframe(summary_by_user[['Full Name', 'user', 'total_records', 'exception_records', 'exception_rate']].sort_values('exception_rate', ascending=False), use_container_width=True, hide_index=True)

    fig_mistakes_summary = None
    if not all_exc_df.empty:
        st.markdown("##### 🛠️ Common Mistake Analysis")
        all_mistakes_df = all_exc_df['exception_reason'].str.split('; ', expand=True).stack().str.strip().value_counts().reset_index()
        all_mistakes_df.columns = ['Mistake Type', 'Count']
        fig_mistakes_summary = px.bar(all_mistakes_df.head(15), x='Mistake Type', y='Count', title="Top 15 Mistake Types Across All Users in Scope", template="plotly_white")
        st.plotly_chart(fig_mistakes_summary, use_container_width=True)

    if not notif_df.empty:
        st.markdown("##### 🔔 Notification Frequency Analysis")
        notif_df['Full Name'] = notif_df['username'].map(full_name_map).fillna(notif_df['username'])
        fig_notif = px.bar(notif_df, x='Full Name', y='count', color='notification_type',
                           title='Notification Counts by User and Type',
                           labels={'Full Name': 'User', 'count': 'Number of Notifications', 'notification_type': 'Notification Type'},
                           barmode='stack')
        st.plotly_chart(fig_notif, use_container_width=True)

    st.markdown("---")
    with st.expander("📧 Email Overall Summary Report"):
        with st.form("email_summary_form"):
            to_input = st.text_input("To (separate multiple emails with a comma)")
            cc_input = st.text_input("CC (separate multiple emails with a comma)")
            subject = st.text_input("Subject", value="Overall User Performance Summary")
            
            summary_table_html = summary_by_user.sort_values('exception_rate', ascending=False).to_html(index=False, border=0, classes="dataframe", float_format='{:.2f}'.format)
            email_body_html = f"""
            <html style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;">
            <body style="background-color: #f4f4f4; margin: 0; padding: 20px;">
                <h2 style="color: #333333;">Overall User Performance Summary</h2>
                {summary_table_html}
                <h3 style="color: #333333; margin-top: 30px;">Top 15 Common Mistake Types</h3>
                <img src="cid:mistake_summary_chart" width="560">
            </body></html>"""
            
            submitted = st.form_submit_button("Send Summary Email")
            if submitted:
                to_recipients_list = [email.strip() for email in to_input.split(',') if email.strip()]
                cc_recipients_list = [email.strip() for email in cc_input.split(',') if email.strip()]
                if not to_recipients_list:
                    st.error("Please provide at least one recipient in the 'To' field.")
                elif fig_mistakes_summary is None:
                    st.error("Cannot send email because the summary chart could not be generated.")
                else:
                    with st.spinner("Preparing and sending email..."):
                        mistakes_summary_bytes = fig_mistakes_summary.to_image(format="png", scale=2)
                        images_to_embed = {"mistake_summary_chart": mistakes_summary_bytes}
                        send_performance_email(to_recipients=to_recipients_list, subject=subject, html_body=email_body_html, cc_recipients=cc_recipients_list, images=images_to_embed)

# We create new helper functions to avoid repeating code    


def show_ledger_summary_page(start_date, end_date):
    st.markdown("### 🧾 Ledger & Sub-Ledger Exception Summary")
    account_names_df, subledger_names_df = load_account_name_mapping(), load_subledger_name_mapping()
    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display page: mapping files could not be loaded."); return
    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]
    if history_df.empty: st.info("No validation runs found for the selected period."); return
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="ledger_summary_run_selector")
    if not selected_run_display: return
    selected_run_id = run_options_with_all[selected_run_display]
    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        run_ids_to_load = history_df['id'].tolist() if selected_run_id == "all" else [selected_run_id]
        exceptions_df = pd.concat([df for df in [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load] if not df.empty], ignore_index=True)
        total_records_in_scope = history_df[history_df['id'].isin(run_ids_to_load)]['total_records'].sum()
    ledger_errors_df = exceptions_df[exceptions_df['Exception Reasons'].str.contains("Incorrect Ledger/Sub-Ledger Combination", na=False)].copy()
    if ledger_errors_df.empty: st.success(f"No 'Incorrect Ledger/Sub-Ledger Combination' exceptions found."); return

    # Merge names
    if 'Account2.Code' in ledger_errors_df.columns:
        ledger_errors_df['merge_key_account'] = ledger_errors_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(ledger_errors_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in merged_df.columns:
            merged_df['Account2.Name'] = merged_df['Account2.Name_map'].fillna(merged_df['Account2.Name'])
            merged_df.drop(columns=['Account2.Name_map'], inplace=True)
    else: merged_df = ledger_errors_df
    
    if 'Sub Ledger.Code' in merged_df.columns:
        merged_df['merge_key_subledger'] = merged_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(merged_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in merged_df.columns:
            merged_df['SubLedger.Name'] = merged_df['SubLedger.Name_map'].fillna(merged_df['SubLedger.Name'])
            merged_df.drop(columns=['SubLedger.Name_map'], inplace=True)

    merged_df.drop(columns=[col for col in ['merge_key_account', 'merge_key_subledger'] if col in merged_df.columns], inplace=True)
    merged_df['Account2.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['SubLedger.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['Department.Name'].fillna('N/A - Department Missing', inplace=True)

    st.markdown("---")
    st.markdown("#### 📈 Metrics for Ledger/Sub-Ledger Exceptions")
    total_ledger_exceptions = len(merged_df)
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    display_metric("Total Transactions in Scope", f"{int(total_records_in_scope):,}", container=metric_col1)
    display_metric("Ledger/Sub-Ledger Exceptions", f"{total_ledger_exceptions:,}", container=metric_col2)
    exception_percentage = (total_ledger_exceptions / total_records_in_scope * 100) if total_records_in_scope > 0 else 0
    display_metric("% of Total", f"{exception_percentage:.2f}%", "of transactions in this scope have this error", container=metric_col3)
    
    st.markdown("#### 📊 Exception Count by Ledger and Department")
    summary_df = merged_df.groupby(['Account2.Name', 'SubLedger.Name', 'Department.Name']).size().reset_index(name='Count').sort_values(by='Count', ascending=False)
    summary_df.columns = ['Ledger Name', 'Sub-Ledger Name', 'Department', 'Count']
    chart_type = st.radio("Select Chart Type", ["Bar Chart", "Pie Chart"], horizontal=True, key="ledger_summary_chart_type")
    
    if chart_type == "Bar Chart":
        fig = px.bar(summary_df, x='Ledger Name', y='Count', color='Department', barmode='group', hover_data=['Sub-Ledger Name', 'Department'], title='Ledger/Sub-Ledger Exception Counts by Department')
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)
    if chart_type == "Pie Chart":
        pie_summary = summary_df.groupby('Ledger Name')['Count'].sum().reset_index().sort_values(by='Count', ascending=False)
        fig = px.pie(pie_summary.head(10), names='Ledger Name', values='Count', title='Top 10 Distribution of Exceptions by Ledger Name')
        st.plotly_chart(fig, use_container_width=True)
        
    st.markdown("#### 📋 Detailed Summary Table")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)


def show_user_ledger_exceptions_page(start_date, end_date):
    st.markdown("### 👤🧾 User-wise Ledger Exception Details")

    # --- 1. Get User Context and load necessary data ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    account_names_df = load_account_name_mapping()
    subledger_names_df = load_subledger_name_mapping()
    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display page: account or sub-ledger name mapping files could not be loaded.")
        return

    # This correctly fetches only the runs the user is allowed to see
    history_df = db_manager.get_validation_history(user_role, username, managed_users)
    if start_date and end_date:
        history_df = history_df[(history_df['upload_time'].dt.date >= start_date) & (history_df['upload_time'].dt.date <= end_date)]

    if history_df.empty:
        st.info("No validation runs found for the selected period.")
        return

    # --- 2. Select Run Scope ---
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}
    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="user_ledger_run_selector")

    if not selected_run_display:
        return
    selected_run_id = run_options_with_all[selected_run_display]

    # --- 3. Load and Filter Master Data ---
    with st.spinner(f"Loading data for '{selected_run_display}'..."):
        run_ids_to_load = history_df['id'].tolist() if selected_run_id == "all" else [selected_run_id]
        all_exceptions = [db_manager.get_exceptions_by_run(run_id) for run_id in run_ids_to_load]
        exceptions_df = pd.concat([df for df in all_exceptions if not df.empty], ignore_index=True)

    ledger_exception_reason = "Incorrect Ledger/Sub-Ledger Combination"
    ledger_errors_df_master = exceptions_df[exceptions_df['Exception Reasons'].str.contains(ledger_exception_reason, na=False)].copy()

    if ledger_errors_df_master.empty:
        st.success(f"No 'Incorrect Ledger/Sub-Ledger Combination' exceptions found for the selected scope.")
        return

    # --- 4. NEW AND CRITICAL: HIERARCHICAL DATA SCOPING BLOCK ---
    # This block filters the data *before* it's used to create the user selection filter.
    ledger_errors_df = pd.DataFrame()
    if user_role == 'User':
        ledger_errors_df = ledger_errors_df_master[ledger_errors_df_master['Created user'].str.lower() == username.lower()].copy()
    elif user_role == 'Manager':
        team_to_filter = [u.lower() for u in managed_users] + [username.lower()]
        ledger_errors_df = ledger_errors_df_master[ledger_errors_df_master['Created user'].str.lower().isin(team_to_filter)].copy()
    else:  # Management and Super User see all ledger exceptions within the selected runs
        ledger_errors_df = ledger_errors_df_master.copy()

    if ledger_errors_df.empty:
        st.warning("No ledger/sub-ledger exceptions found for your specific access scope (user or team).")
        return
    # --- END OF CORRECTION ---

    st.markdown("---")
    st.markdown("#### Filters")

    # The user_list for the dropdown is now created from the correctly scoped 'ledger_errors_df'
    user_list = sorted(ledger_errors_df['Created user'].dropna().unique().tolist())
    if not user_list:
        st.warning("No users found with this type of exception in your scope.")
        return

    # The selectbox will now only show users a manager is allowed to see.
    selected_user = st.selectbox("Select User", options=user_list, key="user_ledger_user_selector")

    if not selected_user:
        st.info("Please select a user to see their specific exceptions.")
        return

    user_specific_errors_df = ledger_errors_df[ledger_errors_df['Created user'] == selected_user].copy()

    # The rest of the function for merging and displaying charts now operates on the correctly filtered data.
    # No changes are needed from this point onwards.
    
    if user_specific_errors_df.empty:
        st.warning(f"No ledger/sub-ledger exceptions found for user '{selected_user}' in this scope.")
        return

    if 'Account2.Code' in user_specific_errors_df.columns:
        user_specific_errors_df['merge_key_account'] = user_specific_errors_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(user_specific_errors_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in merged_df.columns:
            merged_df['Account2.Name'] = merged_df['Account2.Name_map'].fillna(merged_df['Account2.Name'])
            merged_df.drop(columns=['Account2.Name_map'], inplace=True)
    else: merged_df = user_specific_errors_df

    if 'Sub Ledger.Code' in merged_df.columns:
        merged_df['merge_key_subledger'] = merged_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        merged_df = pd.merge(merged_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in merged_df.columns:
            merged_df['SubLedger.Name'] = merged_df['SubLedger.Name_map'].fillna(merged_df['SubLedger.Name'])
            merged_df.drop(columns=['SubLedger.Name_map'], inplace=True)

    merged_df.drop(columns=[col for col in ['merge_key_account', 'merge_key_subledger'] if col in merged_df.columns], inplace=True)
    merged_df['Account2.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['SubLedger.Name'].fillna('N/A - Name Not Mapped', inplace=True)
    merged_df['Department.Name'].fillna('N/A - Department Missing', inplace=True)

    st.markdown("---")
    st.markdown(f"#### 📊 Visual Summary for **{selected_user}**")

    plot_df = merged_df.groupby(['Account2.Name', 'SubLedger.Name']).size().reset_index(name='Count')
    plot_df['Combination'] = plot_df['Account2.Name'] + " / " + plot_df['SubLedger.Name']
    
    fig = px.bar(plot_df.sort_values('Count', ascending=False).head(15), x='Count', y='Combination', orientation='h',
                 title=f"Top 15 Ledger/Sub-Ledger Exception Counts for {selected_user}",
                 labels={'Combination': 'Ledger / Sub-Ledger Combination', 'Count': 'Number of Exceptions'},
                 text='Count', color='Count', color_continuous_scale=px.colors.sequential.Plasma)
    fig.update_layout(yaxis={'categoryorder':'total ascending'}, title_x=0.5, title_font=PLOTLY_TITLE_FONT, font=PLOTLY_FONT,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(l=40, r=20, t=60, b=40))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown(f"#### 📋 Detailed Exceptions by **{selected_user}**")
    display_interactive_exceptions(merged_df, key_prefix="user_ledger_view")


def show_location_expenses_page(start_date, end_date):
    st.markdown("### 📍 Location Expenses")
    st.markdown("Analyze expenses aggregated by Ledger and Sub-Ledger across all locations. Select a single validation run or 'All Runs' for an aggregate view over the selected time period.")

    # 1. Select a Run or "All Runs"
    history_df = db_manager.get_validation_history()
    if start_date and end_date:
        history_df = history_df[
            (history_df['upload_time'].dt.date >= start_date) & 
            (history_df['upload_time'].dt.date <= end_date)
        ]
        
    if history_df.empty:
        st.info("No validation runs found for the selected period. Upload a file to see analytics.")
        return

    # Add "All Runs" option to the beginning of the list
    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    run_options_with_all = {"All Runs (Summary)": "all", **run_options}

    selected_run_display = st.selectbox("Select Validation Run", options=list(run_options_with_all.keys()), key="location_exp_run_selector")

    if not selected_run_display:
        return
    
    selected_run_id = run_options_with_all[selected_run_display]

    # 2. Load Data and Mappings
    with st.spinner(f"Loading expense data for '{selected_run_display}'..."):
        all_dfs = []
        if selected_run_id == "all":
            run_ids_to_load = history_df['id'].tolist()
        else:
            run_ids_to_load = [selected_run_id]

        for run_id in run_ids_to_load:
            df = db_manager.get_exceptions_by_run(run_id)
            if not df.empty:
                all_dfs.append(df)
        
        if not all_dfs:
            st.success(f"No expense data (from exception records) found for the selected scope.")
            return
            
        expense_data_df = pd.concat(all_dfs, ignore_index=True)
        account_names_df = load_account_name_mapping()
        subledger_names_df = load_subledger_name_mapping()

    if account_names_df is None or subledger_names_df is None:
        st.error("Cannot display page because account or sub-ledger name mapping files are missing from the 'reference_data' directory.")
        return

    # 3. Merge to get Names and prepare data
    if 'Net amount' in expense_data_df.columns:
        expense_data_df['Net amount'] = pd.to_numeric(expense_data_df['Net amount'], errors='coerce').fillna(0)
    else:
        st.error("The 'Net amount' column is missing from the data.")
        return
        
    # Merge Account Names
    if 'Account2.Code' in expense_data_df.columns and not expense_data_df.empty:
        expense_data_df['merge_key_account'] = expense_data_df['Account2.Code'].astype(str).str.strip().str.lower()
        account_names_df['merge_key_account'] = account_names_df['Account2.Code'].astype(str).str.strip().str.lower()
        expense_data_df = pd.merge(expense_data_df, account_names_df, on='merge_key_account', how='left', suffixes=('', '_map'))
        if 'Account2.Name_map' in expense_data_df.columns:
            expense_data_df['Account2.Name'] = expense_data_df['Account2.Name'].fillna(expense_data_df['Account2.Name_map'])
            expense_data_df.drop(columns=[col for col in ['Account2.Name_map', 'Account2.Code_map', 'merge_key_account'] if col in expense_data_df.columns], inplace=True)

    # Merge Sub-Ledger Names
    if 'Sub Ledger.Code' in expense_data_df.columns and not expense_data_df.empty:
        expense_data_df['merge_key_subledger'] = expense_data_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        subledger_names_df['merge_key_subledger'] = subledger_names_df['Sub Ledger.Code'].astype(str).str.strip().str.lower()
        expense_data_df = pd.merge(expense_data_df, subledger_names_df, on='merge_key_subledger', how='left', suffixes=('', '_map'))
        if 'SubLedger.Name_map' in expense_data_df.columns:
            expense_data_df['SubLedger.Name'] = expense_data_df['SubLedger.Name'].fillna(expense_data_df['SubLedger.Name_map'])
            expense_data_df.drop(columns=[col for col in ['SubLedger.Name_map', 'Sub Ledger.Code_map', 'merge_key_subledger'] if col in expense_data_df.columns], inplace=True)

    # Fill NA for key fields to ensure they appear in filters and pivots
    for col, fill_val in {
        'Account2.Name': 'N/A - Name Not Mapped',
        'SubLedger.Name': 'N/A - Name Not Mapped',
        'Department.Name': 'N/A - Department Missing',
        'Sub Department.Name': 'N/A',
        'Location.Name': 'N/A - Location Missing'
    }.items():
        if col not in expense_data_df.columns: expense_data_df[col] = fill_val
        else: expense_data_df[col].fillna(fill_val, inplace=True)

    # 4. Filters
    st.markdown("---")
    st.markdown("#### 🔍 Filters")
    
    filter_col1, filter_col2 = st.columns(2)
    
    with filter_col1:
        departments = ["All"] + sorted(expense_data_df['Department.Name'].unique().tolist())
        selected_dept = st.selectbox("Filter by Department", options=departments, key="loc_exp_dept_filter")

    # Filter dataframe based on department selection to populate sub-department options
    df_for_sub_dept_options = expense_data_df[expense_data_df['Department.Name'] == selected_dept] if selected_dept != "All" else expense_data_df

    with filter_col2:
        sub_dept_options = ["All"] + sorted([sd for sd in df_for_sub_dept_options['Sub Department.Name'].unique() if sd != 'N/A'])
        if 'N/A' in df_for_sub_dept_options['Sub Department.Name'].unique():
             sub_dept_options.append('N/A')
        selected_sub_dept = st.selectbox("Filter by Sub-Department", options=sub_dept_options, key="loc_exp_sub_dept_filter")

    # Apply final sub-department filter
    final_filtered_df = df_for_sub_dept_options[df_for_sub_dept_options['Sub Department.Name'] == selected_sub_dept] if selected_sub_dept != "All" else df_for_sub_dept_options
    
    if final_filtered_df.empty:
        st.warning("No data matches the current filter criteria.")
        return

    # 5. Create and Display Pivot Table
    st.markdown("---")
    st.markdown("#### 💰 Ledger Expense Summary by Location")
    st.markdown("This table shows the total expense amount for each Ledger/Sub-Ledger combination, broken down by Location.")
    
    try:
        # Dynamically set the index for the pivot table
        pivot_index = ['Account2.Name', 'SubLedger.Name']
        if selected_dept == "All":
            pivot_index.insert(0, 'Department.Name')
            if selected_sub_dept == "All":
                pivot_index.insert(1, 'Sub Department.Name')
        
        pivot_table = pd.pivot_table(
            final_filtered_df,
            values='Net amount',
            index=pivot_index,
            columns='Location.Name',
            aggfunc=np.sum,
            fill_value=0,
            margins=True, # Add row and column totals (Grand Total)
            margins_name="Grand Total"
        )
        
        st.dataframe(pivot_table.style.format("{:,.2f}"), use_container_width=True)

        # 6. Download Button
        @st.cache_data
        def convert_df_to_excel(df_to_convert):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_to_convert.to_excel(writer, sheet_name='Ledger_Expense_by_Location')
            return output.getvalue()

        excel_data = convert_df_to_excel(pivot_table)
        
        st.download_button(
            label="📥 Download Summary as Excel",
            data=excel_data,
            file_name=f"Ledger_Expense_by_Location_{selected_run_display.replace(':', '')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_loc_exp"
        )

    except Exception as e:
        st.error(f"Could not generate the pivot table. Error: {e}")
        logging.error(f"Error creating pivot table on location expense page: {e}", exc_info=True)

def show_access_control_page():
    st.markdown("### 🔑 Access Control Panel")
    st.caption("This page is only visible to the Super User.")

    # This hardcoded list of all page names is used for the permission settings.
    # IMPORTANT: This list must be maintained if you add/remove/rename pages in the main() function.
    all_pages = [
        "📊 Dashboard Analytics", "📈 Trends & History", "👤📊 User Performance",
        "🗂️ Report Archive", "📝 Correction Status", "📋 Exception Details", 
        "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", 
        "👤📍 User & Location Analysis", "🗑️ Data Management", "⚙️ Settings", "🛠️ User Management",
        "🕵️ Suspicious Tag Control", "📋 Suspicious Transactions","📝 Correction Entries", "📈 Correction Analytics","✉️ Send Notification"
    ]
    
    # Fetch all current permissions from the database once for efficiency
    permissions = db_manager.get_all_permissions()
    all_users_df = db_manager.get_all_users()

    if not permissions or not permissions['roles']:
        st.error("Could not load permissions from the database. Please check the database connection and tables.")
        return

    tab1, tab2 = st.tabs(["🧰 Manage Role Permissions", "👤 Manage User Overrides"])

    with tab1:
        st.subheader("Set General Permissions for Each Role")
        
        # We don't allow editing the Super User role's permissions to prevent lockouts
        roles_to_manage = ["User", "Manager", "Management"]
        
        for role in roles_to_manage:
            with st.expander(f"Permissions for **{role}** Role"):
                current_role_perms = permissions['roles'].get(role, {})
                
                # --- Upload Permission for Role ---
                can_upload_role = current_role_perms.get('can_upload', True)
                new_can_upload_role = st.toggle("Can Upload Files?", value=bool(can_upload_role), key=f"upload_role_{role}")
                
                # --- Disabled Pages for Role ---
                disabled_pages_role_str = current_role_perms.get('disabled_pages', '')
                disabled_pages_role_list = disabled_pages_role_str.split(',') if disabled_pages_role_str else []
                
                new_disabled_pages_role = st.multiselect(
                    "Select Dashboards to DISABLE for this Role:",
                    options=all_pages,
                    default=disabled_pages_role_list,
                    key=f"pages_role_{role}"
                )

                if st.button(f"Update '{role}' Permissions", key=f"update_role_{role}"):
                    success = db_manager.update_role_permissions(role, new_can_upload_role, new_disabled_pages_role)
                    if success:
                        st.success(f"Permissions for the '{role}' role have been updated.")
                        st.rerun()
                    else:
                        st.error("Failed to update role permissions.")

    with tab2:
        st.subheader("Set Specific Permission Overrides for an Individual User")
        st.caption("Settings here will override the user's general role permissions.")
        
        # Exclude the super user from being edited
        users_to_manage = all_users_df[all_users_df['role'] != 'Super User']['username'].tolist()
        
        if not users_to_manage:
            st.info("No users available to manage.")
        else:
            selected_user = st.selectbox("Select a user to manage their specific permissions", options=[""] + sorted(users_to_manage))
            
            if selected_user:
                current_user_perms = permissions['users'].get(selected_user, {})
                
                # --- Upload Permission Override for User ---
                can_upload_user_override = current_user_perms.get('can_upload') # This will be True, False, or None
                
                upload_override_options = ["Inherit from Role", "Always Allow", "Always Deny"]
                current_upload_index = 0 # Default to "Inherit"
                if can_upload_user_override is True:
                    current_upload_index = 1
                elif can_upload_user_override is False:
                    current_upload_index = 2
                
                new_upload_override_selection = st.radio(
                    f"Upload Permission for **{selected_user}**:",
                    options=upload_override_options,
                    index=current_upload_index,
                    key=f"upload_user_{selected_user}",
                    horizontal=True
                )

                # Convert radio selection back to database value (True, False, None)
                new_can_upload_user = None 
                if new_upload_override_selection == "Always Allow":
                    new_can_upload_user = True
                elif new_upload_override_selection == "Always Deny":
                    new_can_upload_user = False

                # --- Disabled Pages Override for User ---
                disabled_pages_user_str = current_user_perms.get('disabled_pages')
                disabled_pages_user_list = disabled_pages_user_str.split(',') if disabled_pages_user_str else []
                
                new_disabled_pages_user = st.multiselect(
                    f"Select Dashboards to specifically DISABLE for **{selected_user}**:",
                    options=all_pages,
                    default=disabled_pages_user_list,
                    key=f"pages_user_{selected_user}"
                )
                st.caption("Note: This list completely overrides the role's disabled pages. An empty list here means the user will inherit their role's page visibility settings.")

                if st.button(f"Update '{selected_user}' Overrides", key=f"update_user_{selected_user}"):
                    success = db_manager.update_user_permissions(selected_user, new_can_upload_user, new_disabled_pages_user)
                    if success:
                        st.success(f"Overrides for '{selected_user}' have been updated.")
                        st.rerun()
                    else:
                        st.error("Failed to update user overrides.")


def show_report_archive_page():
    st.markdown("### 🗂️ Report Archive")
    st.info("Select one or more validation runs to generate a consolidated report. You can then download it or email it directly from the dashboard.")

    # --- 1. Get User Context and Accessible Runs ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    history_df = db_manager.get_validation_history(
        user_role=user_role, username=username, managed_users=managed_users
    )

    if history_df.empty:
        st.warning("No validation reports found for your accessible scope.")
        return

    # --- 2. Multi-Select Widget for Runs ---
    run_options_dict = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    
    if 'archive_multiselect' not in st.session_state:
        st.session_state.archive_multiselect = []

    col1, col2 = st.columns([3, 1])
    with col1:
        selected_run_displays = st.multiselect(
            "STEP 1: Select Report Scope (you can select multiple)",
            options=list(run_options_dict.keys()),
            default=st.session_state.archive_multiselect,
            key="archive_multiselect_widget",
            placeholder="Choose one or more runs..."
        )
        st.session_state.archive_multiselect = selected_run_displays
    
    with col2:
        st.write("") 
        st.write("")
        if st.button("Select All Runs", use_container_width=True):
            st.session_state.archive_multiselect = list(run_options_dict.keys())
            st.rerun()

    if not selected_run_displays:
        st.info("Please select at least one run to generate a report.")
        return

    # --- 3. Fetch Data and Apply Top-Level Security Filter ---
    with st.spinner("Loading and filtering report data based on your access level..."):
        selected_ids = [run_options_dict[display] for display in selected_run_displays]
        
        all_exceptions_list = [db_manager.get_exceptions_by_run(run_id) for run_id in selected_ids]

        non_empty_dfs = [df for df in all_exceptions_list if not df.empty]
        
        if not non_empty_dfs:
            st.success("✅ All exceptions in the selected run(s) have been corrected. There is nothing to report.")
            return

        unfiltered_df = pd.concat([df for df in all_exceptions_list if not df.empty], ignore_index=True)

        if unfiltered_df.empty:
            st.warning("The selected run(s) contain no exception records to report.")
            return
            
        scoped_df = pd.DataFrame()
        if user_role == 'User':
            scoped_df = unfiltered_df[unfiltered_df['Created user'].str.lower() == username.lower()]
        elif user_role == 'Manager':
            team_to_filter = [u.lower() for u in managed_users] + [username.lower()]
            scoped_df = unfiltered_df[unfiltered_df['Created user'].str.lower().isin(team_to_filter)]
        else:
            scoped_df = unfiltered_df.copy()

    if scoped_df.empty:
        st.error("No transactions found within your accessible scope for the selected run(s).")
        return

    # --- 4. User Filter (based on now-scoped data) ---
    st.markdown("---")
    st.markdown("#### STEP 2: Choose How to Filter the Report by User")

    users_in_this_report = sorted(scoped_df['Created user'].dropna().unique())
    user_filter_selection = ""
    filtered_df = pd.DataFrame()

    if user_role == 'User':
        st.success(f"✔️ Your report will be automatically filtered for your user: **{username}**")
        filtered_df = scoped_df.copy()
        user_filter_selection = username
    elif user_role == 'Manager':
        options = ["All My Users"] + users_in_this_report
        user_filter_selection = st.selectbox("Download report for:", options)
        if user_filter_selection == "All My Users":
            filtered_df = scoped_df.copy()
        else:
            filtered_df = scoped_df[scoped_df['Created user'] == user_filter_selection]
    else: # Management and Super User
        filter_level = st.radio(
            "Filter report by:",
            ("All Users", "By Manager", "By Individual User"),
            horizontal=True,
            key="mgmt_filter_level"
        )

        if filter_level == "All Users":
            user_filter_selection = "All_Users_(Unfiltered)"
            filtered_df = scoped_df.copy()

        elif filter_level == "By Manager":
            all_managers = db_manager.get_users_by_role('Manager')
            if not all_managers:
                st.warning("No users with the 'Manager' role exist in the system.")
            else:
                selected_manager = st.selectbox("Select a Manager to view their team's report", [""] + all_managers)
                if selected_manager:
                    user_filter_selection = f"Team_of_{selected_manager}"
                    manager_team_list = db_manager.get_managed_users(selected_manager)
                    team_to_filter = [selected_manager.lower()] + [u.lower() for u in manager_team_list]
                    filtered_df = scoped_df[scoped_df['Created user'].str.lower().isin(team_to_filter)]

        elif filter_level == "By Individual User":
            selected_user = st.selectbox("Select an Individual User", [""] + users_in_this_report)
            if selected_user:
                user_filter_selection = selected_user
                filtered_df = scoped_df[scoped_df['Created user'] == selected_user]

    if filtered_df.empty:
        st.error("No transactions found for the selected filter combination. Please make a selection.")
        return
        
    st.markdown("---")
    st.markdown("#### STEP 3: Generate Report")
    st.markdown(f"Your final report contains **{len(filtered_df)}** exception records based on your filters.")

    # --- 5. Generate and Download/Email Logic ---
    special_cols = ['id', 'run_id', 'Exception Reasons', 'Severity', 'original_row_data']
    original_cols = [col for col in filtered_df.columns if col not in special_cols]
    final_column_order = original_cols + ['Exception Reasons', 'Severity']
    
    final_df_for_excel = filtered_df[[col for col in final_column_order if col in filtered_df.columns]]

    final_report_data = create_excel_report(final_df_for_excel, {})
    
    if final_report_data:
        report_filename = f"Filtered_Report_{user_filter_selection.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="📥 Download Report as Excel",
            data=final_report_data,
            file_name=report_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("📧 Email This Report as an Attachment"):
            with st.form("email_report_form"):
                email_to = st.text_input("To (separate multiple emails with a comma)")
                email_cc = st.text_input("CC (optional)")
                email_subject = st.text_input("Subject", value=f"Data Validation Report: {user_filter_selection}")
                email_body = st.text_area("Email Body", value=f"Please find the attached data validation report for {user_filter_selection}, generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}.")
                
                submitted = st.form_submit_button("Send Email")
                if submitted:
                    to_recipients = [email.strip() for email in email_to.split(',') if email.strip()]
                    cc_recipients = [email.strip() for email in email_cc.split(',') if email.strip()] if email_cc else []
                    
                    if not to_recipients:
                        st.error("Please enter at least one recipient in the 'To' field.")
                    else:
                        with st.spinner("Attaching report and sending email..."):
                            send_report_email_with_attachment(
                                to_recipients=to_recipients,
                                cc_recipients=cc_recipients,
                                subject=email_subject,
                                html_body=f"<p>{email_body}</p>",
                                attachment_data=final_report_data,
                                attachment_filename=report_filename
                            )
    else:
        st.error("An error occurred while generating the Excel report data.")

def show_correction_entries_page():
    st.markdown("### 📝 Correction Entries")
    st.caption("Review and update the status of individual exception records. Use the filters to find specific entries.")

    # --- 1. Get User Context and Permissions ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    manager_can_accept = False
    if user_role == 'Manager':
        user_profile = db_manager.get_user_profile(username)
        if user_profile.get('can_accept_corrections') is not False:
            manager_can_accept = True
            
    show_accept_button = user_role in ['Super User', 'Management'] or manager_can_accept
    
    # --- 2. Notifications ---
    correction_notifications = [
        n for n in db_manager.get_notifications_for_user(username) 
        if n['notification_type'] in ['Correction Required', 'Correction Reminder']
    ]
    if correction_notifications:
        for notif in correction_notifications:
            st.warning(f"**Message from Admin ({notif['created_at'].strftime('%d-%b-%Y')}):** {notif['message']}")
        if st.button("Dismiss Messages", key="dismiss_correction_msgs"):
            notif_ids = [n['id'] for n in correction_notifications]
            db_manager.mark_notifications_as_read(notif_ids)
            st.rerun()

    # --- 3. Filters and Data Fetching ---
    st.markdown("---")
    st.markdown("#### Filters")

    # This logic correctly fetches entries that are not yet Accepted OR marked as 'Yes'
    all_accessible_entries = db_manager.get_correction_entries(
        user_role=user_role,
        username=username,
        managed_users=managed_users
    )

    if all_accessible_entries.empty:
        st.success("✅ No pending correction entries found for your current view.")
        return

    # Filter UI and logic
    filter_col1, filter_col2, filter_col3 = st.columns(3)
    with filter_col1:
        user_options = ["All"] + sorted(all_accessible_entries['created_user'].unique().tolist())
        filter_user_selection = st.selectbox("Filter by User", options=user_options)
    with filter_col2:
        run_options = ["All"] + sorted(all_accessible_entries['run_id'].unique().tolist())
        filter_run_selection = st.selectbox("Filter by Run ID", options=run_options)
    with filter_col3:
        narration_filter = st.text_input("Filter by Narration")
    
    entries_df = all_accessible_entries.copy()
    if filter_user_selection != "All":
        entries_df = entries_df[entries_df['created_user'] == filter_user_selection]
    if filter_run_selection != "All":
        entries_df = entries_df[entries_df['run_id'] == filter_run_selection]
    if narration_filter:
        try:
            mask = entries_df['narration'].str.contains(narration_filter, case=False, na=False)
            entries_df = entries_df[mask]
        except Exception:
            pass

    if entries_df.empty:
        st.warning("No entries match your current filter criteria.")
        return

    st.info(f"Displaying {len(entries_df)} correction entries based on your filters.")

    # --- 4. MODIFIED: New Bulk Actions Section ---
    st.markdown("---")
    st.markdown("#### Bulk Actions for Visible Entries")
    st.caption("Apply an action to all entries currently visible on the page.")
    
    # -- Bulk Status Updates (Available to all roles) --
    st.markdown("##### Update Status")
    status_cols = st.columns(3)
    if status_cols[0].button("Set All to 'Yes'", use_container_width=True):
        with st.spinner("Updating all entries to 'Yes'..."):
            db_manager.batch_update_exception_status(entries_df['id'].tolist(), 'Yes', action_by=username, user_role=user_role)
        st.rerun()

    if status_cols[1].button("Set All to 'No'", use_container_width=True):
        with st.spinner("Updating all entries to 'No'..."):
            db_manager.batch_update_exception_status(entries_df['id'].tolist(), 'No', action_by=username, user_role=user_role)
        st.rerun()
        
    if status_cols[2].button("Set All to 'Pending'", use_container_width=True):
        with st.spinner("Updating all entries to 'Pending'..."):
            db_manager.batch_update_exception_status(entries_df['id'].tolist(), 'Pending', action_by=username, user_role=user_role)
        st.rerun()

    # -- Bulk Acceptance (Privileged roles only) --
    if show_accept_button:
        st.markdown("##### Final Acceptance (Privileged)")
        if st.button("Accept All Visible Entries", type="primary", use_container_width=True):
            with st.spinner("Accepting all visible entries..."):
                db_manager.batch_accept_correction_entries(entries_df['id'].tolist())
            st.success("All visible entries have been accepted and removed from all dashboards.")
            st.rerun()

    st.markdown("---")

    # --- 5. Display Individual Entries ---
    for _, row in entries_df.iterrows():
        exception_id = row['id']
        current_status = row['correction_status']
        
        with st.expander(f"**ID: {exception_id}** | User: **{row.get('created_user')}** | Dept: **{row.get('department')}** | Status: **{current_status}**"):
            
            st.markdown("##### Details")
            metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
            metric_col1.metric("Location", row.get('location', 'N/A'))
            metric_col2.metric("Crop", row.get('crop', 'N/A'))
            metric_col3.metric("Net Amount", f"{row.get('net_amount', 0):,.2f}")
            metric_col4.metric("Severity", row.get('severity', 0))
            
            st.markdown(f"**Exception Reason(s):** {row.get('exception_reason', 'N/A')}")
            st.markdown(f"**Narration:**")
            
            narration_text = row.get('narration')
            if not narration_text: 
                try:
                    original_data = json.loads(row['original_row_data'])
                    narration_key = next((k for k in original_data if k.lower() == 'narration'), None)
                    if narration_key:
                        narration_text = original_data.get(narration_key)
                except (json.JSONDecodeError, TypeError):
                    narration_text = None
            
            st.info(f"{narration_text or 'No narration provided.'}")

            st.markdown("---")
            st.markdown("##### Actions")
            
            action_col1, action_col2 = st.columns([2, 1])
            with action_col1:
                st.write("**Update Status:**")
                status_options = ['Pending', 'Yes', 'No']
                new_status = st.radio(
                    "Correction Status",
                    options=status_options,
                    index=status_options.index(current_status),
                    key=f"status_radio_{exception_id}",
                    horizontal=True,
                    label_visibility="collapsed"
                )
                if new_status != current_status:
                    db_manager.update_exception_status(exception_id, new_status, action_by=username, user_role=user_role)
                    st.rerun()
            with action_col2:
                if show_accept_button:
                    st.write("**Final Acceptance:**")
                    if st.button("Accept", key=f"accept_btn_{exception_id}", type="primary", use_container_width=True):
                        with st.spinner(f"Accepting entry {exception_id}..."):
                            db_manager.accept_correction_entry(exception_id)
                        st.success(f"Entry {exception_id} accepted and closed.")
                        st.rerun()
            
            with st.expander("View Original Row Data"):
                try:
                    original_data = json.loads(row['original_row_data'])
                    st.dataframe(pd.DataFrame([original_data]), use_container_width=True)
                except:
                    st.warning("Could not display original row data.")

def show_correction_analytics_page():
    st.markdown("### 📈 Correction Analytics")

    # This page is exclusively for the Super User
    if st.session_state.get("role") != "Super User":
        st.error("🔒 You do not have permission to view this page.")
        return

    st.caption("Analyze the breakdown of correction statuses across the system or for a specific user.")

    # --- 1. User Selection Filter ---
    all_users_df = db_manager.get_all_users()
    user_list = sorted(all_users_df['username'].tolist()) if not all_users_df.empty else []
    
    view_options = ["Overall Summary"] + user_list
    selected_view = st.selectbox("Select a view", options=view_options)

    # --- 2. Data Fetching ---
    target_user = None if selected_view == "Overall Summary" else selected_view
    
    with st.spinner(f"Loading analytics data for {selected_view}..."):
        status_data = db_manager.get_correction_analytics_data(target_user=target_user)

    # --- 3. Chart Rendering ---
    labels = list(status_data.keys())
    values = list(status_data.values())

    if sum(values) == 0:
        st.info(f"No correction status data found for {selected_view}.")
        return

    title = f"Correction Status for {selected_view}" if target_user else "Overall Correction Status Distribution"

    fig = px.pie(
        names=labels, 
        values=values,
        title=title,
        color=labels,
        color_discrete_map={
            'Yes': '#28a745', 
            'No': '#dc3545', 
            'Pending': '#ffc107'
        },
        hole=0.3
    )
    fig.update_traces(textinfo='percent+label', pull=[0.05, 0.05, 0.05])
    fig.update_layout(title_x=0.5, font=PLOTLY_FONT)
    
    st.plotly_chart(fig, use_container_width=True)

    # --- 4. Display Raw Data ---
    st.markdown("#### Data Summary")
    summary_df = pd.DataFrame([status_data])
    st.dataframe(summary_df, use_container_width=True)

def show_data_management_page():
    st.markdown("### 🗑️ Data Management")
    st.warning("🚨 **Caution:** Actions on this page are permanent and cannot be undone.")

    st.markdown("---")
    st.markdown("#### ❌ Delete a Specific Validation Run")
    st.markdown("Select a validation run to permanently delete it and all its associated data (exceptions, summaries, etc.).")

    history_df = db_manager.get_validation_history()

    if history_df.empty:
        st.info("No validation runs available to delete.")
        return

    run_options = {f"Run {row['id']}: {row['filename']} ({pd.to_datetime(row['upload_time']).strftime('%Y-%m-%d %H:%M')})": row['id'] for _, row in history_df.iterrows()}
    
    if 'run_to_delete_display' not in st.session_state:
        st.session_state.run_to_delete_display = None

    selected_run_display = st.selectbox(
        "Select a run to delete", 
        options=[None] + list(run_options.keys()), 
        key="delete_run_select"
    )

    if selected_run_display:
        selected_run_id = run_options[selected_run_display]
        run_details = history_df[history_df['id'] == selected_run_id].iloc[0]

        with st.expander("⚠️ Review Run Details Before Deleting", expanded=True):
            st.markdown(f"- **Run ID:** `{run_details['id']}`")
            st.markdown(f"- **Filename:** `{run_details['filename']}`")
            st.markdown(f"- **Upload Time:** `{pd.to_datetime(run_details['upload_time']).strftime('%Y-%m-%d %H:%M:%S')}`")
            st.markdown(f"- **Total Records:** `{run_details['total_records']}`")
            st.markdown(f"- **Total Exceptions:** `{run_details['total_exceptions']}`")

            st.error("This action will permanently delete the run entry, all its exceptions, department summaries, and user performance records.")
            
            confirm_delete = st.checkbox("I understand this is permanent and want to delete this run.", key=f"confirm_delete_{selected_run_id}")

            if st.button(f"Permanently Delete Run {selected_run_id}", disabled=not confirm_delete):
                with st.spinner(f"Deleting run {selected_run_id}..."):
                    success = db_manager.delete_run(selected_run_id)
                    if success:
                        st.success(f"Successfully deleted Run ID {selected_run_id}.")
                        st.rerun()
                    else:
                        st.error(f"Failed to delete Run ID {selected_run_id}. Check logs for details.")
    else:
        st.info("Select a run from the dropdown list to manage it.")


def show_settings_page():
    st.markdown("### ⚙️ Settings")

    # --- NEW --- Notification Settings Section (for Super User only)
    if st.session_state.get("role") == "Super User":
        st.markdown("---")
        st.markdown("#### 🔔 Notification Service Settings")
        st.caption("Control the automated notification service that alerts users about unresolved entries.")

        settings = db_manager.get_notification_settings()
        
        # Global on/off switch
        st.markdown("##### Global Service Status")
        global_status_current = settings.get('global', {}).get('automatic_notifications_enabled') == 'true'
        global_status_new = st.toggle("Enable Automatic Notification Service for All Users", value=global_status_current)
        
        if global_status_new != global_status_current:
            db_manager.update_notification_setting('global', 'automatic_notifications_enabled', global_status_new)
            st.success(f"Global notification service has been {'enabled' if global_status_new else 'disabled'}.")
            st.rerun()

        # Per-user on/off switch
        st.markdown("##### Per-User Notification Settings")
        st.caption("You can override the global setting and disable automatic notifications for specific users.")
        
        all_users_df = db_manager.get_all_users()
        # Exclude Super Users from this list as they don't have entries to correct
        users_to_manage = all_users_df[all_users_df['role'] != 'Super User'].copy()

        if not users_to_manage.empty:
            # Default to True if the database value is NULL
            users_to_manage['receive_auto_notifications'] = users_to_manage['receive_auto_notifications'].apply(lambda x: False if x is False else True)
            
            users_to_manage.rename(columns={'receive_auto_notifications': 'Receive Automatic Notifications'}, inplace=True)

            edited_df = st.data_editor(
                users_to_manage[['username', 'role', 'Receive Automatic Notifications']],
                use_container_width=True,
                hide_index=True,
                key="user_notif_settings_editor",
                disabled=['username', 'role'] # Make username and role read-only
            )
            
            if st.button("Save User-Specific Settings"):
                with st.spinner("Saving settings..."):
                    # Find what changed between the original and edited dataframes
                    original_users_map = users_to_manage.set_index('username')['Receive Automatic Notifications'].to_dict()
                    edited_users_map = edited_df.set_index('username')['Receive Automatic Notifications'].to_dict()

                    for user, new_setting in edited_users_map.items():
                        if new_setting != original_users_map.get(user):
                            db_manager.update_notification_setting('user', user, new_setting)
                st.success("User-specific settings have been saved.")
                st.rerun()

    # --- Existing Sections ---
    st.markdown("---")
    st.markdown("#### 🛠 Database Management")
    st.warning("🚨 **Caution:** Clearing the database is irreversible.")

    confirm_clear = st.checkbox("I am absolutely sure I want to delete all validation history and associated data (exceptions, summaries, etc.). User accounts will NOT be deleted.")

    if st.button("🗑️ Clear All Validation Data", type="primary"):
        if confirm_clear:
            conn = db_manager._get_connection()
            try:
                with st.spinner("Clearing all validation and run data..."):
                    with conn.cursor() as cursor:
                        cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
                        tables_to_clear = [
                            "correction_logs", # New table
                            "entry_clarifications", # New table
                            "suspicious_transactions_log",
                            "transaction_fingerprints",
                            "correction_status",
                            "user_performance",
                            "department_summary",
                            "exceptions",
                            "validation_runs"
                        ]
                        for table in tables_to_clear:
                            cursor.execute(f"TRUNCATE TABLE `{table}`")
                            logging.info(f"Cleared table: {table}")
                        cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
                conn.commit()
                st.success("All validation data has been cleared successfully. User accounts were not affected. Please refresh the page.")
                st.rerun()
            except mysql.connector.Error as e:
                st.error(f"Failed to clear database: {e}")
                conn.rollback()
            finally:
                if conn and conn.is_connected():
                    conn.close()
        else:
            st.error("You must check the confirmation box to proceed.")

    st.markdown("---")
    st.markdown("#### ℹ️ About This Dashboard")
    dashboard_version = "4.0.0-mysql"
    st.markdown(f"""**Data Validation Dashboard - Version {dashboard_version}**\n\nThis application is designed to help users validate data from Excel files against a predefined set of business rules...\n\nBuilt with Streamlit, Pandas, Plotly, and MySQL.""")

def show_send_notification_page():
    st.markdown("### ✉️ Send Manual Notification")

    # Role check to ensure only authorized users see this page
    user_role = st.session_state.get("role")
    if user_role not in ["Management", "Super User"]:
        st.error("🔒 You do not have permission to access this page.")
        return

    st.caption("Send a direct message or alert to any user in the system.")
    
    # Fetch all users to populate the dropdown
    all_users_df = db_manager.get_all_users()
    if all_users_df.empty:
        st.warning("There are no users in the system to send notifications to.")
        return
        
    user_list = sorted(all_users_df['username'].tolist())

    with st.form("send_notification_form", clear_on_submit=True):
        st.markdown("##### Compose Your Message")
        
        # Dropdown to select the recipient
        target_user = st.selectbox(
            "Select User to Notify",
            options=user_list,
            index=None,
            placeholder="Choose a recipient..."
        )
        
        # Input for the message title
        notification_title = st.text_input(
            "Notification Title",
            value="Message from Admin"
        )
        
        # Text area for the message body
        notification_message = st.text_area(
            "Message Body",
            height=150
        )
        
        submitted = st.form_submit_button("Send Notification", type="primary")
        
        if submitted:
            # Validation
            if not target_user:
                st.warning("Please select a user to notify.")
            elif not notification_message.strip():
                st.warning("The message body cannot be empty.")
            else:
                # Call the existing database function to create the notification
                with st.spinner(f"Sending notification to {target_user}..."):
                    success = db_manager.create_notification(
                        username=target_user,
                        notif_type=notification_title,
                        message=notification_message
                    )
                
                if success:
                    st.success(f"Notification successfully sent to {target_user}!")
                else:
                    st.error("Failed to send notification. Please check the logs.")

def show_correction_status_page(start_date, end_date):
    st.markdown("### 📝 Correction Status")
    db_manager = get_database_manager()

    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    managed_users = st.session_state.get("managed_users", [])

    # === Part 1: Status Update Section ===
    history_df = db_manager.get_validation_history(user_role, username, managed_users)
    
    runs_for_status_update_df = pd.DataFrame()
    update_caption = ""

    if user_role == "Super User":
        runs_for_status_update_df = history_df
        st.markdown("#### Update Status for Any Run (Super User)")
        update_caption = "As a Super User, you can update the correction status for any validation run from history."
    else:
        today = datetime.now().date()
        history_df['upload_time'] = pd.to_datetime(history_df['upload_time'])
        runs_for_status_update_df = history_df[history_df['upload_time'].dt.date == today].copy()
        st.markdown("#### Update Status for Today's Runs")
        update_caption = "You can only update the status for validation runs that occurred today."
    
    st.caption(update_caption)

    if runs_for_status_update_df.empty:
        st.info("No validation runs are available for you to update in the selected scope.")
    else:
        runs_for_status_update_df['display_name'] = "Run " + runs_for_status_update_df['id'].astype(str) + ": " + runs_for_status_update_df['filename'].fillna('N/A')
        selected_run_display = st.selectbox(
            "Select a run to update status:",
            options=runs_for_status_update_df['display_name'].tolist(),
            index=None,
            placeholder="Choose a run..."
        )

        if selected_run_display:
            selected_run_id = int(runs_for_status_update_df[runs_for_status_update_df['display_name'] == selected_run_display]['id'].iloc[0])

            all_users_with_exceptions_in_run = []
            conn = db_manager._get_connection()
            try:
                query = """
                    SELECT DISTINCT `user` FROM `user_performance` 
                    WHERE run_id = %s AND exception_records > 0 AND `user` IS NOT NULL AND `user` != ''
                """
                df_users = pd.read_sql_query(query, conn, params=(selected_run_id,))
                if not df_users.empty:
                    all_users_with_exceptions_in_run = df_users['user'].tolist()
            finally:
                if conn.is_connected():
                    conn.close()

            users_to_show = []
            if user_role == 'User':
                if username in all_users_with_exceptions_in_run:
                    users_to_show = [username]
            elif user_role == 'Manager':
                team_members_lower = {u.lower() for u in (managed_users + [username])}
                users_to_show = [u for u in all_users_with_exceptions_in_run if u.lower() in team_members_lower]
            else:
                users_to_show = all_users_with_exceptions_in_run
            
            if not users_to_show:
                st.success("✅ No users with exceptions were found in this run for your specific scope.")
            else:
                current_statuses = db_manager.get_correction_status_for_run(selected_run_id)
                st.markdown(f"**Set the correction status for each user ({len(users_to_show)} found). Changes are saved automatically.**")
                
                cols = st.columns(min(len(users_to_show), 3))
                for i, user in enumerate(sorted(users_to_show)):
                    current_status = current_statuses.get(user, "Pending")
                    status_options = ["Yes", "No", "Pending"]
                    
                    with cols[i % 3]:
                        new_status = st.radio(
                            f"Status for **{user}**",
                            options=status_options,
                            index=status_options.index(current_status),
                            key=f"status_{selected_run_id}_{user}",
                            horizontal=True
                        )
                        if new_status != current_status:
                            with st.spinner(f"Updating status for {user}..."):
                                db_manager.add_or_update_correction_status(selected_run_id, user, new_status)
                            st.rerun()

    # === Part 2: Analytics Section ===
    if user_role in ["Manager", "Management", "Super User"]:
        st.markdown("---")
        st.markdown("#### Correction Performance Analytics")
        if not start_date or not end_date:
            st.warning("Please select a date range from the sidebar to view analytics.")
            return
            
        history_df['upload_time'] = pd.to_datetime(history_df['upload_time'])
        all_runs_in_range = history_df[
            (history_df['upload_time'].dt.date >= start_date) &
            (history_df['upload_time'].dt.date <= end_date)
        ]
        if all_runs_in_range.empty:
            st.info("No validation runs found in the selected date range for your scope.")
            return
            
        run_ids_in_scope = all_runs_in_range['id'].tolist()
        
        accessible_users = None
        if user_role == 'Manager':
            accessible_users = managed_users + [username]
        
        # The debug info is removed from the function call
        summary_df, _ = db_manager.get_correction_summary(run_ids_in_scope, accessible_users)
        
        if summary_df.empty:
            st.info("No correction status data to analyze for the selected scope.")
            return
        
        st.markdown("##### Overall Correction Status")
        status_counts = summary_df['status'].value_counts()
        fig_pie = px.pie(
            names=status_counts.index, values=status_counts.values,
            title="Overall Distribution of Correction Statuses",
            color_discrete_map={'Yes': '#28a745', 'No': '#dc3545', 'Pending': '#ffc107'}, hole=0.3
        )
        fig_pie.update_layout(title_x=0.5, font=PLOTLY_FONT)
        st.plotly_chart(fig_pie, use_container_width=True)
        
        st.markdown("##### Performance by User")
        user_summary = summary_df.groupby('user')['status'].value_counts().unstack(fill_value=0)
        for status_col in ['Yes', 'No', 'Pending']:
            if status_col not in user_summary.columns: user_summary[status_col] = 0
        user_summary = user_summary[['Yes', 'No', 'Pending']]
        fig_bar = px.bar(
            user_summary, x=user_summary.index, y=['Yes', 'No', 'Pending'],
            title="User-wise Correction Status Breakdown",
            labels={'x': 'User', 'value': 'Number of Runs', 'variable': 'Status'},
            barmode='stack', color_discrete_map={'Yes': '#28a745', 'No': '#dc3545', 'Pending': '#ffc107'}
        )
        fig_bar.update_layout(title_x=0.5, font=PLOTLY_FONT, xaxis_tickangle=-45)
        st.plotly_chart(fig_bar, use_container_width=True)
        st.dataframe(user_summary, use_container_width=True)


def show_user_management_page():
    st.markdown("### 🛠️ User Management")
    st.info("View all users, create new accounts, and manage existing users.")

    db_manager = get_database_manager()

    # Fetch all necessary user data at the beginning
    all_users_df = db_manager.get_all_users()
    all_users_list = all_users_df['username'].tolist() if not all_users_df.empty else []
    manager_list = all_users_df[all_users_df['role'] == 'Manager']['username'].tolist() if not all_users_df.empty else []
    management_list = db_manager.get_management_users()

    tab1, tab2, tab3 = st.tabs(["👥 View All Users", "➕ Create New User", "✏️ Manage Existing User"])

    with tab1:
        st.markdown("#### Current Users in the System")
        if not all_users_df.empty:
            st.dataframe(
                all_users_df,
                use_container_width=True,
                hide_index=True,
                # --- MODIFIED --- Added can_accept_corrections to the view
                column_order=("id", "username", "role", "full_name", "disabled", "reports_to", "mapped_to_management", "can_accept_corrections"),
                column_config={
                    "id": "ID", "username": "Username", "role": "Role", "full_name": "Full Name",
                    "disabled": st.column_config.CheckboxColumn("Disabled?"),
                    "reports_to": "Reports To", "mapped_to_management": "Mapped To Mgmt",
                    "can_accept_corrections": st.column_config.CheckboxColumn("Can Accept Corrections?"),
                }
            )
        else:
            st.warning("No users found in the database.")

    with tab2:
        # This section remains unchanged
        st.markdown("#### Create a New User Account")
        with st.form("create_user_form"):
            new_username = st.text_input("Username*")
            new_full_name = st.text_input("Full Name (Optional)")
            new_password = st.text_input("Password*", type="password")
            new_role = st.selectbox("Role*", ["User", "Manager", "Management", "Super User"])
            st.markdown("---")
            new_email = st.text_input("Email (for notifications)")
            new_mobile = st.text_input("Mobile Number")
            st.markdown("---")
            send_credentials_checkbox = st.checkbox("Send login details to user via email", value=False)
            
            submitted = st.form_submit_button("Create User")
            if submitted:
                if not new_username or not new_password:
                    st.error("Username and Password cannot be empty.")
                else:
                    result = db_manager.add_user(
                        username=new_username, password=new_password, role=new_role, 
                        full_name=new_full_name, email=new_email, mobile_number=new_mobile
                    )
                    if result is True:
                        st.success(f"User '{new_username}' created successfully!")
                        if send_credentials_checkbox and new_email:
                            pass
                        st.rerun() 
                    else:
                        st.error(result)

    with tab3:
        st.markdown("#### Edit, Map, Disable, or Delete a User")
        if not all_users_list:
            st.warning("No users exist to be managed.")
        else:
            user_to_manage = st.selectbox("Select a User to Manage", [""] + all_users_list, key="user_manage_select")

            if user_to_manage:
                user_data = db_manager.get_user_profile(user_to_manage)
                
                st.markdown("##### Edit Profile Information")
                # This section remains unchanged
                with st.form(f"edit_profile_{user_to_manage}"):
                    full_name_edit = st.text_input("Full Name", value=user_data.get('full_name', '') or '')
                    email_edit = st.text_input("Email", value=user_data.get('email', '') or '')
                    mobile_edit = st.text_input("Mobile Number", value=user_data.get('mobile_number', '') or '')
                    if st.form_submit_button("Update Profile"):
                        if db_manager.update_user_profile(user_to_manage, full_name_edit, email_edit, mobile_edit):
                            st.success("User profile updated successfully."); st.rerun()
                        else:
                            st.error("Failed to update profile.")
                
                st.markdown("---")

                st.markdown("##### Account Settings")
                is_disabled = bool(user_data.get('disabled', False))
                new_disabled_status = st.toggle("Account Disabled", value=is_disabled, key=f"disable_{user_to_manage}")
                if new_disabled_status != is_disabled:
                    if db_manager.set_user_disabled_status(user_to_manage, new_disabled_status):
                        st.success(f"User account has been {'disabled' if new_disabled_status else 'enabled'}."); st.rerun()
                    else:
                        st.error("Failed to update account status.")
                
                current_role = user_data.get('role', 'User')
                role_options = ["User", "Manager", "Management", "Super User"]
                current_role_index = role_options.index(current_role) if current_role in role_options else 0
                new_role_edit = st.selectbox("Change Role", role_options, index=current_role_index, key=f"role_{user_to_manage}")
                if st.button("Update Role"):
                    db_manager.update_user_role(user_to_manage, new_role_edit)
                    st.success(f"Role for '{user_to_manage}' updated to '{new_role_edit}'."); st.rerun()
                
                if new_role_edit == 'User':
                    current_manager = user_data.get('reports_to')
                    manager_options = ["None"] + manager_list
                    current_manager_index = manager_options.index(current_manager) if current_manager in manager_options else 0
                    new_manager_map = st.selectbox(f"Map '{user_to_manage}' to Manager", manager_options, index=current_manager_index)
                    if st.button("Update User's Manager"):
                        db_manager.update_user_mapping(user_to_manage, new_manager_map)
                        st.success(f"'{user_to_manage}' now reports to '{new_manager_map}'."); st.rerun()

                # --- NEW SNIPPET --- This block is new
                if new_role_edit == 'Manager':
                    st.markdown("##### Manager Permissions")
                    # For a manager, True (or Null) means they can accept. Only explicitly False denies permission.
                    can_accept_current = False if user_data.get('can_accept_corrections') is False else True
                    
                    can_accept_new = st.toggle("Allow this manager to accept corrections", value=can_accept_current, key=f"accept_perm_{user_to_manage}")
                    
                    if can_accept_new != can_accept_current:
                        if db_manager.set_manager_acceptance_permission(user_to_manage, can_accept_new):
                            st.success(f"Permission for {user_to_manage} updated.")
                            st.rerun()
                        else:
                            st.error("Failed to update permission.")

                    # This mapping logic is existing, just moved inside the conditional block
                    current_management_mapping = user_data.get('mapped_to_management')
                    management_options = ["None"] + management_list
                    current_map_index = management_options.index(current_management_mapping) if current_management_mapping in management_options else 0
                    new_management_map = st.selectbox(f"Map '{user_to_manage}' to Management User", options=management_options, index=current_map_index)
                    if st.button("Update Manager-to-Management Mapping"):
                        db_manager.update_manager_to_management_mapping(user_to_manage, new_management_map)
                        st.success(f"Successfully mapped '{user_to_manage}' to '{new_management_map}'."); st.rerun()
                # --- END OF NEW SNIPPET ---

                st.markdown("---")

                with st.expander("🔑 Change Password"):
                    # This section remains unchanged
                    with st.form(f"change_password_form_{user_to_manage}"):
                        password_to_change = st.text_input("New Password", type="password", key=f"new_pass_{user_to_manage}")
                        if st.form_submit_button("Set New Password"):
                            if password_to_change:
                                db_manager.update_user_password(user_to_manage, password_to_change)
                                st.success(f"Password for '{user_to_manage}' has been changed.")
                            else:
                                st.warning("Password field cannot be empty.")
                
                with st.expander("🚨 Danger Zone - Permanent Deletion"):
                     # This section remains unchanged
                    st.warning(f"This action will permanently delete the user '{user_to_manage}' and cannot be undone.")
                    confirm_delete_text = st.text_input(f"To confirm, type the username '{user_to_manage}' exactly:")
                    if st.button("Permanently Delete User", type="primary", disabled=(confirm_delete_text != user_to_manage)):
                        result = db_manager.delete_user(user_to_manage)
                        if result is True:
                            st.success(f"User '{user_to_manage}' has been permanently deleted."); st.rerun()
                        else:
                            st.error(f"Could not delete user. Reason: {result}")

def show_suspicious_tag_control_page():
    st.markdown("### Suspicious Transaction Tag Control")
    st.caption("Use this page to define the rules that flag transactions for manual review.")

    # These are the fixed data columns you can apply rules to.
    rule_columns = ['FC-Vertical.Name', 'Location.Name', 'Activity.Name', 'Crop.Name']
    
    tab1, tab2 = st.tabs(["⚙️ Manage Rules", "🗃️ Manage Dropdown Options"])

    # --- TAB 1: Manage and Define Rules ---
    with tab1:
        st.subheader("Rule Matrix for each Sub-Department")
        st.info("Set the values for each column that should be considered 'suspicious'. You can select multiple values. Click 'Save All Rules' at the bottom to apply your changes.")
        
        all_sub_depts = load_sub_departments()
        if not all_sub_depts:
            st.error("Cannot display rules: The Sub-Department list could not be loaded from 'reference_data/SubDepartment.xlsx'.")
            return

        # Fetch all rules and options from the database once for efficiency
        all_rules_df = db_manager.get_all_suspicious_rules()
        all_options = {col: db_manager.get_rule_options(col) for col in rule_columns}

        # Create a dictionary for quick lookup of existing rule settings
        rules_lookup = {}
        if not all_rules_df.empty:
            for _, rule in all_rules_df.iterrows():
                key = (rule['sub_department_name'], rule['rule_column'])
                rules_lookup[key] = rule['rule_values']

        # A single form is used to prevent the page from reloading on every click,
        # allowing you to set all rules before saving.
        with st.form("rules_form"):
            for sub_dept in all_sub_depts:
                st.markdown(f"--- \n**{sub_dept}**")
                cols = st.columns(len(rule_columns))
                
                for i, col_name in enumerate(rule_columns):
                    with cols[i]:
                        # Prepare the list of choices and find the currently saved selections
                        options_list = [opt['option_value'] for opt in all_options.get(col_name, [])]
                        default_selections = rules_lookup.get((sub_dept, col_name), [])
                        
                        st.multiselect(
                            label=col_name,
                            options=options_list,
                            default=default_selections,
                            key=f"rule_{sub_dept}_{col_name}" # Unique key for each widget
                        )
            
            # The single save button at the end of the form
            submitted = st.form_submit_button("Save All Rules")
            if submitted:
                with st.spinner("Saving all rule changes..."):
                    # When saved, iterate through every widget's state and update the DB
                    for sub_dept in all_sub_depts:
                        for col_name in rule_columns:
                            selected_values = st.session_state[f"rule_{sub_dept}_{col_name}"]
                            db_manager.save_suspicious_rule(sub_dept, col_name, selected_values)
                st.success("All rules have been saved successfully!")
                time.sleep(1)
                st.rerun()

    # --- TAB 2: Manage the Options Available in the Dropdowns ---
    with tab2:
        st.subheader("Edit the options available for rule creation")
        
        selected_column_to_edit = st.selectbox("Select a column to manage its options:", rule_columns)
        
        if selected_column_to_edit:
            st.markdown(f"#### Options for **{selected_column_to_edit}**")
            
            # Form to add a new option value
            with st.form(f"add_option_form_{selected_column_to_edit}", clear_on_submit=True):
                new_option_value = st.text_input("New Option Value")
                add_submitted = st.form_submit_button("Add Option")
                if add_submitted and new_option_value:
                    result = db_manager.add_rule_option(selected_column_to_edit, new_option_value)
                    if result is True:
                        st.success(f"Option '{new_option_value}' added.")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(result)

            st.markdown("---")
            st.markdown("##### Existing Options")
            
            # Display current options with delete buttons
            current_options = db_manager.get_rule_options(selected_column_to_edit)
            if not current_options:
                st.info("No custom options are defined for this column.")
            else:
                for option in current_options:
                    option_id = option['id']
                    option_val = option['option_value']
                    col1, col2 = st.columns([4, 1])
                    col1.write(option_val)
                    if col2.button("Delete", key=f"del_opt_{option_id}"):
                        if db_manager.delete_rule_option(option_id):
                            st.success(f"Option '{option_val}' deleted.")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("Failed to delete option.")

def show_suspicious_category_transactions_page():
    st.markdown("### Suspicious Category Transactions")
    
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")

    # --- Section 1: Triage for Admins/Managers ---
    if user_role in ["Super User", "Management", "Manager"]:
        st.subheader("Items Pending Your Review")
        st.caption("Review transactions flagged by custom rules. Accept them or reject and send back for correction.")
        
        # Determine team scope for Managers
        team_list = []
        if user_role == "Manager":
            managed_users = st.session_state.get("managed_users", [])
            team_list = [username.lower()] + [user.lower() for user in managed_users]

        all_pending_df = db_manager.get_suspicious_transactions_for_admin()
        
        # Filter for Managers, show all for Super User/Management
        pending_review_df = all_pending_df
        if user_role == "Manager":
            if not all_pending_df.empty:
                pending_review_df = all_pending_df[all_pending_df['created_user'].str.lower().isin(team_list)].copy()
            else:
                pending_review_df = pd.DataFrame()

        if pending_review_df.empty:
            st.success("✅ There are no new transactions pending review.")
        else:
            st.info(f"You have **{len(pending_review_df)}** transaction(s) pending your review.")
            
            for _, row in pending_review_df.iterrows():
                log_id = row['id']
                with st.expander(f"**ID {log_id}** | User: **{row.get('created_user', 'N/A')}** | Department: **{row.get('Department.Name', 'N/A')}** | Amount: **{row.get('Net amount', 0):,.2f}**"):
                    
                    st.markdown("##### Transaction Details")
                    # Restored detailed metrics display
                    details_cols = st.columns(3)
                    details_cols[0].metric("Sub Department", row.get('Sub Department.Name', 'N/A'))
                    details_cols[1].metric("Location", row.get('Location.Name', 'N/A'))
                    details_cols[2].metric("Crop", row.get('Crop.Name', 'N/A'))
                    st.dataframe(pd.DataFrame([row.to_dict()]), use_container_width=True, hide_index=True)
                    
                    st.markdown("##### Actions")
                    action_cols = st.columns(2)
                    with action_cols[0]:
                        if st.button("Accept", key=f"accept_{log_id}", use_container_width=True):
                            if db_manager.accept_suspicious_transaction(log_id, username):
                                st.success(f"Transaction ID {log_id} has been accepted."); time.sleep(1); st.rerun()
                    with action_cols[1]:
                        with st.form(key=f"reject_form_{log_id}"):
                            rejection_comment = st.text_area("Rejection Comment (Required)", height=100)
                            if st.form_submit_button("Reject and Send for Correction", type="primary"):
                                if rejection_comment.strip():
                                    if db_manager.reject_suspicious_transaction(log_id, username, rejection_comment):
                                        user_to_notify = row.get('created_user')
                                        if user_to_notify:
                                            db_manager.create_notification(username=user_to_notify, notif_type="Correction Required", message=f"Transaction ID {log_id} requires correction.")
                                        st.success(f"Transaction ID {log_id} sent to user."); time.sleep(1); st.rerun()
                                else:
                                    st.warning("A comment is required.")
        
        st.markdown("---")
        # --- Section 2: Tracking for Admins/Managers ---
        st.subheader("Track Items Sent for Correction")
        
        all_rejected_df = db_manager.get_rejected_transactions()
        
        rejected_df_for_view = all_rejected_df
        if user_role == "Manager":
            if not all_rejected_df.empty:
                rejected_df_for_view = all_rejected_df[all_rejected_df['created_user'].str.lower().isin(team_list)].copy()
            else:
                rejected_df_for_view = pd.DataFrame()

        if rejected_df_for_view.empty:
            st.info("No items are currently pending user correction.")
        else:
            for _, row in rejected_df_for_view.iterrows():
                log_id = row['id']
                status = row.get('status')
                reviewed_at_str = pd.to_datetime(row.get('reviewed_at')).strftime('%d-%b-%Y') if pd.notna(row.get('reviewed_at')) else ''
                
                status_color = "orange" if status == "Rejected" else "green"
                status_icon = "⏳" if status == "Rejected" else "✅"
                
                with st.expander(f"**ID {log_id}** | User: **{row.get('created_user')}** | Sent Back By: **{row.get('reviewed_by')}** on {reviewed_at_str}"):
                    st.markdown(f"**Status:** :{status_color}[{status} {status_icon}]")
                    st.info(f"**Admin Comment:** {row.get('admin_comment')}")
                    if status == 'User Corrected':
                        corrected_time = pd.to_datetime(row.get('user_corrected_at')).strftime('%d-%b-%Y %H:%M')
                        st.success(f"User confirmed correction on {corrected_time}")
                    
                    st.markdown("##### Original Transaction Details")
                    # --- RESTORED --- Consistent detailed metrics display
                    details_cols_rej = st.columns(3)
                    details_cols_rej[0].metric("Sub Department", row.get('Sub Department.Name', 'N/A'))
                    details_cols_rej[1].metric("Location", row.get('Location.Name', 'N/A'))
                    details_cols_rej[2].metric("Crop", row.get('Crop.Name', 'N/A'))
                    st.dataframe(pd.DataFrame([row.to_dict()]), use_container_width=True, hide_index=True)
                    
                    # --- ADDED --- "Call Back" button for Super User
                    if user_role == "Super User" and status == "Rejected":
                        st.markdown("---")
                        if st.button("Call Back Request", key=f"callback_{log_id}", help="This will remove the task from the user's queue and place it back into the 'Pending Review' list."):
                            if db_manager.call_back_rejected_transaction(log_id):
                                st.success(f"Transaction ID {log_id} has been successfully called back. It is now pending review again.")
                                user_to_notify = row.get('created_user')
                                if user_to_notify:
                                    db_manager.create_notification(username=user_to_notify, notif_type="Task Recalled", message=f"The correction task for Transaction ID {log_id} has been recalled by an administrator.")
                                time.sleep(2)
                                st.rerun()
                            else:
                                st.error("Failed to call back the transaction.")


    # --- View for Standard Users (Unchanged) ---
    elif user_role == "User":
        st.caption("These are the transactions that were sent back to you for correction by a manager.")
        
        correction_tasks_df = db_manager.get_suspicious_transactions_for_user(username)
        
        if correction_tasks_df.empty:
            st.success("👍 You have no pending correction tasks.")
            return

        st.warning(f"You have **{len(correction_tasks_df)}** transaction(s) that require your attention.")

        for _, row in correction_tasks_df.iterrows():
            log_id = row['id']
            with st.expander(f"**Correction Task ID {log_id}** | Reviewed by: **{row.get('reviewed_by')}**", expanded=True):
                st.warning(f"**Manager's Comment:** {row.get('admin_comment')}")
                st.markdown("##### Original Transaction Details")
                st.dataframe(pd.DataFrame([row.to_dict()]), use_container_width=True, hide_index=True)
                
                if st.button("I have corrected this transaction", key=f"confirm_{log_id}", type="primary"):
                    if db_manager.confirm_user_correction(log_id):
                        st.success("Thank you for confirming. The task has been removed from your list."); time.sleep(1); st.rerun()

def send_performance_email(to_recipients, subject, html_body, cc_recipients=None, images=None):
    """
    Connects to the SMTP server and sends a multipart HTML email with embedded images.

    Args:
        to_recipients (list): A list of email addresses for the 'To' field.
        subject (str): The subject line of the email.
        html_body (str): The HTML content of the email body.
        cc_recipients (list, optional): A list of email addresses for the 'CC' field. Defaults to None.
        images (dict, optional): A dictionary of images to embed. Defaults to None.
    """
    try:
        # Fetch credentials from Streamlit secrets
        sender_email = st.secrets["email_credentials"]["sender_email"]
        password = st.secrets["email_credentials"]["sender_password"]
        smtp_server = st.secrets["email_credentials"]["smtp_server"]
        smtp_port = st.secrets["email_credentials"]["smtp_port"]

        if not to_recipients:
            st.error("No recipients specified in the 'To' field.")
            return
            
        if cc_recipients is None:
            cc_recipients = []

        # Combine all recipients for the mail server
        all_recipients_list = to_recipients + cc_recipients

        # Create the email message
        message = MIMEMultipart("related")
        message["From"] = sender_email
        message["To"] = ", ".join(to_recipients)
        if cc_recipients:
            message["Cc"] = ", ".join(cc_recipients)
        message["Subject"] = subject

        # Attach the HTML body and images
        message.attach(MIMEText(html_body, "html"))
        if images:
            for cid, img_data in images.items():
                img = MIMEImage(img_data, _subtype="png")
                img.add_header('Content-ID', f'<{cid}>')
                message.attach(img)

        # Send the email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, all_recipients_list, message.as_string())
        
        st.success(f"Performance report successfully sent to: {', '.join(all_recipients_list)}")
        logging.info(f"Successfully sent performance email to: {', '.join(all_recipients_list)}")

    except KeyError:
        st.error("Email credentials are not configured in st.secrets. Please check your secrets.toml file.")
        logging.error("Email credentials missing in st.secrets.")
    except Exception as e:
        st.error(f"Failed to send email: {e}")
        logging.error(f"Failed to send email: {e}", exc_info=True)

# ADD THIS NEW FUNCTION TO YOUR SCRIPT

def send_report_email_with_attachment(to_recipients, subject, html_body, attachment_data, attachment_filename, cc_recipients=None):
    """
    Connects to the SMTP server and sends an email with a file attachment.
    """
    try:
        # Fetch credentials from Streamlit secrets
        sender_email = st.secrets["email_credentials"]["sender_email"]
        password = st.secrets["email_credentials"]["sender_password"]
        smtp_server = st.secrets["email_credentials"]["smtp_server"]
        smtp_port = st.secrets["email_credentials"]["smtp_port"]

        if not to_recipients:
            st.error("No recipients specified in the 'To' field.")
            return

        # Create the email message
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = ", ".join(to_recipients)
        if cc_recipients:
            message["Cc"] = ", ".join(cc_recipients)
        message["Subject"] = subject

        # Attach the HTML body
        message.attach(MIMEText(html_body, "html"))

        # Attach the file
        if attachment_data and attachment_filename:
            part = MIMEApplication(attachment_data.getvalue(), Name=attachment_filename)
            part['Content-Disposition'] = f'attachment; filename="{attachment_filename}"'
            message.attach(part)

        # Send the email
        all_recipients_list = to_recipients + (cc_recipients or [])
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, all_recipients_list, message.as_string())
        
        st.success(f"Report successfully sent to: {', '.join(all_recipients_list)}")
        logging.info(f"Successfully sent report email to: {', '.join(all_recipients_list)}")

    except KeyError:
        st.error("Email credentials are not configured in st.secrets. Please check your secrets.toml file.")
    except Exception as e:
        st.error(f"Failed to send email: {e}")
        logging.error(f"Failed to send email: {e}", exc_info=True)

def send_new_user_credentials_email(to_email, new_username, plain_text_password):
    """
    Sends a welcome email to a new user with their login credentials.
    """
    try:
        # Fetch credentials from Streamlit secrets
        sender_email = st.secrets["email_credentials"]["sender_email"]
        password = st.secrets["email_credentials"]["sender_password"]
        smtp_server = st.secrets["email_credentials"]["smtp_server"]
        smtp_port = st.secrets["email_credentials"]["smtp_port"]

        if not to_email:
            st.warning("Could not send welcome email: No email address provided.")
            return

        # Create the email message
        message = MIMEMultipart("alternative")
        message["From"] = sender_email
        message["To"] = to_email
        message["Subject"] = "Welcome! Your Account for the Data Validation Dashboard is Ready"

        # Create the HTML body of the email
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; }}
                .container {{ padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #f9f9f9; max-width: 600px; margin: auto; }}
                .header {{ font-size: 24px; color: #2d3748; border-bottom: 2px solid #667eea; padding-bottom: 10px; }}
                .credential-box {{ background-color: #fff; border: 1px solid #e2e8f0; padding: 15px; margin-top: 20px; border-radius: 5px; }}
                .credential-box p {{ margin: 5px 0; }}
                .credential-box strong {{ color: #4a5568; }}
                .footer {{ font-size: 12px; color: #888; margin-top: 20px; text-align: center; }}
                .warning {{ color: #d9534f; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2 class="header">Welcome to the Dashboard!</h2>
                <p>Hello {new_username},</p>
                <p>An account has been created for you on the Data Validation Dashboard. You can use the following credentials to log in:</p>
                <div class="credential-box">
                    <p><strong>Username:</strong> {new_username}</p>
                    <p><strong>Password:</strong> {plain_text_password}</p>
                </div>
                <p class="warning">For your security, it is recommended that you change your password after your first login.</p>
                <p>Thank you!</p>
                <div class="footer">
                    <p><i>This is an automated message from the Data Validation Dashboard.</i></p>
                </div>
            </div>
        </body>
        </html>
        """

        message.attach(MIMEText(html_body, "html"))
        
        # Send the email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, to_email, message.as_string())
        
        st.success(f"Login credentials successfully sent to {to_email}.")
        logging.info(f"Successfully sent new user credentials to: {to_email}")

    except KeyError:
        st.error("Could not send email: Email credentials are not configured in st.secrets.")
    except Exception as e:
        st.error(f"Failed to send welcome email: {e}")
        logging.error(f"Failed to send new user welcome email: {e}", exc_info=True)

def add_date_filters_to_sidebar(history_df, key_suffix=""):
    """Adds date filter widgets to the sidebar and returns start and end dates."""
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🗓️ Date Filter")

    if history_df.empty:
        st.sidebar.warning("No historical data to filter.")
        return None, None

    history_df['upload_time'] = pd.to_datetime(history_df['upload_time'])
    min_date = history_df['upload_time'].min().date()
    max_date = history_df['upload_time'].max().date()

    filter_type = st.sidebar.radio(
        "Filter by:",
        ("All Time", "Month", "Custom Range"),
        key=f"date_filter_type_{key_suffix}"
    )

    start_date, end_date = None, None

    if filter_type == "Month":
        history_df['month_year'] = history_df['upload_time'].dt.to_period('M')
        month_options = sorted(history_df['month_year'].unique(), reverse=True)
        selected_month = st.sidebar.selectbox(
            "Select Month",
            options=month_options,
            format_func=lambda x: x.strftime('%B %Y'),
            key=f"month_select_{key_suffix}"
        )
        if selected_month:
            start_date = selected_month.start_time.date()
            end_date = selected_month.end_time.date()

    elif filter_type == "Custom Range":
        start_date_input = st.sidebar.date_input("Start date", min_date, min_value=min_date, max_value=max_date, key=f"start_date_{key_suffix}")
        end_date_input = st.sidebar.date_input("End date", max_date, min_value=min_date, max_value=max_date, key=f"end_date_{key_suffix}")
        if start_date_input and end_date_input:
            start_date = start_date_input
            end_date = end_date_input
    else:  # All Time
        start_date = min_date
        end_date = max_date

    return start_date, end_date


def main():
    st.markdown("<h1>🎯 Data Validation Dashboard</h1>", unsafe_allow_html=True)
    
    if not check_password():
        st.stop()

    # --- Get User Context from Session State ---
    user_role = st.session_state.get("role")
    username = st.session_state.get("username_actual")
    
    if user_role == "Manager" and "managed_users" not in st.session_state:
        st.session_state["managed_users"] = db_manager.get_managed_users(username)
    managed_users = st.session_state.get("managed_users", [])

    # --- Run core logic checks and fetch notifications ---
    run_user_session_checks(username)
    user_notifications = db_manager.get_notifications_for_user(username)
    
    st.sidebar.image("assets/logo.png", width=150)
    st.sidebar.info(f"Logged in as: **{st.session_state.get('full_name', username)}**\n\nRole: **{user_role}**")

    with st.sidebar.expander(f"🔔 Notifications ({len(user_notifications)})", expanded=bool(user_notifications)):
        if not user_notifications:
            st.write("No new notifications.")
        else:
            for notif in user_notifications:
                st.info(f"**{notif['notification_type']}** ({notif['created_at'].strftime('%d-%b-%Y')})\n\n{notif['message']}")
            if st.button("Mark all as read"):
                notif_ids = [n['id'] for n in user_notifications]
                db_manager.mark_notifications_as_read(notif_ids)
                st.rerun()
    
    if st.sidebar.button("Logout"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
    
    st.sidebar.markdown("---")
    
    # --- Hierarchical Permission Logic ---
    user_permissions = db_manager.get_user_permissions(username)
    if user_role == "Super User":
        user_permissions = {"can_upload": True, "disabled_pages": []}

    if user_permissions is None:
        st.error("Could not load your permissions. Please contact an administrator.")
        st.stop()
    
    st.session_state['can_upload'] = user_permissions['can_upload']
    user_disabled_pages = user_permissions['disabled_pages']
    
    # --- MODIFIED --- Default page access now includes all new pages for the appropriate roles
    page_access = {
        "Super User": [
            "🏠 Upload & Validate", "📊 Dashboard Analytics", "📈 Trends & History", "👤📊 User Performance", 
            "📝 Correction Entries", "📈 Correction Analytics", "✉️ Send Notification",
            "📋 Suspicious Transactions", "🕵️ Suspicious Tag Control",
            "🗂️ Report Archive", "📝 Correction Status", "📢 Clarification Center", "📋 Exception Details", 
            "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", 
            "👤📍 User & Location Analysis", "🔑 Access Control", "🛠️ User Management", 
            "🗑️ Data Management", "⚙️ Settings"
        ],
        "Management": [
            "🏠 Upload & Validate", "📊 Dashboard Analytics", "📈 Trends & History", "👤📊 User Performance",
            "📝 Correction Entries", "✉️ Send Notification",
            "📋 Suspicious Transactions", "🕵️ Suspicious Tag Control",
            "🗂️ Report Archive", "📝 Correction Status", "📢 Clarification Center", "📋 Exception Details", 
            "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", 
            "👤📍 User & Location Analysis", "🔑 Access Control"
        ],
        "Manager": [
            "🏠 Upload & Validate", "📊 Dashboard Analytics", "📈 Trends & History", "👤📊 User Performance",
            "📝 Correction Entries",
            "📋 Suspicious Transactions",
            "🗂️ Report Archive", "📝 Correction Status", "📢 Clarification Center", "📋 Exception Details", 
            "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", 
            "👤📍 User & Location Analysis"
        ],
        "User": [
            "🏠 Upload & Validate", "👤📊 User Performance",
            "📝 Correction Entries",
            "📋 Suspicious Transactions",
            "🗂️ Report Archive", "📝 Correction Status", "👤🧾 User-wise Ledger Exceptions"
        ]
    }
    
    base_pages_for_role = page_access.get(user_role, [])
    page_navigation_options = [page for page in base_pages_for_role if page not in user_disabled_pages]
    
    # This clarification logic for a different feature remains unchanged
    if st.session_state.get('clarification_required', False):
        st.error("ACTION REQUIRED: Clarification Needed", icon="⚠️")
        run_ids_str = ", ".join(map(str, st.session_state.get('clarification_run_ids', [])))
        st.warning(f"Our records show that you have 3 or more consecutive unresolved validation runs (**Run IDs: {run_ids_str}**). Please provide a clarification below to continue accessing the dashboard.")
        
        with st.form("clarification_form"):
            clarification_text = st.text_area("Please provide your clarification for the delay in resolving these exceptions:")
            submitted = st.form_submit_button("Submit Clarification")
            if submitted:
                if not clarification_text.strip():
                    st.error("Clarification cannot be empty.")
                else:
                    success = db_manager.submit_clarification(username, st.session_state['clarification_run_ids'], clarification_text)
                    if success:
                        st.session_state['clarification_required'] = False
                        st.success("Thank you. Your clarification has been submitted.")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Failed to submit clarification. Please try again.")
        st.stop()

    with st.sidebar:
        if not page_navigation_options:
            st.warning("You do not have access to any dashboards. Please contact your administrator.")
            st.stop()
        selected_page = st.radio("Main Navigation:", page_navigation_options, label_visibility="collapsed")

    st.markdown("<p style='text-align:center;color:#718096;font-size:1.2rem;margin-bottom:2rem;'>Upload expense reports for validation or manage data and users.</p>", unsafe_allow_html=True)
    
    start_date, end_date = None, None
    pages_with_filter = ["📊 Dashboard Analytics", "📈 Trends & History", "👤📊 User Performance", "📝 Correction Status", "📋 Exception Details", "📍 Location Expenses", "🧾 Ledger/Sub-Ledger Summary", "👤🧾 User-wise Ledger Exceptions", "👤📍 User & Location Analysis"]
    
    if selected_page in pages_with_filter:
        history_df = db_manager.get_validation_history(user_role, username, managed_users)
        start_date, end_date = add_date_filters_to_sidebar(history_df, key_suffix=selected_page.replace(" ", "_"))

    # --- MODIFIED --- Page routing now includes the new pages
    if selected_page == "🏠 Upload & Validate": show_upload_page()
    elif selected_page == "📊 Dashboard Analytics": show_analytics_page(start_date, end_date)
    elif selected_page == "📈 Trends & History": show_trends_page(start_date, end_date)
    elif selected_page == "👤📊 User Performance": show_user_performance_page(start_date, end_date)
    elif selected_page == "🗂️ Report Archive": show_report_archive_page()
    elif selected_page == "📝 Correction Status": show_correction_status_page(start_date, end_date)
    elif selected_page == "📢 Clarification Center": show_clarification_center_page()
    elif selected_page == "📋 Exception Details": show_exception_details_page(start_date, end_date)
    elif selected_page == "📍 Location Expenses": show_location_expenses_page(start_date, end_date)
    elif selected_page == "🧾 Ledger/Sub-Ledger Summary": show_ledger_summary_page(start_date, end_date)
    elif selected_page == "👤🧾 User-wise Ledger Exceptions": show_user_ledger_exceptions_page(start_date, end_date)
    elif selected_page == "👤📍 User & Location Analysis": show_user_location_page(start_date, end_date)
    elif selected_page == "🔑 Access Control": show_access_control_page()
    elif selected_page == "🛠️ User Management": show_user_management_page()
    elif selected_page == "🗑️ Data Management": show_data_management_page()
    elif selected_page == "⚙️ Settings": show_settings_page()
    elif selected_page == "🕵️ Suspicious Tag Control": show_suspicious_tag_control_page()
    elif selected_page == "📋 Suspicious Transactions": show_suspicious_category_transactions_page()
    elif selected_page == "📝 Correction Entries": show_correction_entries_page()
    elif selected_page == "📈 Correction Analytics": show_correction_analytics_page()
    elif selected_page == "✉️ Send Notification": show_send_notification_page()


if __name__ == "__main__":
    log_file_path = "dashboard.log"
    try:
        log_dir = os.path.dirname(os.path.abspath(log_file_path)) or '.'
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        if not os.access(log_dir, os.W_OK):
            raise IOError(f"Log file directory '{log_dir}' is not writable.")
        
        with open(log_file_path, "a") as f:
            f.write(f"--- Log session started at {datetime.now()} ---\n")
    except (IOError, OSError) as e:
        print(f"Warning: Log file '{log_file_path}' is not writable or cannot be created. Logging to console only. Error: {e}")
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[logging.StreamHandler()]
        )
    else:
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s [%(filename)s:%(lineno)d] - %(message)s',
            handlers=[
                logging.FileHandler(log_file_path, mode='a', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
    
    logging.info("Dashboard application started.")
    main()